from pathlib import Path
from openpyxl import load_workbook
files = sorted(Path('/app/output').glob('*_filled*.xlsx'), key=lambda p: p.stat().st_mtime)
wb = load_workbook(files[-1], data_only=False)
ws = wb.active
headers = [c.value for c in ws[1]]
want = {
    'Incident / Request / Change Created Date & Time',
    'Description',
    'Requester',
    'Consultant',
    'Actual Response Date & Time',
    'Actual Resolved Date & Time',
}
idxs = [(i+1,h) for i,h in enumerate(headers) if h in want]
print('HEADERS:', idxs)
for row in ws.iter_rows(min_row=2, values_only=False):
    vals = [cell.value for cell in row]
    if 'SF005-->RE: SF005 IDoc Failed at ES PROD --> 05-02-2026' in ''.join('' if v is None else str(v) for v in vals):
        print('ROW', row[0].row)
        for i,h in idxs:
            print(i, h, row[i-1].value)
        print('---')
