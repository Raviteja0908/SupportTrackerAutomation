from pathlib import Path
from openpyxl import load_workbook
files = sorted(Path('/app/output').glob('*_filled*.xlsx'), key=lambda p: p.stat().st_mtime)
print('FILE:', files[-1].name if files else 'NO_FILLED_XLSX')
if not files:
    raise SystemExit(1)
wb = load_workbook(files[-1], data_only=False)
ws = wb.active
for row in ws.iter_rows(values_only=True):
    vals = ['' if v is None else str(v) for v in row]
    line = ' | '.join(vals)
    if 'SF005-->RE: SF005 IDoc Failed at ES PROD --> 05-02-2026' in line:
        print(line)
