import argparse
import csv
import html
import json
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from email import policy
from email.parser import BytesParser
from email.utils import getaddresses, parsedate_to_datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from src.rules.subject_normalizer import extract_subject_from_description, normalize_subject
from src.rules.time_resolver import (
    _classify_reply_kind,
    _is_ess_sender,
    _match_requester,
    _to_ist,
)


@dataclass
class EmailRec:
    subject: str
    sender_name: str
    sender_email: str
    sent_time: datetime | None
    body: str
    body_html: str
    path: Path


def _load_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = []
        for idx, row in enumerate(csv.DictReader(handle), start=2):
            row["_line"] = idx
            rows.append(row)
        return rows


def _get(row: dict, *names: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        if name in row and row[name] not in (None, ""):
            return str(row[name]).strip()
        value = lowered.get(name.lower())
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _parse_eml(path: Path) -> EmailRec | None:
    try:
        with path.open("rb") as handle:
            msg = BytesParser(policy=policy.default).parse(handle)
        plain = ""
        html_body = ""
        if msg.is_multipart():
            for part in msg.walk():
                ctype = (part.get_content_type() or "").lower()
                try:
                    content = part.get_content()
                except Exception:
                    content = None
                if ctype == "text/plain" and isinstance(content, str) and not plain:
                    plain = content
                elif ctype == "text/html" and isinstance(content, str) and not html_body:
                    html_body = content
        else:
            try:
                content = msg.get_content()
            except Exception:
                content = ""
            if isinstance(content, str):
                plain = content
        sender_name = ""
        sender_email = ""
        try:
            addrs = getaddresses([msg.get("from", "")])
            if addrs:
                sender_name, sender_email = addrs[0]
        except Exception:
            pass
        sent_time = None
        try:
            sent_time = parsedate_to_datetime(msg.get("date", ""))
        except Exception:
            sent_time = None
        return EmailRec(
            subject=str(msg.get("subject", "") or ""),
            sender_name=sender_name or "",
            sender_email=(sender_email or "").lower(),
            sent_time=sent_time,
            body=plain or "",
            body_html=html_body or "",
            path=path,
        )
    except Exception:
        return None


def _fmt(dt: datetime | None) -> str:
    if not dt:
        return "-"
    try:
        dt = _to_ist(dt)
    except Exception:
        pass
    return dt.strftime("%d-%m-%Y %H:%M")


def _parse_cell_dt(value: str) -> datetime | None:
    if not value:
        return None
    for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except Exception:
            continue
    return None


def _clean_lines(email_obj: EmailRec) -> list[str]:
    raw = f"{email_obj.body}\n{email_obj.body_html}"
    if not raw:
        return []
    txt = re.sub(r"(?is)<style.*?>.*?</style>", " ", raw)
    txt = re.sub(r"(?is)<script.*?>.*?</script>", " ", txt)
    txt = re.sub(r"(?i)<\s*br\s*/?>", "\n", txt)
    txt = re.sub(r"(?i)</\s*(p|div|tr|td|th|li|h[1-6])\s*>", "\n", txt)
    txt = re.sub(r"(?is)<[^>]+>", " ", txt)
    txt = html.unescape(txt)
    return [ln.strip() for ln in txt.splitlines() if ln and ln.strip()]


def _normalize_quoted_sent_text(value: str) -> str:
    value = re.sub(r"(?i)^sent\s*:\s*", "", value or "").strip()
    value = re.sub(r"(?i)^(mon|tue|wed|thu|fri|sat|sun)\w*,?\s*", "", value).strip()
    value = re.sub(r"\(.*?\)", " ", value).strip()
    value = re.sub(r"(?i)(\d)(am|pm)\b", r"\1 \2", value)
    value = re.sub(r"(?i)\b(a\.m\.|p\.m\.)\b", lambda m: m.group(1).replace(".", "").upper(), value)
    return " ".join(value.replace(",", " ").replace(" at ", " ").split())


def _extract_primary_sent_fragments(*values: str) -> list[str]:
    regexes = [
        r"\b[A-Za-z]+,\s+\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
        r"\b[A-Za-z]+,\s+\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
        r"\b[A-Za-z]+,\s+[A-Za-z]+\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
        r"\b[A-Za-z]+,\s+[A-Za-z]+\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
        r"\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
        r"\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
        r"\b\d{1,2}[-./]\d{1,2}[-./]\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
        r"\b\d{1,2}[-./]\d{1,2}[-./]\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
    ]
    out = []
    seen = set()
    for value in values:
        text = (value or "").strip()
        if not text:
            continue
        matches = []
        for rx in regexes:
            for match in re.finditer(rx, text):
                frag = match.group(0).strip()
                key = frag.lower()
                if key in seen:
                    continue
                matches.append((match.start(), -len(frag), frag))
        matches.sort()
        for _, _, frag in matches:
            key = frag.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(frag)
        if out:
            break
    return out


def _parse_quoted_sent_time(sent_line: str) -> datetime | None:
    if not sent_line:
        return None
    normalized = _normalize_quoted_sent_text(sent_line)
    raw_clean = re.sub(r"(?i)^sent\s*:\s*", "", sent_line or "").strip()
    candidates = []
    for cand in _extract_primary_sent_fragments(raw_clean, normalized):
        if cand and cand not in candidates:
            candidates.append(cand)
    for cand in (normalized, raw_clean):
        if cand and cand not in candidates:
            candidates.append(cand)

    fmts_24h = [
        "%d %B %Y %H:%M:%S",
        "%d %B %Y %H:%M",
        "%d %b %Y %H:%M:%S",
        "%d %b %Y %H:%M",
        "%B %d %Y %H:%M:%S",
        "%B %d %Y %H:%M",
        "%b %d %Y %H:%M:%S",
        "%b %d %Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%A %d %B %Y %H:%M:%S",
        "%A %d %B %Y %H:%M",
        "%A %B %d %Y %H:%M:%S",
        "%A %B %d %Y %H:%M",
    ]
    fmts_12h = [
        "%d %B %Y %I:%M:%S %p",
        "%d %B %Y %I:%M %p",
        "%d %b %Y %I:%M:%S %p",
        "%d %b %Y %I:%M %p",
        "%B %d %Y %I:%M:%S %p",
        "%B %d %Y %I:%M %p",
        "%b %d %Y %I:%M:%S %p",
        "%b %d %Y %I:%M %p",
        "%d-%m-%Y %I:%M:%S %p",
        "%d-%m-%Y %I:%M %p",
        "%d.%m.%Y %I:%M:%S %p",
        "%d.%m.%Y %I:%M %p",
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y %I:%M %p",
        "%Y-%m-%d %I:%M:%S %p",
        "%Y-%m-%d %I:%M %p",
        "%A %d %B %Y %I:%M:%S %p",
        "%A %d %B %Y %I:%M %p",
        "%A %B %d %Y %I:%M:%S %p",
        "%A %B %d %Y %I:%M %p",
    ]
    for cand in candidates:
        has_am_pm = bool(re.search(r"(?i)\b(am|pm|a\.m\.|p\.m\.)\b", cand or ""))
        for fmt in (fmts_12h if has_am_pm else fmts_24h):
            try:
                return datetime.strptime(cand, fmt)
            except Exception:
                continue
        try:
            parsed = parsedate_to_datetime(cand)
            if parsed:
                return parsed
        except Exception:
            pass
    return None


def _extract_quoted_blocks(email_obj: EmailRec):
    lines = _clean_lines(email_obj)
    out = []

    def _label(line: str):
        m = re.match(r"(?i)^(from|sent|to|cc|subject|objet)\b\s*:?\s*(.*)$", line or "")
        return m.group(1).lower() if m else None

    def _value(line: str, label: str) -> str:
        m = re.match(rf"(?i)^{label}\b\s*:?\s*(.*)$", line or "")
        return (m.group(1) if m else "").strip()

    def _looks_like_body_line(line: str) -> bool:
        text = (line or "").strip()
        if not text:
            return False
        if _label(text):
            return False
        if re.match(r"(?i)^(hi|hello|dear|regards|kind regards|best regards|thanks|thank you|please)\b", text):
            return True
        if text.lower().startswith("@"):
            return True
        if len(text.split()) >= 7 and not re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, flags=re.I):
            return True
        return False

    def _looks_like_from_value(text: str) -> bool:
        text = (text or "").strip()
        if not text or _looks_like_body_line(text):
            return False
        if re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, flags=re.I):
            return True
        if re.search(r"<[^>]+>", text):
            return True
        parts = [p for p in re.split(r"[^A-Za-z0-9]+", text) if p]
        return 1 <= len(parts) <= 6

    def _append_meridiem(sent_line: str, start_idx: int) -> str:
        sent_low = (sent_line or "").lower()
        if (" am" in sent_low) or (" pm" in sent_low):
            return sent_line
        for j in range(start_idx + 1, min(start_idx + 4, len(lines))):
            extra = (lines[j] or "").strip()
            if re.fullmatch(r"(?i)(am|pm|a\.m\.|p\.m\.)", extra):
                return f"{sent_line} {extra}"
        return sent_line

    i = 0
    while i < len(lines):
        if not re.search(r"(?i)\bfrom\b\s*:", lines[i] or ""):
            i += 1
            continue
        from_line = ""
        sent_line = ""
        subj_line = ""
        end = i
        for j in range(i, min(i + 16, len(lines))):
            cur = (lines[j] or "").strip()
            if j > i and (re.search(r"(?i)\bfrom\b\s*:", cur or "") or re.match(r"(?i)^[-_]{3,}$", cur)):
                break
            label = _label(cur)
            if label == "from" and not from_line:
                from_line = _value(cur, "from")
                if (not from_line) and (j + 1) < len(lines):
                    next_line = (lines[j + 1] or "").strip()
                    if next_line and (_label(next_line) is None) and _looks_like_from_value(next_line):
                        from_line = next_line
                        end = j + 1
            elif label == "sent" and not sent_line:
                sent_line = cur
                if (not _value(cur, "sent")) and (j + 1) < len(lines):
                    next_line = (lines[j + 1] or "").strip()
                    if next_line and (_label(next_line) is None):
                        sent_line = f"{cur} {next_line}".strip()
                        end = j + 1
            elif label in {"subject", "objet"} and not subj_line:
                subj_line = cur
                if (not _value(cur, label)) and (j + 1) < len(lines):
                    next_line = (lines[j + 1] or "").strip()
                    if next_line and (_label(next_line) is None):
                        subj_line = f"{cur} {next_line}".strip()
                        end = j + 1
            elif j > i and label is None and _looks_like_body_line(cur):
                end = j - 1
                break
            elif j > i and label is None and sent_line and (subj_line or from_line):
                end = j - 1
                break
            end = j
        if sent_line:
            sent_line = _append_meridiem(sent_line, end)
            sent_dt = _parse_quoted_sent_time(sent_line)
            subj = re.sub(r"(?i)^(subject|objet)\b\s*:?\s*", "", subj_line).strip() if subj_line else ""
            out.append((from_line, sent_dt, subj, sent_line))
        i = max(i + 1, end + 1)
    return out


def _family_subject(row: dict) -> str:
    return normalize_subject(extract_subject_from_description(_get(row, "Description")))


def _same_subject(email_subject: str, family_subject: str) -> bool:
    e = normalize_subject(email_subject or "")
    f = normalize_subject(family_subject or "")
    return bool(e and f and (e == f or e in f or f in e))


def _match_row_debug(output_row: dict, debug_rows: list[dict]) -> dict | None:
    desc = _get(output_row, "Description")
    for row in debug_rows:
        if _get(row, "Description") == desc:
            return row
    return None


def _split_note_tags(notes: str) -> list[str]:
    return [part.strip() for part in str(notes or "").split(";") if part and part.strip()]


def _classify_note_tags(notes: str) -> dict:
    tags = _split_note_tags(notes)
    pre = []
    owner = []
    post = []
    for tag in tags:
        low = tag.lower()
        if any(
            marker in low
            for marker in (
                "bluecleared",
                "blueclearedstrict",
                "blueclearedessonly",
                "blueclearedfinalvalidator",
                "match=",
            )
        ):
            post.append(tag)
        elif any(
            marker in low
            for marker in (
                "lanelocalinitialepisode",
                "quotedrequestonlyreanchor",
                "quotedrequestonlydirectreply",
                "quotedrequestonlyhybridliveack",
                "esscontinuationguard[",
                "occurrencelocked",
                "requesterepisoderebaseguard",
                "latespisoderebaseguard",
                "episodespanrebaseguard",
                "episodconsistencyguard",
                "episodeackrefreshguard",
                "ackdelaywindowguard",
                "resolvednonackguard",
                "crosssubjectduplicateguard",
                "quotedrequestrebaseguard",
            )
        ):
            owner.append(tag)
        else:
            pre.append(tag)
    return {"pre": pre, "owner": owner, "post": post}


def _likely_final_owner(notes: str, created_src: str, ack_src: str, resolved_src: str) -> str:
    notes_l = (notes or "").lower()
    created_l = (created_src or "").lower()
    ack_l = (ack_src or "").lower()
    resolved_l = (resolved_src or "").lower()
    if "lanelocalinitialepisode[" in notes_l:
        return "seeded_lane"
    if "esscontinuationguard[" in notes_l:
        return "ess_continuation_guard"
    if "occurrencelocked" in notes_l:
        return "occurrence_locked_lane"
    if created_l == "created_clamped_to_first" or "created retained (response anchor unreliable)" in notes_l:
        return "base_created_retention"
    if created_l.startswith("parsed_from_quoted") or ack_l.startswith("parsed_from_quoted") or resolved_l.startswith("parsed_from_quoted"):
        return "quoted_reanchor_path"
    if any(x in notes_l for x in ("task", "systemnotification", "requester span(")):
        return "late_special_case_path"
    return "unknown"


def _guess_fill_name(rgb: str) -> str:
    rgb_u = (rgb or "").upper()
    if rgb_u == "00BDD7EE":
        return "blue"
    if rgb_u == "00FFF2CC":
        return "yellow"
    if rgb_u == "00FFC7CE":
        return "red"
    if rgb_u in {"", "00000000", "000000", "NONE"}:
        return "clear"
    return rgb_u or "unknown"


def _find_workbook_row(workbook_path: Path, description: str) -> dict | None:
    if load_workbook is None or not workbook_path.exists():
        return None
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = wb["LOG"] if "LOG" in wb.sheetnames else wb.active
        headers = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(1, col).value
            if value is not None:
                headers[str(value).strip().lower()] = col
        desc_col = headers.get("description")
        if not desc_col:
            return None
        for row_idx in range(2, ws.max_row + 1):
            desc = str(ws.cell(row_idx, desc_col).value or "").strip()
            if desc == description:
                fill_rgb = str(getattr(getattr(ws.cell(row_idx, 1).fill, "start_color", None), "rgb", "") or "")
                return {"row": row_idx, "fill_rgb": fill_rgb, "fill_name": _guess_fill_name(fill_rgb)}
    finally:
        wb.close()
    return None


def _quoted_sender_is_ess(from_line: str, ess_team: list[str]) -> bool | None:
    blob = (from_line or "").strip().lower()
    if not blob:
        return None
    temp = EmailRec(subject="", sender_name=from_line or "", sender_email="", sent_time=None, body="", body_html="", path=Path("."))
    if _is_ess_sender(temp, ess_team):
        return True
    email_match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", blob, flags=re.I)
    if email_match:
        return False
    return None


def _infer_lane_episode(reply_msg: EmailRec, reply_ist: datetime, subject_norm: str, thread: list[EmailRec], ess_team: list[str]):
    quoted_blocks = _extract_quoted_blocks(reply_msg)
    lane_blocks = []
    for idx, (from_line, sent_dt, q_subj, sent_line) in enumerate(quoted_blocks):
        if not sent_dt:
            continue
        sent_ist = _to_ist(sent_dt)
        if sent_ist >= reply_ist:
            continue
        if q_subj and not _same_subject(q_subj, subject_norm):
            continue
        lane_blocks.append((idx, from_line or "", sent_ist, q_subj or "", sent_line))
    ack_idx = None
    ack_ist = None
    ack_msg = None
    for idx, from_line, sent_ist, _q_subj, _sent_line in lane_blocks:
        if _quoted_sender_is_ess(from_line, ess_team) is not True:
            continue
        live_ack_matches = []
        for e in thread:
            e_ist = _to_ist(e.sent_time) if e.sent_time else None
            if not e_ist or e_ist >= reply_ist:
                continue
            if not _is_ess_sender(e, ess_team):
                continue
            if not _same_subject(e.subject, subject_norm):
                continue
            cls = _classify_reply_kind(e)
            is_ack_candidate = (
                (cls.get("explicit_ack") or cls.get("short_ess_ack") or cls.get("ack_like"))
                and not cls.get("real_reply")
                and not cls.get("direct_resolution")
            )
            if not is_ack_candidate:
                continue
            delta = abs((e_ist - sent_ist).total_seconds())
            if delta > 300:
                continue
            live_ack_matches.append((delta, e_ist, e))
        if not live_ack_matches:
            continue
        live_ack_matches.sort(key=lambda item: (item[0], item[1]))
        ack_idx = idx
        ack_ist = live_ack_matches[0][1]
        ack_msg = live_ack_matches[0][2]
        break

    req_ist = None
    if ack_idx is not None and ack_ist:
        for next_idx, next_from_line, next_sent_ist, _next_q_subj, _sent_line in lane_blocks:
            if next_idx <= ack_idx:
                continue
            if next_sent_ist >= ack_ist:
                continue
            if _quoted_sender_is_ess(next_from_line, ess_team) is not False:
                continue
            if (ack_ist - next_sent_ist) > timedelta(minutes=16):
                continue
            req_ist = next_sent_ist
            break
    if ack_ist and not req_ist:
        ack_ist = None
        ack_msg = None
    if not req_ist:
        for _idx, from_line, sent_ist, _q_subj, _sent_line in lane_blocks:
            if _quoted_sender_is_ess(from_line, ess_team) is False:
                req_ist = sent_ist
                break
    return {
        "request": req_ist,
        "ack": ack_ist or reply_ist,
        "resolved": reply_ist,
        "ack_msg": ack_msg,
        "quoted_blocks": lane_blocks,
        "mode": "req-ack-reply" if ack_msg is not None else "direct-reply",
    }


def _source_strength_value(src_value: str) -> int:
    src_l = (src_value or "").strip().lower()
    if not src_l:
        return 0
    if src_l.startswith("parsed_from_raw_eml"):
        return 20
    if src_l.startswith("parsed_from_quoted"):
        return 35
    if src_l.startswith("parsed_"):
        return 25
    return 70


def _rewrite_profile(created: datetime | None, response: datetime | None, resolved: datetime | None, notes: str, created_src: str, ack_src: str, resolved_src: str):
    c_ist = _to_ist(created) if created else None
    a_ist = _to_ist(response) if response else None
    r_ist = _to_ist(resolved) if resolved else None
    c_min = c_ist.replace(second=0, microsecond=0) if c_ist else None
    a_min = a_ist.replace(second=0, microsecond=0) if a_ist else None
    r_min = r_ist.replace(second=0, microsecond=0) if r_ist else None
    ordered = bool(c_ist and a_ist and r_ist and c_ist <= a_ist <= r_ist)
    ack_gap = (a_min - c_min) if (c_min and a_min and a_min >= c_min) else None
    all_same = bool(c_min and a_min and r_min and c_min == a_min == r_min)
    notes_l = (notes or "").lower()
    ack_src_l = (ack_src or "").lower()
    strong_live_ack = bool(
        ordered
        and ack_gap is not None
        and ack_gap <= timedelta(minutes=16)
        and (
            "quotedrequestonlypreservedliveack" in notes_l
            or (ack_src and not ack_src_l.startswith("parsed_from_"))
        )
    )
    source_strength = max(
        _source_strength_value(created_src),
        _source_strength_value(ack_src),
        _source_strength_value(resolved_src),
    )
    rewrite_strength = source_strength
    if ordered:
        rewrite_strength = max(rewrite_strength, 45)
    if all_same:
        rewrite_strength = min(rewrite_strength or 20, 30)
    if ack_gap is not None and ack_gap <= timedelta(minutes=16) and not all_same:
        rewrite_strength = max(rewrite_strength, 80 if strong_live_ack else 60)
    if "created retained (response anchor unreliable)" in notes_l:
        rewrite_strength = min(rewrite_strength, 15)
    if "dateanchorignoredstale" in notes_l:
        rewrite_strength = min(rewrite_strength, 20)
    if "quotedrequestonly" in notes_l and not strong_live_ack:
        rewrite_strength = min(rewrite_strength, 25)
    return {
        "ordered": ordered,
        "all_same": all_same,
        "ack_gap": ack_gap,
        "strong_live_ack": strong_live_ack,
        "rewrite_strength": rewrite_strength,
        "low_confidence": rewrite_strength <= 25,
        "c_min": c_min,
        "a_min": a_min,
        "r_min": r_min,
        "notes_l": notes_l,
    }


def _candidate_rewrite_strength(candidate_kind: str, owner_tag: str, cand_c_min, cand_a_min, cand_r_min) -> int:
    kind_key = (candidate_kind or "").strip().lower()
    owner_l = (owner_tag or "").lower()
    strength = {
        "occurrence_ess": 85,
        "system": 72,
        "task": 34,
        "sequence": 36,
        "hybrid": 58,
        "quoted": 50,
        "requester_ack": 42,
        "continuation": 32,
        "monotonic": 16,
        "risk": 20,
    }.get(kind_key, 48)
    if cand_c_min == cand_a_min == cand_r_min:
        strength -= 12
    elif cand_a_min == cand_r_min and cand_c_min < cand_a_min:
        strength += 8
    if "directreply" in owner_l or "reply-anchored" in owner_l:
        strength += 10
    if "systemnotification" in owner_l:
        strength += 6
    if "liverequestanchorguard" in owner_l:
        strength += 10
    return max(0, strength)


def _episode_selection_score(c_min, a_min, r_min, base_strength, notes_l, candidate_kind, owner_tag, live_reply_min):
    if not (c_min and a_min and r_min):
        return -9999
    score = max(0, int(base_strength or 0))
    if c_min < a_min < r_min:
        score += 10
    elif c_min < a_min == r_min:
        score += 7
    elif c_min == a_min == r_min:
        score -= 18
    ack_gap = (a_min - c_min) if a_min >= c_min else None
    if ack_gap is not None:
        if ack_gap <= timedelta(minutes=16):
            score += 12
        elif ack_gap <= timedelta(hours=1):
            score += 4
        elif ack_gap >= timedelta(hours=24):
            score -= 8
    owner_l = (owner_tag or "").lower()
    kind_l = (candidate_kind or "").lower()
    if "directreply" in owner_l or "reply-anchored" in owner_l:
        score += 10
    if kind_l in {"occurrence_ess", "system"}:
        score += 8
    if live_reply_min:
        if r_min == live_reply_min:
            score += 24
            if a_min <= live_reply_min:
                score += 4
        elif kind_l in {"quoted", "hybrid", "risk", "requester_ack"}:
            score -= 20
    if "created retained (response anchor unreliable)" in (notes_l or ""):
        score -= 16
    if "dateanchorignoredstale" in (notes_l or ""):
        score -= 8
    if "quotedrequestonly" in (notes_l or "") and kind_l not in {"quoted", "hybrid"}:
        score -= 6
    return score


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-csv", required=True)
    parser.add_argument("--debug-csv")
    parser.add_argument("--eml-dir", required=True)
    parser.add_argument("--subject", required=True)
    parser.add_argument("--workbook", help="Optional workbook path for fill-color trace")
    args = parser.parse_args()

    output_rows = _load_csv(Path(args.output_csv))
    debug_rows = _load_csv(Path(args.debug_csv)) if args.debug_csv else []
    matches = [r for r in output_rows if args.subject.lower() in _get(r, "Description").lower()]
    if not matches:
        print("No matching rows found.")
        return 1

    config_path = Path("/app/config/ess_team.json")
    if not config_path.exists():
        config_path = Path("config/ess_team.json")
    ess_team = json.loads(config_path.read_text(encoding="utf-8"))

    emails = []
    for path in sorted(Path(args.eml_dir).rglob("*.eml")):
        rec = _parse_eml(path)
        if rec:
            emails.append(rec)

    for row in matches:
        family_subject = _family_subject(row)
        description = _get(row, "Description")
        requester = _get(row, "Requested By", "Requester")
        created = _get(row, "Created Date & Time")
        response = _get(row, "Actual Response Date & Time")
        resolved = _get(row, "Actual Resolved Date & Time")
        debug_row = _match_row_debug(row, debug_rows)
        notes = _get(debug_row or {}, "Notes")
        created_src = _get(debug_row or {}, "CreatedSource")
        ack_src = _get(debug_row or {}, "AckSource")
        resolved_src = _get(debug_row or {}, "ResolvedSource")
        created_dt = _parse_cell_dt(created)
        response_dt = _parse_cell_dt(response)
        resolved_dt = _parse_cell_dt(resolved)
        resolved_min = _to_ist(resolved_dt).replace(second=0, microsecond=0) if resolved_dt else None
        note_groups = _classify_note_tags(notes)
        final_owner = _likely_final_owner(notes, created_src, ack_src, resolved_src)
        workbook_trace = _find_workbook_row(Path(args.workbook), description) if args.workbook else None

        candidates = [e for e in emails if _same_subject(e.subject, family_subject)]
        candidates.sort(key=lambda e: _to_ist(e.sent_time) if e.sent_time else datetime.min)
        live_replies = []
        for e in candidates:
            if not e.sent_time:
                continue
            if requester and not _match_requester(e.sender_name, e.sender_email, requester):
                continue
            live_replies.append(e)

        target_reply = None
        for e in live_replies:
            e_min = _to_ist(e.sent_time).replace(second=0, microsecond=0)
            if resolved_min and e_min == resolved_min:
                target_reply = e
                break
        if target_reply is None and live_replies:
            target_reply = live_replies[-1]

        print("=" * 100)
        print(f"line={row.get('_line')} desc={description}")
        print(f"row_triplet={created} / {response} / {resolved}")
        print(f"notes={notes or '-'}")
        print(f"sources={created_src or '-'} / {ack_src or '-'} / {resolved_src or '-'}")
        if workbook_trace:
            print(
                f"workbook_row={workbook_trace['row']} "
                f"fill={workbook_trace['fill_name']} ({workbook_trace['fill_rgb'] or '-'})"
            )
        print(f"likely_final_owner={final_owner}")
        print(f"pre_tags={note_groups['pre'] or ['-']}")
        print(f"owner_tags={note_groups['owner'] or ['-']}")
        print(f"post_tags={note_groups['post'] or ['-']}")
        print("-" * 100)
        print("live requester replies:")
        for e in live_replies:
            tag = ""
            e_min = _to_ist(e.sent_time).replace(second=0, microsecond=0)
            if resolved_min and e_min == resolved_min:
                tag = "  <-- output resolved"
            print(f"  {_fmt(e.sent_time)} | from={e.sender_email or e.sender_name} | subj={normalize_subject(e.subject)}{tag}")

        if not target_reply:
            print("-" * 100)
            print("diagnosis=no live final reply found")
            continue

        inferred = _infer_lane_episode(target_reply, _to_ist(target_reply.sent_time), family_subject, candidates, ess_team)
        print("-" * 100)
        print(f"target_reply={_fmt(target_reply.sent_time)} | {target_reply.sender_email or target_reply.sender_name}")
        print(f"inferred_mode={inferred['mode']}")
        print(
            "expected_seed_triplet="
            f"{_fmt(inferred['request'])} / {_fmt(inferred['ack'])} / {_fmt(inferred['resolved'])}"
        )
        profile = _rewrite_profile(created_dt, response_dt, resolved_dt, notes, created_src, ack_src, resolved_src)
        risky_markers = (
            "quotedrequestonly",
            "quotedpairgap>16m",
            "created retained (response anchor unreliable)",
            "dateanchorafter",
            "dateanchormissing",
            "dateanchorignoredstale",
            "norequesterthreadrecovery",
            "ambiguousresolvedbyrequester",
            "requester follow-up",
            "requester span(",
            "esscontinuationguard[",
            "ess-only; no non-ess request",
            "hybrid",
        )
        strict_all_live = (
            profile["ordered"]
            and not profile["low_confidence"]
            and all([created_src, ack_src, resolved_src])
            and not any(str(src).strip().lower().startswith("parsed_") for src in (created_src, ack_src, resolved_src))
            and not any(str(src).strip().lower() == "ack not found" for src in (created_src, ack_src, resolved_src))
            and resolved_min == _to_ist(target_reply.sent_time).replace(second=0, microsecond=0)
            and (not requester or _match_requester(resolved_src, resolved_src, requester))
            and not any(marker in (notes or "").lower() for marker in risky_markers)
        )
        seed_primary = False
        seed_reason = "no qualifying quoted block"
        for idx, (_block_idx, _from_line, sent_ist, q_subj, _sent_line) in enumerate(inferred["quoted_blocks"]):
            if not sent_ist or sent_ist >= _to_ist(target_reply.sent_time):
                continue
            if (_to_ist(target_reply.sent_time) - sent_ist) > timedelta(hours=48):
                continue
            if q_subj and not _same_subject(q_subj, family_subject):
                continue
            seed_primary = not strict_all_live
            seed_reason = f"block#{idx+1} qualifies"
            break
        live_reply_min = _to_ist(target_reply.sent_time).replace(second=0, microsecond=0)
        current_score = _episode_selection_score(
            profile["c_min"],
            profile["a_min"],
            profile["r_min"],
            profile["rewrite_strength"],
            profile["notes_l"],
            "current",
            "current",
            live_reply_min,
        )
        cand_c = inferred["request"].replace(second=0, microsecond=0) if inferred.get("request") else None
        cand_a = inferred["ack"].replace(second=0, microsecond=0) if inferred.get("ack") else None
        cand_r = inferred["resolved"].replace(second=0, microsecond=0) if inferred.get("resolved") else None
        candidate_strength = _candidate_rewrite_strength(
            "quoted" if inferred["mode"] == "direct-reply" else "hybrid",
            f"LaneLocalInitialEpisode[{inferred['mode']}]",
            cand_c,
            cand_a,
            cand_r,
        ) if cand_c and cand_a and cand_r else -9999
        candidate_score = _episode_selection_score(
            cand_c,
            cand_a,
            cand_r,
            candidate_strength,
            "",
            "quoted" if inferred["mode"] == "direct-reply" else "hybrid",
            f"LaneLocalInitialEpisode[{inferred['mode']}]",
            live_reply_min,
        ) if cand_c and cand_a and cand_r else -9999
        would_allow = False
        if cand_c and cand_a and cand_r and cand_c <= cand_a <= cand_r:
            if not profile["ordered"]:
                would_allow = True
            elif not profile["low_confidence"] and current_score >= (candidate_score + 10):
                would_allow = False
            elif profile["low_confidence"] and candidate_score >= (current_score + 6):
                would_allow = True
            else:
                would_allow = True
        print(f"strict_all_live={strict_all_live}")
        print(f"seed_primary_gate={seed_primary} ({seed_reason})")
        print(
            "rewrite_gate="
            f"{would_allow} "
            f"(current_strength={profile['rewrite_strength']}, current_score={current_score}, "
            f"candidate_strength={candidate_strength}, candidate_score={candidate_score}, "
            f"low_confidence={profile['low_confidence']})"
        )

        print("-" * 100)
        print("quoted blocks under target reply:")
        if not inferred["quoted_blocks"]:
            print("  no lane-matching quoted blocks found")
        else:
            for idx, from_line, sent_ist, q_subj, sent_line in inferred["quoted_blocks"]:
                who = "ESS" if _quoted_sender_is_ess(from_line, ess_team) is True else "NON-ESS"
                print(f"  {idx + 1}. {_fmt(sent_ist)} | {who} | from={from_line or '-'} | sent_line={sent_line}")
                print(f"     subj={q_subj or '-'}")

        created_dt = _parse_cell_dt(created)
        response_dt = _parse_cell_dt(response)
        resolved_dt = _parse_cell_dt(resolved)
        final_triplet = (
            _fmt(created_dt),
            _fmt(response_dt),
            _fmt(resolved_dt),
        )
        expected_triplet = (
            _fmt(inferred["request"]),
            _fmt(inferred["ack"]),
            _fmt(inferred["resolved"]),
        )

        print("-" * 100)
        print("lifecycle:")
        print(
            "  pre_state="
            f"{_fmt(created_dt)} / {_fmt(response_dt)} / {_fmt(resolved_dt)} "
            f"| low_confidence={profile['low_confidence']} all_same={profile['all_same']} "
            f"strong_live_ack={profile['strong_live_ack']}"
        )
        print(
            "  seed_candidate="
            f"{_fmt(inferred['request'])} / {_fmt(inferred['ack'])} / {_fmt(inferred['resolved'])} "
            f"| mode={inferred['mode']}"
        )
        print(
            "  final_owner_guess="
            f"{final_owner} | created_src={created_src or '-'} ack_src={ack_src or '-'} resolved_src={resolved_src or '-'}"
        )
        if workbook_trace:
            print(f"  post_fill={workbook_trace['fill_name']}")

        print("-" * 100)
        if final_triplet == expected_triplet:
            print("diagnosis=final row matches inferred reply-anchored lane")
        elif expected_triplet[0] != final_triplet[0] and expected_triplet[1:] == final_triplet[1:]:
            print("diagnosis=promotion_failed_created_only")
        elif expected_triplet[2] == final_triplet[2] and expected_triplet[0] != final_triplet[0]:
            print("diagnosis=mixed_lane_created_not_rebased")
        elif expected_triplet[2] != final_triplet[2]:
            print("diagnosis=final_reply_anchor_mismatch")
        else:
            print("diagnosis=seed_vs_final_mismatch")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
