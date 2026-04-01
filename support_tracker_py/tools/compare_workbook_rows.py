import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DISPLAY_COLUMNS = [
    "Description",
    "Requester",
    "Consultant",
    "Service No",
    "Created Date & Time",
    "Actual Response Date & Time",
    "Actual Resolved Date & Time",
    "ServiceRequest/Incident?",
    "ServiceRequest/Incident type?",
    "Interface Code",
]


def _fmt(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y %H:%M")
    s = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%d-%m-%Y %H:%M")
        except Exception:
            pass
    return s


def _header_map(ws):
    best = None
    best_count = -1
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        current = {}
        count = 0
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row_idx, col_idx).value
            if not v:
                continue
            name = str(v).strip()
            current[name] = col_idx
            if name in DISPLAY_COLUMNS:
                count += 1
        if count > best_count:
            best_count = count
            best = (row_idx, current)
    if not best:
        raise RuntimeError("Could not detect header row")
    return best


def _load_rows(path: Path, sheet_name: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
    except KeyError:
        wb.close()
        raise RuntimeError(f"Worksheet not found: {sheet_name}")

    header_row, headers = _header_map(ws)
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {"_row": row_idx}
        empty = True
        for name, col_idx in headers.items():
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                empty = False
            row[name] = value
        if not empty:
            rows.append(row)
    wb.close()
    return rows


def _matches(row, created_times, descriptions):
    created = _fmt(row.get("Created Date & Time"))
    desc = str(row.get("Description") or "")
    if created_times and created in created_times:
        return True
    if descriptions:
        desc_l = desc.lower()
        for needle in descriptions:
            if needle.lower() in desc_l:
                return True
    return False


def _display_row(row):
    out = [f"row={row.get('_row', '')}"]
    for name in DISPLAY_COLUMNS:
        if name in row:
            out.append(f"{name}={_fmt(row.get(name))}")
    return " | ".join(out)


def _compare_rows(left_rows, right_rows):
    left_map = {
        str(r.get("Description") or "").strip().lower(): r
        for r in left_rows
        if str(r.get("Description") or "").strip()
    }
    right_map = {
        str(r.get("Description") or "").strip().lower(): r
        for r in right_rows
        if str(r.get("Description") or "").strip()
    }
    keys = []
    seen = set()
    for d in list(left_map.keys()) + list(right_map.keys()):
        if d not in seen:
            seen.add(d)
            keys.append(d)

    for key in keys:
        left = left_map.get(key)
        right = right_map.get(key)
        print("=" * 100)
        print(key)
        if left:
            print("LEFT :", _display_row(left))
        else:
            print("LEFT : MISSING")
        if right:
            print("RIGHT:", _display_row(right))
        else:
            print("RIGHT: MISSING")


def main():
    parser = argparse.ArgumentParser(description="Compare Support Tracker workbook rows")
    parser.add_argument("--left", required=True, help="Left workbook path")
    parser.add_argument("--right", help="Right workbook path")
    parser.add_argument("--sheet", default="LOG", help="Worksheet name")
    parser.add_argument("--time", action="append", default=[], help="Created time to match, e.g. 26-02-2026 14:19")
    parser.add_argument("--desc", action="append", default=[], help="Description substring to match")
    args = parser.parse_args()

    left_rows = [r for r in _load_rows(Path(args.left), args.sheet) if _matches(r, set(args.time), args.desc)]
    if args.right:
        right_rows = [r for r in _load_rows(Path(args.right), args.sheet) if _matches(r, set(args.time), args.desc)]
        _compare_rows(left_rows, right_rows)
        return

    for row in left_rows:
        print(_display_row(row))


if __name__ == "__main__":
    main()
