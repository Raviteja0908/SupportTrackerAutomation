import os
from types import SimpleNamespace
import sys
import re
import html
import time
import math
from email import policy
from email.parser import BytesParser
from email.utils import parsedate_to_datetime
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import PatternFill

from src.output.run_logger import RunLogger
from src.pst_reader import read_pst_emails
from src.rules.subject_normalizer import normalize_subject, extract_subject_from_description, normalize_subject_for_match
from src.rules.environment import resolve_environment, resolve_environment_thread_fallback
from src.rules.interface import resolve_interface_code
from src.rules.service_request import resolve_service_request
from src.rules.incident_type import resolve_incident_type
from src.rules.time_resolver import (
    resolve_times_with_debug,
    _extract_canonical_message_lines,
    _extract_canonical_current_text,
    _extract_canonical_quoted_header_candidates,
    _match_requester,
    _is_ess_sender,
    _is_ack_like_reply,
    _classify_reply_kind as _shared_reply_classification,
    _is_thanks_info_reply,
    _email_has_explicit_ack_signal,
    _email_stable_key,
    _is_force_same_time_subject,
    _is_nonfinal_followup_reply as _shared_nonfinal_followup_reply,
    _is_real_reply_candidate as _shared_real_reply_candidate,
    _extract_request_time_from_email,
    _format_time,
    _to_ist,
    TimeResult,
    TimeDebug,
)
from src.excel.template_filler import fill_template, EXPECTED_HEADERS, select_target_sheet
from src.output.csv_writer import write_csv
from src.output.run_logger import MarkingReason
from src.utils import (
    load_json_list,
    load_aspose_license,
    load_subject_exclusions,
)


def _normalize_template_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (name or "").lower()).strip()


def _detect_workbook_kind(template_name: str) -> str:
    norm = _normalize_template_name(template_name)
    if "incident" in norm and "self" in norm and "service" in norm:
        return "incident_self_service"
    if "task" in norm and "business" in norm:
        return "task_business"
    if "incident" in norm and "business" in norm and "self" not in norm:
        return "incident_business"
    return "unknown"


def _workbook_label(kind: str) -> str:
    return {
        "incident_business": "Incident Business",
        "task_business": "Task Business",
        "incident_self_service": "Incident Self Service",
    }.get(kind, "Unknown")


def _template_output_stem(name: str) -> str:
    stem = Path(name or "").stem.lower()
    stem = re.sub(r"[^a-z0-9]+", "_", stem).strip("_")
    return stem or "automation_output"


def _resolve_self_service_category_type(category_type: str, subject: str) -> str:
    cat = (category_type or "").strip()
    if cat:
        return cat

    subj = (subject or "").lower()
    deployment_tokens = [
        "deployment",
        "deployed",
        "property config",
        "property configuration",
        "account rotation",
        "new us admin",
        "new admin",
        "wikis pending",
        "access",
    ]
    if any(tok in subj for tok in deployment_tokens):
        return "ES - Deployment/Property config"

    if "exception external" in subj:
        return "ES - exception external"
    if "exception internal" in subj:
        return "ES - exception internal"

    return "ES - exception internal"


_SOURCE_LOCKED_SAME_TIME_TOKENS = (
    "force prod subject; all times same",
    "latestconsultantreplyallsame",
    "maintenancepatchallthreesame",
    "selfserviceallthreesame",
    "systemnotificationdirectreplyallsame",
)


# DL addresses used when an ESS member intentionally reroutes their own
# follow-up/closure reply to the shared ESS distribution list instead of
# the original requester. When a reply matches this pattern, it must NOT
# be treated as an "ESS replied over ESS" same-time collapse signal for
# the Business Incident sheet.
_ESS_DL_ONLY_RECIPIENTS = (
    "enterprise-services-support@umusic.com",
    "enterprise-services-support@inveniolsi.com",
)


def _is_ess_dl_only_reroute(email_record, ess_team) -> bool:
    """
    Returns True when an email should be EXCLUDED from the ESS-over-ESS
    all-three-same collapse pool because the sender (an ESS member)
    intentionally rerouted the reply to only the shared ESS DL.

    The To-recipients of this email, once any address belonging to the
    sender themselves is excluded, must resolve to ONLY one of the known
    ESS DL addresses (no other recipient present). Cc is intentionally
    NOT checked, to avoid false positives from someone who simply cc'd
    the DL while still genuinely replying to the requester.

    This function is intentionally conservative: if recipient data is
    missing or ambiguous, it returns False (does NOT exclude), so
    existing behavior is preserved whenever there is any doubt.
    """
    if not email_record:
        return False

    to_recipients = getattr(email_record, "to_recipients", None)
    if to_recipients:
        sender_email = (getattr(email_record, "sender_email", "") or "").strip().lower()
        other_recipients = {
            addr for addr in to_recipients
            if addr and addr != sender_email
        }
        if other_recipients and other_recipients.issubset(set(_ESS_DL_ONLY_RECIPIENTS)):
            return True

    return False


def _has_source_locked_same_time(notes_text: str) -> bool:
    notes_l = (notes_text or "").lower()
    return any(token in notes_l for token in _SOURCE_LOCKED_SAME_TIME_TOKENS)


def _append_note_token(notes_text: str, token: str) -> str:
    parts = [part.strip() for part in (notes_text or "").split(";") if part.strip()]
    if token not in parts:
        parts.append(token)
    return "; ".join(parts)


def _is_same_time_category(category_type_value: str) -> bool:
    cat = (category_type_value or "").strip().lower()
    if not cat:
        return False
    if cat in {"maintenance", "patch"}:
        return True
    return cat in {"maintenance/patch", "es - maintenance/patch", "es-maintenance/patch"}


def _is_system_sender_from_headers(email_record) -> bool:
    sender_email = (getattr(email_record, "sender_email", "") or "").lower()
    sender_name = (getattr(email_record, "sender_name", "") or "").lower()
    tokens = (
        "eai-system-notification",
        "system-notification",
        "system notification",
        "no-reply",
        "noreply",
        "do-not-reply",
        "donotreply",
    )
    return any(token in sender_email or token in sender_name for token in tokens)


def _collapse_all_times_same(times: TimeResult, debug: TimeDebug, same_time: str, note_token: str, source: str = ""):
    if not same_time:
        return times, debug
    src = source or debug.resolved_src or debug.ack_src or debug.created_src
    return (
        TimeResult(same_time, same_time, same_time),
        TimeDebug(src, src, src, _append_note_token(debug.notes, note_token)),
    )


def _parse_output_time_value(value: str):
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y %I:%M %p"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _is_system_notification_direct_reply_case(thread, requester_name: str, ess_team, times: TimeResult) -> bool:
    if not thread or not requester_name or not times:
        return False

    response_dt = _parse_output_time_value(times.response)
    resolved_dt = _parse_output_time_value(times.resolved)
    if not response_dt or not resolved_dt:
        return False

    response_ist = _to_ist(response_dt)
    resolved_ist = _to_ist(resolved_dt)
    if response_ist.replace(second=0, microsecond=0) != resolved_ist.replace(second=0, microsecond=0):
        return False

    matched = []
    for e in thread:
        sent_time = getattr(e, "sent_time", None)
        if not sent_time:
            continue
        try:
            sent_ist = _to_ist(sent_time)
        except Exception:
            continue
        matched.append((sent_ist, e))
    if not matched:
        return False
    matched.sort(key=lambda item: item[0])

    system_msgs = [(sent_ist, e) for sent_ist, e in matched if _is_system_sender_from_headers(e) and sent_ist <= response_ist]
    if not system_msgs:
        return False
    latest_system_ist = system_msgs[-1][0]

    requester_after_system = [
        (sent_ist, e)
        for sent_ist, e in matched
        if sent_ist >= latest_system_ist
        and sent_ist <= response_ist
        and _match_requester(getattr(e, "sender_name", ""), getattr(e, "sender_email", ""), requester_name)
        and not _is_system_sender_from_headers(e)
    ]
    if not requester_after_system:
        return False
    if requester_after_system[-1][0].replace(second=0, microsecond=0) != response_ist.replace(second=0, microsecond=0):
        return False

    for sent_ist, e in matched:
        if sent_ist <= latest_system_ist or sent_ist >= response_ist:
            continue
        if _is_system_sender_from_headers(e):
            continue
        if _match_requester(getattr(e, "sender_name", ""), getattr(e, "sender_email", ""), requester_name):
            continue
        if _is_ess_sender(e, ess_team) and _is_ack_like_reply(e):
            return False
        return False

    for sent_ist, e in matched:
        if sent_ist <= latest_system_ist or sent_ist > response_ist:
            continue
        if _is_ess_sender(e, ess_team) and _is_ack_like_reply(e):
            return False

    return True


def _apply_same_time_normalization(
    *,
    thread,
    requester_name: str,
    ess_team,
    subject_norm: str,
    category_type_value: str,
    times: TimeResult,
    debug: TimeDebug,
):
    if not times or not debug:
        return times, debug

    if _is_force_same_time_subject(subject_norm, thread):
        dated_thread = [e for e in (thread or []) if getattr(e, "sent_time", None)]
        first = min(dated_thread, key=lambda e: _to_ist(e.sent_time)) if dated_thread else None
        if first and _format_time(first.sent_time):
            same_time = _format_time(first.sent_time)
            src = first.sender_email or first.sender_name or debug.created_src
            times, debug = _collapse_all_times_same(
                times,
                debug,
                same_time,
                "Force PROD subject; all times same",
                source=src,
            )

    if _is_same_time_category(category_type_value):
        same_time = times.resolved or times.response or times.created
        if same_time:
            times, debug = _collapse_all_times_same(
                times,
                debug,
                same_time,
                "MaintenancePatchAllThreeSame",
            )

    if _is_system_notification_direct_reply_case(thread, requester_name, ess_team, times):
        same_time = times.resolved or times.response or times.created
        if same_time:
            times, debug = _collapse_all_times_same(
                times,
                debug,
                same_time,
                "SystemNotificationDirectReplyAllSame",
            )

    return times, debug


def _should_latest_consultant_reply_collapse_all_same(
    notes_text: str,
    created_src: str,
    requester_name: str,
) -> bool:
    notes_l = (notes_text or "").lower()
    if "latestconsultantreplyallsame" in notes_l:
        return True
    if "latestconsultantreply" in notes_l:
        return False

    same_time_hint = any(
        token in notes_l
        for token in (
            "ess initiated; no ack; no consultant reply after request",
            "failed subject; ess initiated; no ack phrase",
            "ess-only; no non-ess request",
        )
    )
    if not same_time_hint:
        return False

    created_src = (created_src or "").strip()
    if not created_src:
        return False
    if created_src.upper().startswith("PARSED_FROM_"):
        return False

    if requester_name and created_src and _match_requester(created_src, created_src, requester_name):
        return False

    return True


def _row_has_force_same_time_lock(state, notes_text: str = "") -> bool:
    if _has_source_locked_same_time(notes_text):
        return True
    subject_norm = state.get("subject_norm") if state else ""
    thread = state.get("thread") if state else []
    return _is_force_same_time_subject(subject_norm, thread)


def _is_shared_real_reply_candidate(email_record) -> bool:
    return _shared_reply_flags(email_record=email_record)["substantive_reply"]


def _shared_reply_flags(*, email_record=None, cls=None) -> dict:
    if cls is None:
        cls = _shared_reply_classification(email_record)
    direct_resolution = bool(cls.get("direct_resolution"))
    real_reply = bool(cls.get("real_reply"))
    # direct_resolution is a strong explicit helper, not a requirement; the
    # broader substantive bucket is real_reply, which already includes direct
    # resolution when present.
    thanks_info = bool(cls.get("thanks_info"))
    nonfinal_followup = bool(cls.get("nonfinal_followup"))
    explicit_ack = bool(cls.get("explicit_ack"))
    short_ess_ack = bool(cls.get("short_ess_ack"))
    ack_like = bool(cls.get("ack_like"))
    ignore_reply = thanks_info or nonfinal_followup
    ack_candidate = bool(
        (explicit_ack or short_ess_ack or ack_like)
        and not real_reply
        and not direct_resolution
        and not ignore_reply
    )
    return {
        "direct_resolution": direct_resolution,
        "real_reply": real_reply,
        "substantive_reply": real_reply,
        "thanks_info": thanks_info,
        "nonfinal_followup": nonfinal_followup,
        "ignore_reply": ignore_reply,
        "explicit_ack": explicit_ack,
        "short_ess_ack": short_ess_ack,
        "ack_like": ack_like,
        "ack_candidate": ack_candidate,
        "kind": cls.get("kind") or "",
    }


def _is_shared_ignore_reply(email_record) -> bool:
    return _shared_reply_flags(email_record=email_record)["ignore_reply"]


def _is_shared_ack_candidate(email_record) -> bool:
    return _shared_reply_flags(email_record=email_record)["ack_candidate"]


def _is_shared_ess_ack_candidate(email_record) -> bool:
    return _is_shared_ack_candidate(email_record)


def _shared_resolution_candidates(candidates):
    ordered = [
        e for e in (candidates or [])
        if e is not None and getattr(e, "sent_time", None) is not None
    ]
    preferred = [e for e in ordered if _is_shared_real_reply_candidate(e)]
    return preferred


def _seed_clean_message_lines(email_record) -> list[str]:
    return _extract_canonical_message_lines(email_record)


def _seed_current_html_text(email_record, max_lines: int = 48) -> str:
    return _extract_canonical_current_text(email_record, max_lines=max_lines)


def _seed_quoted_sender_is_ess(from_line: str, ess_team) -> bool | None:
    blob = (from_line or "").strip()
    if not blob:
        return None
    temp = SimpleNamespace(sender_name=blob, sender_email="", body="", body_html="")
    if _is_ess_sender(temp, ess_team):
        return True
    if re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", blob, flags=re.I):
        return False
    return None


def _seed_quoted_sender_matches_live(from_line: str, email_obj) -> bool:
    sender_email = (getattr(email_obj, "sender_email", "") or "").strip().lower()
    sender_name = (getattr(email_obj, "sender_name", "") or "").strip().lower()
    blob = (from_line or "").strip().lower()
    if sender_email and sender_email in blob:
        return True
    if not sender_name:
        return False
    stop = {"ess", "enterprise", "services", "support", "team", "umg", "music"}
    quoted_tokens = {tok for tok in re.split(r"[^a-z0-9]+", blob) if len(tok) >= 3 and tok not in stop}
    sender_tokens = {tok for tok in re.split(r"[^a-z0-9]+", sender_name) if len(tok) >= 3 and tok not in stop}
    overlap = quoted_tokens & sender_tokens
    return len(overlap) >= 2 or any(len(tok) >= 6 for tok in overlap)

def _post_ack_reply_rank(email_record, requester_name: str = ""):
    flags = _shared_reply_flags(email_record=email_record)
    requester_match = bool(
        requester_name
        and _match_requester(
            getattr(email_record, "sender_name", ""),
            getattr(email_record, "sender_email", ""),
            requester_name,
        )
    )
    return (
        0 if requester_match else 1,
        0 if flags["direct_resolution"] else 1,
        0 if flags["substantive_reply"] else 1,
        1 if flags["ack_candidate"] else 0,
        _to_ist(getattr(email_record, "sent_time", None)) or datetime.max,
    )


def _remove_note_tokens(notes_text: str, tokens_to_remove) -> str:
    parts = [part.strip() for part in (notes_text or "").split(";")]
    keep = []
    remove_set = {tok.strip().lower() for tok in (tokens_to_remove or []) if tok}
    for part in parts:
        if not part:
            continue
        if part.lower() in remove_set:
            continue
        keep.append(part)
    return "; ".join(keep)


def main() -> int:
    input_dir = Path(os.getenv("INPUT_DIR", "/app/input"))
    output_dir = Path(os.getenv("OUTPUT_DIR", "/app/output"))
    config_dir = Path(os.getenv("CONFIG_DIR", "/app/config"))

    output_dir.mkdir(parents=True, exist_ok=True)
    logger = RunLogger(output_dir / "processing.log")

    load_aspose_license(logger)

    pst_files = list(input_dir.glob("*.pst"))
    if not pst_files:
        logger.log("[ERROR] No PST files found.")
        return 1

    ess_team = load_json_list(config_dir / "ess_team.json")
    subject_exclusions = load_subject_exclusions(config_dir / "subject_exclusions.json")

    logger.log(f"[INFO] Found {len(pst_files)} PST file(s).")

    eml_root = output_dir / "eml"
    eml_root.mkdir(parents=True, exist_ok=True)

    emails = []
    for pst_path in pst_files:
        logger.log(f"[INFO] Reading PST: {pst_path}")
        try:
            emails.extend(read_pst_emails(pst_path, logger, eml_root))
        except Exception as exc:
            logger.log(f"[ERROR] PST extraction failed for {pst_path}: {exc}")
            return 1

    logger.log(f"[INFO] Total emails extracted: {len(emails)}")
    if not emails:
        logger.log("[ERROR] No emails extracted from PST files; stopping before workbook fill.")
        return 1

    threads = {}
    alt_index = {}
    inc_index = {}
    for email in emails:
        subject_norm = normalize_subject(email.subject)
        if not subject_norm:
            continue
        if subject_exclusions and any(x in subject_norm.lower() for x in subject_exclusions):
            # Keep maintenance subjects in threads so they can be filled.
            if "maintenance" not in subject_norm.lower():
                continue
        threads.setdefault(subject_norm, []).append(email)
    for key in threads.keys():
        alt_key = normalize_subject_for_match(key)
        if alt_key and alt_key != key:
            alt_index.setdefault(alt_key, []).append(key)

    # Build INC index from subject + body (including HTML) for safe fallback
    for key, thread in threads.items():
        incs = set()
        for e in thread:
            text = f"{e.subject or ''}\n{e.body or ''}\n{getattr(e, 'body_html', '') or ''}"
            for m in re.findall(r"\bINC\d{6,}\b", text, flags=re.IGNORECASE):
                incs.add(m.upper())
        for inc in incs:
            inc_index.setdefault(inc, []).append(key)

    logger.log(f"[INFO] Total threads: {len(threads)}")
    write_csv(
        output_dir / "thread_keys.csv",
        [{"SubjectKey": k} for k in sorted(threads.keys())],
        ["SubjectKey"],
    )

    # Load Excel template from output folder.
    # Priority: explicit TEMPLATE_PATH env var -> newest non-filled .xlsx in output_dir.
    template_path_env = os.getenv("TEMPLATE_PATH", "").strip()
    if template_path_env:
        template_path = Path(template_path_env)
        if not template_path.exists():
            logger.log(f"[ERROR] TEMPLATE_PATH not found: {template_path}")
            return 1
        logger.log(f"[INFO] Using template (TEMPLATE_PATH): {template_path.name}")
    else:
        candidates = [
            p for p in output_dir.glob("*.xlsx")
            if not p.name.lower().endswith("_filled.xlsx")
            and "_filled_" not in p.name.lower()
        ]
        if not candidates:
            logger.log("[ERROR] No template .xlsx found in output folder.")
            return 1
        # Pick most recently modified template by default
        template_path = max(candidates, key=lambda p: p.stat().st_mtime)
        logger.log(f"[INFO] Using template: {template_path.name}")

    workbook_kind = _detect_workbook_kind(template_path.name)
    if workbook_kind == "unknown":
        logger.log(f"[ERROR] Unrecognized workbook type from filename: {template_path.name}")
        return 1
    logger.log(f"[INFO] Workbook type: {_workbook_label(workbook_kind)}")

    def _normalize_header(text: str) -> str:
        if text is None:
            return ""
        return " ".join(str(text).replace("\n", " ").split()).strip().lower()

    def _find_header_row(ws):
        expected = {_normalize_header(h) for h in EXPECTED_HEADERS}
        best_row = 1
        best_hits = 0
        for row in range(1, min(15, ws.max_row) + 1):
            row_values = [
                _normalize_header(ws.cell(row, c).value)
                for c in range(1, ws.max_column + 1)
            ]
            hits = sum(1 for v in row_values if v in expected and v)
            if hits > best_hits:
                best_hits = hits
                best_row = row
        return best_row

    def _build_col_map(ws, header_row):
        col_map = {}
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(header_row, col).value
            if raw is None:
                continue
            key = _normalize_header(raw)
            if key and key not in col_map:
                col_map[key] = col
        return col_map

    # Prepare CSV outputs for auditing
    automation_rows = []
    debug_rows = []
    same_time_rows = []
    row_states = []
    _match_tokens_cache = {}
    _id_like_tokens_cache = {}
    _interface_tokens_cache = {}
    _inc_tokens_cache = {}
    _sig_num_tokens_cache = {}
    _part_tokens_cache = {}
    _subject_for_match_cache = {}
    _subject_family_similarity_cache = {}
    _same_subject_family_cache = {}
    _strong_identity_overlap_cache = {}
    _fresh_picker_subject_safe_cache = {}
    _strict_subject_identity_gate_cache = {}
    _subject_date_value_cache = {}
    _soft_subject_similarity_cache = {}
    _subject_fingerprint_cache = {}
    _subject_token_idf_cache = {}
    _subject_token_df_cache = {"df": None}
    _subject_fingerprint_norm_cache = {}
    _thread_requester_presence_cache = {}
    _thread_consultant_date_cache = {}
    _thread_date_token_presence_cache = {}
    _thread_non_ess_cache = {}
    _thread_baseline_delta_cache = {}
    _family_subject_score_cache = {}
    _distinctive_subject_tokens_cache = {}
    _quoted_subject_variant_cache = {}
    _subject_short_variant_tokens_cache = {}
    _raw_match_token_cache = {}

    def _match_tokens(text: str):
        key = text or ""
        cached = _match_tokens_cache.get(key)
        if cached is not None:
            return cached
        if not text:
            out = set()
        else:
            t = re.sub(r"[^a-z0-9]+", " ", text.lower())
            out = {p for p in t.split() if p}
        _match_tokens_cache[key] = out
        return out

    def _id_like_tokens(text: str) -> set:
        _ck = text or ""
        if _ck in _id_like_tokens_cache:
            return _id_like_tokens_cache[_ck]
        if not text:
            tokens = set()
            _id_like_tokens_cache[_ck] = tokens
            return tokens
        text = re.sub(r"[Ã¢â‚¬ÂÃ¢â‚¬â€˜Ã¢â‚¬â€™Ã¢â‚¬â€œÃ¢â‚¬â€Ã¢â‚¬â€¢Ã¢Ë†â€™Ã¯Â¹Â£Ã¯Â¼Â\u00ad]", "-", text)
        text = re.sub(r"[\u200b\u200c\u200d\u2060\ufeff]", "", text)
        tokens = {
            m.group(0).lower()
            for m in re.finditer(r"(?<![a-z0-9])[a-z]{2,}[a-z0-9\-]*\d[a-z0-9\-]*(?![a-z0-9])", text.lower())
        }
        if not tokens:
            for part in re.split(r"[^a-z0-9\-]+", text.lower()):
                if not part:
                    continue
                if any(c.isalpha() for c in part) and any(c.isdigit() for c in part):
                    tokens.add(part)
        _id_like_tokens_cache[_ck] = tokens
        return tokens

    # Pre-warm the EML header summary cache from already-loaded EmailRecord objects.
    # Prevents _get_eml_header_summary() from re-opening every .eml file from disk.
    _eml_header_prewarm = {}
    for _pw_e in emails:
        _pw_path = str(getattr(_pw_e, "path", "") or "")
        if not _pw_path:
            continue
        _pw_subject_raw = getattr(_pw_e, "subject", "") or ""
        _pw_subject_norm = normalize_subject(_pw_subject_raw)
        _eml_header_prewarm[_pw_path] = {
            "sent_dt": getattr(_pw_e, "sent_time", None),
            "subject_raw": _pw_subject_raw,
            "subject_norm": _pw_subject_norm,
            "subject_ids": _id_like_tokens(_pw_subject_norm),
            "sender_email": (getattr(_pw_e, "sender_email", "") or "").lower(),
            "sender_name": getattr(_pw_e, "sender_name", "") or "",
        }
    logger.log(f"[INFO] EML header cache pre-warmed: {len(_eml_header_prewarm)} entries")

    def _task_subject_core(text: str) -> str:
        norm = normalize_subject(text or "")
        if not norm:
            return ""
        core = norm
        changed = True
        while changed:
            changed = False
            for prefix in ("ess - ", "re: ", "fw: ", "fwd: "):
                if core.startswith(prefix):
                    core = core[len(prefix):].strip()
                    changed = True
        core = re.sub(r"\s*-\s*task\d+\b", "", core, flags=re.IGNORECASE).strip()
        core = re.sub(r"\s*(?:--?>|=>)\s*\d{1,2}[-/.]\d{1,2}(?:[-/.]\d{2,4})?\s*$", "", core, flags=re.IGNORECASE).strip()
        core = re.sub(r"\s*(?:--?>|=>)\s*\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\s*$", "", core, flags=re.IGNORECASE).strip()
        core = re.sub(r"\s+", " ", core).strip(" -:")
        return core.lower()

    def _part_tokens(text: str):
        key = text or ""
        cached = _part_tokens_cache.get(key)
        if cached is not None:
            return cached
        if not text:
            tokens = set()
        else:
            t = text.lower()
            tokens = set()
            for m in re.findall(r"\bpt\s*\d+\b", t):
                tokens.add(m.replace(" ", ""))
            for m in re.findall(r"\bpart\s*\d+\b", t):
                tokens.add(m.replace(" ", ""))
            for m in re.findall(r"\bpt\d+\b", t):
                tokens.add(m)
        _part_tokens_cache[key] = tokens
        return tokens

    def _extract_dr_ids(text: str):
        if not text:
            return set()
        ids = set()
        for m in re.findall(r"\bDR\s*ID\s*[:#]?\s*(DR-\d{6,}-\d+)\b", text, flags=re.IGNORECASE):
            ids.add(m.upper())
        for m in re.findall(r"\b(DR-\d{6,}-\d+)\b", text, flags=re.IGNORECASE):
            ids.add(m.upper())
        return ids

    def _thread_dr_ids(thread):
        ids = set()
        for e in thread:
            ids |= _extract_dr_ids(e.subject or "")
            ids |= _extract_dr_ids(e.body or "")
            ids |= _extract_dr_ids(getattr(e, "body_html", "") or "")
        return ids

    def _is_deployment_request_subject(text: str) -> bool:
        if not text:
            return False
        s = text.lower()
        return "deployment request" in s

    def _is_deployment_success_subject(text: str) -> bool:
        if not text:
            return False
        s = text.lower()
        return "success:" in s and "deployment" in s

    def _iface_tokens(text: str) -> set:
        if not text:
            return set()
        tokens = re.findall(r"\b[a-z]{1,5}\d{2,}\b", text, flags=re.IGNORECASE)
        return {t.lower() for t in tokens}

    ARROW_SPLIT_RE = r"\s*(?:--?>|->|→|➔|➡|=>)\s*"

    def _looks_like_date_only(text: str) -> bool:
        if not text:
            return True
        s = text.strip()
        date_patterns = [
            r"^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}$",
            r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$",
            r"^\d{1,2}[-/.]\d{1,2}$",
        ]
        return any(re.match(p, s) for p in date_patterns)

    def _extract_date_tokens(text: str) -> list[str]:
        if not text:
            return []
        tokens = []
        tokens += re.findall(r"\b\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}\b", text)
        tokens += re.findall(r"\b\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\b", text)
        # de-dup, keep order
        seen = set()
        out = []
        for t in tokens:
            if t not in seen:
                seen.add(t)
                out.append(t)
        return out

    def _extract_date_tokens_from_description(description: str) -> list[str]:
        if not description:
            return []
        if not re.search(ARROW_SPLIT_RE, description):
            return []
        parts = re.split(ARROW_SPLIT_RE, description)
        if not parts:
            return []
        tail = parts[-1].strip()
        if not tail:
            return []
        return _extract_date_tokens(tail)

    def _has_explicit_date_marker(text: str) -> bool:
        if not text:
            return False
        if not re.search(ARROW_SPLIT_RE, text):
            return False
        parts = re.split(ARROW_SPLIT_RE, text)
        if not parts:
            return False
        tail = parts[-1].strip()
        if not tail:
            return False
        return bool(_extract_date_tokens(tail))

    def _thread_has_date_token(thread, date_tokens: list[str]) -> bool:
        if not date_tokens or not thread:
            return False
        parsed_dates = []
        for t in date_tokens:
            dt = _parse_date_token(t)
            if dt:
                parsed_dates.append(dt)
        for e in thread:
            if parsed_dates:
                sent_d = None
                try:
                    sent_d = _to_ist(e.sent_time).date()
                except Exception:
                    try:
                        sent_d = e.sent_time.date()
                    except Exception:
                        sent_d = None
                if sent_d and sent_d in parsed_dates:
                    return True
            text = f"{e.subject or ''}\n{e.body or ''}\n{getattr(e, 'body_html', '') or ''}"
            for t in date_tokens:
                if t in text:
                    return True
        return False

    def _interface_prefix(text: str) -> str:
        if not text:
            return ""
        # Match leading interface-like tokens (e.g., CS001, ID082, VMI001)
        m = re.match(r"^([a-z]{2,}\d{2,})", text.strip(), flags=re.IGNORECASE)
        return (m.group(1).lower() if m else "")

    def _interface_tokens(text: str) -> set:
        key = text or ""
        cached = _interface_tokens_cache.get(key)
        if cached is not None:
            return cached
        if not text:
            out = set()
        else:
            tokens = re.findall(r"\b[a-z]{1,5}\d{2,}\b", text, flags=re.IGNORECASE)
            out = {t.lower() for t in tokens}
        _interface_tokens_cache[key] = out
        return out

    def _inc_tokens(text: str) -> set:
        key = text or ""
        cached = _inc_tokens_cache.get(key)
        if cached is not None:
            return cached
        if not text:
            out = set()
        else:
            tokens = re.findall(r"\binc\d{6,}\b", text, flags=re.IGNORECASE)
            out = {t.lower() for t in tokens}
        _inc_tokens_cache[key] = out
        return out

    def _sig_num_tokens(text: str) -> set:
        """Return significant numeric tokens for disambiguation (exclude years)."""
        key = text or ""
        cached = _sig_num_tokens_cache.get(key)
        if cached is not None:
            return cached
        if not text:
            out = set()
        else:
            out = set()
            for n in re.findall(r"\b\d{3,5}\b", text):
                try:
                    iv = int(n)
                except Exception:
                    continue
                if 1900 <= iv <= 2099:
                    continue
                out.add(n)
        _sig_num_tokens_cache[key] = out
        return out

    _subject_variant_stopwords = {
        "re", "fw", "fwd", "the", "and", "for", "from", "with", "without",
        "into", "onto", "this", "that", "these", "those", "file", "files",
        "testfile", "failed", "failure", "process", "processing", "data",
        "check", "month", "prod", "uat", "fct", "sap", "es", "at", "on",
        "to", "in", "of", "by", "due", "missing", "skipped",
    }

    _subject_short_variant_stopwords = {
        "re", "fw", "fwd", "aw", "wg", "sv",
        "in", "on", "at", "to", "of", "for", "not",
        "and", "or", "the", "a", "an", "is", "it",
        "by", "as", "if", "no", "we", "us",
        "es", "api", "sap", "uat", "fct",
    }

    def _distinctive_subject_tokens(text: str) -> set:
        key = text or ""
        cached = _distinctive_subject_tokens_cache.get(key)
        if cached is not None:
            return cached
        norm = _normalize_subject_for_match_cached(text or "")
        tokens = {
            tok for tok in _match_tokens(norm)
            if len(tok) >= 4 and tok not in _subject_variant_stopwords
        }
        tokens -= _interface_tokens(norm)
        tokens -= _inc_tokens(norm)
        tokens -= _part_tokens(norm)
        tokens -= _sig_num_tokens(norm)
        tokens = {tok for tok in tokens if not tok.isdigit()}
        _distinctive_subject_tokens_cache[key] = tokens
        return tokens

    def _subject_short_variant_tokens(text: str) -> set:
        key = text or ""
        cached = _subject_short_variant_tokens_cache.get(key)
        if cached is not None:
            return cached
        norm = normalize_subject(key)
        if not norm:
            out = set()
        else:
            out = {
                tok
                for tok in _match_tokens(norm)
                if len(tok) <= 3 and tok.isalpha() and tok not in _subject_short_variant_stopwords
            }
        _subject_short_variant_tokens_cache[key] = out
        return out

    def _quoted_subject_short_variant_ok(row_subject: str, candidate_subject: str) -> bool:
        row_short = _subject_short_variant_tokens(row_subject or "")
        if not row_short:
            return True
        cand_short = _subject_short_variant_tokens(candidate_subject or "")
        if not cand_short:
            return False
        return not row_short.isdisjoint(cand_short)

    def _quoted_subject_variant_owns_row(row_subject: str, candidate_subject: str) -> bool:
        cache_key = (row_subject or "", candidate_subject or "")
        cached = _quoted_subject_variant_cache.get(cache_key)
        if cached is not None:
            return cached
        row_norm = normalize_subject(row_subject or "")
        cand_norm = normalize_subject(candidate_subject or "")
        if not row_norm or not cand_norm:
            out = False
        elif row_norm == cand_norm:
            out = True
        else:
            row_distinctive = _distinctive_subject_tokens(row_norm)
            cand_distinctive = _distinctive_subject_tokens(cand_norm)
            row_only = row_distinctive - cand_distinctive
            cand_only = cand_distinctive - row_distinctive
            if row_only and cand_only:
                out = False
            else:
                out = True
        _quoted_subject_variant_cache[cache_key] = out
        _quoted_subject_variant_cache[(cache_key[1], cache_key[0])] = out
        return out

    def _normalize_subject_for_match_cached(text: str) -> str:
        key = text or ""
        cached = _subject_for_match_cache.get(key)
        if cached is not None:
            return cached
        out = normalize_subject_for_match(key)
        _subject_for_match_cache[key] = out
        return out

    def _subject_family_similarity(a: str, b: str) -> float:
        cache_key = (a or "", b or "")
        cached = _subject_family_similarity_cache.get(cache_key)
        if cached is not None:
            return cached
        a_norm = _normalize_subject_for_match_cached(a or "")
        b_norm = _normalize_subject_for_match_cached(b or "")
        a_tokens = _match_tokens(a_norm)
        b_tokens = _match_tokens(b_norm)
        if not a_tokens or not b_tokens:
            out = 0.0
        else:
            out = _token_overlap_score(a_tokens, b_tokens)
        _subject_family_similarity_cache[cache_key] = out
        _subject_family_similarity_cache[(cache_key[1], cache_key[0])] = out
        return out

    def _subject_fingerprint_tokens(text: str) -> set[str]:
        key = text or ""
        cached = _subject_fingerprint_cache.get(key)
        if cached is not None:
            return cached
        norm = normalize_subject(key)
        if not norm:
            out = set()
        else:
            out = set()
            out |= {f"inc:{tok}" for tok in _inc_tokens(norm)}
            out |= {f"iface:{tok}" for tok in _interface_tokens(norm)}
            out |= {f"id:{tok}" for tok in _id_like_tokens(norm)}
            out |= {f"num:{tok}" for tok in _sig_num_tokens(norm)}
            out |= {f"part:{tok}" for tok in _part_tokens(norm)}
            out |= {f"date:{tok}" for tok in _subject_date_value_set(norm)}
            out |= {f"tok:{tok}" for tok in _distinctive_subject_tokens(norm)}
            if not out:
                out = {f"tok:{tok}" for tok in _match_tokens(_normalize_subject_for_match_cached(norm)) if len(tok) >= 4}
        _subject_fingerprint_cache[key] = out
        return out

    def _subject_token_idf(token: str) -> float:
        cached = _subject_token_idf_cache.get(token)
        if cached is not None:
            return cached
        df_cache = _subject_token_df_cache.get("df")
        if df_cache is None:
            df = {}
            try:
                subject_keys = list(threads.keys())
            except Exception:
                subject_keys = []
            for subject_key in subject_keys:
                for tok in _subject_fingerprint_tokens(subject_key):
                    df[tok] = df.get(tok, 0) + 1
            _subject_token_df_cache["df"] = df
            df_cache = df
        try:
            corpus_size = max(len(threads), 1)
        except Exception:
            corpus_size = 1
        df_val = (df_cache or {}).get(token, 0)
        weight = 1.0 + math.log((corpus_size + 1.0) / (df_val + 1.0))
        _subject_token_idf_cache[token] = weight
        return weight

    def _subject_fingerprint_norm(tokens: set[str]) -> float:
        if not tokens:
            return 0.0
        key = tuple(sorted(tokens))
        cached = _subject_fingerprint_norm_cache.get(key)
        if cached is not None:
            return cached
        out = math.sqrt(sum((_subject_token_idf(tok) ** 2) for tok in tokens))
        _subject_fingerprint_norm_cache[key] = out
        return out

    def _soft_subject_similarity(a: str, b: str) -> float:
        cache_key = (a or "", b or "")
        cached = _soft_subject_similarity_cache.get(cache_key)
        if cached is not None:
            return cached
        a_norm = _normalize_subject_for_match_cached(a or "")
        b_norm = _normalize_subject_for_match_cached(b or "")
        if not a_norm or not b_norm:
            out = 0.0
        else:
            out = _subject_family_similarity(a_norm, b_norm)
            fp_a = _subject_fingerprint_tokens(a_norm)
            fp_b = _subject_fingerprint_tokens(b_norm)
            if fp_a and fp_b:
                shared = fp_a & fp_b
                if shared:
                    num = sum((_subject_token_idf(tok) ** 2) for tok in shared)
                    den_a = _subject_fingerprint_norm(fp_a)
                    den_b = _subject_fingerprint_norm(fp_b)
                    if den_a > 0 and den_b > 0:
                        out = max(out, num / (den_a * den_b))
        _soft_subject_similarity_cache[cache_key] = out
        _soft_subject_similarity_cache[(cache_key[1], cache_key[0])] = out
        return out

    def _same_subject_family(a: str, b: str) -> bool:
        cache_key = (a or "", b or "")
        cached = _same_subject_family_cache.get(cache_key)
        if cached is not None:
            return cached
        a_norm = _normalize_subject_for_match_cached(a or "")
        b_norm = _normalize_subject_for_match_cached(b or "")
        if not a_norm or not b_norm:
            out = False
        elif a_norm == b_norm:
            out = True
        elif (a_norm in b_norm or b_norm in a_norm) and _subject_family_similarity(a_norm, b_norm) >= 0.45:
            out = True
        else:
            out = False
        _same_subject_family_cache[cache_key] = out
        _same_subject_family_cache[(cache_key[1], cache_key[0])] = out
        return out

    def _strong_identity_overlap(a: str, b: str) -> bool:
        cache_key = (a or "", b or "")
        cached = _strong_identity_overlap_cache.get(cache_key)
        if cached is not None:
            return cached
        if not a or not b:
            out = False
            _strong_identity_overlap_cache[cache_key] = out
            _strong_identity_overlap_cache[(cache_key[1], cache_key[0])] = out
            return out
        signal_pairs = (
            (_inc_tokens(a), _inc_tokens(b)),
            (_interface_tokens(a), _interface_tokens(b)),
            (_id_like_tokens(a), _id_like_tokens(b)),
            (_sig_num_tokens(a), _sig_num_tokens(b)),
            (_part_tokens(a), _part_tokens(b)),
        )
        out = False
        for left, right in signal_pairs:
            if left and right and not left.isdisjoint(right):
                out = True
                break
        _strong_identity_overlap_cache[cache_key] = out
        _strong_identity_overlap_cache[(cache_key[1], cache_key[0])] = out
        return out

    def _fresh_picker_subject_safe(subject_norm: str, key: str, iface_tokens=None, allow_added_inc: bool = False) -> bool:
        if not subject_norm or not key:
            return False
        cache_key = (
            subject_norm or "",
            key or "",
            tuple(sorted(iface_tokens)) if iface_tokens else (),
            bool(allow_added_inc),
        )
        cached = _fresh_picker_subject_safe_cache.get(cache_key)
        if cached is not None:
            return cached
        subj_norm = _normalize_subject_for_match_cached(subject_norm)
        key_norm = _normalize_subject_for_match_cached(key)
        if not _strict_subject_identity_gate(
            subj_norm,
            key_norm,
            iface_tokens=iface_tokens,
        ):
            out = False
            _fresh_picker_subject_safe_cache[cache_key] = out
            return out
        subj_inc_set = _inc_tokens(subj_norm)
        key_inc_set = _inc_tokens(key_norm)
        if not allow_added_inc and _defer_added_inc_identity(subj_inc_set, key_inc_set):
            out = False
            _fresh_picker_subject_safe_cache[cache_key] = out
            return out
        if _same_subject_family(subj_norm, key_norm):
            out = True
            _fresh_picker_subject_safe_cache[cache_key] = out
            return out
        if not _strong_identity_overlap(subj_norm, key_norm):
            out = False
            _fresh_picker_subject_safe_cache[cache_key] = out
            return out
        # Interface hints are only supporting signals. If the row subject itself
        # does not explicitly carry interface tokens, do not let inferred hints
        # become a hard rejection gate.
        subj_iface_set = _interface_tokens(subj_norm)
        row_has_explicit_iface = bool(subj_iface_set)
        hinted_iface_set = set(iface_tokens) if iface_tokens else set()
        key_iface_set = _interface_tokens(key_norm)
        if row_has_explicit_iface and key_iface_set and subj_iface_set.isdisjoint(key_iface_set):
            out = False
            _fresh_picker_subject_safe_cache[cache_key] = out
            return out
        similarity = _soft_subject_similarity(subj_norm, key_norm)
        if similarity >= 0.55:
            out = True
            _fresh_picker_subject_safe_cache[cache_key] = out
            return out
        shared_num = _sig_num_tokens(subj_norm) & _sig_num_tokens(key_norm)
        shared_part = _part_tokens(subj_norm) & _part_tokens(key_norm)
        shared_iface = set()
        if row_has_explicit_iface and key_iface_set:
            shared_iface = subj_iface_set & key_iface_set
        elif hinted_iface_set and key_iface_set:
            shared_iface = hinted_iface_set & key_iface_set
        shared_id = _id_like_tokens(subj_norm) & _id_like_tokens(key_norm)
        out = similarity >= 0.35 and bool(shared_num or shared_part or shared_iface or shared_id)
        _fresh_picker_subject_safe_cache[cache_key] = out
        return out

    def _find_hidden_subject_rescue(subject_norm: str, requester: str, date_tokens=None, iface_tokens=None, baseline_date=None):
        if not subject_norm or not requester:
            return None
        iface_tokens = iface_tokens or set()
        subj_norm = normalize_subject_for_match(subject_norm)
        subj_tokens = _match_tokens(subj_norm)
        if not subj_tokens:
            return None
        subj_inc_set = _inc_tokens(subj_norm)
        subj_iface_set = _interface_tokens(subj_norm) or set(iface_tokens)
        ranked = []
        for key, thread in threads.items():
            if not thread:
                continue
            has_requester = _thread_has_requester(thread, requester)
            has_consultant_date = bool(date_tokens and _thread_has_consultant_on_or_near_date(thread, date_tokens, requester))
            has_thread_date = bool(date_tokens and _thread_has_date_token(thread, date_tokens))
            if not has_requester and not has_consultant_date:
                continue
            if date_tokens and not (has_thread_date or has_consultant_date):
                continue
            best_hidden_subject = None
            best_hidden_score = None
            best_hidden_similarity = 0.0
            best_hidden_id_overlap = False
            for e in thread:
                subj_raw = getattr(e, "subject", "") or ""
                hidden_norm = normalize_subject_for_match(subj_raw)
                if not hidden_norm:
                    continue
                same_family = _same_subject_family(subj_norm, hidden_norm)
                strong_id = _strong_identity_overlap(subj_norm, hidden_norm)
                if not same_family and not strong_id:
                    continue
                hidden_iface_set = _interface_tokens(hidden_norm)
                if subj_iface_set and hidden_iface_set and subj_iface_set.isdisjoint(hidden_iface_set):
                    continue
                hidden_inc_set = _inc_tokens(hidden_norm)
                if subj_inc_set and hidden_inc_set and subj_inc_set.isdisjoint(hidden_inc_set):
                    continue
                similarity = _subject_family_similarity(subj_norm, hidden_norm)
                if similarity < 0.45 and not strong_id:
                    continue
                score = similarity
                if same_family:
                    score += 0.18
                if strong_id:
                    score += 0.16
                if has_requester:
                    score += 0.10
                if has_thread_date:
                    score += 0.08
                if has_consultant_date:
                    score += 0.12
                if baseline_date:
                    delta = _requester_min_delta_days(thread, requester, baseline_date)
                    if delta is not None:
                        score += max(0.0, 0.08 - min(0.08, 0.02 * delta))
                if best_hidden_score is None or score > best_hidden_score:
                    best_hidden_score = score
                    best_hidden_subject = subj_raw
                    best_hidden_similarity = similarity
                    best_hidden_id_overlap = strong_id
            if best_hidden_score is None:
                continue
            ranked.append((
                -best_hidden_score,
                -best_hidden_similarity,
                0 if best_hidden_id_overlap else 1,
                key,
                best_hidden_subject,
                thread,
            ))
        if not ranked:
            return None
        ranked.sort()
        best = ranked[0]
        second = ranked[1] if len(ranked) > 1 else None
        best_score = -best[0]
        second_score = -second[0] if second else -1.0
        if best_score < 0.78:
            return None
        if second and (best_score - second_score) < 0.08:
            return None
        return best[5], f"HiddenSubjectRescue:{best[4]}"

    def _baseline_refine_safe(subject_norm: str, current_match_note: str, refined_key: str) -> bool:
        if not refined_key:
            return False
        if _same_subject_family(subject_norm, refined_key):
            return True
        match_note_l = (current_match_note or "").lower()
        exact_like = (
            "rowexact" in match_note_l
            or match_note_l.startswith("exact")
            or match_note_l.startswith("altexact")
        )
        if exact_like:
            return False
        if not _strong_identity_overlap(subject_norm, refined_key):
            return False
        return _subject_family_similarity(subject_norm, refined_key) >= 0.35

    def _subject_for_description(description: str) -> str:
        subject_text = extract_subject_from_description(description or "")
        if description and re.search(r"(?:--\.?>|â†’|âž”|âž¡|=>)", description):
            parts = re.split(r"\s*(?:--\.?>|â†’|âž”|âž¡|=>)\s*", description, maxsplit=1)
            if len(parts) >= 2:
                _right = parts[1].strip()
                _right_l = _right.lower()
                if "deployment request" in _right_l or ("success:" in _right_l and "deployment" in _right_l):
                    subject_text = _right
        return subject_text

    def _requester_key(value: str) -> str:
        if not value:
            return ""
        s = str(value).strip().lower()
        return re.sub(r"[^a-z0-9]", "", s)

    def _service_no_key(value: str) -> str:
        if not value:
            return ""
        s = str(value).strip().upper()
        return re.sub(r"\s+", "", s)

    def _parse_time_str(value: str):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%d-%m-%Y %H:%M")
        except Exception:
            return None

    def _ack_missing(times: TimeResult | None, debug: TimeDebug | None) -> bool:
        if not times or not debug:
            return True
        notes = (debug.notes or "")
        if "ACK NOT FOUND" in notes:
            return True
        if times.created and times.response and times.created == times.response:
            return True
        if times.created and times.response and times.resolved:
            return times.created == times.response == times.resolved
        return False

    def _thread_has_consultant_on_date(thread, date_tokens: list[str], requester: str) -> bool:
        if not date_tokens or not thread or not requester:
            return False
        for e in thread:
            if _match_requester(e.sender_name, e.sender_email, requester):
                if _email_date_matches(e.sent_time, date_tokens):
                    return True
        return False

    def _thread_has_consultant_on_or_near_date(
        thread,
        date_tokens: list[str],
        requester: str,
        grace_hours: int = 12,
    ) -> bool:
        if not date_tokens or not thread or not requester:
            return False
        anchor_date = _anchor_date(date_tokens)
        if not anchor_date:
            return False
        anchor_start = _to_ist(datetime(anchor_date.year, anchor_date.month, anchor_date.day))
        window_start = anchor_start - timedelta(hours=grace_hours)
        for e in thread:
            if not _match_requester(e.sender_name, e.sender_email, requester):
                continue
            try:
                sent = _to_ist(e.sent_time)
            except Exception:
                continue
            if sent.date() == anchor_date or (window_start <= sent < anchor_start):
                return True
        return False

    def _thread_has_requester(thread, requester: str) -> bool:
        if not thread or not requester:
            return False
        for e in thread:
            if _match_requester(e.sender_name, e.sender_email, requester):
                return True
        return False

    def _pick_reply_after_ack(consultant_after, ack_ist, requester_name: str = "", grace_minutes: int = 16):
        if not consultant_after:
            return None
        # Prefer the earliest non-ack/non-reminder reply after ack.
        non_ack = _shared_resolution_candidates(consultant_after)
        if non_ack:
            non_ack.sort(
                key=lambda e: (
                    _post_ack_reply_rank(e, requester_name),
                    e.sent_time,
                )
            )
            for e in non_ack:
                try:
                    if _to_ist(e.sent_time) > ack_ist:
                        return e
                except Exception:
                    continue
        if len(consultant_after) >= 2:
            first_dt = _to_ist(consultant_after[0].sent_time)
            if first_dt <= ack_ist + timedelta(minutes=grace_minutes) and _is_ack_like_reply(consultant_after[0]):
                # Skip to the next non-ack reply if available.
                for e in consultant_after[1:]:
                    if _is_shared_real_reply_candidate(e):
                        return e
                return consultant_after[1]
        return consultant_after[0]

    def _defer_added_inc_identity(subj_inc_set, key_inc_set) -> bool:
        return bool(key_inc_set) and (not subj_inc_set or subj_inc_set.isdisjoint(key_inc_set))

    def _find_unique_requester_date_thread(subject_norm, requester, date_tokens, iface_tokens=None):
        if not subject_norm or not requester or not date_tokens:
            return None
        iface_tokens = iface_tokens or set()
        subj_tokens = _match_tokens(subject_norm)
        if not subj_tokens:
            return None
        subj_inc_set = _inc_tokens(subject_norm)
        subj_num_set = _sig_num_tokens(subject_norm)
        hits = []
        # Keep soft RE/FW-family expansion limited to the main finder. These
        # requester/date helpers should behave like the older broader logic and
        # operate on normalized thread keys only.
        pool = list(threads.items())
        for key, thread in pool:
            if not _thread_has_requester(thread, requester):
                continue
            if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=iface_tokens):
                continue
            key_tokens = _match_tokens(key)
            if not key_tokens:
                continue
            if subj_inc_set:
                key_inc_set = _inc_tokens(key)
                if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                    continue
            else:
                key_inc_set = _inc_tokens(key)
                if _defer_added_inc_identity(subj_inc_set, key_inc_set):
                    continue
            key_num_set = _sig_num_tokens(key)
            if subj_num_set and key_num_set and subj_num_set.isdisjoint(key_num_set):
                continue
            if iface_tokens:
                key_iface_set = _interface_tokens(key)
                if not key_iface_set or key_iface_set.isdisjoint(iface_tokens):
                    continue
            score = _token_overlap_score(subj_tokens, key_tokens)
            contains = (subject_norm in key or key in subject_norm)
            if score < 0.45 and not contains:
                continue
            if not (
                _thread_has_date_token(thread, date_tokens)
                or _thread_has_consultant_on_or_near_date(thread, date_tokens, requester)
            ):
                continue
            hits.append((score, key, thread))
        if len(hits) == 1:
            return hits[0][2], f"RequesterDateUnique:{hits[0][1]}"
        return None

    def _find_best_requester_date_thread(subject_norm, requester, date_tokens, iface_tokens=None):
        if not subject_norm or not requester or not date_tokens:
            return None
        anchor_date = _anchor_date(date_tokens)
        if not anchor_date:
            return None
        iface_tokens = iface_tokens or set()
        anchor_start = _to_ist(datetime(anchor_date.year, anchor_date.month, anchor_date.day))
        subj_tokens = _match_tokens(subject_norm)
        if not subj_tokens:
            return None
        subj_inc_set = _inc_tokens(subject_norm)
        subj_num_set = _sig_num_tokens(subject_norm)
        best = None
        # Keep soft RE/FW-family expansion limited to the main finder. These
        # requester/date helpers should behave like the older broader logic and
        # operate on normalized thread keys only.
        pool = list(threads.items())
        for key, thread in pool:
            if not _thread_has_requester(thread, requester):
                continue
            if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=iface_tokens):
                continue
            key_tokens = _match_tokens(key)
            if not key_tokens:
                continue
            if subj_inc_set:
                key_inc_set = _inc_tokens(key)
                if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                    continue
            else:
                key_inc_set = _inc_tokens(key)
                if _defer_added_inc_identity(subj_inc_set, key_inc_set):
                    continue
            key_num_set = _sig_num_tokens(key)
            if subj_num_set and key_num_set and subj_num_set.isdisjoint(key_num_set):
                continue
            if iface_tokens:
                key_iface_set = _interface_tokens(key)
                if not key_iface_set or key_iface_set.isdisjoint(iface_tokens):
                    continue
            score = _token_overlap_score(subj_tokens, key_tokens)
            contains = (subject_norm in key or key in subject_norm)
            if score < 0.35 and not contains:
                continue

            # Require a date signal (token in thread or consultant on/near anchor date).
            date_signal = (
                _thread_has_date_token(thread, date_tokens)
                or _thread_has_consultant_on_or_near_date(thread, date_tokens, requester)
            )
            if not date_signal:
                if not contains and score < 0.55:
                    continue

            # Find closest consultant reply to anchor start.
            consultant_replies = [
                e for e in thread
                if e.sent_time and _match_requester(e.sender_name, e.sender_email, requester)
            ]
            if not consultant_replies:
                continue
            consultant_non_ack = _shared_resolution_candidates(consultant_replies)
            if not consultant_non_ack:
                continue
            reply_pool = consultant_non_ack
            min_delta = None
            for e in reply_pool:
                try:
                    delta = abs(_to_ist(e.sent_time) - anchor_start)
                except Exception:
                    continue
                if min_delta is None or delta < min_delta:
                    min_delta = delta
            if min_delta is None:
                continue
            # Too far from anchor date → skip.
            if min_delta > timedelta(days=5):
                continue

            # Prefer smaller delta, then higher score, then date token presence.
            token_bonus = 0.2 if _thread_has_date_token(thread, date_tokens) else 0.0
            score = min(1.0, score + (0.1 if contains else 0.0) + token_bonus)
            candidate = (min_delta, -score, key, thread)
            if best is None or candidate < best:
                best = candidate

        if best:
            _, _, key, thread = best
            return thread, f"RequesterDateBest:{key}"
        return None

    def _union_alt_threads(alt_subject, requester=None, subject_norm=None, iface_tokens=None):
        if not alt_subject:
            return None
        keys = alt_index.get(alt_subject, [])
        if not keys:
            return None
        merged = []
        seen = set()
        has_requester = False
        for k in keys:
            if subject_norm and not _fresh_picker_subject_safe(subject_norm, k, iface_tokens=iface_tokens):
                continue
            for e in threads.get(k, []):
                dedup_key = (e.subject, e.sender_email, e.sent_time)
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)
                merged.append(e)
                if requester and _match_requester(e.sender_name, e.sender_email, requester):
                    has_requester = True
        if requester and not has_requester:
            return None
        return merged

    def _parse_date_token(token: str):
        if not token:
            return None
        t = token.strip()
        try:
            # yyyy-mm-dd / yyyy/mm/dd / yyyy.mm.dd
            if re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$", t):
                parts = re.split(r"[-/.]", t)
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                return datetime(y, m, d).date()
            # dd-mm-yyyy or dd-mm-yy (day-first)
            if re.match(r"^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}$", t):
                parts = re.split(r"[-/.]", t)
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 100:
                    y = 2000 + y
                return datetime(y, m, d).date()
        except Exception:
            return None
        return None

    def _subject_date_value_set(text: str) -> set:
        key = text or ""
        cached = _subject_date_value_cache.get(key)
        if cached is not None:
            return cached
        out = set()
        for token in _extract_date_tokens(text or ""):
            dt = _parse_date_token(token)
            if dt:
                out.add(dt.isoformat())
        _subject_date_value_cache[key] = out
        return out

    def _strict_subject_identity_gate(subject_norm: str, candidate_norm: str, iface_tokens=None) -> bool:
        if not subject_norm or not candidate_norm:
            return bool(subject_norm and candidate_norm)
        cache_key = (
            subject_norm or "",
            candidate_norm or "",
            tuple(sorted(iface_tokens)) if iface_tokens else (),
        )
        cached = _strict_subject_identity_gate_cache.get(cache_key)
        if cached is not None:
            return cached

        row_inc_set = _inc_tokens(subject_norm)
        cand_inc_set = _inc_tokens(candidate_norm)
        if row_inc_set and cand_inc_set and row_inc_set.isdisjoint(cand_inc_set):
            _strict_subject_identity_gate_cache[cache_key] = False
            return False

        row_iface_set = _interface_tokens(subject_norm)
        cand_iface_set = _interface_tokens(candidate_norm)
        hinted_iface_set = set(iface_tokens) if iface_tokens else set()
        if row_iface_set and cand_iface_set and row_iface_set.isdisjoint(cand_iface_set):
            _strict_subject_identity_gate_cache[cache_key] = False
            return False
        if (not row_iface_set) and hinted_iface_set and cand_iface_set and row_inc_set and hinted_iface_set.isdisjoint(cand_iface_set):
            _strict_subject_identity_gate_cache[cache_key] = False
            return False

        row_dates = _subject_date_value_set(subject_norm)
        cand_dates = _subject_date_value_set(candidate_norm)
        if (
            row_dates
            and cand_dates
            and _has_explicit_date_marker(subject_norm)
            and _has_explicit_date_marker(candidate_norm)
            and row_dates.isdisjoint(cand_dates)
        ):
            _strict_subject_identity_gate_cache[cache_key] = False
            return False

        _strict_subject_identity_gate_cache[cache_key] = True
        return True

    def _thread_has_requester_cached(thread_key: str, thread, requester: str) -> bool:
        thread_sig = tuple(id(e) for e in (thread or ()))
        cache_key = (thread_key or "", thread_sig, (requester or "").strip().lower())
        cached = _thread_requester_presence_cache.get(cache_key)
        if cached is not None:
            return cached
        out = _thread_has_requester(thread, requester)
        _thread_requester_presence_cache[cache_key] = out
        return out

    def _thread_has_consultant_on_or_near_date_cached(
        thread_key: str,
        thread,
        date_tokens: list[str],
        requester: str,
    ) -> bool:
        token_key = tuple(date_tokens or ())
        thread_sig = tuple(id(e) for e in (thread or ()))
        cache_key = (thread_key or "", thread_sig, token_key, (requester or "").strip().lower())
        cached = _thread_consultant_date_cache.get(cache_key)
        if cached is not None:
            return cached
        out = _thread_has_consultant_on_or_near_date(thread, list(token_key), requester)
        _thread_consultant_date_cache[cache_key] = out
        return out

    def _thread_has_date_token_cached(thread_key: str, thread, date_tokens: list[str]) -> bool:
        token_key = tuple(date_tokens or ())
        thread_sig = tuple(id(e) for e in (thread or ()))
        cache_key = (thread_key or "", thread_sig, token_key)
        cached = _thread_date_token_presence_cache.get(cache_key)
        if cached is not None:
            return cached
        out = _thread_has_date_token(thread, list(token_key))
        _thread_date_token_presence_cache[cache_key] = out
        return out

    def _thread_has_non_ess_cached(thread_key: str, thread) -> bool:
        thread_sig = tuple(id(e) for e in (thread or ()))
        cache_key = (thread_key or "", thread_sig)
        cached = _thread_non_ess_cache.get(cache_key)
        if cached is not None:
            return cached
        out = any(not _is_ess_sender(e, ess_team) for e in (thread or []))
        _thread_non_ess_cache[cache_key] = out
        return out

    def _thread_baseline_delta_cached(thread_key: str, thread, requester: str, baseline_date):
        baseline_key = baseline_date.isoformat() if hasattr(baseline_date, "isoformat") else str(baseline_date or "")
        thread_sig = tuple(id(e) for e in (thread or ()))
        cache_key = (thread_key or "", thread_sig, (requester or "").strip().lower(), baseline_key)
        cached = _thread_baseline_delta_cache.get(cache_key)
        if cached is not None:
            return cached
        out = _requester_min_delta_days(thread, requester, baseline_date)
        _thread_baseline_delta_cache[cache_key] = out
        return out

    def _family_subject_score(
        subject_norm: str,
        thread_key: str,
        thread,
        requester: str = "",
        date_tokens=None,
        iface_tokens=None,
        baseline_date=None,
    ) -> float:
        token_key = tuple(date_tokens or ())
        iface_key = tuple(sorted(iface_tokens or ()))
        baseline_key = baseline_date.isoformat() if hasattr(baseline_date, "isoformat") else str(baseline_date or "")
        cache_key = (
            subject_norm or "",
            thread_key or "",
            (requester or "").strip().lower(),
            token_key,
            iface_key,
            baseline_key,
        )
        cached = _family_subject_score_cache.get(cache_key)
        if cached is not None:
            return cached

        score = max(
            _subject_family_similarity(subject_norm, thread_key),
            _soft_subject_similarity(subject_norm, thread_key),
        )
        same_family = _same_subject_family(subject_norm, thread_key)
        strong_identity = _strong_identity_overlap(subject_norm, thread_key)
        row_inc_set = _inc_tokens(subject_norm)
        key_inc_set = _inc_tokens(thread_key)
        row_iface_set = _interface_tokens(subject_norm)
        hinted_iface_set = set(iface_tokens or ())
        key_iface_set = _interface_tokens(thread_key)
        row_dates = _subject_date_value_set(subject_norm)
        key_dates = _subject_date_value_set(thread_key)

        if row_inc_set and key_inc_set and not row_inc_set.isdisjoint(key_inc_set):
            score += 0.18
        if row_iface_set and key_iface_set and not row_iface_set.isdisjoint(key_iface_set):
            score += 0.12
        elif (not row_iface_set) and hinted_iface_set and key_iface_set and not hinted_iface_set.isdisjoint(key_iface_set):
            score += 0.08
        if row_dates and key_dates and not row_dates.isdisjoint(key_dates):
            score += 0.08
        if subject_norm and thread_key and (subject_norm in thread_key or thread_key in subject_norm):
            score += 0.08

        has_requester = bool(requester and _thread_has_requester_cached(thread_key, thread, requester))
        has_consultant_date = bool(
            requester and date_tokens and _thread_has_consultant_on_or_near_date_cached(thread_key, thread, date_tokens, requester)
        )
        has_thread_date = bool(date_tokens and _thread_has_date_token_cached(thread_key, thread, date_tokens))
        has_non_ess = _thread_has_non_ess_cached(thread_key, thread)

        if has_requester and (same_family or strong_identity):
            score += 0.08
        if has_consultant_date and same_family:
            score += 0.10
        if has_thread_date and (same_family or strong_identity):
            score += 0.06
        if has_non_ess and same_family:
            score += 0.03
        if requester and baseline_date and (same_family or strong_identity):
            delta = _thread_baseline_delta_cached(thread_key, thread, requester, baseline_date)
            if delta is not None:
                score += max(0.0, 0.06 - min(0.06, 0.01 * delta))

        out = min(1.0, score)
        _family_subject_score_cache[cache_key] = out
        return out

    def _email_date_matches(sent_time, date_tokens: list[str]) -> bool:
        if not sent_time or not date_tokens:
            return False
        try:
            sent_d = _to_ist(sent_time).date()
        except Exception:
            try:
                sent_d = sent_time.date()
            except Exception:
                return False
        for t in date_tokens:
            dt = _parse_date_token(t)
            if dt and dt == sent_d:
                return True
        return False

    def _anchor_date(date_tokens: list[str]):
        if not date_tokens:
            return None
        for t in date_tokens:
            dt = _parse_date_token(t)
            if dt:
                return dt
        return None

    def _coerce_row_created_date(value):
        if value is None:
            return None
        if isinstance(value, datetime):
            try:
                return _to_ist(value).date()
            except Exception:
                return value.date()
        # openpyxl may return date objects for date-only cells.
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day") and not isinstance(value, str):
            try:
                return datetime(int(value.year), int(value.month), int(value.day)).date()
            except Exception:
                pass
        s = str(value).strip()
        if not s:
            return None
        # Common datetime string formats seen in exports/templates.
        for fmt in (
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %H:%M",
            "%d.%m.%Y %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%d-%m-%Y %I:%M %p",
            "%d/%m/%Y %I:%M %p",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
        ):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        # Try to pull a date token from mixed strings, e.g. "2026-01-06 10:22:11 UTC".
        m = re.search(r"\b\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\b", s)
        if m:
            d = _parse_date_token(m.group(0))
            if d:
                return d
        m = re.search(r"\b\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}\b", s)
        if m:
            d = _parse_date_token(m.group(0))
            if d:
                return d
        dt = _parse_time_str(s)
        if dt:
            try:
                return _to_ist(dt).date()
            except Exception:
                return dt.date()
        try:
            d = _parse_date_token(s)
            return d
        except Exception:
            return None

    def _is_stale_anchor_date(date_tokens: list[str], row_created_value, max_days: int = 14) -> bool:
        anchor = _anchor_date(date_tokens)
        if not anchor:
            return False
        created_d = _coerce_row_created_date(row_created_value)
        if not created_d:
            return False
        return abs((created_d - anchor).days) > max_days

    def _anchor_relevant_to_thread(anchor_date, thread, max_days: int = 21) -> bool:
        if not anchor_date or not thread:
            return False
        thread_dates = []
        for e in thread:
            if not e.sent_time:
                continue
            try:
                thread_dates.append(_to_ist(e.sent_time).date())
            except Exception:
                continue
        if not thread_dates:
            return False
        min_d = min(thread_dates)
        max_d = max(thread_dates)
        return (min_d - timedelta(days=max_days)) <= anchor_date <= (max_d + timedelta(days=max_days))

    def _baseline_row_created_value(row_context: dict):
        # Prefer immutable ServiceNow baseline if present.
        # "SLA Start Date & Time" is not overwritten by our resolver.
        sla_start = row_context.get("Sla Start Date & Time") or row_context.get("SLA Start Date & Time")
        if _coerce_row_created_date(sla_start):
            return sla_start
        return row_context.get("Created Date & Time")

    def _requester_min_delta_days(thread, requester: str, baseline_date):
        if not thread or not requester or not baseline_date:
            return None
        best = None
        for e in thread:
            if not e.sent_time:
                continue
            if not _match_requester(e.sender_name, e.sender_email, requester):
                continue
            try:
                sent_d = _to_ist(e.sent_time).date()
            except Exception:
                try:
                    sent_d = e.sent_time.date()
                except Exception:
                    continue
            delta = abs((sent_d - baseline_date).days)
            if best is None or delta < best:
                best = delta
        return best

    def _find_best_thread_near_baseline(subject_norm: str, requester: str, baseline_date, iface_tokens=None):
        if not subject_norm or not requester or not baseline_date:
            return None
        iface_tokens = iface_tokens or set()
        subj_tokens = _match_tokens(subject_norm)
        if not subj_tokens:
            return None
        subj_inc_set = _inc_tokens(subject_norm)
        subj_iface_set = _interface_tokens(subject_norm)
        if not subj_iface_set and iface_tokens:
            subj_iface_set = iface_tokens

        best = None
        # Keep soft RE/FW-family expansion limited to the main finder. Baseline
        # refinement should stay on normalized thread keys only.
        pool = list(threads.items())
        for key, thread in pool:
            if not _thread_has_requester(thread, requester):
                continue
            if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=iface_tokens):
                continue
            key_tokens = _match_tokens(key)
            if not key_tokens:
                continue
            if subj_inc_set:
                key_inc_set = _inc_tokens(key)
                if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                    continue
            else:
                key_inc_set = _inc_tokens(key)
                if _defer_added_inc_identity(subj_inc_set, key_inc_set):
                    continue
            if subj_iface_set:
                key_iface_set = _interface_tokens(key)
                if key_iface_set and subj_iface_set.isdisjoint(key_iface_set):
                    continue

            score = _token_overlap_score(subj_tokens, key_tokens)
            contains = (subject_norm in key or key in subject_norm)
            if score < 0.35 and not contains:
                continue

            delta_days = _requester_min_delta_days(thread, requester, baseline_date)
            if delta_days is None:
                continue
            # Keep this bounded so we don't jump to unrelated long-history threads.
            if delta_days > 14:
                continue

            candidate = (delta_days, -score, -len(key_tokens), key, thread)
            if best is None or candidate < best:
                best = candidate

        if best:
            delta_days, _neg_score, _neg_len, key, thread = best
            return thread, f"BaselineDateRefined:{key}", delta_days
        return None

    def _email_on_or_after_date(sent_time, anchor_date) -> bool:
        if not sent_time or not anchor_date:
            return False
        try:
            sent_d = _to_ist(sent_time).date()
        except Exception:
            try:
                sent_d = sent_time.date()
            except Exception:
                return False
        return sent_d >= anchor_date

    def _thread_has_non_ess_near_time(thread, target_dt, hours=24) -> bool:
        if not thread or not target_dt:
            return False
        try:
            target_ist = _to_ist(target_dt)
        except Exception:
            return False
        window = timedelta(hours=hours)
        for e in thread:
            if _is_ess_sender(e, ess_team):
                continue
            if not e.sent_time:
                continue
            try:
                sent_ist = _to_ist(e.sent_time)
            except Exception:
                continue
            if abs(sent_ist - target_ist) <= window:
                return True
        return False

    def _latest_non_ess_before(thread, cutoff_dt):
        if not thread or not cutoff_dt:
            return None
        out = None
        cutoff_ist = _to_ist(cutoff_dt)
        for e in thread:
            if _is_ess_sender(e, ess_team):
                continue
            if not e.sent_time:
                continue
            try:
                sent_ist = _to_ist(e.sent_time)
            except Exception:
                continue
            if sent_ist <= cutoff_ist and (out is None or sent_ist > _to_ist(out.sent_time)):
                out = e
        return out

    def _apply_final_time_guards(
        thread,
        requester,
        times,
        debug,
        base_times,
        base_debug,
        baseline_created_date=None,
        is_dep_req=False,
        is_dep_succ=False,
    ):
        if not times or not debug or not thread:
            return times, debug
        if is_dep_req or is_dep_succ:
            return times, debug

        notes = debug.notes or ""
        if "Maintenance override" in notes:
            return times, debug

        created_dt = _parse_time_str(times.created)
        resp_dt = _parse_time_str(times.response)
        res_dt = _parse_time_str(times.resolved)
        created_far_from_baseline = False
        if created_dt and baseline_created_date:
            try:
                created_far_from_baseline = abs((_to_ist(created_dt).date() - baseline_created_date).days) > 14
            except Exception:
                created_far_from_baseline = False

        # Guard 1: when Created comes from parsed quote and there is no nearby
        # non-ESS evidence, anchor Created back to latest non-ESS request.
        if (
            created_dt
            and isinstance(debug.created_src, str)
            and debug.created_src.startswith("PARSED_")
            and any(not _is_ess_sender(e, ess_team) for e in thread)
            and (
                not _thread_has_non_ess_near_time(thread, created_dt, hours=24)
                or created_far_from_baseline
            )
        ):
            cutoff = resp_dt or res_dt
            if not cutoff:
                consultant_times = [
                    e.sent_time for e in thread
                    if e.sent_time and _match_requester(e.sender_name, e.sender_email, requester)
                ]
                if consultant_times:
                    try:
                        cutoff = max(consultant_times, key=_to_ist)
                    except Exception:
                        cutoff = None
            req = _latest_non_ess_before(thread, cutoff) if cutoff else None
            if req:
                times = TimeResult(_format_time(req.sent_time), times.response, times.resolved)
                debug = TimeDebug(
                    req.sender_email or req.sender_name,
                    debug.ack_src,
                    debug.resolved_src,
                    f"{notes}; GuardCreatedFromNonESS{'Baseline' if created_far_from_baseline else ''}",
                )
                notes = debug.notes
                created_dt = _parse_time_str(times.created)

        # Guard 2: enforce Created <= Response using non-ESS request fallback.
        if created_dt and resp_dt and _to_ist(created_dt) > _to_ist(resp_dt):
            req = _latest_non_ess_before(thread, resp_dt)
            if req:
                times = TimeResult(_format_time(req.sent_time), times.response, times.resolved)
                debug = TimeDebug(
                    req.sender_email or req.sender_name,
                    debug.ack_src,
                    debug.resolved_src,
                    f"{notes}; GuardCreatedBeforeResponse",
                )
                notes = debug.notes
                created_dt = _parse_time_str(times.created)
            elif base_times:
                base_created_dt = _parse_time_str(base_times.created)
                if base_created_dt and _to_ist(base_created_dt) <= _to_ist(resp_dt):
                    times = TimeResult(base_times.created, times.response, times.resolved)
                    src = base_debug.created_src if base_debug else debug.created_src
                    debug = TimeDebug(
                        src,
                        debug.ack_src,
                        debug.resolved_src,
                        f"{notes}; GuardCreatedRevertBase",
                    )
                    notes = debug.notes
                    created_dt = _parse_time_str(times.created)

        # Guard 3: enforce Resolved >= Response when both exist.
        if resp_dt and res_dt and _to_ist(res_dt) < _to_ist(resp_dt):
            if _has_source_locked_same_time(notes) or _is_force_same_time_subject(None, thread):
                return times, debug
            ack_ist = _to_ist(resp_dt)
            window_end = ack_ist + timedelta(hours=48)
            consultant_after = [
                e for e in thread
                if _match_requester(e.sender_name, e.sender_email, requester)
                and e.sent_time
                and _to_ist(e.sent_time) > ack_ist
                and _to_ist(e.sent_time) <= window_end
                and _is_shared_real_reply_candidate(e)
            ]
            consultant_after.sort(key=lambda e: e.sent_time)
            if consultant_after:
                pick = consultant_after[0]
                times = TimeResult(times.created, times.response, _format_time(pick.sent_time))
                debug = TimeDebug(
                    debug.created_src,
                    debug.ack_src,
                    pick.sender_email or pick.sender_name,
                    f"{notes}; GuardResolvedAfterResponse",
                )
            elif base_times:
                base_res_dt = _parse_time_str(base_times.resolved)
                if base_res_dt and _to_ist(base_res_dt) >= _to_ist(resp_dt):
                    src = base_debug.resolved_src if base_debug else debug.resolved_src
                    times = TimeResult(times.created, times.response, base_times.resolved)
                    debug = TimeDebug(
                        debug.created_src,
                        debug.ack_src,
                        src,
                        f"{notes}; GuardResolvedRevertBase",
                    )

        # Guard 4 (ESS-only baseline alignment): when row baseline day is far from
        # current created day, prefer requester reply nearest to baseline day.
        if (
            baseline_created_date
            and thread
            and requester
            and "ESS-only; no non-ESS request" in notes
        ):
            created_dt = _parse_time_str(times.created)
            created_ist = _to_ist(created_dt) if created_dt else None
            current_delta = abs((created_ist.date() - baseline_created_date).days) if created_ist else 9999
            if current_delta > 2:
                requester_replies = [
                    e for e in thread
                    if e.sent_time and _match_requester(e.sender_name, e.sender_email, requester)
                ]
                if requester_replies:
                    requester_replies.sort(key=lambda e: e.sent_time)
                    requester_non_ack = _shared_resolution_candidates(requester_replies)
                    if not requester_non_ack:
                        return times, debug
                    reply_pool = requester_non_ack
                    # Do not baseline-force across multi-episode threads.
                    # If requester has a much later episode, keep current timing.
                    latest_reply = reply_pool[-1]

                    baseline_mid = _to_ist(datetime(
                        baseline_created_date.year,
                        baseline_created_date.month,
                        baseline_created_date.day,
                    ))
                    def _rank(e):
                        sent = _to_ist(e.sent_time)
                        day_delta = abs((sent.date() - baseline_created_date).days)
                        return (day_delta, abs(sent - baseline_mid))
                    best = min(reply_pool, key=_rank)
                    best_day_delta = abs((_to_ist(best.sent_time).date() - baseline_created_date).days)
                    allow_baseline_force = (_to_ist(latest_reply.sent_time) - _to_ist(best.sent_time)) <= timedelta(days=2)
                    if allow_baseline_force and best_day_delta <= 2 and best_day_delta < current_delta:
                        t = _format_time(best.sent_time)
                        if t:
                            times = TimeResult(t, t, t)
                            debug = TimeDebug(
                                best.sender_email or best.sender_name,
                                best.sender_email or best.sender_name,
                                best.sender_email or best.sender_name,
                                f"{notes}; GuardBaselineRequesterSpan",
                            )
        return times, debug

    def _token_overlap_score(a: set, b: set) -> float:
        if not a or not b:
            return 0.0
        inter = len(a & b)
        return (2 * inter) / (len(a) + len(b))

    _seeded_first_resolution_cache = {}

    def _seed_subject_match_candidate_local(subject_norm_value: str, row_tokens: set, row_id_tokens: set, candidate_subject: str) -> bool:
        cand_norm = normalize_subject(candidate_subject or "")
        if not cand_norm:
            return False
        if not _strict_subject_identity_gate(
            subject_norm_value or "",
            cand_norm,
            iface_tokens=_interface_tokens(subject_norm_value or ""),
        ):
            return False
        if row_id_tokens:
            cand_ids = _id_like_tokens(cand_norm)
            if not cand_ids or row_id_tokens.isdisjoint(cand_ids):
                return False
        if not row_tokens:
            return True
        cand_tokens = _match_tokens(cand_norm)
        score = _token_overlap_score(row_tokens, cand_tokens) if cand_tokens else 0.0
        contains = bool(subject_norm_value and cand_norm and (subject_norm_value in cand_norm or cand_norm in subject_norm_value))
        return score >= 0.45 or contains or _fresh_picker_subject_safe(
            subject_norm_value,
            cand_norm,
            iface_tokens=_interface_tokens(subject_norm_value or ""),
            allow_added_inc=True,
        )

    def _clone_seeded_resolution_local(seed_result):
        if not seed_result:
            return None
        return {
            "times": TimeResult(
                seed_result["times"].created,
                seed_result["times"].response,
                seed_result["times"].resolved,
            ),
            "debug": TimeDebug(
                seed_result["debug"].created_src,
                seed_result["debug"].ack_src,
                seed_result["debug"].resolved_src,
                seed_result["debug"].notes,
            ),
            "episode": dict(seed_result["episode"]) if seed_result.get("episode") else None,
        }

    def _seed_lane_episode_from_thread_local(thread, requester_name: str, ess_team, subject_norm_value: str, description: str = ""):
        if not thread or not requester_name or not subject_norm_value:
            return None
        row_tokens = _match_tokens(subject_norm_value)
        row_id_tokens = _id_like_tokens(subject_norm_value)
        if not row_id_tokens:
            row_id_tokens = _id_like_tokens(description or "")

        reply_candidates = []
        for e in thread:
            if not getattr(e, "sent_time", None):
                continue
            if not _match_requester(e.sender_name, e.sender_email, requester_name):
                continue
            if not _is_shared_real_reply_candidate(e):
                continue
            if not _seed_subject_match_candidate_local(subject_norm_value, row_tokens, row_id_tokens, getattr(e, "subject", "") or ""):
                continue
            reply_candidates.append(e)
        if not reply_candidates:
            return None
        reply_candidates.sort(key=lambda e: e.sent_time)
        reply_msg = reply_candidates[-1]
        reply_ist = _to_ist(reply_msg.sent_time)

        quoted_candidates = _extract_canonical_quoted_header_candidates(reply_msg, allow_relaxed=False)
        if not quoted_candidates:
            quoted_candidates = _extract_canonical_quoted_header_candidates(reply_msg, allow_relaxed=True)
        if not quoted_candidates:
            return None

        lane_blocks = []
        for idx, candidate in enumerate(quoted_candidates):
            sent_ist = _to_ist(candidate.sent_dt) if candidate.sent_dt else None
            if not sent_ist or sent_ist >= reply_ist:
                continue
            if (reply_ist - sent_ist) > timedelta(hours=48):
                continue
            q_subj = re.sub(r"(?i)^(subject|objet)\b\s*:?\s*", "", candidate.subject_line or "").strip()
            subj_match = bool(q_subj) and _seed_subject_match_candidate_local(subject_norm_value, row_tokens, row_id_tokens, q_subj)
            allow_blank = (not q_subj) and idx <= 2
            if not subj_match and not allow_blank:
                continue
            lane_blocks.append((idx, candidate.from_line or "", sent_ist, q_subj or "", candidate.sent_line or ""))
        if not lane_blocks:
            return None

        live_ack_candidates = []
        for e in thread:
            e_ist = _to_ist(e.sent_time) if getattr(e, "sent_time", None) else None
            if not e_ist or e_ist >= reply_ist:
                continue
            if not _is_ess_sender(e, ess_team):
                continue
            if not _seed_subject_match_candidate_local(subject_norm_value, row_tokens, row_id_tokens, getattr(e, "subject", "") or ""):
                continue
            if _is_shared_ess_ack_candidate(e):
                live_ack_candidates.append((e_ist, e))

        ack_idx = None
        ack_ist = None
        ack_msg = None
        for idx, from_line, sent_ist, _q_subj, _sent_line in lane_blocks:
            if _seed_quoted_sender_is_ess(from_line, ess_team) is not True:
                continue
            matches = []
            for e_ist, e in live_ack_candidates:
                delta = abs((e_ist - sent_ist).total_seconds())
                if delta > 300:
                    continue
                sender_score = 1 if _seed_quoted_sender_matches_live(from_line, e) else 0
                matches.append((sender_score, -delta, e_ist, id(e), e))
            if not matches:
                continue
            matches.sort(reverse=True)
            ack_idx = idx
            ack_ist = matches[0][2]
            ack_msg = matches[0][4]
            break

        req_ist = None
        if ack_idx is not None and ack_ist:
            for next_idx, next_from_line, next_sent_ist, _next_q_subj, _sent_line in lane_blocks:
                if next_idx <= ack_idx:
                    continue
                if next_sent_ist >= ack_ist:
                    continue
                if _seed_quoted_sender_is_ess(next_from_line, ess_team) is not False:
                    continue
                if (ack_ist - next_sent_ist) > timedelta(minutes=16):
                    continue
                req_ist = next_sent_ist
                break
        if ack_ist and not req_ist:
            ack_ist = None
            ack_msg = None
        if not req_ist:
            for _idx, from_line, sent_ist, _q_subj, _sent_line in lane_blocks:
                if _seed_quoted_sender_is_ess(from_line, ess_team) is False:
                    req_ist = sent_ist
                    break
        if not req_ist:
            return None

        direct_reply_mode = ack_msg is None or ack_ist is None
        ack_pick = reply_ist if direct_reply_mode else ack_ist
        if not (req_ist <= ack_pick <= reply_ist):
            return None

        created_src = "PARSED_FROM_QUOTED_REQUEST"
        ack_src = (
            (reply_msg.sender_email or reply_msg.sender_name or "PARSED_FROM_QUOTED_REPLY")
            if direct_reply_mode
            else (ack_msg.sender_email or ack_msg.sender_name or "ESS_LANE_LOCAL_ACK")
        )
        resolved_src = reply_msg.sender_email or reply_msg.sender_name or "LANE_LOCAL_REPLY"
        mode = "direct-reply" if direct_reply_mode else "req-ack-reply"
        note = f"SeedPrimaryResolver[{mode}]"
        return {
            "times": TimeResult(
                _format_time(req_ist),
                _format_time(ack_pick),
                _format_time(reply_ist),
            ),
            "debug": TimeDebug(created_src, ack_src, resolved_src, note),
            "episode": {
                "created": req_ist,
                "response": ack_pick,
                "resolved": reply_ist,
                "mode": mode,
                "created_src": created_src,
                "ack_src": ack_src,
                "resolved_src": resolved_src,
                "authoritative": True,
            },
        }

    def _resolve_times_seeded_first(thread, requester_name, ess_team, subject_norm: str | None = None, description: str = ""):
        thread_sig = tuple(id(e) for e in (thread or ()))
        cache_key = (
            thread_sig,
            requester_name or "",
            subject_norm or "",
            description or "",
        )
        cached_hit = cache_key in _seeded_first_resolution_cache
        seeded = _seeded_first_resolution_cache.get(cache_key)
        if seeded is None:
            seeded = _seed_lane_episode_from_thread_local(
                thread,
                requester_name=requester_name,
                ess_team=ess_team,
                subject_norm_value=subject_norm or "",
                description=description or "",
            )
            _seeded_first_resolution_cache[cache_key] = seeded or False
        elif seeded is False:
            seeded = None
        else:
            seeded = _clone_seeded_resolution_local(seeded)
        if seeded:
            if not cached_hit:
                seeded = _clone_seeded_resolution_local(seeded)
            return seeded["times"], seeded["debug"], seeded["episode"]
        times, debug = resolve_times_with_debug(
            thread=thread,
            requester_name=requester_name,
            ess_team=ess_team,
            subject_norm=subject_norm,
        )
        return times, debug, None

    # Pre-scan template to count repeated subjects per consultant (safe, read-only)
    group_counts = {}
    group_counts_by_service = {}
    subject_service_keys = {}
    subject_family_rows = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(template_path)
        ws = select_target_sheet(wb, logger, preferred_sheet_name="LOG")
        header_row = _find_header_row(ws)
        col_map = _build_col_map(ws, header_row)
        desc_col = col_map.get("description")
        consultant_col = col_map.get("consultant") or col_map.get("requester")
        service_no_col = col_map.get("service no")
        if desc_col and consultant_col:
            for row in range(header_row + 1, ws.max_row + 1):
                desc_val = ws.cell(row, desc_col).value
                cons_val = ws.cell(row, consultant_col).value
                if not desc_val or not cons_val:
                    continue
                subject_text = _subject_for_description(str(desc_val))
                subject_norm = normalize_subject(subject_text)
                subject_key = (subject_norm or "").lower()
                if not subject_key:
                    continue
                if subject_exclusions and any(x in subject_key for x in subject_exclusions):
                    if "maintenance" not in subject_key:
                        continue
                service_key = _service_no_key(ws.cell(row, service_no_col).value) if service_no_col else ""
                key = (subject_key, _requester_key(cons_val))
                group_counts[key] = group_counts.get(key, 0) + 1
                subject_family_rows.setdefault(subject_key, []).append((row, service_key))
                if service_key:
                    subject_service_keys.setdefault(subject_key, set()).add(service_key)
                    service_key_tuple = (subject_key, _requester_key(cons_val), service_key)
                    group_counts_by_service[service_key_tuple] = group_counts_by_service.get(service_key_tuple, 0) + 1
        wb.close()
    except Exception as e:
        logger.log(f"[WARNING] Pre-scan for repeated subjects failed: {e}")
        group_counts = {}
        group_counts_by_service = {}
        subject_service_keys = {}
        subject_family_rows = {}

    def _subject_has_multiple_service_nos(subject_norm_value: str) -> bool:
        subject_key = (subject_norm_value or "").lower()
        return len(subject_service_keys.get(subject_key, set())) >= 2

    def _subject_service_bucket(subject_norm_value: str, service_no_value: str = "") -> str:
        subject_key = (subject_norm_value or "").lower()
        service_key = _service_no_key(service_no_value)
        if service_key and _subject_has_multiple_service_nos(subject_key):
            return service_key
        return ""

    def _occurrence_group_key(subject_norm_value: str, requester_value: str, service_no_value: str = ""):
        subject_key = (subject_norm_value or "").lower()
        requester_key = _requester_key(requester_value)
        service_bucket = _subject_service_bucket(subject_key, service_no_value)
        if service_bucket:
            return (subject_key, requester_key, service_bucket)
        return (subject_key, requester_key)

    def _occurrence_group_total(subject_norm_value: str, requester_value: str, service_no_value: str = "") -> int:
        subject_key = (subject_norm_value or "").lower()
        requester_key = _requester_key(requester_value)
        service_bucket = _subject_service_bucket(subject_key, service_no_value)
        if service_bucket:
            return group_counts_by_service.get((subject_key, requester_key, service_bucket), 0)
        return group_counts.get((subject_key, requester_key), 0)

    def _pre_resolve_family_slot(subject_norm_value: str, row_index_value) -> tuple[int, int]:
        subject_key = (subject_norm_value or "").lower()
        family_rows = sorted(subject_family_rows.get(subject_key, []), key=lambda x: x[0])
        if len({svc for _, svc in family_rows if svc}) < 2:
            return 0, 1
        slot_index = 0
        for idx, (row_num, _svc) in enumerate(family_rows):
            if row_num == row_index_value:
                slot_index = idx
                break
        return slot_index, len(family_rows) or 1

    def _bind_initial_thread_to_occurrence(thread, requester_value: str, subject_norm_value: str, row_index_value):
        def _bind_email_ist(email_record):
            sent = getattr(email_record, "sent_time", None)
            return _to_ist(sent) if sent else None
        if not thread or not requester_value or not _subject_has_multiple_service_nos(subject_norm_value):
            return thread, ""
        slot_index, family_total = _pre_resolve_family_slot(subject_norm_value, row_index_value)
        if family_total < 2:
            return thread, ""

        consultant_replies = [
            e for e in thread
            if getattr(e, "sent_time", None)
            and _match_requester(e.sender_name, e.sender_email, requester_value)
            and not _is_ack_like_reply(e)
            and not _is_thanks_info_reply(e)
        ]
        consultant_replies.sort(key=lambda e: e.sent_time)
        if len(consultant_replies) < 2:
            return thread, ""

        # Safest generic split: prefer distinct reply dates first. This catches
        # repeated same-subject rows that belong to different day-lanes without
        # inventing same-day episode separation.
        date_buckets = {}
        for e in consultant_replies:
            e_ist = _bind_email_ist(e)
            if not e_ist:
                continue
            date_buckets.setdefault(e_ist.date(), []).append(e)
        ordered_dates = sorted(date_buckets)
        if len(ordered_dates) >= family_total:
            chosen_date = ordered_dates[min(slot_index, len(ordered_dates) - 1)]
            bucket = date_buckets.get(chosen_date) or []
            if bucket:
                cutoff = bucket[-1].sent_time
                sliced = [e for e in thread if not getattr(e, "sent_time", None) or e.sent_time <= cutoff]
                if sliced:
                    return sliced, f"InitialOccurrenceLaneDate#{slot_index + 1}"

        unique_reply_minutes = []
        seen_minutes = set()
        for e in consultant_replies:
            e_ist = _bind_email_ist(e)
            if not e_ist:
                continue
            minute_key = e_ist.replace(second=0, microsecond=0)
            if minute_key in seen_minutes:
                continue
            seen_minutes.add(minute_key)
            unique_reply_minutes.append((minute_key, e))
        if len(unique_reply_minutes) >= family_total:
            pick_ist, pick_msg = unique_reply_minutes[min(slot_index, len(unique_reply_minutes) - 1)]
            cutoff = pick_msg.sent_time
            sliced = [e for e in thread if not getattr(e, "sent_time", None) or e.sent_time <= cutoff]
            if sliced:
                return sliced, f"InitialOccurrenceLaneReply#{slot_index + 1}"

        # Fallback: use wide chronology clusters only when they are clearly
        # separated by long gaps, which is still safe for repeated day-based lanes.
        clusters = []
        current = []
        prev_ist = None
        for e in consultant_replies:
            e_ist = _bind_email_ist(e)
            if not e_ist:
                continue
            if prev_ist and (e_ist - prev_ist) > timedelta(hours=48):
                if current:
                    clusters.append(current)
                current = [e]
            else:
                current.append(e)
            prev_ist = e_ist
        if current:
            clusters.append(current)
        if len(clusters) >= family_total:
            cluster = clusters[min(slot_index, len(clusters) - 1)]
            cutoff = cluster[-1].sent_time
            sliced = [e for e in thread if not getattr(e, "sent_time", None) or e.sent_time <= cutoff]
            if sliced:
                return sliced, f"InitialOccurrenceLaneCluster#{slot_index + 1}"

        return thread, ""

    episode_counters = {}
    duplicate_group_state = {}
    created_history = []
    env_cache = {}
    _env_consultant_text_cache = {}
    _env_thread_text_cache = {}
    stage_times_enabled = os.getenv("DEBUG_STAGE_TIMES", "0") == "1"
    stage_time_stats = {}

    def _stage_timer_start():
        return time.perf_counter() if stage_times_enabled else None

    def _stage_timer_stop(name: str, started_at, *, items: int = 0):
        if not stage_times_enabled or started_at is None:
            return
        stat = stage_time_stats.setdefault(name, {"seconds": 0.0, "calls": 0, "items": 0})
        stat["seconds"] += max(0.0, time.perf_counter() - started_at)
        stat["calls"] += 1
        stat["items"] += max(0, items)

    # Build deployment request/success index by DR ID (safe override case only)
    deployment_index = {}
    for key, thread in threads.items():
        subj_l = (key or "").lower()
        is_req = _is_deployment_request_subject(subj_l)
        is_succ = _is_deployment_success_subject(subj_l)
        if not (is_req or is_succ):
            continue
        dr_ids = _thread_dr_ids(thread)
        if not dr_ids:
            continue
        info = {
            "thread": thread,
            "subject_key": key,
            "iface": _iface_tokens(key),
        }
        for dr in dr_ids:
            bucket = deployment_index.setdefault(dr, {"request": [], "success": []})
            if is_req:
                bucket["request"].append(info)
            else:
                bucket["success"].append(info)

    def _find_iface_specific_thread(subject_norm, requester, date_tokens=None, iface_tokens=None, baseline_date=None):
        iface_tokens = iface_tokens or set()
        if not subject_norm or not iface_tokens:
            return None
        subj_tokens = _match_tokens(subject_norm)
        if not subj_tokens:
            return None
        short_stop = {
            "re", "fw", "fwd", "aw", "wg", "sv",
            "in", "on", "at", "to", "of", "for", "not",
            "and", "or", "the", "a", "an", "is", "it",
            "by", "as", "if", "no", "we", "us",
            "today", "below", "details", "file", "files",
            "input", "output", "received", "failed",
        }
        subj_short = {t for t in subj_tokens if len(t) <= 3 and t.isalpha() and t not in short_stop}

        ranked = []
        for key, thread in threads.items():
            key_iface_set = _interface_tokens(key)
            if not key_iface_set or key_iface_set.isdisjoint(iface_tokens):
                continue
            key_tokens = _match_tokens(key)
            if not key_tokens:
                continue
            score = _token_overlap_score(subj_tokens, key_tokens)
            contains = (subject_norm in key or key in subject_norm)
            if score < 0.35 and not contains:
                continue
            short_overlap = len(subj_short & key_tokens) if subj_short else 0

            has_consultant_date = False
            if date_tokens and requester:
                has_consultant_date = _thread_has_consultant_on_or_near_date_cached(key, thread, date_tokens, requester)

            delta_days = 9999
            if baseline_date and requester:
                d = _thread_baseline_delta_cached(key, thread, requester, baseline_date)
                if d is not None:
                    delta_days = d

            ranked.append((0 if has_consultant_date else 1, -short_overlap, delta_days, -score, key, thread))

        if not ranked:
            return None
        if subj_short:
            max_short = max((-r[1]) for r in ranked)
            if max_short > 0:
                ranked = [r for r in ranked if (-r[1]) == max_short]
        ranked.sort()
        best = ranked[0]
        return best[5], f"IfaceSpecific:{best[4]}"

    def find_thread(subject_norm, requester, date_tokens=None, prefer_consultant_date=False, iface_tokens=None, baseline_date=None):
        if not subject_norm:
            return [], "Empty subject"
        iface_tokens = iface_tokens or set()
        alt_subject = normalize_subject_for_match(subject_norm)
        enable_safe_iface_hint = os.getenv("SAFE_IFACE_HINT", "1") == "1"

        def _strong_exact_signal(thread_key, thread_items):
            if not thread_items:
                return False
            same_family = _same_subject_family(subject_norm, thread_key)
            strong_identity = _strong_identity_overlap(subject_norm, thread_key)
            if not same_family and not strong_identity:
                return False
            if requester and _thread_has_requester_cached(thread_key, thread_items, requester):
                return True
            if date_tokens and requester and _thread_has_consultant_on_or_near_date_cached(thread_key, thread_items, date_tokens, requester):
                return True
            if baseline_date and requester:
                d = _thread_baseline_delta_cached(thread_key, thread_items, requester, baseline_date)
                if d is not None and d <= 2:
                    return True
            return False

        def _prefer_iface_pick(exact_t, iface_t):
            if not exact_t or not iface_t:
                return False
            if requester:
                exact_has_req = _thread_has_requester(exact_t, requester)
                iface_has_req = _thread_has_requester(iface_t, requester)
                if iface_has_req and not exact_has_req:
                    return True
            if date_tokens and requester:
                exact_has_date = _thread_has_consultant_on_or_near_date(exact_t, date_tokens, requester)
                iface_has_date = _thread_has_consultant_on_or_near_date(iface_t, date_tokens, requester)
                if iface_has_date and not exact_has_date:
                    return True
            if baseline_date and requester:
                exact_delta = _requester_min_delta_days(exact_t, requester, baseline_date)
                iface_delta = _requester_min_delta_days(iface_t, requester, baseline_date)
                if iface_delta is not None and (exact_delta is None or (iface_delta + 1) < exact_delta):
                    return True
            return False

        def _filter_thread_subject_identity(thread_items):
            if not thread_items:
                return thread_items
            filtered = []
            comparable = 0
            for e in thread_items:
                e_subject_norm = normalize_subject(getattr(e, "subject", "") or "")
                if not e_subject_norm:
                    filtered.append(e)
                    continue
                comparable += 1
                if _strict_subject_identity_gate(
                    subject_norm,
                    e_subject_norm,
                    iface_tokens=iface_tokens,
                ):
                    filtered.append(e)
            if filtered:
                return filtered
            if comparable <= 0:
                return thread_items
            return []

        if subject_norm in threads:
            t = _filter_thread_subject_identity(threads[subject_norm])
            key_iface_set = _interface_tokens(subject_norm)
            if iface_tokens and not key_iface_set and enable_safe_iface_hint:
                iface_pick = _find_iface_specific_thread(
                    subject_norm,
                    requester,
                    date_tokens=date_tokens,
                    iface_tokens=iface_tokens,
                    baseline_date=baseline_date,
                )
                if iface_pick:
                    iface_thread, _iface_note = iface_pick
                    if (not _strong_exact_signal(subject_norm, t)) or _prefer_iface_pick(t, iface_thread):
                        return iface_pick
            # Interface token is a hint, not a blocker for exact subject match.
            # Keep exact matches even when the subject has no interface token.
            if (not iface_tokens) or (not key_iface_set) or (not iface_tokens.isdisjoint(key_iface_set)):
                if prefer_consultant_date and date_tokens and requester and not _thread_has_requester(t, requester):
                    alt = _find_unique_requester_date_thread(subject_norm, requester, date_tokens, iface_tokens)
                    if alt:
                        return alt
                    alt = _find_best_requester_date_thread(subject_norm, requester, date_tokens, iface_tokens)
                    if alt:
                        return alt
                    # Last-resort union of alt-subject threads when requester is missing
                    if alt_subject:
                        union = _union_alt_threads(alt_subject, requester, subject_norm=subject_norm, iface_tokens=iface_tokens)
                        if union:
                            return union, f"AltUnion:{alt_subject}"
            if t:
                return t, "Exact"
        if alt_subject in threads:
            t = _filter_thread_subject_identity(threads[alt_subject])
            key_iface_set = _interface_tokens(alt_subject)
            if iface_tokens and not key_iface_set and enable_safe_iface_hint:
                iface_pick = _find_iface_specific_thread(
                    subject_norm,
                    requester,
                    date_tokens=date_tokens,
                    iface_tokens=iface_tokens,
                    baseline_date=baseline_date,
                )
                if iface_pick:
                    iface_thread, _iface_note = iface_pick
                    if (not _strong_exact_signal(alt_subject, t)) or _prefer_iface_pick(t, iface_thread):
                        return iface_pick
            # Interface token is a hint, not a blocker for exact alt-subject match.
            if (not iface_tokens) or (not key_iface_set) or (not iface_tokens.isdisjoint(key_iface_set)):
                if prefer_consultant_date and date_tokens and requester and not _thread_has_requester(t, requester):
                    alt = _find_unique_requester_date_thread(subject_norm, requester, date_tokens, iface_tokens)
                    if alt:
                        return alt
                    alt = _find_best_requester_date_thread(subject_norm, requester, date_tokens, iface_tokens)
                    if alt:
                        return alt
                    # Last-resort union of alt-subject threads when requester is missing
                    if alt_subject:
                        union = _union_alt_threads(alt_subject, requester, subject_norm=subject_norm, iface_tokens=iface_tokens)
                        if union:
                            return union, f"AltUnion:{alt_subject}"
            if t:
                return t, "AltExact"

        subj_tokens = _match_tokens(subject_norm)
        if not subj_tokens:
            return [], "No tokens"

        date_only = _looks_like_date_only(subject_norm)
        subj_prefix = _interface_prefix(subject_norm)
        subj_iface_set = _interface_tokens(subject_norm)
        if not subj_iface_set and iface_tokens:
            subj_iface_set = iface_tokens
        subj_inc_set = _inc_tokens(subject_norm)
        subj_num_set = _sig_num_tokens(subject_norm)

        subj_part_set = _part_tokens(subject_norm)
        short_stop = {
            "re", "fw", "fwd", "aw", "wg", "sv",
            "in", "on", "at", "to", "of", "for", "not",
            "and", "or", "the", "a", "an", "is", "it",
            "by", "as", "if", "no", "we", "us",
        }
        short_tokens = {t for t in subj_tokens if len(t) <= 3 and t.isalpha() and t not in short_stop}

        candidates = []
        for key, thread in threads.items():
            key_tokens = _match_tokens(key)
            if not key_tokens:
                continue
            if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=subj_iface_set or iface_tokens):
                continue
            thread = _filter_thread_subject_identity(thread)
            candidate_anchor = _strong_exact_signal(key, thread)

            key_prefix = _interface_prefix(key)
            if subj_prefix and key_prefix and subj_prefix != key_prefix:
                continue

            key_iface_set = _interface_tokens(key)
            if subj_iface_set and key_iface_set and subj_iface_set.isdisjoint(key_iface_set):
                continue

            key_inc_set = _inc_tokens(key)
            if subj_inc_set:
                # Require INC match when subject has INC
                if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                    continue
            elif _defer_added_inc_identity(subj_inc_set, key_inc_set):
                continue
            key_num_set = _sig_num_tokens(key)
            if subj_num_set and key_num_set and subj_num_set.isdisjoint(key_num_set):
                continue

            score = _token_overlap_score(subj_tokens, key_tokens)
            soft_score = _soft_subject_similarity(subject_norm, key)
            family_score = _family_subject_score(
                subject_norm,
                key,
                thread,
                requester=requester,
                date_tokens=date_tokens,
                iface_tokens=subj_iface_set or iface_tokens,
                baseline_date=baseline_date,
            )
            contains = (subject_norm in key or key in subject_norm)
            boost = 0.0
            if subj_inc_set and key_inc_set and not subj_inc_set.isdisjoint(key_inc_set):
                boost += 0.2
            if subj_iface_set and key_iface_set and not subj_iface_set.isdisjoint(key_iface_set):
                boost += 0.2
            if contains:
                boost += 0.1
            if short_tokens:
                short_hit = len(short_tokens & key_tokens)
                if short_hit:
                    boost += min(0.15, 0.05 * short_hit)
            score = min(1.0, max(score, soft_score, family_score) + boost)
            if score < 0.55 and not contains and not candidate_anchor:
                continue

            if date_only:
                # Date-only subjects are high-risk; require stronger signal
                if score >= 0.8 or (subj_prefix and key_prefix == subj_prefix and score >= 0.5):
                    candidates.append((score, key, thread, "Score"))
                continue

            if score >= 0.55:
                candidates.append((score, key, thread, "Score"))
            elif short_tokens and (short_tokens & key_tokens) and score >= 0.35:
                # Keep near-miss siblings that match distinguishing short tokens
                # (e.g., INT/GB) so VariantTok can disambiguate correctly.
                candidates.append((score, key, thread, "ScoreShort"))
            elif contains and not date_only and len(subject_norm) >= 12 and score >= 0.35:
                candidates.append((score, key, thread, "Contains"))

        if alt_subject and alt_subject in alt_index:
            for key in alt_index[alt_subject]:
                if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=subj_iface_set or iface_tokens):
                    continue
                thread = _filter_thread_subject_identity(threads.get(key, []))
                candidates.append((0.95, key, thread, "AltKey"))

        # If the subject has INC IDs, prefer exact INC set matches.
        # This prevents matching a superset INC subject when an exact one exists.
        if subj_inc_set and candidates:
            exact_inc = []
            for score, key, thread, note in candidates:
                key_inc_set = _inc_tokens(key)
                if key_inc_set == subj_inc_set:
                    exact_inc.append((score, key, thread, note + "+IncExact"))
            if exact_inc:
                candidates = exact_inc

        # If we have an interface prefix hint (from description), prefer
        # threads that include those interface tokens.
        if iface_tokens and candidates:
            iface_candidates = []
            for score, key, thread, note in candidates:
                key_iface_set = _interface_tokens(key)
                if key_iface_set and not iface_tokens.isdisjoint(key_iface_set):
                    iface_candidates.append((score, key, thread, note + "+Iface"))
            if iface_candidates:
                candidates = iface_candidates

        if subj_part_set and candidates:
            part_candidates = []
            for score, key, thread, note in candidates:
                if _part_tokens(key) & subj_part_set:
                    part_candidates.append((score, key, thread, note + '+Part'))
            if part_candidates:
                candidates = part_candidates

        # Variant-token disambiguation for sibling subjects (e.g., INT vs GB).
        # Keep this non-destructive: boost scores and only hard-pick when one
        # candidate is uniquely best on variant-token overlap.
        if len(candidates) > 1:
            variant_stop = {
                "re", "fw", "fwd", "aw", "wg", "sv",
                "in", "on", "at", "to", "of", "for", "not",
                "and", "or", "the", "a", "an", "is", "it",
                "by", "as", "if", "no", "we", "us",
                "today", "below", "details", "file", "files",
                "input", "output", "received", "failed",
            }
            cand_with_tokens = []
            token_counts = {}
            for score, key, thread, note in candidates:
                key_tokens = _match_tokens(key)
                cand_with_tokens.append((score, key, thread, note, key_tokens))
                for t in key_tokens:
                    token_counts[t] = token_counts.get(t, 0) + 1

            varying_tokens = {
                t for t in subj_tokens
                if t not in variant_stop and 0 < token_counts.get(t, 0) < len(candidates)
            }
            if varying_tokens:
                best_overlap = 0
                winners = []
                boosted = []
                for score, key, thread, note, key_tokens in cand_with_tokens:
                    overlap = len(varying_tokens & key_tokens)
                    if overlap > best_overlap:
                        best_overlap = overlap
                        winners = [(score, key, thread, note)]
                    elif overlap == best_overlap:
                        winners.append((score, key, thread, note))

                    bonus = min(0.18, 0.06 * overlap) if overlap > 0 else 0.0
                    boosted_score = min(1.0, score + bonus)
                    boosted_note = note + "+VariantTok" if overlap > 0 else note
                    boosted.append((boosted_score, key, thread, boosted_note))

                # Only force a pick when variant-token evidence has a unique winner.
                if best_overlap > 0 and len(winners) == 1:
                    _, key, thread, note = winners[0]
                    return thread, f"{note}+VariantTokUnique:{key}"

                # Otherwise keep all candidates and only adjust scores/notes.
                candidates = boosted

        # (Short-token disambiguation is handled via score boosting above)

        # If subject carries distinguishing short tokens but none of the current
        # candidates contain them, run a global variant scan to avoid sibling drift
        # (e.g., selecting GB thread for an INT subject).
        if short_tokens and candidates:
            cand_has_variant = False
            for _, key, _, _ in candidates:
                if short_tokens & _match_tokens(key):
                    cand_has_variant = True
                    break
            if not cand_has_variant:
                variant_global = []
                for key, thread in threads.items():
                    key_tokens = _match_tokens(key)
                    if not (short_tokens & key_tokens):
                        continue

                    key_prefix = _interface_prefix(key)
                    if subj_prefix and key_prefix and subj_prefix != key_prefix:
                        continue

                    key_iface_set = _interface_tokens(key)
                    if subj_iface_set and key_iface_set and subj_iface_set.isdisjoint(key_iface_set):
                        continue

                    key_inc_set = _inc_tokens(key)
                    if subj_inc_set and (not key_inc_set or subj_inc_set.isdisjoint(key_inc_set)):
                        continue
                    if (not subj_inc_set) and _defer_added_inc_identity(subj_inc_set, key_inc_set):
                        continue

                    score = _token_overlap_score(subj_tokens, key_tokens)
                    contains = (subject_norm in key or key in subject_norm)
                    if score < 0.35 and not contains:
                        continue
                    variant_overlap = len(short_tokens & key_tokens)
                    variant_global.append((variant_overlap, score, key, thread))

                if variant_global:
                    variant_global.sort(key=lambda x: (-x[0], -x[1], -len(x[2])))
                    best = variant_global[0]
                    second = variant_global[1] if len(variant_global) > 1 else None
                    if second is None or best[0] > second[0] or (best[0] == second[0] and (best[1] - second[1]) >= 0.08):
                        return best[3], f"VariantGlobal:{best[2]}"

        if not candidates and subj_inc_set:
            inc_upper = {i.upper() for i in subj_inc_set}
            inc_threads = []
            for inc in inc_upper:
                for key in inc_index.get(inc, []):
                    thread = threads.get(key, [])
                    if thread:
                        inc_threads.append((key, thread))

            # Only accept if there is a single unique thread (safe fallback)
            unique = {(key, id(thread)): thread for key, thread in inc_threads}
            if len(unique) == 1:
                only_thread = next(iter(unique.values()))
                return only_thread, "INCBodyOnly"

            if len(unique) > 1 and requester:
                req = (requester or "").strip().lower()
                for key, thread in inc_threads:
                    for e in thread:
                        if req in (e.sender_name or "").lower() or req in (e.sender_email or "").lower():
                            return thread, f"INCBodyOnly+Requester:{key}"

        if not candidates and not subj_inc_set:
            inc_fallback = []
            for key, thread in threads.items():
                key_tokens = _match_tokens(key)
                if not key_tokens:
                    continue
                key_inc_set = _inc_tokens(key)
                if not key_inc_set:
                    continue

                key_prefix = _interface_prefix(key)
                if subj_prefix and key_prefix and subj_prefix != key_prefix:
                    continue

                key_iface_set = _interface_tokens(key)
                if subj_iface_set and key_iface_set and subj_iface_set.isdisjoint(key_iface_set):
                    continue

                key_num_set = _sig_num_tokens(key)
                if subj_num_set and key_num_set and subj_num_set.isdisjoint(key_num_set):
                    continue

                score = _token_overlap_score(subj_tokens, key_tokens)
                contains = (subject_norm in key or key in subject_norm)
                if score < 0.55 and not contains:
                    continue
                inc_fallback.append((score, key, thread, "IncFallback"))
            if inc_fallback:
                candidates = inc_fallback

        if not candidates:
            hidden_rescue = _find_hidden_subject_rescue(
                subject_norm,
                requester,
                date_tokens=date_tokens,
                iface_tokens=subj_iface_set or iface_tokens,
                baseline_date=baseline_date,
            )
            if hidden_rescue:
                return hidden_rescue
            return [], "No match"

        # Prefer candidates that actually contain a non-ESS request email.
        # This avoids matching ESS-only threads when a real request exists elsewhere.
        def _thread_has_non_ess(thread_key_or_thread, t=None):
            thread_key = thread_key_or_thread if t is not None else ""
            thread = t if t is not None else thread_key_or_thread
            if not thread:
                return False
            return _thread_has_non_ess_cached(thread_key, thread)

        non_ess_candidates = [c for c in candidates if _thread_has_non_ess(c[1], c[2])]
        if non_ess_candidates:
            # When the row has an explicit date marker and consultant, keep
            # ESS-only candidates so we don't drop a valid consultant reply
            # thread for date-anchored rows.
            if not (prefer_consultant_date and date_tokens and requester):
                candidates = non_ess_candidates

        # If we still don't have a consultant-on-date candidate, do a global
        # scan for explicit date markers and pick the best thread that has a
        # requester reply on/near that date.
        if prefer_consultant_date and date_tokens and requester:
            has_consultant_date = False
            for _, _, thread, _ in candidates:
                if _thread_has_consultant_on_or_near_date(thread, date_tokens, requester):
                    has_consultant_date = True
                    break
            if not has_consultant_date:
                subj_tokens = _match_tokens(subject_norm)
                subj_inc_set = _inc_tokens(subject_norm)
                subj_num_set = _sig_num_tokens(subject_norm)
                best = None
                # Keep soft RE/FW-family expansion limited to the main finder.
                # This later global rescue should stay on normalized thread
                # keys, otherwise fallback selection gets wider than intended.
                pool = list(threads.items())
                for key, thread in pool:
                    if not _thread_has_consultant_on_or_near_date(thread, date_tokens, requester):
                        continue
                    if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=subj_iface_set or iface_tokens):
                        continue
                    key_tokens = _match_tokens(key)
                    if not key_tokens:
                        continue
                    if subj_inc_set:
                        key_inc_set = _inc_tokens(key)
                        if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                            continue
                    else:
                        key_inc_set = _inc_tokens(key)
                        if _defer_added_inc_identity(subj_inc_set, key_inc_set):
                            continue
                    key_num_set = _sig_num_tokens(key)
                    if subj_num_set and key_num_set and subj_num_set.isdisjoint(key_num_set):
                        continue
                    score = _token_overlap_score(subj_tokens, key_tokens)
                    contains = (subject_norm in key or key in subject_norm)
                    if score < 0.4 and not contains:
                        continue
                    candidate = (-score, key, thread)
                    if best is None or candidate < best:
                        best = candidate
                if best:
                    _, key, thread = best
                    return thread, f"ConsultantDateGlobal:{key}"

        # If we have an explicit date marker and a requester, prefer threads that
        # contain a consultant reply on that date (whole mail chain search).
        if prefer_consultant_date and date_tokens and requester and len(candidates) > 1:
            consultant_date_candidates = []
            for score, key, thread, note in candidates:
                if _thread_has_consultant_on_or_near_date(thread, date_tokens, requester):
                    consultant_date_candidates.append((score, key, thread, f"{note}+ConsultantDate"))
            if consultant_date_candidates:
                candidates = consultant_date_candidates

        # If the subject contains an explicit date and we still have multiple candidates,
        # prefer threads that contain the same date token in subject/body.
        if date_tokens and len(candidates) > 1:
            date_matches = []
            for score, key, thread, note in candidates:
                if _thread_has_date_token(thread, date_tokens):
                    date_matches.append((score, key, thread, f"{note}+Date"))
            if date_matches:
                candidates = date_matches

        # Keep only the strongest remaining families after the date/requester
        # narrowing above, so we don't drop good candidates too early.
        if len(candidates) > 8:
            ranked_candidates = []
            for score, key, thread, note in candidates:
                family_score = _family_subject_score(
                    subject_norm,
                    key,
                    thread,
                    requester=requester,
                    date_tokens=date_tokens,
                    iface_tokens=subj_iface_set or iface_tokens,
                    baseline_date=baseline_date,
                )
                ranked_candidates.append((family_score, score, key, thread, note))
            ranked_candidates.sort(key=lambda x: (-x[0], -x[1], -len(x[2])))
            best_family_score = ranked_candidates[0][0]
            shortlist = [
                (score, key, thread, f"{note}+Family")
                for family_score, score, key, thread, note in ranked_candidates
                if family_score >= (best_family_score - 0.10)
            ]
            if len(shortlist) < min(len(ranked_candidates), 8):
                shortlist = [
                    (score, key, thread, f"{note}+Family")
                    for family_score, score, key, thread, note in ranked_candidates[:8]
                ]
            candidates = shortlist

        # Compute distinguishing subject tokens over the current candidate set.
        # These tokens are used as compatibility signals for requester-based
        # resolution and low-score fallback disambiguation.
        variant_stop2 = {
            "re", "fw", "fwd", "aw", "wg", "sv",
            "in", "on", "at", "to", "of", "for", "not",
            "and", "or", "the", "a", "an", "is", "it",
            "by", "as", "if", "no", "we", "us",
            "today", "below", "details", "file", "files",
            "input", "output", "received", "failed",
        }
        cand_token_map = {}
        cand_counts = {}
        for _, key, _, _ in candidates:
            toks = _match_tokens(key)
            cand_token_map[key] = toks
            for t in toks:
                cand_counts[t] = cand_counts.get(t, 0) + 1
        discr_tokens = {
            t for t in subj_tokens
            if t not in variant_stop2 and 0 < cand_counts.get(t, 0) < len(candidates)
        }
        max_discr_overlap = 0
        if discr_tokens:
            for _, key, _, _ in candidates:
                ov = len(discr_tokens & cand_token_map.get(key, set()))
                if ov > max_discr_overlap:
                    max_discr_overlap = ov
        max_short_overlap = 0
        if short_tokens:
            for _, key, _, _ in candidates:
                ov = len(short_tokens & cand_token_map.get(key, set()))
                if ov > max_short_overlap:
                    max_short_overlap = ov

        # Resolve by requester only if it appears in exactly one compatible candidate.
        req = (requester or "").strip().lower()
        if req:
            req_hits = []
            for _, key, thread, _ in sorted(candidates, key=lambda x: (-x[0], -len(x[1]))):
                if short_tokens and max_short_overlap > 0:
                    ov_short = len(short_tokens & cand_token_map.get(key, set()))
                    if ov_short < max_short_overlap:
                        continue
                if any(req in (e.sender_name or "").lower() or req in (e.sender_email or "").lower() for e in thread):
                    if discr_tokens and max_discr_overlap > 0:
                        ov = len(discr_tokens & cand_token_map.get(key, set()))
                        if ov < max_discr_overlap:
                            continue
                    req_hits.append((key, thread))
            if len(req_hits) == 1:
                key, thread = req_hits[0]
                return thread, f"AmbiguousResolvedByRequester:{key}"

        # Pick best scored candidate
        candidates.sort(key=lambda x: (-x[0], -len(x[1])))
        top = candidates[0]
        if len(candidates) == 1:
            return top[2], f"{top[3]}:{top[1]}"

        # Smart low-score fallback: prefer candidate with best combined evidence
        # instead of returning unknown too aggressively.
        if len(candidates) > 1 and top[0] < 0.75:
            def _evidence(c):
                score, key, thread, _note = c
                ev = score
                if req and any(req in (e.sender_name or "").lower() or req in (e.sender_email or "").lower() for e in thread):
                    ev += 0.08
                if date_tokens and _thread_has_date_token(thread, date_tokens):
                    ev += 0.06
                if discr_tokens and max_discr_overlap > 0:
                    ov = len(discr_tokens & cand_token_map.get(key, set()))
                    ev += (0.10 * ov / max_discr_overlap)
                if _thread_has_non_ess(thread):
                    ev += 0.02
                return ev

            ranked = sorted(candidates, key=lambda c: (-_evidence(c), -len(c[1])))
            best = ranked[0]
            second = ranked[1] if len(ranked) > 1 else None
            best_ev = _evidence(best)
            second_ev = _evidence(second) if second else -1

            if (best_ev - second_ev) >= 0.05:
                return best[2], f"{best[3]}+SmartFallback:{best[1]}"

            if discr_tokens and max_discr_overlap > 0:
                best_ov = len(discr_tokens & cand_token_map.get(best[1], set()))
                second_ov = len(discr_tokens & cand_token_map.get(second[1], set())) if second else -1
                if best_ov > second_ov:
                    return best[2], f"{best[3]}+SmartFallback:{best[1]}"

            return [], f"Ambiguous:{len(candidates)}"

        return top[2], f"{top[3]}:{top[1]}"

    def resolve_row(row_context):
        nonlocal created_history
        skip_history_update = False
        description = row_context.get("Description", "")
        requester = row_context.get("Consultant", "") or row_context.get("Requester", "")
        service_no = row_context.get("Service No", "") or ""
        category_raw = row_context.get("Category", "")
        category_type_raw = row_context.get("Category Type", "")
        category_type = category_type_raw
        row_index = row_context.get("RowIndex")
        # Defaults used in row_states even when deployment override fires
        date_anchor_missing = False
        date_anchor_after = False
        base_times = None
        base_debug = None
        _seeded_won = False
        group_key = ("", "")
        group_total = 0

        subject_text = _subject_for_description(description)
        group_key = _occurrence_group_key(normalize_subject(subject_text), requester, service_no)
        group_total = _occurrence_group_total(normalize_subject(subject_text), requester, service_no)
        # Interface tokens from description prefix help disambiguate
        # when the subject text itself loses the interface prefix.
        desc_prefix = ""
        if description and re.search(r"(?:--\.?>|â†’|âž”|âž¡|Ã¢â€ â€™|Ã¢Å¾â€|Ã¢Å¾Â¡|=>)", description):
            parts = re.split(r"\s*(?:--\.?>|â†’|âž”|âž¡|Ã¢â€ â€™|Ã¢Å¾â€|Ã¢Å¾Â¡|=>)\s*", description, maxsplit=1)
            if len(parts) >= 2:
                desc_prefix = parts[0].strip()
        iface_hint_tokens = _interface_tokens(desc_prefix) if desc_prefix else set()
        subject_norm = normalize_subject(subject_text)
        raw_subject_norm = normalize_subject(description or "")
        date_tokens = _extract_date_tokens(subject_text)
        date_tokens += _extract_date_tokens_from_description(description)
        # de-dup while preserving order
        if date_tokens:
            seen = set()
            deduped = []
            for t in date_tokens:
                if t in seen:
                    continue
                seen.add(t)
                deduped.append(t)
            date_tokens = deduped
        explicit_marker = _has_explicit_date_marker(subject_text) or _has_explicit_date_marker(description)
        stale_anchor = _is_stale_anchor_date(date_tokens, _baseline_row_created_value(row_context))
        if stale_anchor:
            # Subject contains a historical date token (common in long mail chains).
            # Ignore it for matching/anchoring to avoid jumping to an old sibling thread.
            date_tokens = []
            explicit_marker = False
        baseline_created_date = _coerce_row_created_date(_baseline_row_created_value(row_context))
        thread = None
        match_note = ""
        if raw_subject_norm and raw_subject_norm != subject_norm and raw_subject_norm in threads:
            raw_thread = threads.get(raw_subject_norm) or []
            if (not requester) or _thread_has_requester(raw_thread, requester):
                thread = raw_thread
                match_note = "RowExact"
        deployment_like_row = _is_deployment_request_subject(subject_text or "") or _is_deployment_success_subject(subject_text or "")
        if thread is None:
            thread, match_note = find_thread(
                subject_norm,
                requester,
                date_tokens=date_tokens,
                prefer_consultant_date=explicit_marker,
                iface_tokens=iface_hint_tokens,
                baseline_date=baseline_created_date,
            )
        # Thread safety refinement:
        # if selected thread does not contain requester replies, try a requester-backed
        # candidate before resolving times. This prevents "No ESS/requester replies"
        # on wrong sibling threads.
        if (not deployment_like_row) and thread and requester and not _thread_has_requester(thread, requester):
            refined_thread = None
            refined_note = ""
            if date_tokens:
                alt = _find_unique_requester_date_thread(
                    subject_norm,
                    requester,
                    date_tokens,
                    iface_tokens=iface_hint_tokens,
                )
                if alt:
                    refined_thread, refined_note = alt
            if not refined_thread and date_tokens:
                alt = _find_best_requester_date_thread(
                    subject_norm,
                    requester,
                    date_tokens,
                    iface_tokens=iface_hint_tokens,
                )
                if alt:
                    refined_thread, refined_note = alt
            if not refined_thread and baseline_created_date:
                refined = _find_best_thread_near_baseline(
                    subject_norm,
                    requester,
                    baseline_created_date,
                    iface_tokens=iface_hint_tokens,
                )
                if refined:
                    refined_thread, refined_note, _ = refined
            if not refined_thread and requester:
                subj_tokens = _match_tokens(subject_norm)
                subj_inc_set = _inc_tokens(subject_norm)
                best = None
                for key, cand_thread in threads.items():
                    if not _thread_has_requester(cand_thread, requester):
                        continue
                    if not _fresh_picker_subject_safe(subject_norm, key, iface_tokens=iface_hint_tokens):
                        continue
                    key_tokens = _match_tokens(key)
                    if not key_tokens:
                        continue
                    if subj_inc_set:
                        key_inc_set = _inc_tokens(key)
                        if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                            continue
                    else:
                        key_inc_set = _inc_tokens(key)
                        if _defer_added_inc_identity(subj_inc_set, key_inc_set):
                            continue
                    if iface_hint_tokens:
                        key_iface_set = _interface_tokens(key)
                        if key_iface_set and iface_hint_tokens.isdisjoint(key_iface_set):
                            continue
                    score = _token_overlap_score(subj_tokens, key_tokens) if subj_tokens else 0.0
                    contains = bool(subject_norm and (subject_norm in key or key in subject_norm))
                    # Keep this less strict than initial thread match because this
                    # block runs only after we detected "no requester in selected
                    # thread" and must recover from sibling drift.
                    if score < 0.30 and not contains:
                        continue
                    delta = _requester_min_delta_days(cand_thread, requester, baseline_created_date) if baseline_created_date else 9999
                    if delta is None:
                        delta = 9999
                    candidate = (delta, -score, -len(key_tokens), key, cand_thread)
                    if best is None or candidate < best:
                        best = candidate
                if best:
                    _delta, _neg_score, _neg_len, best_key, best_thread = best
                    refined_thread = best_thread
                    refined_note = f"RequesterRefined:{best_key}"
            if refined_thread and refined_thread is not thread:
                thread = refined_thread
                match_note = f"{match_note}; {refined_note}" if match_note else refined_note
            elif not refined_thread:
                # Keep explicit trace when requester-backed recovery was not found.
                match_note = f"{match_note}; NoRequesterThreadRecovery" if match_note else "NoRequesterThreadRecovery"
        # Baseline thread refinement (safe):
        # If selected thread's requester episode is far from ServiceNow baseline,
        # prefer a closer requester-backed sibling thread.
        # Keep this narrow to avoid disturbing correctly matched rows.
        if (not deployment_like_row) and thread and requester and baseline_created_date:
            current_delta = _requester_min_delta_days(thread, requester, baseline_created_date)
            match_note_l = (match_note or "").lower()
            ambiguous_match = (
                ("score" in match_note_l)
                or ("ambiguous" in match_note_l)
                or ("norequesterthreadrecovery" in match_note_l)
            )
            needs_tight_refine = bool(explicit_marker or date_tokens or ambiguous_match)
            # Default threshold remains conservative, but tighten to >1 day for
            # ambiguous/date-driven rows where 1-2 day drift is still a real miss.
            refine_needed = (
                current_delta is None
                or current_delta > 7
                or (needs_tight_refine and current_delta > 1)
            )
            if refine_needed:
                refined = _find_best_thread_near_baseline(
                    subject_norm,
                    requester,
                    baseline_created_date,
                    iface_tokens=iface_hint_tokens,
                )
                if refined:
                    new_thread, refine_note, new_delta = refined
                    refined_key = refine_note.split(":", 1)[1] if ":" in refine_note else ""
                    if (current_delta is None or new_delta + 1 < current_delta) and _baseline_refine_safe(subject_norm, match_note, refined_key):
                        thread = new_thread
                        match_note = f"{match_note}; BaselineThreadRefined:{refine_note}" if match_note else f"BaselineThreadRefined:{refine_note}"

        # When a stale date marker existed, do one stricter baseline refinement.
        if (not deployment_like_row) and thread and requester and baseline_created_date and stale_anchor:
            current_delta = _requester_min_delta_days(thread, requester, baseline_created_date)
            if current_delta is None or current_delta > 3:
                refined = _find_best_thread_near_baseline(
                    subject_norm,
                    requester,
                    baseline_created_date,
                    iface_tokens=iface_hint_tokens,
                )
                if refined:
                    new_thread, refine_note, new_delta = refined
                    refined_key = refine_note.split(":", 1)[1] if ":" in refine_note else ""
                    if (current_delta is None or new_delta + 1 < current_delta) and _baseline_refine_safe(subject_norm, match_note, refined_key):
                        thread = new_thread
                        match_note = f"{match_note}; {refine_note}"

        full_thread = thread
        initial_occurrence_note = ""
        if thread and requester and row_index:
            bound_thread, bind_note = _bind_initial_thread_to_occurrence(
                thread,
                requester,
                subject_norm,
                row_index,
            )
            if bound_thread:
                thread = bound_thread
            initial_occurrence_note = bind_note or ""

        thread_key = (
            id(full_thread) if full_thread else 0,
            len(full_thread) if full_thread else 0,
        )
        env_cache_key = (
            (subject_text or "").strip().lower(),
            _requester_key(requester),
            thread_key,
            (description or "").strip().lower() if not full_thread else "",
        )
        env = env_cache.get(env_cache_key)
        if env is None:
            # Environment: subject -> consultant replies -> description ->
            # broader selected thread -> final default PROD.
            env = resolve_environment(subject_text, "")
            if (not env) and full_thread and requester:
                consultant_text_key = (thread_key, _requester_key(requester))
                consultant_text = _env_consultant_text_cache.get(consultant_text_key)
                if consultant_text is None:
                    consultant_bodies = []
                    for e in full_thread:
                        if not _match_requester(e.sender_name, e.sender_email, requester):
                            continue
                        # Read consultant content only when subject alone did not resolve env.
                        if e.body:
                            consultant_bodies.append(e.body)
                        if getattr(e, "body_html", None):
                            consultant_bodies.append(e.body_html)
                    consultant_text = "\n".join(consultant_bodies)
                    _env_consultant_text_cache[consultant_text_key] = consultant_text
                if consultant_text:
                    env = resolve_environment(subject_text, consultant_text)
            if not env:
                env = resolve_environment(subject_text, description or "")
            if (not env) and full_thread:
                thread_text = _env_thread_text_cache.get(thread_key)
                if thread_text is None:
                    thread_parts = []
                    for e in full_thread:
                        subj = getattr(e, "subject", "") or ""
                        if subj:
                            thread_parts.append(subj)
                        if e.body:
                            thread_parts.append(e.body)
                        if getattr(e, "body_html", None):
                            thread_parts.append(e.body_html)
                    thread_text = "\n".join(thread_parts)
                    _env_thread_text_cache[thread_key] = thread_text
                if thread_text:
                    env = resolve_environment_thread_fallback(thread_text)
            if not env:
                env = "PROD"
            env_cache[env_cache_key] = env
        interface_code = resolve_interface_code(description)
        if workbook_kind == "task_business":
            category_type = "Business"
            service_request = "ServiceRequest"
            incident_type = "SR-File process"
        else:
            if workbook_kind == "incident_self_service":
                category_type = _resolve_self_service_category_type(category_type_raw, subject_text or description)
            incident_type = resolve_incident_type(category_type, description)
            service_request = resolve_service_request(category_type)
        # Honor explicit sheet hint first when present. This protects rows where
        # category type is wrong in source data, without hardcoding any subject.
        sr_hint = str(
            row_context.get("ServiceRequest/Incident?", "")
            or row_context.get("Servicerequest/incident?", "")
            or ""
        ).strip().lower()
        cat_l = (category_type or "").lower()

        if "file process/data check" in cat_l:
            service_request = "ServiceRequest"
            incident_type = "SR-File process"
        elif sr_hint.startswith("service"):
            service_request = "ServiceRequest"
        elif sr_hint.startswith("incident"):
            service_request = "Incident"
            if not (isinstance(incident_type, str) and incident_type.startswith("INC-")):
                incident_type = "INC-Exceptions"
        else:
            if isinstance(incident_type, str) and incident_type.startswith("INC-"):
                service_request = "Incident"

        deployment_override = None
        is_dep_req = _is_deployment_request_subject(subject_text)
        is_dep_succ = _is_deployment_success_subject(subject_text)
        if is_dep_req or is_dep_succ:
            row_iface = _iface_tokens(subject_text)
            def _dep_env(text: str) -> str:
                if not text:
                    return ""
                s = text.lower()
                if "qa" in s:
                    return "qa"
                if "prod" in s or "production" in s:
                    return "prod"
                if "uat" in s:
                    return "uat"
                return ""
            def _thread_has_sender(thread, email: str) -> bool:
                if not thread or not email:
                    return False
                target = email.lower()
                for e in thread:
                    if (e.sender_email or "").lower() == target:
                        return True
                return False
            row_env = _dep_env(subject_text)
            row_dr_ids = set()
            if thread:
                row_dr_ids |= _thread_dr_ids(thread)
            # Always include DR IDs directly from the row text; thread match can drift.
            row_dr_ids |= _extract_dr_ids(description or "")
            row_dr_ids |= _extract_dr_ids(subject_text or "")

            for dr in row_dr_ids:
                bucket = deployment_index.get(dr)
                if not bucket:
                    continue
                req_candidates = bucket["request"]
                succ_candidates = bucket["success"]
                if not req_candidates:
                    continue

                # Require environment match (PROD/UAT) when present on the row.
                if row_env:
                    req_candidates = [c for c in req_candidates if _dep_env(c["subject_key"]) == row_env]
                    succ_candidates = [c for c in succ_candidates if _dep_env(c["subject_key"]) == row_env]
                    if not req_candidates:
                        continue

                if row_iface:
                    req_candidates = [c for c in req_candidates if row_iface & c["iface"]] or req_candidates
                    succ_candidates = [c for c in succ_candidates if row_iface & c["iface"]] or succ_candidates

                # For ai-assist driven deployments, prefer threads actually sent by ai-assist.
                ai_sender = "ai-assist@umusic.com"
                if any(_thread_has_sender(c["thread"], ai_sender) for c in req_candidates + succ_candidates):
                    req_candidates = [c for c in req_candidates if _thread_has_sender(c["thread"], ai_sender)] or req_candidates
                    succ_candidates = [c for c in succ_candidates if _thread_has_sender(c["thread"], ai_sender)] or succ_candidates

                if is_dep_req:
                    # Always anchor request rows from DR-matched request candidates.
                    req_thread = min(req_candidates, key=lambda c: min(e.sent_time for e in c["thread"]))["thread"]
                    succ_thread = (
                        min(succ_candidates, key=lambda c: min(e.sent_time for e in c["thread"]))["thread"]
                        if succ_candidates else None
                    )
                else:
                    if not succ_candidates:
                        continue
                    succ_thread = thread if thread else min(succ_candidates, key=lambda c: min(e.sent_time for e in c["thread"]))["thread"]
                    req_thread = min(req_candidates, key=lambda c: min(e.sent_time for e in c["thread"]))["thread"]

                req_email = min(req_thread, key=lambda e: e.sent_time)
                succ_email = min(succ_thread, key=lambda e: e.sent_time) if succ_thread else None
                times = TimeResult(
                    _format_time(req_email.sent_time),
                    _format_time(req_email.sent_time),
                    _format_time(succ_email.sent_time) if succ_email else _format_time(req_email.sent_time),
                )
                debug = TimeDebug(
                    req_email.sender_email or req_email.sender_name,
                    req_email.sender_email or req_email.sender_name,
                    (succ_email.sender_email or succ_email.sender_name) if succ_email else (req_email.sender_email or req_email.sender_name),
                    f"{'DeploymentPair' if succ_email else 'DeploymentRequestOnly'} DR={dr}; Match={match_note}",
                )
                deployment_override = (times, debug)
                break

        if deployment_override:
            times, debug = deployment_override
            _seeded_won = False
            if times and debug:
                base_times = TimeResult(times.created, times.response, times.resolved)
                base_debug = TimeDebug(debug.created_src, debug.ack_src, debug.resolved_src, debug.notes)
        else:
            times = None
            debug = None
            has_row_ids = bool(_inc_tokens(subject_text) or _extract_dr_ids(subject_text))
            date_anchor_missing = False
            date_anchor_after = False
            # Date-anchor override (narrow): if row has a date token and the
            # consultant replied on that date, resolve using thread up to that reply.
            if thread and requester and date_tokens:
                pick = None
                date_anchor_exact = False
                date_anchor_near = False
                date_anchor_crossover = False
                anchor_date = _anchor_date(date_tokens)
                # Guard: ignore stale explicit date markers that are far away from
                # the actual matched thread date range.
                if anchor_date and not _anchor_relevant_to_thread(anchor_date, thread):
                    anchor_date = None

                # Prefer same-day consultant replies first.
                consultant_on_date = [
                    e for e in thread
                    if _match_requester(e.sender_name, e.sender_email, requester)
                    and _email_date_matches(e.sent_time, date_tokens)
                ]
                consultant_on_date.sort(key=lambda e: e.sent_time)
                if consultant_on_date:
                    pick = consultant_on_date[-1]
                    date_anchor_exact = True
                # Midnight crossover guard (narrow): if the row has an explicit date marker,
                # prefer the consultant reply closest to midnight (±2h) around the anchor date.
                elif anchor_date and explicit_marker:
                    anchor_start = _to_ist(datetime(anchor_date.year, anchor_date.month, anchor_date.day))
                    crossover_window = timedelta(hours=2)
                    crossover_candidates = [
                        e for e in thread
                        if _match_requester(e.sender_name, e.sender_email, requester)
                        and abs(_to_ist(e.sent_time) - anchor_start) <= crossover_window
                    ]
                    if crossover_candidates:
                        crossover_candidates.sort(
                            key=lambda e: (
                                abs(_to_ist(e.sent_time) - anchor_start),
                                _to_ist(e.sent_time),
                            )
                        )
                        pick = crossover_candidates[0]
                        date_anchor_near = True
                        date_anchor_crossover = True

                if not pick and anchor_date:
                    # Grace window before the anchor date (e.g., replies just before midnight)
                    anchor_start = _to_ist(datetime(anchor_date.year, anchor_date.month, anchor_date.day))
                    grace_hours = 12
                    window_start = anchor_start - timedelta(hours=grace_hours)
                    consultant_near_before = [
                        e for e in thread
                        if _match_requester(e.sender_name, e.sender_email, requester)
                        and window_start <= _to_ist(e.sent_time) < anchor_start
                    ]
                    consultant_near_before.sort(key=lambda e: e.sent_time)
                    if consultant_near_before:
                        pick = consultant_near_before[-1]
                        date_anchor_near = True
                    else:
                        consultant_after = [
                            e for e in thread
                            if _match_requester(e.sender_name, e.sender_email, requester)
                            and _email_on_or_after_date(e.sent_time, anchor_date)
                        ]
                        consultant_after.sort(key=lambda e: e.sent_time)
                        if consultant_after:
                            pick = consultant_after[0]
                            date_anchor_after = True
                if pick:
                    sliced_thread = [e for e in thread if e.sent_time <= pick.sent_time]
                    if sliced_thread:
                        times, debug, _episode = _resolve_times_seeded_first(
                            thread=sliced_thread,
                            requester_name=requester,
                            ess_team=ess_team,
                            subject_norm=subject_norm,
                            description=description,
                        )
                        _seeded_won = _episode is not None
                        debug = TimeDebug(
                            debug.created_src,
                            debug.ack_src,
                            debug.resolved_src,
                            f"{debug.notes}; "
                            f"{'DateAnchor' if date_anchor_exact else ('DateAnchorCrossover' if date_anchor_crossover else ('DateAnchorNear' if date_anchor_near else 'DateAnchorAfter'))}",
                        )
                        # If the row has an explicit date marker (e.g. "-->07-01-2026"),
                        # keep date-specific rows ordered, but do NOT overwrite a
                        # valid request-created time.
                        if explicit_marker and pick:
                            anchor_dt = _to_ist(pick.sent_time)
                            created_dt = _parse_time_str(times.created)
                            created_dt_ist = _to_ist(created_dt) if created_dt else None
                            if not created_dt_ist or created_dt_ist.date() != anchor_dt.date():
                                t = _format_time(pick.sent_time)
                                if t:
                                    # Preserve created if present; only enforce
                                    # response/resolved not before anchored reply.
                                    resp_dt = _parse_time_str(times.response)
                                    res_dt = _parse_time_str(times.resolved)
                                    resp_dt_ist = _to_ist(resp_dt) if resp_dt else None
                                    res_dt_ist = _to_ist(res_dt) if res_dt else None
                                    created_from_parse = isinstance(debug.created_src, str) and debug.created_src.startswith("PARSED_FROM_")
                                    deployment_like = "deployment request" in (subject_text or "").lower()
                                    force_created_to_anchor = bool(created_from_parse)
                                    if (
                                        not force_created_to_anchor
                                        and deployment_like
                                        and created_dt_ist
                                        and abs((anchor_dt.date() - created_dt_ist.date()).days) >= 1
                                    ):
                                        # For explicit-date deployment-like rows, avoid
                                        # carrying stale created dates from quoted history.
                                        force_created_to_anchor = True
                                    new_created = t if force_created_to_anchor else (times.created or t)
                                    new_response = times.response
                                    new_resolved = times.resolved
                                    if not resp_dt_ist or resp_dt_ist < anchor_dt:
                                        new_response = t
                                    if not res_dt_ist or res_dt_ist < anchor_dt:
                                        new_resolved = t
                                    times = TimeResult(new_created, new_response, new_resolved)
                                    note_extra = "; DateAnchorCreatedAdjusted" if force_created_to_anchor else ""
                                    debug = TimeDebug(
                                        (pick.sender_email or pick.sender_name) if force_created_to_anchor else (debug.created_src or (pick.sender_email or pick.sender_name)),
                                        (pick.sender_email or pick.sender_name) if (new_response == t) else debug.ack_src,
                                        (pick.sender_email or pick.sender_name) if (new_resolved == t) else debug.resolved_src,
                                        f"{debug.notes}; DateAnchorOccurrence{note_extra}",
                                    )
                    # DateAnchor ESS-only override (strict):
                    # Only if this subject+consultant occurs once AND the
                    # anchored reply is the latest consultant reply overall.
                    # Additionally, only force all-three-same when the
                    # anchored slice has a single ESS mail. If there are
                    # multiple ESS mails, keep normal span logic to avoid
                    # collapsing all three times.
                    if group_total == 1:
                        consultant_all = [
                            e for e in thread
                            if _match_requester(e.sender_name, e.sender_email, requester)
                        ]
                        consultant_all.sort(key=lambda e: e.sent_time)
                        if consultant_all:
                            consultant_non_ack = _shared_resolution_candidates(consultant_all)
                            if consultant_non_ack:
                                latest_all = consultant_non_ack[-1]
                            else:
                                latest_all = None
                            if latest_all and _to_ist(latest_all.sent_time) == _to_ist(pick.sent_time):
                                non_ess_present = any(
                                    not _is_ess_sender(e, ess_team) for e in thread
                                )
                                if not non_ess_present:
                                    ess_in_slice = [
                                        e for e in sliced_thread
                                        if _is_ess_sender(e, ess_team)
                                    ]
                                    if len(ess_in_slice) <= 1:
                                        t = _format_time(pick.sent_time)
                                        if t:
                                            times = TimeResult(t, t, t)
                                            debug = TimeDebug(
                                                pick.sender_email or pick.sender_name,
                                                pick.sender_email or pick.sender_name,
                                                pick.sender_email or pick.sender_name,
                                                f"{debug.notes}; DateAnchorESSOnly",
                                            )
                else:
                    date_anchor_missing = True
            if times is None:
                times, debug, _episode = _resolve_times_seeded_first(
                    thread=thread,
                    requester_name=requester,
                    ess_team=ess_team,
                    subject_norm=subject_norm,
                    description=description,
                )
                _seeded_won = _episode is not None
            if initial_occurrence_note and times and debug:
                debug = TimeDebug(
                    debug.created_src,
                    debug.ack_src,
                    debug.resolved_src,
                    f"{debug.notes}; {initial_occurrence_note}",
                )
            if date_anchor_missing and times and debug:
                debug = TimeDebug(
                    debug.created_src,
                    debug.ack_src,
                    debug.resolved_src,
                    f"{debug.notes}; DateAnchorMissing",
                )
            if stale_anchor and times and debug:
                debug = TimeDebug(
                    debug.created_src,
                    debug.ack_src,
                    debug.resolved_src,
                    f"{debug.notes}; DateAnchorIgnoredStale",
                )
            if times and debug and base_times is None:
                base_times = TimeResult(times.created, times.response, times.resolved)
                base_debug = TimeDebug(debug.created_src, debug.ack_src, debug.resolved_src, debug.notes)

            # Episode override (narrow): repeated same-subject for same consultant
            # Only apply when base resolution lacks a reliable ack and we have
            # enough distinct consultant replies to map to each occurrence.
            if (
                group_total >= 2
                and not date_tokens
                and not has_row_ids
                and thread
                and requester
                and _ack_missing(times, debug)
            ):
                consultant_replies_all = [
                    e for e in thread
                    if _match_requester(e.sender_name, e.sender_email, requester)
                ]
                consultant_replies = _shared_resolution_candidates(consultant_replies_all)
                preferred_consultant_replies = _shared_resolution_candidates(consultant_replies)
                if preferred_consultant_replies:
                    consultant_replies = preferred_consultant_replies
                consultant_replies.sort(key=lambda e: e.sent_time)
                if len(consultant_replies) >= group_total and len(consultant_replies) >= 2:
                    idx = episode_counters.get(group_key, 0)
                    occ_meta = _shared_occurrence_pick(
                        {
                            "occurrence_key": group_key,
                            "group_total": group_total,
                            "multi_service_subject": _subject_has_multiple_service_nos(subject_norm),
                            "subject_norm": subject_norm,
                            "requester": requester,
                            "service_no": service_no,
                            "list_index": None,
                        },
                        subject_norm_value=subject_norm,
                        requester_value=requester,
                        current_created_ist=None,
                        current_ack_ist=None,
                        current_resolved_ist=None,
                        default_idx=idx,
                    )
                    episode_counters[group_key] = idx + 1
                    pick = consultant_replies[min(occ_meta["pick_idx"], len(consultant_replies) - 1)]
                    cutoff = pick.sent_time
                    sliced_thread = [e for e in thread if e.sent_time <= cutoff]
                    if sliced_thread:
                        new_times, new_debug, _episode = _resolve_times_seeded_first(
                            thread=sliced_thread,
                            requester_name=requester,
                            ess_team=ess_team,
                            subject_norm=subject_norm,
                            description=description,
                        )
                        _seeded_won = _episode is not None
                        orig_all_same = (
                            times.created
                            and times.response
                            and times.resolved
                            and times.created == times.response == times.resolved
                        )
                        new_all_same = (
                            new_times.created
                            and new_times.response
                            and new_times.resolved
                            and new_times.created == new_times.response == new_times.resolved
                        )
                        if not orig_all_same or not new_all_same:
                            times = new_times
                            debug = TimeDebug(
                                new_debug.created_src,
                                new_debug.ack_src,
                                new_debug.resolved_src,
                                f"{new_debug.notes}; EpisodeOverride#{idx + 1}",
                            )

            # Single-occurrence latest-reply override (narrow):
            # If there's no date/id hint and no ack was found, use the latest
            # consultant reply for response/resolved. For ESS/no-ack rows whose
            # current created time came from a stale non-requester source, collapse
            # all three times to that latest reply instead of preserving the stale
            # created value.
            if (
                times
                and debug
                and thread
                and requester
                and group_total == 1
                and not date_tokens
                and not has_row_ids
                and _ack_missing(times, debug)
            ):
                consultant_replies_all = [
                    e for e in thread
                    if _match_requester(e.sender_name, e.sender_email, requester)
                ]
                consultant_replies = [e for e in consultant_replies_all if not _is_ack_like_reply(e)]
                consultant_replies = _shared_resolution_candidates(consultant_replies)
                consultant_replies.sort(key=lambda e: e.sent_time)
                if consultant_replies:
                    latest = consultant_replies[-1]
                    created_dt = _parse_time_str(times.created)
                    latest_dt = _to_ist(latest.sent_time)
                    created_dt_ist = _to_ist(created_dt) if created_dt else None
                    max_span = timedelta(days=7)
                    within_span = True
                    if created_dt_ist:
                        within_span = (latest_dt - created_dt_ist) <= max_span
                    if within_span and (not created_dt_ist or latest_dt >= created_dt_ist):
                        latest_str = _format_time(latest.sent_time)
                        if latest_str:
                            collapse_all_same = _should_latest_consultant_reply_collapse_all_same(
                                debug.notes,
                                debug.created_src,
                                requester,
                            )
                            if collapse_all_same:
                                latest_src = latest.sender_email or latest.sender_name
                                times = TimeResult(latest_str, latest_str, latest_str)
                                debug = TimeDebug(
                                    latest_src,
                                    latest_src,
                                    latest_src,
                                    f"{debug.notes}; LatestConsultantReplyAllSame",
                                )
                            else:
                                times = TimeResult(times.created, latest_str, latest_str)
                                debug = TimeDebug(
                                    debug.created_src,
                                    latest.sender_email or latest.sender_name,
                                    latest.sender_email or latest.sender_name,
                                    f"{debug.notes}; LatestConsultantReply",
                                )

            # Duplicate de-collision (narrow):
            # If repeated rows for same (subject, consultant) still resolve to an
            # identical time triplet, advance by occurrence using consultant reply
            # sequence inside the same matched thread.
            # This specifically protects repeated long-chain rows where date-marker
            # branches can collapse multiple occurrences to the same pick.
            if (
                times
                and debug
                and thread
                and requester
                and group_total >= 2
                and (date_tokens or explicit_marker or stale_anchor)
                and not is_dep_req
                and not is_dep_succ
                and "maintenance" not in (description or "").lower()
            ):
                current_triplet = (times.created, times.response, times.resolved)
                state = duplicate_group_state.get(group_key, {"seen": 0, "last_triplet": None})
                seen = state.get("seen", 0)
                last_triplet = state.get("last_triplet")

                if seen >= 1 and last_triplet == current_triplet:
                    consultant_replies_all = [
                        e for e in thread
                        if e.sent_time and _match_requester(e.sender_name, e.sender_email, requester)
                    ]
                    consultant_replies = _shared_resolution_candidates(consultant_replies_all)
                    preferred_consultant_replies = _shared_resolution_candidates(consultant_replies)
                    if preferred_consultant_replies:
                        consultant_replies = preferred_consultant_replies
                    consultant_replies.sort(key=lambda e: e.sent_time)

                    if len(consultant_replies) > seen:
                        pick = consultant_replies[seen]
                        if _is_shared_real_reply_candidate(pick):
                            sliced_thread = [e for e in thread if e.sent_time <= pick.sent_time]
                        else:
                            sliced_thread = []
                        if sliced_thread:
                            new_times, new_debug, _episode = _resolve_times_seeded_first(
                                thread=sliced_thread,
                                requester_name=requester,
                                ess_team=ess_team,
                                subject_norm=subject_norm,
                                description=description,
                            )
                            _seeded_won = _episode is not None
                            new_triplet = (new_times.created, new_times.response, new_times.resolved)
                            if new_triplet != current_triplet:
                                times = new_times
                                debug = TimeDebug(
                                    new_debug.created_src,
                                    new_debug.ack_src,
                                    new_debug.resolved_src,
                                    f"{new_debug.notes}; DuplicateOccurrenceAdjust#{seen + 1}",
                                )
                                current_triplet = new_triplet

                duplicate_group_state[group_key] = {
                    "seen": seen + 1,
                    "last_triplet": current_triplet,
                }

            # Outlier-only ordering guard (narrow): adjust only when clearly off from
            # the local flow (median of last 3 created times).
            row_has_ids = has_row_ids or bool(_inc_tokens(description) or _extract_dr_ids(description))
            date_tokens_match_thread = bool(date_tokens) and _thread_has_date_token(thread, date_tokens) if thread else False
            skip_history_update = False
            if (
                times
                and debug
                and thread
                and requester
                and created_history
                and (not date_tokens or date_anchor_missing or date_anchor_after or not date_tokens_match_thread)
                and (not row_has_ids or date_anchor_missing or date_anchor_after)
                and not is_dep_req
                and not is_dep_succ
                and "maintenance" not in (description or "").lower()
            ):
                ess_only_no_request = "ESS-only; no non-ESS request" in (debug.notes or "")
                created_dt = _parse_time_str(times.created)
                created_dt_ist = _to_ist(created_dt) if created_dt else None
                if created_dt_ist:
                    orig_all_same = (
                        times.created
                        and times.response
                        and times.resolved
                        and times.created == times.response == times.resolved
                    )
                    recent = created_history[-3:]
                    # If recent history is already monotonic, anchor to the latest
                    # timestamp to keep ordering tight. Otherwise use the median
                    # to avoid reacting to a single outlier.
                    expected_created = None
                    is_monotonic = False
                    if len(recent) >= 2:
                        is_monotonic = all(recent[i] <= recent[i + 1] for i in range(len(recent) - 1))
                        if is_monotonic:
                            expected_created = recent[-1]
                    if expected_created is None:
                        recent_sorted = sorted(recent)
                        expected_created = recent_sorted[len(recent_sorted) // 2]
                    expected_day_start = expected_created.replace(hour=0, minute=0, second=0, microsecond=0)
                    expected_day_end = expected_day_start + timedelta(days=1)
                    # Tighten the guard when recent history is already monotonic.
                    if is_monotonic:
                        grace_window = timedelta(hours=6)
                        max_forward_window = timedelta(days=1)
                    else:
                        grace_window = timedelta(hours=12)
                        max_forward_window = timedelta(days=2)
                    anchor_min = expected_created - grace_window
                    anchor_max = expected_created + max_forward_window

                    adjusted = False
                    # 1) Day-based correction: if the created date is outside the expected day,
                    # try to align to a consultant reply within the expected day.
                    if created_dt_ist.date() != expected_created.date():
                        day_thread = []
                        for e in thread:
                            if not e.sent_time:
                                continue
                            try:
                                sent_ist = _to_ist(e.sent_time)
                            except Exception:
                                continue
                            if expected_day_start <= sent_ist < expected_day_end:
                                day_thread.append(e)
                        consultant_in_day = [
                            e for e in day_thread
                            if _match_requester(e.sender_name, e.sender_email, requester)
                        ]
                        candidate_in_day = consultant_in_day
                        # For INC/DR rows, require a consultant reply in the window.
                        allow_ess_fallback = not row_has_ids
                        if not candidate_in_day and ess_only_no_request and allow_ess_fallback:
                            candidate_in_day = [
                                e for e in day_thread
                                if _is_ess_sender(e, ess_team)
                            ]
                        if candidate_in_day:
                            pick = min(
                                candidate_in_day,
                                key=lambda e: abs(_to_ist(e.sent_time) - expected_created),
                            )
                            sliced = [e for e in day_thread if e.sent_time <= pick.sent_time]
                            if sliced:
                                new_times, new_debug, _episode = _resolve_times_seeded_first(
                                    thread=sliced,
                                    requester_name=requester,
                                    ess_team=ess_team,
                                    subject_norm=subject_norm,
                                    description=description,
                                )
                                _seeded_won = _episode is not None
                                new_created_dt = _parse_time_str(new_times.created)
                                new_created_dt_ist = _to_ist(new_created_dt) if new_created_dt else None
                                if new_created_dt_ist and new_created_dt_ist.date() == expected_created.date():
                                    orig_delta = abs(created_dt_ist - expected_created)
                                    new_delta = abs(new_created_dt_ist - expected_created)
                                    new_all_same = (
                                        new_times.created
                                        and new_times.response
                                        and new_times.resolved
                                        and new_times.created == new_times.response == new_times.resolved
                                    )
                                    if (not orig_all_same and new_all_same):
                                        # Skip adjustment if it collapses all three times
                                        # when the original did not.
                                        pass
                                    elif new_delta < orig_delta:
                                        times = new_times
                                        debug = TimeDebug(
                                            new_debug.created_src,
                                            new_debug.ack_src,
                                            new_debug.resolved_src,
                                            f"{new_debug.notes}; OrderOutlierAdjustedDay",
                                        )
                                        adjusted = True

                    # 2) Window-based correction (fallback)
                    if not adjusted and (created_dt_ist < anchor_min or created_dt_ist > anchor_max):
                        window_thread = []
                        for e in thread:
                            if not e.sent_time:
                                continue
                            try:
                                sent_ist = _to_ist(e.sent_time)
                            except Exception:
                                continue
                            if anchor_min <= sent_ist <= anchor_max:
                                window_thread.append(e)
                        consultant_in_window = [
                            e for e in window_thread
                            if _match_requester(e.sender_name, e.sender_email, requester)
                        ]
                        candidate_in_window = consultant_in_window
                        # For INC/DR rows, require a consultant reply in the window.
                        allow_ess_fallback = not row_has_ids
                        if not candidate_in_window and ess_only_no_request and allow_ess_fallback:
                            candidate_in_window = [
                                e for e in window_thread
                                if _is_ess_sender(e, ess_team)
                            ]
                        if candidate_in_window:
                            pick = min(
                                candidate_in_window,
                                key=lambda e: abs(_to_ist(e.sent_time) - expected_created),
                            )
                            sliced = [e for e in window_thread if e.sent_time <= pick.sent_time]
                            if sliced:
                                new_times, new_debug, _episode = _resolve_times_seeded_first(
                                    thread=sliced,
                                    requester_name=requester,
                                    ess_team=ess_team,
                                    subject_norm=subject_norm,
                                    description=description,
                                )
                                _seeded_won = _episode is not None
                                new_created_dt = _parse_time_str(new_times.created)
                                new_created_dt_ist = _to_ist(new_created_dt) if new_created_dt else None
                                if new_created_dt_ist and anchor_min <= new_created_dt_ist <= anchor_max:
                                    orig_delta = abs(created_dt_ist - expected_created)
                                    new_delta = abs(new_created_dt_ist - expected_created)
                                    new_all_same = (
                                        new_times.created
                                        and new_times.response
                                        and new_times.resolved
                                        and new_times.created == new_times.response == new_times.resolved
                                    )
                                    if (not orig_all_same and new_all_same):
                                        # Skip adjustment if it collapses all three times
                                        # when the original did not.
                                        pass
                                    elif new_delta < orig_delta:
                                        times = new_times
                                        debug = TimeDebug(
                                            new_debug.created_src,
                                            new_debug.ack_src,
                                            new_debug.resolved_src,
                                            f"{new_debug.notes}; OrderOutlierAdjusted",
                                        )
                                        adjusted = True

                    if (
                        is_monotonic
                        and not adjusted
                        and (created_dt_ist < anchor_min or created_dt_ist > anchor_max)
                    ):
                        # Don't let a clear outlier poison the monotonic anchor.
                        skip_history_update = True
                        debug = TimeDebug(
                            debug.created_src,
                            debug.ack_src,
                            debug.resolved_src,
                            f"{debug.notes}; OrderOutlierSkippedHistory",
                        )

        # Post-ack resolution update (narrow): if a later consultant reply exists,
        # move only "Resolved" to the first reply after ack.
        if (
            times
            and debug
            and thread
            and requester
            and not is_dep_req
            and not is_dep_succ
            and "Maintenance override" not in (debug.notes or "")
        ):
            # Include related threads with the same normalized subject-for-match
            # to catch consultant replies that landed under a prefixed subject.
            related_threads = [thread]
            used_related = False
            try:
                alt_subject = normalize_subject_for_match(subject_norm)
                if alt_subject and alt_subject in alt_index:
                    for k in alt_index[alt_subject]:
                        t = threads.get(k)
                        if t and t is not thread:
                            related_threads.append(t)
                            used_related = True
            except Exception:
                pass

            merged = []
            seen = set()
            for t in related_threads:
                for e in t:
                    dedup_key = (e.subject, e.sender_email, e.sent_time)
                    if dedup_key in seen:
                        continue
                    seen.add(dedup_key)
                    merged.append(e)

            consultant_replies = [
                e for e in merged
                if _match_requester(e.sender_name, e.sender_email, requester)
            ]
            notes_l_now = ((debug.notes or "").lower() if debug else "")
            skip_resolved_after_ack = (
                "failed subject; no ack phrase" in notes_l_now
                or "no ess or requester replies" in notes_l_now
            )
            if consultant_replies and not skip_resolved_after_ack:
                ack_dt = _parse_time_str(times.response)
                res_dt = _parse_time_str(times.resolved)
                if ack_dt:
                    enable_postack_fallback_30m = os.getenv("POSTACK_FALLBACK_30M", "0") == "1"
                    ack_ist = _to_ist(ack_dt)
                    res_ist = _to_ist(res_dt) if res_dt else None
                    # If resolved is missing or effectively equals ack (same/very close),
                    # but there is a later consultant reply, move resolved to that reply.
                    if (not res_ist) or (res_ist <= (ack_ist + timedelta(minutes=20))):
                        window_hours = 48
                        window_end = ack_ist + timedelta(hours=window_hours)
                        anchor_date = _anchor_date(date_tokens or [])
                        if anchor_date:
                            anchor_start = _to_ist(datetime(anchor_date.year, anchor_date.month, anchor_date.day))
                            window_end = max(window_end, anchor_start + timedelta(hours=36))
                        if _row_has_force_same_time_lock({"subject_norm": subject_norm, "thread": merged}, debug.notes):
                            consultant_after = []
                        else:
                            consultant_after = [
                                e for e in consultant_replies
                                if _to_ist(e.sent_time) > ack_ist
                                and _to_ist(e.sent_time) <= window_end
                                and _is_shared_real_reply_candidate(e)
                            ]
                        if anchor_date:
                            on_anchor = [
                                e for e in consultant_after
                                if _to_ist(e.sent_time).date() == anchor_date
                            ]
                            if on_anchor:
                                consultant_after = on_anchor
                        consultant_after.sort(key=lambda e: e.sent_time)
                        fallback_30m_used = False
                        # If no consultant reply found, broaden to related threads
                        # with strong subject overlap that also contain the requester.
                        if not consultant_after:
                            try:
                                min_score = 0.72
                                if explicit_marker and anchor_date:
                                    min_score = 0.60
                                subj_tokens = _match_tokens(subject_norm)
                                subj_inc_set = _inc_tokens(subject_norm)
                                for key, t in threads.items():
                                    if t in related_threads:
                                        continue
                                    if not _thread_has_requester(t, requester):
                                        continue
                                    key_tokens = _match_tokens(key)
                                    if not key_tokens:
                                        continue
                                    if subj_inc_set:
                                        key_inc_set = _inc_tokens(key)
                                        if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                                            continue
                                    score = _token_overlap_score(subj_tokens, key_tokens)
                                    contains = (subject_norm in key or key in subject_norm)
                                    if score < min_score and not contains:
                                        continue
                                    related_threads.append(t)
                                    used_related = True
                                if used_related:
                                    merged = []
                                    seen = set()
                                    for t in related_threads:
                                        for e in t:
                                            dedup_key = (e.subject, e.sender_email, e.sent_time)
                                            if dedup_key in seen:
                                                continue
                                            seen.add(dedup_key)
                                            merged.append(e)
                                    consultant_replies = [
                                        e for e in merged
                                        if _match_requester(e.sender_name, e.sender_email, requester)
                                    ]
                                    consultant_after = [
                                        e for e in consultant_replies
                                        if _to_ist(e.sent_time) > ack_ist
                                        and _to_ist(e.sent_time) <= window_end
                                        and _is_shared_real_reply_candidate(e)
                                    ]
                                    if anchor_date:
                                        on_anchor = [
                                            e for e in consultant_after
                                            if _to_ist(e.sent_time).date() == anchor_date
                                        ]
                                        if on_anchor:
                                            consultant_after = on_anchor
                                    consultant_after.sort(key=lambda e: e.sent_time)
                            except Exception:
                                pass

                        # Conservative fallback: only when strict non-ack filter yields
                        # no candidate; choose first requester reply >=30m after ack.
                        if not consultant_after and enable_postack_fallback_30m:
                            fallback_after = [
                                e for e in consultant_replies
                                if _to_ist(e.sent_time) >= (ack_ist + timedelta(minutes=30))
                                and _to_ist(e.sent_time) <= window_end
                                and _is_shared_real_reply_candidate(e)
                            ]
                            if anchor_date:
                                on_anchor = [
                                    e for e in fallback_after
                                    if _to_ist(e.sent_time).date() == anchor_date
                                ]
                                if on_anchor:
                                    fallback_after = on_anchor
                            fallback_after.sort(key=lambda e: e.sent_time)
                            if fallback_after:
                                consultant_after = fallback_after
                                fallback_30m_used = True

                        if consultant_after:
                            next_reply = _pick_reply_after_ack(consultant_after, ack_ist, requester)
                            next_str = _format_time(next_reply.sent_time)
                            if next_str:
                                times = TimeResult(times.created, times.response, next_str)
                                debug = TimeDebug(
                                    debug.created_src,
                                    debug.ack_src,
                                    next_reply.sender_email or next_reply.sender_name,
                                    f"{debug.notes}; "
                                    f"{'ResolvedAfterAckFallback30m' if fallback_30m_used else ('ResolvedAfterAckRelated' if used_related else 'ResolvedAfterAck')}",
                                )

        # Maintenance override: if consultant replied, set all three times to the
        # consultant's last reply time. Do not mark red.
        if "maintenance" in (description or "").lower() and thread and requester:
            consultant_replies = [
                e for e in thread
                if _match_requester(e.sender_name, e.sender_email, requester)
            ]
            if consultant_replies:
                last_reply = max(consultant_replies, key=lambda e: e.sent_time)
                times = TimeResult(
                    _format_time(last_reply.sent_time),
                    _format_time(last_reply.sent_time),
                    _format_time(last_reply.sent_time),
                )
                debug = TimeDebug(
                    last_reply.sender_email or last_reply.sender_name,
                    last_reply.sender_email or last_reply.sender_name,
                    last_reply.sender_email or last_reply.sender_name,
                    "Maintenance override; consultant reply",
                )

        if times and debug:
            times, debug = _apply_same_time_normalization(
                thread=thread,
                requester_name=requester,
                ess_team=ess_team,
                subject_norm=subject_norm,
                category_type_value=(category_type or category_type_raw or ""),
                times=times,
                debug=debug,
            )
        if times and debug:
            times, debug = _apply_final_time_guards(
                thread=thread,
                requester=requester,
                times=times,
                debug=debug,
                base_times=base_times,
                base_debug=base_debug,
                baseline_created_date=baseline_created_date,
                is_dep_req=is_dep_req,
                is_dep_succ=is_dep_succ,
            )
        if times and debug:
            times, debug = _apply_same_time_normalization(
                thread=thread,
                requester_name=requester,
                ess_team=ess_team,
                subject_norm=subject_norm,
                category_type_value=(category_type or category_type_raw or ""),
                times=times,
                debug=debug,
            )

        if workbook_kind == "incident_self_service":
            if times.response:
                same_time = times.response
                same_src = debug.ack_src or debug.resolved_src or debug.created_src
                note_suffix = "SelfServiceAllThreeSame"
            elif times.resolved or times.created:
                same_time = times.resolved or times.created
                same_src = debug.resolved_src or debug.created_src or debug.ack_src
                note_suffix = "SelfServiceAllThreeSame(FallbackNoAck)"
            else:
                same_time = None
                same_src = None
                note_suffix = ""
            if same_time and same_src:
                times = TimeResult(same_time, same_time, same_time)
                debug = TimeDebug(
                    same_src,
                    same_src,
                    same_src,
                    f"{debug.notes}; {note_suffix}",
                )

        final_created_dt = _parse_time_str(times.created)
        final_created_dt_ist = _to_ist(final_created_dt) if final_created_dt else None
        if final_created_dt_ist and not skip_history_update:
            created_history.append(final_created_dt_ist)

        if os.getenv("DEBUG_TIMES", "0") == "1":
            logger.log(
                "[DEBUG] "
                f"Requester={requester} | "
                f"SubjectKey={subject_norm} | "
                f"Match={match_note} | "
                f"Created={times.created} ({debug.created_src}) | "
                f"Response={times.response} ({debug.ack_src}) | "
                f"Resolved={times.resolved} ({debug.resolved_src}) | "
                f"Notes={debug.notes}"
            )
        automation_rows.append(
            {
                "Description": description,
                "Requester": requester,
                "Environment": env,
                "Module": "Mule",
                "Issue Type": "Module",
                "Issue occurred in": "ESS",
                "Location/ Branch": "Hyderabad",
                "Category Type": category_type,
                "Created Date & Time": times.created,
                "Actual Response Date & Time": times.response,
                "Actual Resolved Date & Time": times.resolved,
                "ServiceRequest/Incident?": service_request,
                "ServiceRequest/Incident type?": incident_type,
                "Interface Code": interface_code,
            }
        )

        debug_rows.append(
            {
                "Description": description,
                "Requester": requester,
                "SubjectKey": subject_norm,
                "MatchFound": "Y" if thread else "N",
                "ThreadSize": len(thread),
                "CreatedSource": debug.created_src,
                "AckSource": debug.ack_src,
                "ResolvedSource": debug.resolved_src,
                "Notes": f"{debug.notes}; Match={match_note}",
            }
        )

        state_row_has_ids = bool(
            _inc_tokens(subject_text)
            or _extract_dr_ids(subject_text)
            or _inc_tokens(description)
            or _extract_dr_ids(description)
        )
        state_base_created_dt = _parse_time_str(times.created) if times and times.created else None
        state_base_ack_dt = _parse_time_str(times.response) if times and times.response else None
        state_base_resolved_dt = _parse_time_str(times.resolved) if times and times.resolved else None
        row_states.append(
            {
                "row_index": row_index,
                "list_index": len(automation_rows) - 1,
                "description": description,
                "category_type": category_type,
                "requester": requester,
                "service_no": service_no,
                "service_no_key": _service_no_key(service_no),
                "service_bucket": _subject_service_bucket(subject_norm, service_no),
                "multi_service_subject": _subject_has_multiple_service_nos(subject_norm),
                "occurrence_key": group_key,
                "occurrence_family_subject_norm": normalize_subject(_subject_for_description(description or "")) or subject_norm,
                "occurrence_initial_notes_l": (f"{debug.notes}; Match={match_note}" or "").lower(),
                "occurrence_initial_all_ack_family": (
                    "requester span(all-ack->ess)" in (f"{debug.notes}; Match={match_note}" or "").lower()
                    and "ess-only; no non-ess request" in (f"{debug.notes}; Match={match_note}" or "").lower()
                ),
                "occurrence_initial_family_candidate": (
                    (
                        "requester span(all-ack->ess)" in (f"{debug.notes}; Match={match_note}" or "").lower()
                        and "ess-only; no non-ess request" in (f"{debug.notes}; Match={match_note}" or "").lower()
                    )
                    or (
                        "ess-only; no non-ess request" in (f"{debug.notes}; Match={match_note}" or "").lower()
                        and "ackwindowguard" in (f"{debug.notes}; Match={match_note}" or "").lower()
                        and (
                            "dateanchoroccurrence" in (f"{debug.notes}; Match={match_note}" or "").lower()
                            or "quotedrequestonly" in (f"{debug.notes}; Match={match_note}" or "").lower()
                            or "quotedrequestonlynopair" in (f"{debug.notes}; Match={match_note}" or "").lower()
                        )
                    )
                ),
                "subject_norm": subject_norm,
                "date_tokens": date_tokens,
                "explicit_marker": explicit_marker,
                "date_anchor_missing": date_anchor_missing,
                "date_anchor_after": date_anchor_after,
                "stale_anchor": stale_anchor,
                "baseline_created_date": baseline_created_date,
                "group_total": group_total,
                "date_tokens_match_thread": bool(date_tokens) and _thread_has_date_token(thread, date_tokens),
                "thread": full_thread or thread,
                "times": times,
                "debug": debug,
                "shared_decision": None,
                "initial_lane_episode": None,
                "seed_locked": _seeded_won,
                "occurrence_locked": False,
                "occurrence_lock_triplet": None,
                "row_has_ids": state_row_has_ids,
                "is_dep_req": is_dep_req,
                "is_dep_succ": is_dep_succ,
                "match_note": match_note,
                "base_candidate": {
                    "created": _to_ist(state_base_created_dt) if state_base_created_dt else None,
                    "response": _to_ist(state_base_ack_dt) if state_base_ack_dt else None,
                    "resolved": _to_ist(state_base_resolved_dt) if state_base_resolved_dt else None,
                    "created_src": debug.created_src if debug else "",
                    "ack_src": debug.ack_src if debug else "",
                    "resolved_src": debug.resolved_src if debug else "",
                    "notes_l": ((debug.notes or "").lower() if debug else ""),
                },
            }
        )

        if times.created and times.response and times.resolved:
            if (
                times.created == times.response
                and times.response == times.resolved
            ):
                same_time_rows.append(
                    {
                        "Description": description,
                        "Requester": requester,
                        "SubjectKey": subject_norm,
                        "MatchFound": "Y" if thread else "N",
                        "ThreadSize": len(thread),
                        "Created": times.created,
                        "Response": times.response,
                        "Resolved": times.resolved,
                        "CreatedSource": debug.created_src,
                        "AckSource": debug.ack_src,
                        "ResolvedSource": debug.resolved_src,
                        "Notes": f"{debug.notes}; Match={match_note}",
                    }
                )

        resolved_values = {
            "Environment": env,
            "Module": "Mule",
            "Issue Type": "Module",
            "Issue occurred in": "ESS",
            "Location/ Branch": "Hyderabad",
            "Category Type": category_type,
            "Created Date & Time": times.created,
            "Actual Response Date & Time": times.response,
            "Actual Resolved Date & Time": times.resolved,
            "ServiceRequest/Incident?": service_request,
            "ServiceRequest/Incident type?": incident_type,
            "Interface Code": interface_code,
        }
        if "Ack delayed" in debug.notes:
            resolved_values["_MarkBlue"] = True
        return resolved_values

    def _sequence_ordering_pass(ws, col_map, header_row):
        created_col = col_map.get("created date & time")
        response_col = col_map.get("actual response date & time")
        resolved_col = col_map.get("actual resolved date & time")
        env_col = col_map.get("environment")
        if not created_col or not response_col or not resolved_col:
            return

        # Performance caches (no logic change).
        _ist_cache = {}
        _req_match_cache = {}
        _is_ess_cache = {}
        _ack_like_cache = {}
        _ack_like_text_cache = {}
        _emails_for_requester_cache = {}
        _row_subject_match_cache = {}
        _id_token_cache = {}
        _quoted_sent_parse_cache = {}

        def _email_ist(e):
            key = id(e)
            if key in _ist_cache:
                return _ist_cache[key]
            try:
                v = _to_ist(e.sent_time) if getattr(e, "sent_time", None) else None
            except Exception:
                v = None
            _ist_cache[key] = v
            return v

        def _req_match(e, requester_name):
            key = (id(e), requester_name or "")
            if key in _req_match_cache:
                return _req_match_cache[key]
            v = _match_requester(e.sender_name, e.sender_email, requester_name)
            _req_match_cache[key] = v
            return v

        def _ess_sender(e):
            key = id(e)
            if key in _is_ess_cache:
                return _is_ess_cache[key]
            v = _is_ess_sender(e, ess_team)
            _is_ess_cache[key] = v
            return v

        def _ack_like(e):
            key = id(e)
            if key in _ack_like_cache:
                return _ack_like_cache[key]
            cls = _shared_reply_classification(e)
            v = bool(cls.get("ack_like") or cls.get("explicit_ack"))
            _ack_like_cache[key] = v
            return v

        def _ack_like_text_fallback(e):
            key = id(e)
            if key in _ack_like_text_cache:
                return _ack_like_text_cache[key]
            cls = _shared_reply_classification(e)
            out = bool(cls.get("nonfinal_followup") or cls.get("thanks_info"))
            _ack_like_text_cache[key] = out
            return out

        def _is_real_reply_candidate(e):
            return _shared_real_reply_candidate(e)

        def _has_source_locked_same_time(notes_text: str) -> bool:
            notes_l = (notes_text or "").lower()
            return any(token in notes_l for token in _SOURCE_LOCKED_SAME_TIME_TOKENS)

        def _parse_quoted_sent_time(sent_line: str):
            if not sent_line:
                return None
            if sent_line in _quoted_sent_parse_cache:
                return _quoted_sent_parse_cache[sent_line]
            def _normalize_quoted_sent_text(value: str) -> str:
                value = re.sub(r"(?i)^sent\s*:\s*", "", value or "").strip()
                value = re.sub(r"(?i)^(mon|tue|wed|thu|fri|sat|sun)\w*,?\s*", "", value).strip()
                value = re.sub(r"\(.*?\)", " ", value).strip()
                value = re.sub(r"(?i)(\d)(am|pm)\b", r"\1 \2", value)
                value = re.sub(r"(?i)\b(a\.m\.|p\.m\.)\b", lambda m: m.group(1).replace(".", "").upper(), value)
                return " ".join(value.replace(",", " ").replace(" at ", " ").split())

            def _extract_primary_sent_fragments(*values: str):
                regexes = [
                    r"\b[A-Za-z]+,\s+\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
                    r"\b[A-Za-z]+,\s+\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
                    r"\b[A-Za-z]+,\s+[A-Za-z]+\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
                    r"\b[A-Za-z]+,\s+[A-Za-z]+\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
                    r"\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
                    r"\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
                    r"\b\d{1,2}[-./]\d{1,2}[-./]\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
                    r"\b\d{1,2}[-./]\d{1,2}[-./]\d{4}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
                    r"\b\d{4}[-./]\d{1,2}[-./]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?\s*[APMapm]{2}\b",
                    r"\b\d{4}[-./]\d{1,2}[-./]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?\b",
                ]
                out = []
                seen = set()
                for value in values:
                    text = (value or "").strip()
                    if not text:
                        continue
                    matches = []
                    for rx in regexes:
                        for match in re.finditer(rx, text):
                            frag = match.group(0).strip()
                            key = frag.lower()
                            if key in seen:
                                continue
                            matches.append((match.start(), -len(frag), frag))
                    matches.sort()
                    for _, _, frag in matches:
                        key = frag.lower()
                        if key in seen:
                            continue
                        seen.add(key)
                        out.append(frag)
                    if out:
                        break
                return out

            normalized = _normalize_quoted_sent_text(sent_line)
            raw_clean = re.sub(r"(?i)^sent\s*:\s*", "", sent_line or "").strip()
            candidates = []
            for cand in _extract_primary_sent_fragments(raw_clean, normalized):
                if cand and cand not in candidates:
                    candidates.append(cand)
            for cand in (normalized, raw_clean):
                if cand and cand not in candidates:
                    candidates.append(cand)

            fmts_24h = [
                "%d %B %Y %H:%M:%S",
                "%d %B %Y %H:%M",
                "%d %b %Y %H:%M:%S",
                "%d %b %Y %H:%M",
                "%B %d %Y %H:%M:%S",
                "%B %d %Y %H:%M",
                "%b %d %Y %H:%M:%S",
                "%b %d %Y %H:%M",
                "%d-%m-%Y %H:%M:%S",
                "%d-%m-%Y %H:%M",
                "%d.%m.%Y %H:%M:%S",
                "%d.%m.%Y %H:%M",
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
            ]
            fmts_12h = [
                "%d %B %Y %I:%M:%S %p",
                "%d %B %Y %I:%M %p",
                "%d %b %Y %I:%M:%S %p",
                "%d %b %Y %I:%M %p",
                "%B %d %Y %I:%M:%S %p",
                "%B %d %Y %I:%M %p",
                "%b %d %Y %I:%M:%S %p",
                "%b %d %Y %I:%M %p",
                "%d-%m-%Y %I:%M:%S %p",
                "%d-%m-%Y %I:%M %p",
                "%d.%m.%Y %I:%M:%S %p",
                "%d.%m.%Y %I:%M %p",
                "%d/%m/%Y %I:%M:%S %p",
                "%d/%m/%Y %I:%M %p",
                "%Y-%m-%d %I:%M:%S %p",
                "%Y-%m-%d %I:%M %p",
                "%A %d %B %Y %I:%M:%S %p",
                "%A %d %B %Y %I:%M %p",
                "%A %B %d %Y %I:%M:%S %p",
                "%A %B %d %Y %I:%M %p",
            ]
            for cand in candidates:
                cand_has_am_pm = bool(re.search(r"(?i)\b(am|pm|a\.m\.|p\.m\.)\b", cand or ""))
                parse_fmts = fmts_12h if cand_has_am_pm else fmts_24h
                for fmt in parse_fmts:
                    try:
                        parsed = datetime.strptime(cand, fmt)
                        _quoted_sent_parse_cache[sent_line] = parsed
                        return parsed
                    except Exception:
                        continue
            for cand in candidates:
                try:
                    parsed = parsedate_to_datetime(cand)
                    if parsed:
                        _quoted_sent_parse_cache[sent_line] = parsed
                        return parsed
                except Exception:
                    pass
                if not cand_has_am_pm:
                    for fmt in fmts_12h:
                        try:
                            parsed = datetime.strptime(cand, fmt)
                            _quoted_sent_parse_cache[sent_line] = parsed
                            return parsed
                        except Exception:
                            continue
            _quoted_sent_parse_cache[sent_line] = None
            return None

        _quoted_requester_reply_cache = {}
        _quoted_request_before_cache = {}
        _quoted_blocks_for_subject_cache = {}

        def _extract_quoted_requester_reply_ist(email_obj, requester_name: str, subject_norm_value: str, lower_ist, upper_ist):
            cache_key = (id(email_obj), requester_name or "", subject_norm_value or "")
            cands = _quoted_requester_reply_cache.get(cache_key)
            if cands is None:
                row_tokens = _match_tokens(subject_norm_value or "")
                row_id_tokens = _id_like_tokens(subject_norm_value or "")
                cands = []
                for from_line, sent_ist, subj_text in _extract_quoted_blocks_with_subject(email_obj):
                    if requester_name and not _match_requester(from_line, from_line, requester_name):
                        continue
                    if subj_text and row_tokens:
                        subj_norm = normalize_subject(subj_text)
                        subj_ids = _id_like_tokens(subj_norm)
                        subj_tokens = _match_tokens(subj_norm)
                        if not _quoted_subject_confirms_row(
                            subj_norm,
                            subj_ids,
                            subj_tokens,
                            subject_norm_value,
                            row_tokens,
                            row_id_tokens,
                        ):
                            continue
                    elif row_tokens or row_id_tokens:
                        continue
                    cands.append(sent_ist)
                cands.sort()
                _quoted_requester_reply_cache[cache_key] = cands
            if not cands:
                return None
            eligible = []
            lower_bound = (lower_ist - timedelta(days=5)) if lower_ist else None
            for cand in cands:
                if lower_bound and cand < lower_bound:
                    continue
                if upper_ist and cand >= upper_ist:
                    continue
                eligible.append(cand)
            if not eligible:
                return None
            return eligible[-1]

        def _extract_quoted_request_before_ist(email_obj, subject_norm_value: str, upper_ist):
            cache_key = (id(email_obj), subject_norm_value or "")
            cands = _quoted_request_before_cache.get(cache_key)
            if cands is None:
                row_tokens = _match_tokens(subject_norm_value or "")
                row_id_tokens = _id_like_tokens(subject_norm_value or "")
                cands = []
                quoted_blocks = _extract_quoted_blocks_with_subject(email_obj)
                if not quoted_blocks:
                    quoted_blocks = _extract_quoted_blocks_relaxed(email_obj)
                for from_line, sent_ist, q_subj in quoted_blocks:
                    if not sent_ist:
                        continue
                    if q_subj:
                        q_norm = normalize_subject(q_subj or "")
                        q_ids = _id_like_tokens(q_norm)
                        q_tokens = _match_tokens(q_norm)
                        if not _quoted_subject_confirms_row(
                            q_norm,
                            q_ids,
                            q_tokens,
                            subject_norm_value,
                            row_tokens,
                            row_id_tokens,
                        ):
                            continue
                    elif row_tokens or row_id_tokens:
                        continue
                    is_ess = _quoted_from_line_is_ess_shared(from_line)
                    if is_ess:
                        continue
                    cands.append(sent_ist)
                cands.sort()
                _quoted_request_before_cache[cache_key] = cands
            if not cands:
                return None
            if not upper_ist:
                return cands[-1]
            eligible = [cand for cand in cands if cand < upper_ist]
            if not eligible:
                return None
            return eligible[-1]

        def _extract_quoted_blocks(email_obj, subject_norm_value: str):
            cache_key = (id(email_obj), subject_norm_value or "")
            cached = _quoted_blocks_for_subject_cache.get(cache_key)
            if cached is not None:
                return list(cached)
            row_tokens = _match_tokens(subject_norm_value or "")
            blocks = []
            for from_line, sent_ist, subj_text in _extract_bounded_quoted_header_candidates(email_obj, allow_relaxed=False):
                if subj_text and row_tokens:
                    subj_norm = normalize_subject(subj_text)
                    subj_tokens = _match_tokens(subj_norm)
                    score = _token_overlap_score(row_tokens, subj_tokens) if subj_tokens else 0.0
                    contains = bool(subject_norm_value and subj_norm and (subject_norm_value in subj_norm or subj_norm in subject_norm_value))
                    if score < 0.45 and not contains:
                        continue
                blocks.append((from_line, sent_ist))
            _quoted_blocks_for_subject_cache[cache_key] = tuple(blocks)
            return list(_quoted_blocks_for_subject_cache[cache_key])

        def _emails_for_requester(requester_name):
            k = requester_name or ""
            if k in _emails_for_requester_cache:
                return _emails_for_requester_cache[k]
            out = [e for e in emails if getattr(e, "sent_time", None) and _req_match(e, requester_name)]
            _emails_for_requester_cache[k] = out
            return out

        _expanded_thread_cache = {}
        _requester_pool_cache = {}
        _requester_pool_full_cache = {}
        strict_all_live_builder_cache = {}
        row_match_context_cache = {}
        ess_only_msg_partition_cache = {}
        subject_norm_cache = {}
        # Blue-only post resolver mode:
        # run strict post adjustments only in blue-marked section and keep
        # normal rows untouched in post-pass logic.
        # Default ON:
        # keep normal rows untouched in post-pass and run strict resolver only
        # on blue-marked rows.
        blue_only_post_resolver = os.getenv("BLUE_ONLY_POST_RESOLVER", "1") == "1"
        nonblue_row_states = [] if blue_only_post_resolver else row_states
        state_by_list_index = {
            s.get("list_index"): s
            for s in row_states
            if s.get("list_index") is not None
        }
        _subject_family_list_indexes_cache = {}
        trace_row_subject_filter = (os.getenv("TRACE_ROW_SUBJECT") or "").strip().lower()

        def _is_trace_focus_row(state=None, row_vals=None, list_index=None) -> bool:
            if not trace_row_subject_filter:
                return False
            desc = ""
            subj = ""
            if row_vals:
                desc = str(row_vals.get("Description") or "")
            if state:
                subj = str(state.get("subject_norm") or "")
                if not desc:
                    desc = str(state.get("description") or "")
            elif list_index is not None and list_index < len(automation_rows):
                desc = str(automation_rows[list_index].get("Description") or "")
            blob = f"{desc}\n{subj}".lower()
            return trace_row_subject_filter in blob

        def _trace_focus_row(phase: str, *, state=None, row_vals=None, list_index=None, **fields):
            return

        def _subject_family_list_indexes(subject_norm_value: str):
            subject_key = (subject_norm_value or "").lower()
            cached = _subject_family_list_indexes_cache.get(subject_key)
            if cached is not None:
                return cached
            out = [
                s.get("list_index")
                for s in row_states
                if s.get("list_index") is not None
                and ((s.get("subject_norm") or "").lower() == subject_key)
            ]
            out.sort()
            _subject_family_list_indexes_cache[subject_key] = out
            return out

        def _subject_norm_cached(subject: str) -> str:
            key = subject or ""
            if key in subject_norm_cache:
                return subject_norm_cache[key]
            norm = normalize_subject(key)
            subject_norm_cache[key] = norm
            return norm

        def _subject_has_id_token(subject: str, row_id_tokens_set: set) -> bool:
            if not row_id_tokens_set:
                return True
            subj_tokens = _id_like_tokens(subject or "")
            return bool(subj_tokens & row_id_tokens_set)

        def _expanded_thread(subject_norm_value, base_thread, requester_name, include_non_ess=False, reference_ist=None):
            if not base_thread:
                return []
            ref_day = ""
            if reference_ist:
                try:
                    ref_day = reference_ist.date().isoformat()
                except Exception:
                    ref_day = ""
            cache_key = (subject_norm_value or "", requester_name or "", id(base_thread), bool(include_non_ess), ref_day)
            if cache_key in _expanded_thread_cache:
                return _expanded_thread_cache[cache_key]

            merged = list(base_thread)
            try:
                alt_subject = normalize_subject_for_match(subject_norm_value or "")
                if alt_subject and alt_subject in alt_index:
                    for key in alt_index[alt_subject]:
                        t = threads.get(key) or []
                        if not t or t is base_thread:
                            continue
                        if requester_name and not any(_req_match(e, requester_name) for e in t):
                            if include_non_ess:
                                has_non_ess = any((not _ess_sender(e)) for e in t)
                                if not has_non_ess:
                                    continue
                            else:
                                continue
                        if include_non_ess and reference_ist:
                            near = False
                            for e in t:
                                e_ist = _email_ist(e)
                                if not e_ist:
                                    continue
                                if abs((e_ist - reference_ist).total_seconds()) <= (14 * 24 * 3600):
                                    near = True
                                    break
                            if not near:
                                continue
                        merged.extend(t)
            except Exception:
                pass

            dedup = {}
            for e in merged:
                k = (e.subject, e.sender_email, e.sender_name, e.sent_time)
                dedup[k] = e
            out = list(dedup.values())
            out.sort(key=lambda e: (0 if getattr(e, "sent_time", None) else 1, getattr(e, "sent_time", datetime.max)))
            _expanded_thread_cache[cache_key] = out
            return out

        def _is_occurrence_managed_notes(notes_l: str) -> bool:
            notes_l = (notes_l or "").lower()
            return (
                "dateanchoroccurrence" in notes_l
                or "ess-only; no non-ess request" in notes_l
                or "requester follow-up" in notes_l
                or "esscontinuationguard[" in notes_l
                or "quotedrequestonlynopair" in notes_l
            )

        def _is_all_ack_to_ess_notes(notes_l: str) -> bool:
            notes_l = (notes_l or "").lower()
            return (
                "requester span(all-ack->ess)" in notes_l
                and "ess-only; no non-ess request" in notes_l
            )

        def _is_occurrence_all_ack_family_candidate_notes(notes_l: str) -> bool:
            notes_l = (notes_l or "").lower()
            if _is_all_ack_to_ess_notes(notes_l):
                return True
            return (
                "ess-only; no non-ess request" in notes_l
                and "ackwindowguard" in notes_l
                and (
                    "dateanchoroccurrence" in notes_l
                    or "quotedrequestonly" in notes_l
                    or "quotedrequestonlynopair" in notes_l
                )
            )

        def _state_occurrence_notes_l(state) -> str:
            return (state.get("occurrence_initial_notes_l") or "").lower() if state else ""

        def _state_occurrence_family_subject_norm(state) -> str:
            if not state:
                return ""
            return (
                state.get("occurrence_family_subject_norm")
                or state.get("subject_norm")
                or ""
            ).lower()

        def _state_is_all_ack_to_ess(state) -> bool:
            if not state:
                return False
            if state.get("occurrence_initial_all_ack_family"):
                return True
            return _is_all_ack_to_ess_notes(_state_occurrence_notes_l(state))

        def _state_is_occurrence_family_candidate(state) -> bool:
            if not state:
                return False
            if state.get("occurrence_initial_family_candidate"):
                return True
            return _is_occurrence_all_ack_family_candidate_notes(_state_occurrence_notes_l(state))

        def _is_authoritative_occurrence_lane(lane_kind: str) -> bool:
            return (lane_kind or "") in {"ess_over_ess", "ess_acky_sequence"}

        def _shared_occurrence_lane_plan(state):
            requester = state.get("requester") or ""
            subject_norm_value = _state_occurrence_family_subject_norm(state)
            service_no = state.get("service_no") or ""
            list_index = state.get("list_index")
            row_group_total = state.get("group_total") or 0
            if list_index is None or not requester or not subject_norm_value:
                return None

            current_occ_key = state.get("occurrence_key") or _occurrence_group_key(subject_norm_value, requester, service_no)
            current_notes_l = _state_occurrence_notes_l(state)
            current_is_all_ack_to_ess = _state_is_all_ack_to_ess(state)
            current_service_bucket = state.get("service_bucket") or _subject_service_bucket(subject_norm_value, service_no)
            current_iface_tokens = _interface_tokens(subject_norm_value)

            def _group_subject_match(other_subject_norm: str) -> bool:
                other_norm = (other_subject_norm or "").lower()
                if not other_norm:
                    return False
                return _fresh_picker_subject_safe(
                    subject_norm_value,
                    other_norm,
                    iface_tokens=current_iface_tokens,
                    allow_added_inc=True,
                )

            def _collect_group(subject_wide: bool):
                group = []
                group_requesters = set()
                allow_acky_local = False
                for s in row_states:
                    other_li = s.get("list_index")
                    if other_li is None or other_li >= len(automation_rows):
                        continue
                    other_notes_l = _state_occurrence_notes_l(s)
                    if not _is_occurrence_managed_notes(other_notes_l):
                        continue
                    if subject_wide:
                        if current_is_all_ack_to_ess and not _state_is_occurrence_family_candidate(s):
                            continue
                        other_subject_norm = _state_occurrence_family_subject_norm(s)
                        if not _group_subject_match(other_subject_norm):
                            continue
                        other_service_bucket = s.get("service_bucket") or _subject_service_bucket(
                            other_subject_norm,
                            s.get("service_no") or "",
                        )
                        if current_service_bucket and other_service_bucket and other_service_bucket != current_service_bucket:
                            continue
                    else:
                        other_occ_key = s.get("occurrence_key") or _occurrence_group_key(
                            (s.get("subject_norm") or "").lower(),
                            s.get("requester") or "",
                            s.get("service_no") or "",
                        )
                        if other_occ_key != current_occ_key:
                            continue
                    allow_acky_local = allow_acky_local or ("requester span(all-ack->ess)" in other_notes_l)
                    group.append(s)
                    other_requester = (s.get("requester") or "").strip()
                    if other_requester:
                        group_requesters.add(other_requester)
                group.sort(key=lambda x: x.get("row_index") or 10**9)
                return group, group_requesters, allow_acky_local

            requester_group, requester_group_requesters, requester_allow_acky = _collect_group(False)
            use_subject_wide_ess = (
                current_is_all_ack_to_ess
                and _is_occurrence_managed_notes(current_notes_l)
            )
            subject_group = []
            subject_group_requesters = set()
            subject_allow_acky = False
            if use_subject_wide_ess:
                subject_group, subject_group_requesters, subject_allow_acky = _collect_group(True)

            if len(subject_group) >= 2 and len(subject_group) > len(requester_group):
                group_sorted = subject_group
                group_requesters = subject_group_requesters
                allow_acky = subject_allow_acky
                selected_scope = "subject_wide_ess"
            else:
                group_sorted = requester_group
                group_requesters = requester_group_requesters
                allow_acky = requester_allow_acky
                selected_scope = "requester"

            if len(group_sorted) < 2 and row_group_total < 2:
                return None
            if len(group_sorted) < 2:
                return None

            group_sorted.sort(key=lambda x: x.get("row_index") or 10**9)
            slot_index = None
            for idx, s in enumerate(group_sorted):
                if s.get("list_index") == list_index:
                    slot_index = idx
                    break
            if slot_index is None:
                return None

            merged = []
            for s in group_sorted:
                base_thread = s.get("thread") or []
                merged.extend(base_thread)
                expanded = _expanded_thread(
                    subject_norm_value,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=None,
                )
                merged.extend(expanded or [])

            dedup = {}
            for e in merged:
                sent_time = getattr(e, "sent_time", None)
                if not sent_time:
                    continue
                key = (
                    getattr(e, "subject", "") or "",
                    getattr(e, "sender_email", "") or "",
                    getattr(e, "sender_name", "") or "",
                    sent_time,
                )
                dedup[key] = e
            merged = list(dedup.values())
            merged.sort(key=lambda e: e.sent_time)
            if not merged:
                return None

            row_tokens = _match_tokens(subject_norm_value)
            requester_names = tuple(sorted(group_requesters)) if group_requesters else ((requester,) if requester else tuple())

            def _group_req_match(e):
                if not requester_names:
                    return False
                return any(_match_requester(e.sender_name, e.sender_email, req_name) for req_name in requester_names if req_name)

            def _collect_pool(allow_acky_local: bool, use_ess_pool: bool):
                out = []
                for e in merged:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if use_ess_pool:
                        if not _ess_sender(e):
                            continue
                        flags = _shared_reply_flags(email_record=e)
                        if _system_like_sender(e) or flags["ignore_reply"]:
                            continue
                        if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                            continue
                    else:
                        if not _group_req_match(e):
                            continue
                        flags = _shared_reply_flags(email_record=e)
                        if not allow_acky_local and not flags["substantive_reply"]:
                            continue
                    if row_tokens:
                        s_norm = normalize_subject(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(
                            subject_norm_value and s_norm and (
                                subject_norm_value in s_norm or s_norm in subject_norm_value
                            )
                        )
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)

                return _dedupe_reply_minutes_prefer_consultant(
                    out,
                    requester,
                    created_role=not use_ess_pool,
                    all_same=use_ess_pool,
                )

            def _collect_non_ess_ack_pool():
                out = []
                for e in merged:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if not _group_req_match(e):
                        continue
                    if not _is_shared_ack_candidate(e):
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(
                            subject_norm_value and s_norm and (
                                subject_norm_value in s_norm or s_norm in subject_norm_value
                            )
                        )
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)

                return _dedupe_reply_minutes_prefer_consultant(
                    out,
                    requester,
                    created_role=False,
                    all_same=False,
                )

            def _collect_direct_resolution_pool():
                out = []
                for e in merged:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if not _group_req_match(e):
                        continue
                    if not _shared_reply_classification(e).get("direct_resolution"):
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(
                            subject_norm_value and s_norm and (
                                subject_norm_value in s_norm or s_norm in subject_norm_value
                            )
                        )
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)
                return _dedupe_reply_minutes_prefer_consultant(
                    out,
                    requester,
                    created_role=True,
                    all_same=False,
                )

            def _collect_consultant_ess_pool():
                out = []
                for e in merged:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if not _ess_sender(e):
                        continue
                    if not _group_req_match(e):
                        continue
                    flags = _shared_reply_flags(email_record=e)
                    if _system_like_sender(e) or flags["ignore_reply"]:
                        continue
                    if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(
                            subject_norm_value and s_norm and (
                                subject_norm_value in s_norm or s_norm in subject_norm_value
                            )
                        )
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)
                return _dedupe_reply_minutes_prefer_consultant(
                    out,
                    requester,
                    created_role=False,
                    all_same=True,
                )

            def _collect_seeded_reply_pool():
                out = []
                seen = set()
                for slot_state in group_sorted:
                    seeded = slot_state.get("initial_lane_episode") or {}
                    resolved_ist = seeded.get("resolved")
                    if not resolved_ist:
                        continue
                    minute_key = resolved_ist.replace(second=0, microsecond=0)
                    if minute_key in seen:
                        continue

                    best = None
                    for e in merged:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if abs((e_ist - resolved_ist).total_seconds()) > 300:
                            continue
                        if row_tokens:
                            s_norm = normalize_subject(getattr(e, "subject", "") or "")
                            s_tokens = _match_tokens(s_norm)
                            score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                            contains = bool(
                                subject_norm_value and s_norm and (
                                    subject_norm_value in s_norm or s_norm in subject_norm_value
                                )
                            )
                            if score < 0.45 and not contains:
                                continue
                        flags = _shared_reply_flags(email_record=e)
                        if flags["ignore_reply"]:
                            continue
                        if not flags["substantive_reply"]:
                            continue

                        requester_match = 1 if _group_req_match(e) else 0
                        ess_match = 1 if _ess_sender(e) else 0
                        cand = (
                            requester_match,
                            ess_match,
                            -abs((e_ist - resolved_ist).total_seconds()),
                            e_ist,
                            id(e),
                            e,
                        )
                        if best is None or cand > best:
                            best = cand

                    if best is None:
                        continue

                    seen.add(minute_key)
                    out.append(best[5])

                out.sort(key=lambda e: _email_ist(e) or datetime.max)
                return out

            reply_pool = _collect_pool(False, False)
            acky_pool = _collect_pool(True, False) if allow_acky else list(reply_pool)
            ess_pool = _collect_pool(True, True)
            ack_pool = _collect_non_ess_ack_pool()
            direct_pool = _collect_direct_resolution_pool()
            consultant_ess_pool = _collect_consultant_ess_pool()
            seeded_reply_pool = _collect_seeded_reply_pool()

            anchor_li = group_sorted[0].get("list_index")
            a_dt = _parse_time_str(automation_rows[anchor_li].get("Actual Response Date & Time")) if anchor_li is not None else None
            a_ist = _to_ist(a_dt) if a_dt else None
            def _same_month_pool(pool_in):
                if not a_ist:
                    return list(pool_in)
                same_month = []
                for e in pool_in:
                    e_ist = _email_ist(e)
                    if e_ist and e_ist.year == a_ist.year and e_ist.month == a_ist.month:
                        same_month.append(e)
                if len(same_month) >= len(group_sorted):
                    return same_month
                return list(pool_in)

            reply_pool = _same_month_pool(reply_pool)
            acky_pool = _same_month_pool(acky_pool)
            ess_pool = _same_month_pool(ess_pool)
            ack_pool = _same_month_pool(ack_pool)
            direct_pool = _same_month_pool(direct_pool)
            consultant_ess_pool = _same_month_pool(consultant_ess_pool)
            seeded_reply_pool = _same_month_pool(seeded_reply_pool)

            group_indexes = {
                s.get("list_index")
                for s in group_sorted
                if s.get("list_index") is not None
            }
            used_outside_group = set()
            for other_state in row_states:
                other_li = other_state.get("list_index")
                if other_li is None or other_li in group_indexes or other_li >= len(automation_rows):
                    continue
                if selected_scope == "subject_wide_ess":
                    other_subject_norm = (other_state.get("subject_norm") or "").lower()
                    if not _group_subject_match(other_subject_norm):
                        continue
                    other_service_bucket = other_state.get("service_bucket") or _subject_service_bucket(
                        other_subject_norm,
                        other_state.get("service_no") or "",
                    )
                    if current_service_bucket and other_service_bucket and other_service_bucket != current_service_bucket:
                        continue
                else:
                    other_occ_key = other_state.get("occurrence_key") or _occurrence_group_key(
                        (other_state.get("subject_norm") or "").lower(),
                        other_state.get("requester") or "",
                        other_state.get("service_no") or "",
                    )
                    if other_occ_key != current_occ_key:
                        continue
                other_a_dt = _parse_time_str(automation_rows[other_li].get("Actual Response Date & Time"))
                other_a_ist = _to_ist(other_a_dt) if other_a_dt else None
                if other_a_ist:
                    used_outside_group.add(other_a_ist.replace(second=0, microsecond=0))

            def _drop_used(pool_in):
                pool_without_used = []
                for e in pool_in:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    minute_key = e_ist.replace(second=0, microsecond=0)
                    if minute_key in used_outside_group:
                        continue
                    pool_without_used.append(e)
                if len(pool_without_used) >= len(group_sorted):
                    return pool_without_used
                return list(pool_in)

            reply_pool = _drop_used(reply_pool)
            acky_pool = _drop_used(acky_pool)
            ess_pool = _drop_used(ess_pool)
            ack_pool = _drop_used(ack_pool)
            direct_pool = _drop_used(direct_pool)
            consultant_ess_pool = _drop_used(consultant_ess_pool)
            seeded_reply_pool = _drop_used(seeded_reply_pool)

            default_pool = (
                seeded_reply_pool if len(seeded_reply_pool) >= len(group_sorted)
                else reply_pool if len(reply_pool) >= len(group_sorted)
                else acky_pool if len(acky_pool) >= len(group_sorted)
                else consultant_ess_pool if len(consultant_ess_pool) >= len(group_sorted)
                else ess_pool
            )
            if len(default_pool) < len(group_sorted):
                return None

            pick = default_pool[min(slot_index, len(default_pool) - 1)]
            pick_ist = _email_ist(pick)
            if not pick_ist:
                return None

            return {
                "occ_key": current_occ_key,
                "group": group_sorted,
                "group_size": len(group_sorted),
                "slot_index": slot_index,
                "pool": default_pool,
                "reply_pool": reply_pool,
                "seeded_reply_pool": seeded_reply_pool,
                "acky_pool": acky_pool,
                "ess_pool": ess_pool,
                "consultant_ess_pool": consultant_ess_pool,
                "ack_pool": ack_pool,
                "direct_pool": direct_pool,
                "pick": pick,
                "pick_when": pick_ist.replace(second=0, microsecond=0),
                "lane_kind": "reply",
                "merged": merged,
                "notes_l": (debug_rows[list_index].get("Notes", "") or "").lower() if list_index < len(debug_rows) else "",
                "subject_norm_value": subject_norm_value,
                "scope": selected_scope,
            }

        def _shared_occurrence_fill_plan(state, *, quoted_sources=None, c_ist=None):
            lane_plan = _shared_occurrence_lane_plan(state)
            if not lane_plan:
                return None

            def _remember_decision(plan_obj):
                if not state or not plan_obj:
                    return plan_obj
                lane_kind = plan_obj.get("lane_kind") or "reply"
                lane_when = plan_obj.get("pick_when")
                slot_index = plan_obj.get("slot_index", 0)
                fill_style = "all_three_same" if _is_authoritative_occurrence_lane(lane_kind) else "lane_guided"
                confidence = "strong" if _is_authoritative_occurrence_lane(lane_kind) else "moderate"
                decision = {
                    "owner": "shared_occurrence",
                    "row_type": lane_kind,
                    "occurrence_slot": slot_index,
                    "lane_time": lane_when,
                    "fill_style": fill_style,
                    "confidence": confidence,
                }
                if lane_when and fill_style == "all_three_same":
                    decision["triplet"] = (lane_when, lane_when, lane_when)
                existing = state.get("shared_decision") or {}
                if (
                    existing.get("owner") == "shared_occurrence"
                    and _is_authoritative_occurrence_lane(existing.get("row_type") or "")
                    and existing.get("fill_style") == "all_three_same"
                    and existing.get("confidence") == "strong"
                ):
                    existing_triplet = existing.get("triplet")
                    new_triplet = decision.get("triplet")
                    # Do not let later weaker/local recomputation downgrade a
                    # previously established strong occurrence-owned ESS lane.
                    if not _is_authoritative_occurrence_lane(lane_kind):
                        return plan_obj
                    if existing_triplet and new_triplet and existing_triplet != new_triplet:
                        return plan_obj
                state["shared_decision"] = decision
                return plan_obj

            notes_l = lane_plan.get("notes_l", "")
            subject_norm_value = lane_plan.get("subject_norm_value", "")
            merged = lane_plan.get("merged") or []
            row_tokens = _match_tokens(subject_norm_value)
            group_size = lane_plan.get("group_size", 0)
            slot_index = lane_plan.get("slot_index", 0)

            def _pick_from(pool_in, lane_kind: str):
                if not pool_in or len(pool_in) < group_size:
                    return None
                pick = pool_in[min(slot_index, len(pool_in) - 1)]
                pick_ist = _email_ist(pick)
                if not pick_ist:
                    return None
                return _remember_decision({
                    **lane_plan,
                    "pool": pool_in,
                    "pick": pick,
                    "pick_when": pick_ist.replace(second=0, microsecond=0),
                    "lane_kind": lane_kind,
                })

            strong_non_ess_live = False
            if "requester span(all-ack->ess)" in notes_l and "ess-only; no non-ess request" in notes_l:
                # Treat non-ESS live evidence as strong only when it can
                # actually furnish a lane for this repeated family, not merely
                # because some unrelated older non-ESS reply exists somewhere in
                # the merged thread history.
                reply_pool = lane_plan.get("reply_pool") or []
                ack_pool = lane_plan.get("ack_pool") or []
                direct_pool = lane_plan.get("direct_pool") or []
                if slot_index < len(direct_pool):
                    strong_non_ess_live = True
                elif (
                    len(reply_pool) >= max(1, group_size)
                    and slot_index < len(reply_pool)
                    and slot_index < len(ack_pool)
                ):
                    strong_non_ess_live = True

                strong_non_ess_quoted = False
                if quoted_sources and c_ist:
                    ess_email_set = {e.strip().lower() for e in ess_team or []}
                    day_start = datetime(c_ist.year, c_ist.month, c_ist.day, tzinfo=c_ist.tzinfo)
                    day_end = day_start + timedelta(days=1)
                    quoted_non_ess = []
                    quoted_ess = []
                    for e in quoted_sources:
                        for from_line, sent_ist in _extract_quoted_blocks(e, subject_norm_value):
                            if sent_ist < day_start or sent_ist >= day_end:
                                continue
                            addr_hits = re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\\.[A-Z]{2,}", from_line, flags=re.I)
                            if not addr_hits:
                                continue
                            emails_l = [em.lower() for em in addr_hits]
                            if any(em in ess_email_set for em in emails_l):
                                quoted_ess.append(sent_ist)
                            else:
                                quoted_non_ess.append(sent_ist)
                    if quoted_non_ess and quoted_ess:
                        quoted_non_ess.sort()
                        quoted_ess.sort()
                        valid_pairs = []
                        seen_ack_minutes = set()
                        for ack_ist in quoted_ess:
                            reqs = [r for r in quoted_non_ess if r < ack_ist]
                            if not reqs:
                                continue
                            req_ist = reqs[-1]
                            if (ack_ist - req_ist) <= timedelta(minutes=16):
                                ack_minute = ack_ist.replace(second=0, microsecond=0)
                                if ack_minute in seen_ack_minutes:
                                    continue
                                seen_ack_minutes.add(ack_minute)
                                valid_pairs.append((req_ist, ack_ist))
                        if slot_index < len(valid_pairs):
                            strong_non_ess_quoted = True

                if (not strong_non_ess_live) and (not strong_non_ess_quoted):
                    plan = _pick_from(lane_plan.get("consultant_ess_pool") or [], "ess_over_ess")
                    if not plan:
                        plan = _pick_from(lane_plan.get("ess_pool") or [], "ess_over_ess")
                    if plan:
                        return plan

            seeded_plan = _pick_from(lane_plan.get("seeded_reply_pool") or [], "reply")
            if seeded_plan:
                return seeded_plan
            plan = _pick_from(lane_plan.get("reply_pool") or [], "reply")
            if plan:
                return plan
            plan = _pick_from(lane_plan.get("acky_pool") or [], "ess_acky")
            if plan:
                return plan
            plan = _pick_from(lane_plan.get("consultant_ess_pool") or [], "ess_over_ess")
            if not plan:
                plan = _pick_from(lane_plan.get("ess_pool") or [], "ess_over_ess")
            if plan:
                return plan
            return _remember_decision(lane_plan)

        _all_ack_ess_assignment_cache = {}

        def _occurrence_row_anchor_ist(slot_state):
            if not slot_state:
                return None
            slot_list_index = slot_state.get("list_index")
            if slot_list_index is None or slot_list_index >= len(automation_rows):
                return None
            row_vals = automation_rows[slot_list_index]
            triplet = _current_row_triplet_ist(row_vals)
            if triplet:
                return triplet[2] or triplet[1] or triplet[0]
            for col_name in (
                "Actual Resolved Date & Time",
                "Actual Response Date & Time",
                "Created Date & Time",
            ):
                dt = _parse_time_str(row_vals.get(col_name) or "")
                ist = _to_ist(dt) if dt else None
                if ist:
                    return ist.replace(second=0, microsecond=0)
            return None

        def _occurrence_group_anchor_ists(group_sorted):
            anchors = []
            seen = set()
            for slot_state in group_sorted or []:
                anchor_ist = _occurrence_row_anchor_ist(slot_state)
                if not anchor_ist:
                    continue
                minute_key = anchor_ist.replace(second=0, microsecond=0)
                if minute_key in seen:
                    continue
                seen.add(minute_key)
                anchors.append(minute_key)
            anchors.sort()
            return anchors

        def _select_occurrence_anchor_window(pool_in, anchor_ists, rows_needed):
            if not pool_in:
                return [], "empty"
            local_pool = []
            if anchor_ists:
                lower = anchor_ists[0] - timedelta(hours=48)
                upper = anchor_ists[-1] + timedelta(hours=24)
                for email_obj in pool_in:
                    email_ist = _email_ist(email_obj)
                    if email_ist and lower <= email_ist <= upper:
                        local_pool.append(email_obj)
            if local_pool:
                return (
                    local_pool,
                    "family_anchor_window" if len(local_pool) >= rows_needed else "family_anchor_window_partial",
                )
            if len(pool_in) >= rows_needed:
                return list(pool_in[-rows_needed:]), "latest_n_pool"
            return list(pool_in), "full_pool"

        def _assign_occurrence_rows_by_anchor(group_sorted, pool_in):
            if not group_sorted or not pool_in:
                return None
            anchor_ists = _occurrence_group_anchor_ists(group_sorted)
            selected_pool, pool_label = _select_occurrence_anchor_window(pool_in, anchor_ists, len(group_sorted))
            if not selected_pool:
                return None

            row_anchor_pairs = []
            for slot_state in group_sorted:
                row_anchor_pairs.append((slot_state, _occurrence_row_anchor_ist(slot_state)))

            distinct_anchors = {
                anchor.replace(second=0, microsecond=0)
                for _slot_state, anchor in row_anchor_pairs
                if anchor is not None
            }
            use_unique_anchor_match = len(distinct_anchors) >= len(group_sorted) and len(selected_pool) >= len(group_sorted)

            assignments = {}
            available = list(selected_pool)
            if use_unique_anchor_match:
                sorted_pairs = sorted(
                    row_anchor_pairs,
                    key=lambda item: (item[1] is None, item[1] or datetime.max),
                )
                for slot_state, anchor_ist in sorted_pairs:
                    if not available:
                        return None
                    if anchor_ist is None:
                        chosen = available.pop(0)
                    else:
                        chosen_idx = min(
                            range(len(available)),
                            key=lambda idx: abs(((_email_ist(available[idx]) or anchor_ist) - anchor_ist).total_seconds()),
                        )
                        chosen = available.pop(chosen_idx)
                    slot_list_index = slot_state.get("list_index")
                    if slot_list_index is None:
                        continue
                    assignments[slot_list_index] = chosen
            else:
                for slot_state, anchor_ist in row_anchor_pairs:
                    if anchor_ist is None:
                        chosen = selected_pool[-1]
                    else:
                        chosen = min(
                            selected_pool,
                            key=lambda email_obj: abs(((_email_ist(email_obj) or anchor_ist) - anchor_ist).total_seconds()),
                        )
                    slot_list_index = slot_state.get("list_index")
                    if slot_list_index is None:
                        continue
                    assignments[slot_list_index] = chosen

            if len(assignments) < len(group_sorted):
                return None
            return {
                "assignments": assignments,
                "pool": selected_pool,
                "pool_label": pool_label,
            }

        def _occurrence_slot_shape_hints(state):
            if not state:
                return {
                    "same_time_hint": False,
                    "has_local_triplet": False,
                    "has_local_same_time": False,
                    "notes_l": "",
                }
            list_index = state.get("list_index")
            notes_l = ""
            created_src_l = ""
            if list_index is not None and list_index < len(debug_rows):
                notes_l = (debug_rows[list_index].get("Notes", "") or "").lower()
                created_src_l = (debug_rows[list_index].get("CreatedSource", "") or "").lower()

            initial_lane_episode = state.get("initial_lane_episode") or {}
            lane_mode_l = (initial_lane_episode.get("mode") or "").lower()

            has_local_same_time = bool(
                lane_mode_l == "all-three-same"
                or "lanelocalinitialepisode[all-three-same]" in notes_l
                or "force prod subject; all times same" in notes_l
            )
            has_local_triplet = bool(
                lane_mode_l in {"direct-reply", "req-ack-reply"}
                or "lanelocalinitialepisode[direct-reply]" in notes_l
                or "lanelocalinitialepisode[req-ack-reply]" in notes_l
                or created_src_l == "parsed_from_quoted_request"
            )
            same_time_hint = bool(
                has_local_same_time
                or "ess initiated; no ack; no consultant reply after request" in notes_l
                or "failed subject; ess initiated; no ack phrase" in notes_l
                or "ess-only; no non-ess request" in notes_l
            )
            return {
                "same_time_hint": same_time_hint,
                "has_local_triplet": has_local_triplet,
                "has_local_same_time": has_local_same_time,
                "notes_l": notes_l,
            }

        def _occurrence_slot_authoritative_same_time_plan(state):
            lane_plan = _shared_occurrence_lane_plan(state)
            if not lane_plan or (lane_plan.get("group_size") or 0) < 2:
                return None

            shape_hints = _occurrence_slot_shape_hints(state)
            if shape_hints["has_local_triplet"] and not shape_hints["has_local_same_time"]:
                return None
            if not shape_hints["same_time_hint"]:
                return None

            slot_index = lane_plan.get("slot_index", 0)
            group_size = lane_plan.get("group_size", 0)
            for pool_name, lane_kind in (
                ("consultant_ess_pool", "ess_over_ess"),
                ("ess_pool", "ess_over_ess"),
            ):
                pool = lane_plan.get(pool_name) or []
                if len(pool) < group_size:
                    continue
                pick = pool[min(slot_index, len(pool) - 1)]
                pick_ist = _email_ist(pick)
                if not pick_ist:
                    continue
                return {
                    **lane_plan,
                    "pool": pool,
                    "pick": pick,
                    "pick_when": pick_ist.replace(second=0, microsecond=0),
                    "lane_kind": lane_kind,
                    "slot_shape": "all_three_same",
                }
            return None

        def _is_occurrence_acky_candidate(e, requester_names) -> bool:
            e_ist = _email_ist(e)
            if not e_ist:
                return False
            if not requester_names:
                return False
            if not any(
                _match_requester(e.sender_name, e.sender_email, req_name)
                for req_name in requester_names
                if req_name
            ):
                return False
            flags = _shared_reply_flags(email_record=e)
            if _system_like_sender(e) or flags["ignore_reply"]:
                return False
            return bool(flags["substantive_reply"] or flags["ack_candidate"])

        def _subject_wide_all_ack_ess_assignment_map(state):
            if not state:
                return {}
            list_index = state.get("list_index")
            requester = state.get("requester") or ""
            subject_norm_value = _state_occurrence_family_subject_norm(state)
            service_no = state.get("service_no") or ""
            current_service_bucket = state.get("service_bucket") or _subject_service_bucket(subject_norm_value, service_no)
            if list_index is None or not requester or not subject_norm_value:
                return {}

            notes_l = _state_occurrence_notes_l(state)
            if not _state_is_all_ack_to_ess(state):
                return {}

            def _family_subject_norm(other_state) -> str:
                return _state_occurrence_family_subject_norm(other_state)

            def _group_subject_match(other_state) -> bool:
                other_norm = _family_subject_norm(other_state)
                if not subject_norm_value or not other_norm:
                    return False
                if subject_norm_value == other_norm:
                    return True
                return subject_norm_value in other_norm or other_norm in subject_norm_value

            group_sorted = []
            group_requesters = set()
            for s in row_states:
                other_li = s.get("list_index")
                if other_li is None or other_li >= len(automation_rows):
                    continue
                if not _state_is_occurrence_family_candidate(s):
                    continue
                if not _group_subject_match(s):
                    continue
                group_sorted.append(s)
                other_requester = (s.get("requester") or "").strip()
                if other_requester:
                    group_requesters.add(other_requester)

            if len(group_sorted) < 2:
                fallback_group = []
                fallback_requesters = set()
                for s in row_states:
                    other_li = s.get("list_index")
                    if other_li is None or other_li >= len(automation_rows):
                        continue
                    if s.get("is_dep_req") or s.get("is_dep_succ"):
                        continue
                    if not _group_subject_match(s):
                        continue
                    other_subject_norm = _family_subject_norm(s)
                    other_service_bucket = s.get("service_bucket") or _subject_service_bucket(
                        other_subject_norm,
                        s.get("service_no") or "",
                    )
                    if current_service_bucket and other_service_bucket and other_service_bucket != current_service_bucket:
                        continue
                    fallback_group.append(s)
                    other_requester = (s.get("requester") or "").strip()
                    if other_requester:
                        fallback_requesters.add(other_requester)
                if len(fallback_group) >= 2:
                    group_sorted = fallback_group
                    group_requesters = fallback_requesters

            if len(group_sorted) < 2:
                return {}

            group_sorted.sort(key=lambda x: x.get("row_index") or 10**9)
            group_anchor_ists = _occurrence_group_anchor_ists(group_sorted)
            cache_key = (
                tuple(s.get("list_index") for s in group_sorted),
                subject_norm_value,
                tuple((s.get("list_index"), _occurrence_row_anchor_ist(s)) for s in group_sorted),
            )
            cached = _all_ack_ess_assignment_cache.get(cache_key)
            if cached is not None:
                return cached

            requester_names = tuple(sorted(group_requesters)) if group_requesters else ((requester,) if requester else tuple())
            current_iface_tokens = _interface_tokens(subject_norm_value)

            def _group_req_match(email_obj):
                if not requester_names:
                    return False
                return any(
                    _match_requester(email_obj.sender_name, email_obj.sender_email, req_name)
                    for req_name in requester_names
                    if req_name
                )

            def _subject_match_override(email_subject: str) -> bool:
                e_norm = normalize_subject(email_subject or "")
                if not subject_norm_value or not e_norm:
                    return False
                return _fresh_picker_subject_safe(
                    subject_norm_value,
                    e_norm,
                    iface_tokens=current_iface_tokens,
                    allow_added_inc=True,
                )

            def _dedupe_group_ess_minutes(items):
                buckets = {}
                for item in items:
                    item_ist = _email_ist(item)
                    if not item_ist:
                        continue
                    minute_key = item_ist.replace(second=0, microsecond=0)
                    buckets.setdefault(minute_key, []).append(item)

                out = []
                for minute_key in sorted(buckets):
                    bucket = buckets[minute_key]
                    bucket.sort(
                        key=lambda email_obj: (
                            0 if _group_req_match(email_obj) else 1,
                            0 if (_shared_reply_classification(email_obj).get("direct_resolution")) else 1,
                            0 if (_shared_reply_classification(email_obj).get("real_reply")) else 1,
                            _email_ist(email_obj) or datetime.max,
                        )
                    )
                    out.append(bucket[0])
                return out

            consultant_ess_pool = []
            ess_pool = []
            occurrence_acky_pool = []
            for e in emails:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if not _subject_match_override(getattr(e, "subject", "") or ""):
                    continue
                flags = _shared_reply_flags(email_record=e)
                if _ess_sender(e):
                    if _system_like_sender(e) or flags["ignore_reply"]:
                        continue
                    if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                        continue
                    ess_pool.append(e)
                    if _group_req_match(e):
                        consultant_ess_pool.append(e)
                if _is_occurrence_acky_candidate(e, requester_names):
                    occurrence_acky_pool.append(e)

            consultant_ess_pool = _dedupe_group_ess_minutes(consultant_ess_pool)
            ess_pool = _dedupe_group_ess_minutes(ess_pool)
            occurrence_acky_pool = _dedupe_group_ess_minutes(occurrence_acky_pool)

            anchor_li = group_sorted[0].get("list_index")
            a_dt = _parse_time_str(automation_rows[anchor_li].get("Actual Response Date & Time")) if anchor_li is not None else None
            a_ist = _to_ist(a_dt) if a_dt else None

            def _same_month_pool(pool_in):
                if not a_ist:
                    return list(pool_in)
                same_month = []
                for e in pool_in:
                    e_ist = _email_ist(e)
                    if e_ist and e_ist.year == a_ist.year and e_ist.month == a_ist.month:
                        same_month.append(e)
                if len(same_month) >= len(group_sorted):
                    return same_month
                return list(pool_in)

            consultant_ess_pool = _same_month_pool(consultant_ess_pool)
            ess_pool = _same_month_pool(ess_pool)
            occurrence_acky_pool = _same_month_pool(occurrence_acky_pool)

            selected_assignment = None
            lane_kind = ""
            for candidate_pool, candidate_lane_kind in (
                (consultant_ess_pool, "ess_over_ess"),
                (ess_pool, "ess_over_ess"),
                (occurrence_acky_pool, "ess_acky_sequence"),
            ):
                selected_assignment = _assign_occurrence_rows_by_anchor(group_sorted, candidate_pool)
                if selected_assignment:
                    lane_kind = candidate_lane_kind
                    break
            if not selected_assignment:
                _all_ack_ess_assignment_cache[cache_key] = {}
                return {}

            assignments_by_line = selected_assignment["assignments"]
            pool = selected_assignment["pool"]
            ordered_group = sorted(
                group_sorted,
                key=lambda slot_state: (
                    (_email_ist(assignments_by_line.get(slot_state.get("list_index"))) is None),
                    _email_ist(assignments_by_line.get(slot_state.get("list_index"))) or datetime.max,
                    slot_state.get("row_index") or 10**9,
                ),
            )

            assignment_map = {}
            occ_key_default = state.get("occurrence_key") or _occurrence_group_key(subject_norm_value, requester, service_no)
            for slot_index, slot_state in enumerate(ordered_group):
                slot_list_index = slot_state.get("list_index")
                if slot_list_index is None:
                    continue
                pick = assignments_by_line.get(slot_list_index)
                pick_ist = _email_ist(pick)
                if not pick_ist:
                    continue
                assignment_map[slot_list_index] = {
                    "occ_key": slot_state.get("occurrence_key") or occ_key_default,
                    "group": group_sorted,
                    "group_size": len(group_sorted),
                    "slot_index": slot_index,
                    "pick": pick,
                    "pick_when": pick_ist.replace(second=0, microsecond=0),
                    "lane_kind": lane_kind,
                    "scope": "subject_wide_ess_override",
                    "pool": pool,
                    "pool_label": selected_assignment.get("pool_label") or "",
                    "group_anchor_ists": group_anchor_ists,
                    "consultant_ess_pool": consultant_ess_pool,
                    "ess_pool": ess_pool,
                    "occurrence_acky_pool": occurrence_acky_pool,
                }
            _all_ack_ess_assignment_cache[cache_key] = assignment_map
            return assignment_map

        def _subject_wide_all_ack_ess_override_plan(state):
            if not state:
                return None
            list_index = state.get("list_index")
            if list_index is None:
                return None
            assignment_map = _subject_wide_all_ack_ess_assignment_map(state)
            return assignment_map.get(list_index)

        def _preferred_shared_occurrence_plan(state, *, quoted_sources=None, c_ist=None, require_override_for_all_ack: bool = False):
            if not state:
                return None
            if _state_is_all_ack_to_ess(state):
                override_plan = _subject_wide_all_ack_ess_override_plan(state)
                if override_plan:
                    return override_plan
                if require_override_for_all_ack:
                    return None
            return _shared_occurrence_fill_plan(state, quoted_sources=quoted_sources, c_ist=c_ist)

        def _system_like_sender(e):
            sender = f"{getattr(e, 'sender_email', '') or ''} {getattr(e, 'sender_name', '') or ''}".lower()
            if not sender.strip():
                return False
            markers = (
                "system-notification",
                "system notification",
                "no-reply",
                "noreply",
                "do-not-reply",
                "donotreply",
                "mailer-daemon",
                "postmaster",
            )
            return any(m in sender for m in markers)

        def _can_use_reply_as_created_source(e, requester_name: str, *, all_same: bool = False) -> bool:
            if not e:
                return False
            if all_same:
                return True
            # Outside deliberate all-three-same collapse, do not let an ESS
            # ack-like mail become the Created anchor.
            if _ess_sender(e) and _is_shared_ack_candidate(e):
                return False
            return True

        def _reply_choice_rank(e, requester_name: str, *, created_role: bool = False, all_same: bool = False):
            flags = _shared_reply_flags(email_record=e)
            requester_match = _req_match(e, requester_name)
            system_like = _system_like_sender(e)
            created_ok = _can_use_reply_as_created_source(
                e,
                requester_name,
                all_same=all_same,
            )
            return (
                0 if requester_match else 1,
                0 if (not created_role or created_ok) else 1,
                0 if flags["direct_resolution"] else 1,
                0 if flags["substantive_reply"] else 1,
                1 if flags["ack_candidate"] else 0,
                1 if system_like else 0,
                _email_ist(e) or datetime.max,
            )

        def _dedupe_reply_minutes_prefer_consultant(items, requester_name: str, *, created_role: bool = False, all_same: bool = False):
            buckets = {}
            for item in items:
                item_ist = _email_ist(item)
                if not item_ist:
                    continue
                minute_key = item_ist.replace(second=0, microsecond=0)
                buckets.setdefault(minute_key, []).append(item)
            out = []
            for minute_key in sorted(buckets):
                bucket = buckets[minute_key]
                bucket.sort(
                    key=lambda e: _reply_choice_rank(
                        e,
                        requester_name,
                        created_role=created_role,
                        all_same=all_same,
                    )
                )
                out.append(bucket[0])
            return out

        def _requester_pool(subject_norm_value, requester_name, center_ist=None, day_window=21):
            center_key = ""
            if center_ist:
                try:
                    center_key = center_ist.date().isoformat()
                except Exception:
                    center_key = ""
            key = (subject_norm_value or "", requester_name or "", center_key, int(day_window))
            if key in _requester_pool_cache:
                return _requester_pool_cache[key]

            full_key = (subject_norm_value or "", requester_name or "")
            full_pool = _requester_pool_full_cache.get(full_key)
            if full_pool is None:
                row_tokens = _match_tokens(subject_norm_value or "")
                full_pool = []
                for e in emails:
                    if not getattr(e, "sent_time", None):
                        continue
                    if requester_name and not _req_match(e, requester_name):
                        continue
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(e.subject or "")
                        if not _strict_subject_identity_gate(
                            subject_norm_value or "",
                            s_norm,
                            iface_tokens=_interface_tokens(subject_norm_value or ""),
                        ):
                            continue
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm_value and s_norm and (subject_norm_value in s_norm or s_norm in subject_norm_value))
                        if score < 0.45 and not contains and not _fresh_picker_subject_safe(
                            subject_norm_value,
                            s_norm,
                            iface_tokens=_interface_tokens(subject_norm_value or ""),
                            allow_added_inc=True,
                        ):
                            continue
                    full_pool.append(e)
                full_pool.sort(key=lambda e: e.sent_time)
                _requester_pool_full_cache[full_key] = full_pool

            if center_ist:
                max_delta_seconds = day_window * 24 * 3600
                pool = []
                for e in full_pool:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if abs((e_ist - center_ist).total_seconds()) <= max_delta_seconds:
                        pool.append(e)
            else:
                pool = list(full_pool)
            _requester_pool_cache[key] = pool
            return pool

        def _row_subject_match_email(e, subject_norm_value: str, row_tokens: set, row_id_tokens: set) -> bool:
            cache_key = (
                id(e),
                subject_norm_value or "",
                tuple(sorted(row_tokens)) if row_tokens else (),
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
            )
            cached = _row_subject_match_cache.get(cache_key)
            if cached is not None:
                return cached
            s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
            if s_norm and not _strict_subject_identity_gate(
                subject_norm_value or "",
                s_norm,
                iface_tokens=_interface_tokens(subject_norm_value or ""),
            ):
                _row_subject_match_cache[cache_key] = False
                return False
            if row_tokens and s_norm and not _quoted_subject_variant_owns_row(subject_norm_value, s_norm):
                _row_subject_match_cache[cache_key] = False
                return False
            score = _token_overlap_score(row_tokens, _match_tokens(s_norm)) if row_tokens else 0.0
            contains = bool(subject_norm_value and s_norm and (subject_norm_value in s_norm or s_norm in subject_norm_value))
            if row_id_tokens:
                s_ids = _id_token_cache.get(s_norm)
                if s_ids is None:
                    s_ids = _id_like_tokens(s_norm)
                    _id_token_cache[s_norm] = s_ids
                has_id_overlap = bool(s_ids and not row_id_tokens.isdisjoint(s_ids))
                if not has_id_overlap:
                    has_id_overlap = _subject_has_id_token(getattr(e, "subject", "") or "", row_id_tokens)
                if not has_id_overlap:
                    _row_subject_match_cache[cache_key] = False
                    return False
                if not row_tokens:
                    _row_subject_match_cache[cache_key] = True
                    return True
                out = (
                    score >= 0.45
                    or contains
                    or _fresh_picker_subject_safe(
                        subject_norm_value,
                        s_norm,
                        iface_tokens=_interface_tokens(subject_norm_value),
                        allow_added_inc=True,
                    )
                )
                _row_subject_match_cache[cache_key] = out
                return out
            if not row_tokens:
                _row_subject_match_cache[cache_key] = True
                return True
            out = score >= 0.45 or contains
            _row_subject_match_cache[cache_key] = out
            return out

        def _row_subject_match_email_quoted(e, subject_norm_value: str, row_tokens: set, row_id_tokens: set) -> bool:
            if not _row_subject_match_email(e, subject_norm_value, row_tokens, row_id_tokens):
                return False
            candidate_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
            if not candidate_norm:
                return False
            return _quoted_subject_short_variant_ok(subject_norm_value, candidate_norm)

        _ess_only_short_ack_cache = {}
        def _ess_only_short_ack(e) -> bool:
            key = id(e)
            if key in _ess_only_short_ack_cache:
                return _ess_only_short_ack_cache[key]
            cls = _shared_reply_classification(e)
            out = bool(cls.get("short_ess_ack"))
            _ess_only_short_ack_cache[key] = out
            return out

        def _risk_guard_notes_match(notes_l: str) -> bool:
            return any(
                marker in notes_l
                for marker in (
                    "quotedrequestonly",
                    "quotedrequestonlyhybridliveack",
                    "ess-only; no non-ess request",
                    "requester span(",
                    "esscontinuationguard[",
                )
            )

        _risk_guard_precheck_cache = {}
        _episode_candidate_cache = {}
        _rewrite_guard_profile_cache = {}
        _live_reply_lane_cache = {}
        _live_reply_message_cache = {}

        def _risk_guard_precheck(state, row_vals, list_index):
            notes_l = (debug_rows[list_index].get("Notes", "") or "").lower() if list_index < len(debug_rows) else ""
            c_raw = row_vals.get("Created Date & Time") or ""
            a_raw = row_vals.get("Actual Response Date & Time") or ""
            r_raw = row_vals.get("Actual Resolved Date & Time") or ""
            cache_key = (list_index, notes_l, c_raw, a_raw, r_raw)
            cached = _risk_guard_precheck_cache.get(cache_key)
            if cached is not None:
                return cached

            c_dt = _parse_time_str(c_raw) if c_raw else None
            a_dt = _parse_time_str(a_raw) if a_raw else None
            r_dt = _parse_time_str(r_raw) if r_raw else None
            c_ist = _to_ist(c_dt) if c_dt else None
            a_ist = _to_ist(a_dt) if a_dt else None
            r_ist = _to_ist(r_dt) if r_dt else None

            suspicious_all_same = bool(c_ist and a_ist and r_ist and c_ist == a_ist == r_ist)
            suspicious_created_ack = bool(c_ist and a_ist and r_ist and c_ist == a_ist and r_ist > a_ist)
            out = {
                "notes_l": notes_l,
                "has_risky_notes": _risk_guard_notes_match(notes_l),
                "c_ist": c_ist,
                "a_ist": a_ist,
                "r_ist": r_ist,
                "suspicious_all_same": suspicious_all_same,
                "suspicious_created_ack": suspicious_created_ack,
            }
            _risk_guard_precheck_cache[cache_key] = out
            return out

        def _source_strength_value(src_value: str) -> int:
            src_l = (src_value or "").strip().lower()
            if not src_l:
                return 0
            if src_l.startswith("parsed_from_raw_eml"):
                return 20
            if src_l.startswith("parsed_from_quoted"):
                return 35
            if src_l.startswith("parsed_"):
                return 25
            return 70

        def _rewrite_guard_profile(row_vals, list_index):
            notes_l = (debug_rows[list_index].get("Notes", "") or "").lower() if list_index < len(debug_rows) else ""
            created_src = (debug_rows[list_index].get("CreatedSource") or "") if list_index < len(debug_rows) else ""
            ack_src = (debug_rows[list_index].get("AckSource") or "") if list_index < len(debug_rows) else ""
            resolved_src = (debug_rows[list_index].get("ResolvedSource") or "") if list_index < len(debug_rows) else ""
            c_raw = row_vals.get("Created Date & Time") or ""
            a_raw = row_vals.get("Actual Response Date & Time") or ""
            r_raw = row_vals.get("Actual Resolved Date & Time") or ""
            cache_key = (list_index, notes_l, created_src, ack_src, resolved_src, c_raw, a_raw, r_raw)
            cached = _rewrite_guard_profile_cache.get(cache_key)
            if cached is not None:
                return cached

            c_dt = _parse_time_str(c_raw) if c_raw else None
            a_dt = _parse_time_str(a_raw) if a_raw else None
            r_dt = _parse_time_str(r_raw) if r_raw else None
            c_ist = _to_ist(c_dt) if c_dt else None
            a_ist = _to_ist(a_dt) if a_dt else None
            r_ist = _to_ist(r_dt) if r_dt else None
            c_min = c_ist.replace(second=0, microsecond=0) if c_ist else None
            a_min = a_ist.replace(second=0, microsecond=0) if a_ist else None
            r_min = r_ist.replace(second=0, microsecond=0) if r_ist else None
            ordered = bool(c_ist and a_ist and r_ist and c_ist <= a_ist <= r_ist)
            ack_gap = (a_min - c_min) if (c_min and a_min and a_min >= c_min) else None
            all_same = bool(c_min and a_min and r_min and c_min == a_min == r_min)
            ack_src_l = ack_src.lower()
            strong_live_ack = bool(
                ordered
                and not all_same
                and ack_gap is not None
                and ack_gap <= timedelta(minutes=16)
                and c_min is not None
                and a_min is not None
                and c_min < a_min
                and (
                    "quotedrequestonlypreservedliveack" in notes_l
                    or (
                        ack_src
                        and not ack_src_l.startswith("parsed_from_")
                    )
                )
            )
            source_strength = max(
                _source_strength_value(created_src),
                _source_strength_value(ack_src),
                _source_strength_value(resolved_src),
            )
            rewrite_strength = source_strength
            if ordered:
                rewrite_strength = max(rewrite_strength, 45)
            if all_same:
                rewrite_strength = min(rewrite_strength or 20, 30)
            if ack_gap is not None and ack_gap <= timedelta(minutes=16) and not all_same:
                rewrite_strength = max(rewrite_strength, 80 if strong_live_ack else 60)
            if "created retained (response anchor unreliable)" in notes_l:
                rewrite_strength = min(rewrite_strength, 15)
            if "failedcreatedfromrequesterfirst" in notes_l:
                rewrite_strength = min(rewrite_strength, 18)
            if "dateanchorignoredstale" in notes_l:
                rewrite_strength = min(rewrite_strength, 20)
            if "quotedrequestonly" in notes_l and not strong_live_ack:
                rewrite_strength = min(rewrite_strength, 25)
            if "latestconsultantreply" in notes_l and created_src.lower().startswith("parsed_from_"):
                rewrite_strength = min(rewrite_strength, 25)
            low_confidence = rewrite_strength <= 25
            out = {
                "notes_l": notes_l,
                "created_src": created_src,
                "ack_src": ack_src,
                "resolved_src": resolved_src,
                "c_ist": c_ist,
                "a_ist": a_ist,
                "r_ist": r_ist,
                "c_min": c_min,
                "a_min": a_min,
                "r_min": r_min,
                "ordered": ordered,
                "ack_gap": ack_gap,
                "all_same": all_same,
                "strong_live_ack": strong_live_ack,
                "rewrite_strength": rewrite_strength,
                "low_confidence": low_confidence,
            }
            _rewrite_guard_profile_cache[cache_key] = out
            return out

        def _candidate_rewrite_strength(candidate_kind: str, owner_tag: str, cand_c_min, cand_a_min, cand_r_min) -> int:
            kind_key = (candidate_kind or "").strip().lower()
            owner_l = (owner_tag or "").lower()
            strength = {
                "occurrence_ess": 85,
                "system": 72,
                "seed_same_time": 70,
                "task": 34,
                "cleanup": 40,
                "sequence": 36,
                "hybrid": 58,
                "quoted": 50,
                "requester_ack": 42,
                "continuation": 32,
                "monotonic": 16,
                "risk": 20,
            }.get(kind_key, 48)
            if cand_c_min == cand_a_min == cand_r_min:
                strength -= 12
            elif cand_a_min == cand_r_min and cand_c_min < cand_a_min:
                strength += 8
            if "directreply" in owner_l or "reply-anchored" in owner_l:
                strength += 10
            if "systemnotification" in owner_l:
                strength += 6
            if "liverequestanchorguard" in owner_l:
                strength += 10
            return max(0, strength)

        def _live_reply_cache_key(list_index):
            state = state_by_list_index.get(list_index)
            row_vals = automation_rows[list_index] if list_index is not None and list_index < len(automation_rows) else {}
            created_raw = row_vals.get("Created Date & Time") or ""
            response_raw = row_vals.get("Actual Response Date & Time") or ""
            resolved_raw = row_vals.get("Actual Resolved Date & Time") or ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index is not None and list_index < len(debug_rows) else ""
            notes_l = (debug_rows[list_index].get("Notes", "") or "").lower() if list_index is not None and list_index < len(debug_rows) else ""
            requester = (state or {}).get("requester") or ""
            subject_norm_value = ((state or {}).get("subject_norm") or "").lower()
            thread = (state or {}).get("thread") or []
            thread_sig = tuple(id(e) for e in thread)
            return (
                list_index,
                requester,
                subject_norm_value,
                created_raw,
                response_raw,
                resolved_raw,
                resolved_src_now,
                notes_l,
                thread_sig,
            )

        def _live_reply_lane_for_row(list_index):
            cache_key = _live_reply_cache_key(list_index)
            cached = _live_reply_lane_cache.get(cache_key)
            if cached is not None:
                return cached
            state = state_by_list_index.get(list_index)
            if not state:
                _live_reply_lane_cache[cache_key] = None
                return None
            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            subject_norm_value = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time") or "")
            r_ist = _to_ist(r_dt) if r_dt else None
            if not (requester and subject_norm_value and base_thread and r_ist):
                _live_reply_lane_cache[cache_key] = None
                return None
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            row_tokens = _match_tokens(subject_norm_value)
            row_id_tokens = _id_like_tokens(subject_norm_value)
            expanded_thread = _expanded_thread(
                subject_norm_value,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=r_ist,
            )
            def _best_reply_from(thread_in):
                if not thread_in:
                    return None
                best_local = None
                for e in thread_in:
                    if not e.sent_time or not _req_match(e, requester):
                        continue
                    if not _is_real_reply_candidate(e):
                        continue
                    if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                        continue
                    if not _row_subject_match_email(e, subject_norm_value, row_tokens, row_id_tokens):
                        continue
                    e_ist = _email_ist(e)
                    if not e_ist or abs((e_ist - r_ist).total_seconds()) > 300:
                        continue
                    src_match = 0
                    if resolved_src_now and _match_requester(resolved_src_now, resolved_src_now, requester):
                        sender_src = e.sender_email or e.sender_name or ""
                        if sender_src and resolved_src_now and sender_src.lower() == str(resolved_src_now).lower():
                            src_match = 1
                    cand = (src_match, -abs((e_ist - r_ist).total_seconds()), e_ist, id(e), e)
                    if best_local is None or cand > best_local:
                        best_local = cand
                return best_local

            best = _best_reply_from(base_thread)
            if best is None and expanded_thread is not base_thread:
                best = _best_reply_from(expanded_thread)
            out = best[2] if best else None
            _live_reply_lane_cache[cache_key] = out
            return out

        def _live_reply_message_for_row(list_index):
            cache_key = _live_reply_cache_key(list_index)
            cached = _live_reply_message_cache.get(cache_key)
            if cached is not None:
                return cached

            live_reply_ist = _live_reply_lane_for_row(list_index)
            state = state_by_list_index.get(list_index)
            if not (live_reply_ist and state):
                _live_reply_message_cache[cache_key] = None
                return None

            requester = state.get("requester") or ""
            subject_norm_value = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            row_tokens = _match_tokens(subject_norm_value)
            row_id_tokens = _id_like_tokens(subject_norm_value)
            if not (requester and subject_norm_value and base_thread):
                _live_reply_message_cache[cache_key] = None
                return None

            expanded_thread = _expanded_thread(
                subject_norm_value,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=live_reply_ist,
            )
            def _best_message_from(thread_in):
                if not thread_in:
                    return None
                best_local = None
                for e in thread_in:
                    if not _req_match(e, requester):
                        continue
                    if not _is_real_reply_candidate(e):
                        continue
                    if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                        continue
                    if not _row_subject_match_email(e, subject_norm_value, row_tokens, row_id_tokens):
                        continue
                    e_ist = _email_ist(e)
                    if not e_ist or abs((e_ist - live_reply_ist).total_seconds()) > 300:
                        continue
                    cand = (-abs((e_ist - live_reply_ist).total_seconds()), e_ist, id(e), e)
                    if best_local is None or cand > best_local:
                        best_local = cand
                return best_local

            best = _best_message_from(base_thread)
            if best is None and expanded_thread is not base_thread:
                best = _best_message_from(expanded_thread)

            out = best[3] if best else None
            _live_reply_message_cache[cache_key] = out
            return out

        def _episode_selection_score(
            *,
            list_index,
            c_min,
            a_min,
            r_min,
            base_strength: int,
            created_src: str = "",
            ack_src: str = "",
            resolved_src: str = "",
            notes_l: str = "",
            candidate_kind: str = "",
            owner_tag: str = "",
        ) -> int:
            if not (c_min and a_min and r_min):
                return -9999

            score = max(0, int(base_strength or 0))
            score = max(
                score,
                _source_strength_value(created_src),
                _source_strength_value(ack_src),
                _source_strength_value(resolved_src),
            )

            if c_min < a_min < r_min:
                score += 10
            elif c_min < a_min == r_min:
                score += 7
            elif c_min == a_min == r_min:
                score -= 18

            ack_gap = (a_min - c_min) if a_min >= c_min else None
            if ack_gap is not None:
                if ack_gap <= timedelta(minutes=16):
                    score += 12
                elif ack_gap <= timedelta(hours=1):
                    score += 4
                elif ack_gap >= timedelta(hours=24):
                    score -= 8

            owner_l = (owner_tag or "").lower()
            kind_l = (candidate_kind or "").lower()
            if "directreply" in owner_l or "reply-anchored" in owner_l:
                score += 10
            if kind_l in {"occurrence_ess", "system"}:
                score += 8

            live_reply_ist = _live_reply_lane_for_row(list_index)
            if live_reply_ist:
                live_reply_min = live_reply_ist.replace(second=0, microsecond=0)
                if r_min == live_reply_min:
                    # Presence of the real consultant reply lane is the strongest
                    # stability signal for row-owned episodes.
                    score += 24
                    if a_min <= live_reply_min:
                        score += 4
                elif kind_l in {"quoted", "hybrid", "risk", "requester_ack"}:
                    score -= 20

            notes_l = (notes_l or "").lower()
            if "created retained (response anchor unreliable)" in notes_l:
                score -= 16
            if "dateanchorignoredstale" in notes_l:
                score -= 8
            if "quotedrequestonly" in notes_l and kind_l not in {"quoted", "hybrid"}:
                score -= 6

            return score

        def _allow_guard_rewrite(
            row_vals,
            list_index,
            cand_c_ist,
            cand_a_ist,
            cand_r_ist,
            owner_tag: str,
            candidate_kind: str,
            *,
            created_src: str = "",
            ack_src: str = "",
            resolved_src: str = "",
        ) -> bool:
            def _trace_decision(decision: bool, reason: str) -> bool:
                return decision

            if not (cand_c_ist and cand_a_ist and cand_r_ist):
                return _trace_decision(False, "missing_candidate_time")
            cand_c_min = cand_c_ist.replace(second=0, microsecond=0)
            cand_a_min = cand_a_ist.replace(second=0, microsecond=0)
            cand_r_min = cand_r_ist.replace(second=0, microsecond=0)
            if not (cand_c_min <= cand_a_min <= cand_r_min):
                return _trace_decision(False, "candidate_not_ordered")

            state = state_by_list_index.get(list_index)
            if state and state.get("occurrence_locked"):
                locked_triplet = state.get("occurrence_lock_triplet")
                if locked_triplet:
                    return _trace_decision((cand_c_min, cand_a_min, cand_r_min) == tuple(locked_triplet), "occurrence_locked_triplet")
                return _trace_decision(False, "occurrence_locked_no_triplet")

            profile = _rewrite_guard_profile(row_vals, list_index)
            if not profile.get("ordered"):
                return _trace_decision(True, "current_not_ordered")

            cur_c_min = profile.get("c_min")
            cur_a_min = profile.get("a_min")
            cur_r_min = profile.get("r_min")
            if not (cur_c_min and cur_a_min and cur_r_min):
                return True
            notes_l = profile.get("notes_l", "")
            current_created_src = profile.get("created_src", "") or ""
            current_ack_src = profile.get("ack_src", "") or ""
            current_resolved_src = profile.get("resolved_src", "") or ""
            current_strength = profile.get("rewrite_strength", 0)
            low_confidence = profile.get("low_confidence", False)
            candidate_strength = _candidate_rewrite_strength(
                candidate_kind,
                owner_tag,
                cand_c_min,
                cand_a_min,
                cand_r_min,
            )
            repair_only_candidate_kinds = {
                "cleanup",
                "requester_ack",
                "continuation",
                "sequence",
                "risk",
                "monotonic",
            }
            stronger_owner_kinds = {"occurrence_ess", "system"}
            current_episode_score = _episode_selection_score(
                list_index=list_index,
                c_min=cur_c_min,
                a_min=cur_a_min,
                r_min=cur_r_min,
                base_strength=current_strength,
                created_src=current_created_src,
                ack_src=current_ack_src,
                resolved_src=current_resolved_src,
                notes_l=notes_l,
                candidate_kind="current",
                owner_tag="current",
            )
            candidate_episode_score = _episode_selection_score(
                list_index=list_index,
                c_min=cand_c_min,
                a_min=cand_a_min,
                r_min=cand_r_min,
                base_strength=candidate_strength,
                created_src=created_src,
                ack_src=ack_src,
                resolved_src=resolved_src,
                candidate_kind=candidate_kind,
                owner_tag=owner_tag,
            )
            initial_lane_episode = (state or {}).get("initial_lane_episode") if state else None
            candidate_owner_l = (owner_tag or "").lower()
            candidate_is_seeded_owner = (
                "lanelocalinitialepisode" in candidate_owner_l
                or "seedprimaryresolver" in candidate_owner_l
            )
            candidate_is_quoted_family = (
                candidate_kind in {"quoted", "hybrid"}
                or "quoted" in candidate_owner_l
                or "lanelocalinitialepisode" in candidate_owner_l
            )
            candidate_is_deep_proven_pair = bool(
                candidate_kind == "hybrid"
                and created_src == "PARSED_FROM_QUOTED_REQUEST"
                and ack_src == "PARSED_FROM_QUOTED_ACK"
                and cand_c_min < cand_a_min <= cand_r_min
            )
            candidate_is_shallow_fallback = bool(
                candidate_kind in {"quoted", "hybrid"}
                and created_src == "PARSED_FROM_QUOTED_REQUEST"
                and not str(ack_src or "").startswith("PARSED_FROM_QUOTED_ACK")
                and cand_c_min < cand_a_min == cand_r_min
            )
            current_is_deep_proven_pair = bool(
                current_created_src == "PARSED_FROM_QUOTED_REQUEST"
                and current_ack_src == "PARSED_FROM_QUOTED_ACK"
                and cur_c_min < cur_a_min <= cur_r_min
            )
            current_is_shallow_fallback = bool(
                current_created_src == "PARSED_FROM_QUOTED_REQUEST"
                and not str(current_ack_src or "").startswith("PARSED_FROM_QUOTED_ACK")
                and cur_c_min < cur_a_min == cur_r_min
            )
            base_c_min = None
            base_a_min = None
            base_r_min = None
            baseline_mode = ""
            baseline_kind = ""
            baseline_owner = ""
            baseline_episode_score = None
            baseline_is_authoritative = False
            if initial_lane_episode:
                base_c = initial_lane_episode.get("created")
                base_a = initial_lane_episode.get("response")
                base_r = initial_lane_episode.get("resolved")
                if base_c and base_a and base_r:
                    base_c_min = base_c.replace(second=0, microsecond=0)
                    base_a_min = base_a.replace(second=0, microsecond=0)
                    base_r_min = base_r.replace(second=0, microsecond=0)
                    baseline_mode = initial_lane_episode.get("mode") or ""
                    baseline_kind = "quoted" if baseline_mode == "direct-reply" else "hybrid"
                    baseline_owner = f"LaneLocalInitialEpisode[{baseline_mode or 'baseline'}]"
                    baseline_strength = _candidate_rewrite_strength(
                        baseline_kind,
                        baseline_owner,
                        base_c_min,
                        base_a_min,
                        base_r_min,
                    )
                    baseline_episode_score = _episode_selection_score(
                        list_index=list_index,
                        c_min=base_c_min,
                        a_min=base_a_min,
                        r_min=base_r_min,
                        base_strength=baseline_strength,
                        created_src=initial_lane_episode.get("created_src", ""),
                        ack_src=initial_lane_episode.get("ack_src", ""),
                        resolved_src=initial_lane_episode.get("resolved_src", ""),
                        candidate_kind=baseline_kind,
                        owner_tag=baseline_owner,
                    )
                    baseline_is_authoritative = bool(initial_lane_episode.get("authoritative"))
            seed_locked = bool((state or {}).get("seed_locked")) and baseline_is_authoritative

            if candidate_is_deep_proven_pair and cand_r_min == cur_r_min and current_is_shallow_fallback:
                return _trace_decision(True, "deep_proven_pair_beats_shallow_fallback")

            if current_is_deep_proven_pair and cand_r_min == cur_r_min and candidate_is_shallow_fallback:
                return _trace_decision(False, "preserve_deep_proven_pair")

            if (
                candidate_is_seeded_owner
                and cand_r_min == cur_r_min
                and cand_c_min <= cand_a_min <= cand_r_min
                and (
                    low_confidence
                    or profile.get("all_same")
                    or "created retained (response anchor unreliable)" in notes_l
                )
                and candidate_episode_score >= (current_episode_score + 8)
            ):
                return _trace_decision(True, "seeded_primary_takeover")

            if baseline_episode_score is not None and candidate_is_quoted_family:
                if (
                    (cand_c_min, cand_a_min, cand_r_min) != (base_c_min, base_a_min, base_r_min)
                    and candidate_episode_score < (baseline_episode_score - 6)
                ):
                    return _trace_decision(False, "below_initial_lane_baseline")

            if baseline_is_authoritative and base_c_min and base_a_min and base_r_min:
                baseline_triplet = (base_c_min, base_a_min, base_r_min)
                candidate_triplet = (cand_c_min, cand_a_min, cand_r_min)
                current_triplet = (cur_c_min, cur_a_min, cur_r_min)
                if candidate_triplet == baseline_triplet:
                    return _trace_decision(True, "authoritative_seed_exact_match")
                if seed_locked and candidate_kind not in stronger_owner_kinds:
                    return _trace_decision(False, "seed_locked_primary")
                if (
                    current_triplet == baseline_triplet
                    and candidate_triplet != baseline_triplet
                    and candidate_kind not in stronger_owner_kinds
                ):
                    return _trace_decision(False, "authoritative_seed_current_lock")
                if (
                    candidate_triplet != baseline_triplet
                    and candidate_kind in repair_only_candidate_kinds
                    and candidate_kind not in stronger_owner_kinds
                ):
                    return _trace_decision(False, "authoritative_seed_repair_only")
                if candidate_triplet != baseline_triplet and candidate_kind not in stronger_owner_kinds:
                    if cand_r_min != base_r_min:
                        return _trace_decision(False, "authoritative_seed_resolved_lane")
                    if baseline_mode == "direct-reply" and cand_a_min != cand_r_min:
                        return _trace_decision(False, "authoritative_seed_direct_reply_shape")
                    if (
                        baseline_mode in {"direct-reply", "req-ack-reply"}
                        and cand_c_min == cand_a_min == cand_r_min
                    ):
                        return _trace_decision(False, "authoritative_seed_prevent_recollapse")
                    if baseline_episode_score is not None and candidate_episode_score < (baseline_episode_score - 4):
                        return _trace_decision(False, "authoritative_seed_stronger_baseline")

            # Protect already-good live/local windows from later continuation or
            # quoted-style guards. These are the rows that tend to drift without
            # becoming suspicious enough for risky validation.
            if profile.get("strong_live_ack") and candidate_kind in {"quoted", "hybrid", "requester_ack", "continuation", "risk"}:
                return _trace_decision(False, "strong_live_ack_protection")

            # Preserve intentional source-owned collapse rules. These rows are
            # meant to stay all-three-same unless the initial resolver itself
            # changes, so later repair/continuation guards should not reshape
            # them.
            if (
                _has_source_locked_same_time(notes_l)
                and candidate_kind in {"quoted", "hybrid", "requester_ack", "continuation", "risk"}
            ):
                return _trace_decision(
                    cand_c_min == cur_c_min and cand_a_min == cur_a_min and cand_r_min == cur_r_min,
                    "source_locked_all_same_protection",
                )

            occurrence_plan = _preferred_shared_occurrence_plan(state) if state else None
            shared_decision = (state or {}).get("shared_decision") if state else None
            if shared_decision:
                decision_triplet = shared_decision.get("triplet")
                if decision_triplet and (cand_c_min, cand_a_min, cand_r_min) != decision_triplet:
                    confidence = shared_decision.get("confidence") or ""
                    fill_style = shared_decision.get("fill_style") or ""
                    row_type = shared_decision.get("row_type") or ""
                    if confidence == "strong" and fill_style == "all_three_same" and _is_authoritative_occurrence_lane(row_type):
                        if candidate_kind not in {"quoted", "hybrid", "risk"}:
                            return _trace_decision(False, "shared_decision_protection_kind")
                        # Stronger proof must move away from the all-three-same
                        # lane into a proper ordered span, not another collapse.
                        if cand_c_min == cand_a_min == cand_r_min:
                            return _trace_decision(False, "shared_decision_no_recollapse")

            # Do not let all-three collapse owners flatten a row that is already a
            # proper ordered span, except when shared occurrence has explicitly
            # classified this row as an ESS-over-ESS lane and assigned a unique
            # occurrence reply for it.
            if (
                candidate_kind in {"requester_ack", "continuation"}
                and cand_c_min == cand_a_min == cand_r_min
                and not profile.get("all_same")
            ):
                if not (
                    candidate_kind == "occurrence_ess"
                    and occurrence_plan
                    and _is_authoritative_occurrence_lane(occurrence_plan.get("lane_kind") or "")
                ):
                    return _trace_decision(False, "prevent_all_three_collapse")

            # Quoted-style owners should not move a row backward to an older/weaker
            # local episode when the current row already has a valid request->ack
            # span, even if that current span is not explicitly tagged as preserved.
            cur_gap = profile.get("ack_gap")
            if (
                candidate_kind in {"quoted", "hybrid"}
                and cur_gap is not None
                and cur_gap <= timedelta(minutes=16)
                and not profile.get("all_same")
                and cur_c_min < cur_a_min
                and cand_r_min == cur_r_min
                and cand_a_min <= cur_a_min
                and cand_c_min <= cur_c_min
                and not candidate_is_seeded_owner
            ):
                return _trace_decision(False, "protect_existing_valid_span")

            # For repeated occurrence-managed families, do not allow a later
            # guard to converge multiple distinct rows onto the exact same
            # triplet. This keeps one occurrence lane from stealing another
            # lane's episode after shared occurrence has already separated them.
            if occurrence_plan:
                cand_triplet = (cand_c_min, cand_a_min, cand_r_min)
                cur_triplet = (cur_c_min, cur_a_min, cur_r_min)
                for other_state in occurrence_plan["group"]:
                    other_li = other_state.get("list_index")
                    if other_li is None or other_li == list_index or other_li >= len(automation_rows):
                        continue
                    other_row_vals = automation_rows[other_li]
                    other_c_dt = _parse_time_str(other_row_vals.get("Created Date & Time") or "")
                    other_a_dt = _parse_time_str(other_row_vals.get("Actual Response Date & Time") or "")
                    other_r_dt = _parse_time_str(other_row_vals.get("Actual Resolved Date & Time") or "")
                    if not (other_c_dt and other_a_dt and other_r_dt):
                        continue
                    other_triplet = (
                        _to_ist(other_c_dt).replace(second=0, microsecond=0),
                        _to_ist(other_a_dt).replace(second=0, microsecond=0),
                        _to_ist(other_r_dt).replace(second=0, microsecond=0),
                    )
                    if cand_triplet == other_triplet and cur_triplet != cand_triplet:
                        return _trace_decision(False, "occurrence_duplicate_triplet")

            if (
                not low_confidence
                and current_episode_score >= (candidate_episode_score + 10)
                and (cand_c_min, cand_a_min, cand_r_min) != (cur_c_min, cur_a_min, cur_r_min)
            ):
                return _trace_decision(False, "current_episode_stronger")

            if (
                low_confidence
                and candidate_episode_score >= (current_episode_score + 6)
            ):
                return _trace_decision(True, "candidate_beats_low_confidence_current")

            live_reply_ist = _live_reply_lane_for_row(list_index)
            if live_reply_ist and candidate_kind in {"quoted", "hybrid"}:
                live_reply_min = live_reply_ist.replace(second=0, microsecond=0)
                # Once a real live requester reply lane is known, quoted/hybrid
                # episodes must stay on that lane. They may provide an earlier
                # ack in-lane, or collapse to direct-reply, but they must not
                # drift to a different resolved lane or a later sibling-cycle ack.
                if "lanelocalinitialepisode[direct-reply]" in candidate_owner_l and (state.get("group_total") or 0) <= 1:
                    if cand_r_min < live_reply_min:
                        return _trace_decision(False, "quoted_direct_reply_before_live_reply")
                    if cand_a_min != cand_r_min:
                        return _trace_decision(False, "quoted_direct_reply_ack_not_equal_reply")
                else:
                    if cand_r_min != live_reply_min:
                        return _trace_decision(False, "quoted_resolved_not_on_live_reply_lane")
                    if cand_a_min > live_reply_min:
                        return _trace_decision(False, "quoted_ack_after_live_reply")

            return _trace_decision(True, "default_allow")

        def _apply_guarded_episode_update(
            state,
            row_vals,
            list_index,
            row_idx,
            cand_c_ist,
            cand_a_ist,
            cand_r_ist,
            *,
            owner_tag: str,
            candidate_kind: str,
            created_src: str = "",
            ack_src: str = "",
            resolved_src: str = "",
            note_suffix: str = "",
            clear_blue: bool = False,
        ) -> bool:
            _trace_focus_row(
                "apply_guarded_episode_update:start",
                state=state,
                row_vals=row_vals,
                list_index=list_index,
                owner=owner_tag,
                kind=candidate_kind,
                candidate=(
                    f"{_format_time(cand_c_ist) or '-'} / "
                    f"{_format_time(cand_a_ist) or '-'} / "
                    f"{_format_time(cand_r_ist) or '-'}"
                ),
            )
            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                cand_c_ist,
                cand_a_ist,
                cand_r_ist,
                owner_tag,
                candidate_kind,
                created_src=created_src,
                ack_src=ack_src,
                resolved_src=resolved_src,
            ):
                _trace_focus_row(
                    "apply_guarded_episode_update:rejected",
                    state=state,
                    row_vals=row_vals,
                    list_index=list_index,
                    owner=owner_tag,
                    kind=candidate_kind,
                )
                return False

            owner_l = (owner_tag or "").lower()
            if not _commit_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                cand_c_ist,
                cand_a_ist,
                cand_r_ist,
                created_src=created_src,
                ack_src=ack_src,
                resolved_src=resolved_src,
                note_suffix=note_suffix,
                clear_blue=clear_blue,
                seed_owner=(
                    "lanelocalinitialepisode" in owner_l
                    or "seedprimaryresolver" in owner_l
                ),
            ):
                return False

            _trace_focus_row(
                "apply_guarded_episode_update:applied",
                state=state,
                row_vals=row_vals,
                list_index=list_index,
                owner=owner_tag,
                kind=candidate_kind,
            )
            return True

        def _occurrence_expected_reply_ist(state):
            shared_plan = _preferred_shared_occurrence_plan(state)
            if shared_plan:
                return {
                    "when": shared_plan["pick_when"],
                    "slot_index": shared_plan["slot_index"],
                    "group_size": shared_plan["group_size"],
                }

            requester = state.get("requester") or ""
            service_no = state.get("service_no") or ""
            subject_norm_value = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            row_group_total = state.get("group_total") or 0
            list_index = state.get("list_index")
            if row_group_total < 2 or list_index is None or not requester or not subject_norm_value or not base_thread:
                return None
            current_occ_key = state.get("occurrence_key") or _occurrence_group_key(subject_norm_value, requester, service_no)

            def _occurrence_notes_match(notes_l: str) -> bool:
                return (
                    "ess-only; no non-ess request" in notes_l
                    or "requester follow-up" in notes_l
                    or "esscontinuationguard[" in notes_l
                )

            group_sorted = []
            for s in row_states:
                other_li = s.get("list_index")
                if other_li is None or other_li >= len(automation_rows):
                    continue
                other_occ_key = s.get("occurrence_key") or _occurrence_group_key(
                    (s.get("subject_norm") or "").lower(),
                    s.get("requester") or "",
                    s.get("service_no") or "",
                )
                if other_occ_key != current_occ_key:
                    continue
                other_notes_l = (debug_rows[other_li].get("Notes", "") or "").lower() if other_li < len(debug_rows) else ""
                if not _occurrence_notes_match(other_notes_l):
                    continue
                group_sorted.append(s)

            if len(group_sorted) < 2:
                return None

            group_sorted.sort(key=lambda x: x.get("row_index") or 10**9)
            slot_index = None
            for idx, s in enumerate(group_sorted):
                if s.get("list_index") == list_index:
                    slot_index = idx
                    break
            if slot_index is None:
                return None

            row_tokens = _match_tokens(subject_norm_value)
            reply_pool = []
            for e in base_thread:
                if not _req_match(e, requester):
                    continue
                if _ack_like(e) or _ack_like_text_fallback(e):
                    continue
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if row_tokens:
                    s_norm = normalize_subject(getattr(e, "subject", "") or "")
                    s_tokens = _match_tokens(s_norm)
                    score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                    contains = bool(subject_norm_value and s_norm and (subject_norm_value in s_norm or s_norm in subject_norm_value))
                    if score < 0.45 and not contains:
                        continue
                reply_pool.append(e)
            reply_pool.sort(key=lambda e: e.sent_time)

            unique_pool = []
            seen_ts = set()
            for e in reply_pool:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                key = e_ist.replace(second=0, microsecond=0)
                if key in seen_ts:
                    continue
                seen_ts.add(key)
                unique_pool.append(e)

            anchor_li = group_sorted[0].get("list_index")
            a_dt = _parse_time_str(automation_rows[anchor_li].get("Actual Response Date & Time")) if anchor_li is not None else None
            a_ist = _to_ist(a_dt) if a_dt else None
            month_pool = []
            if a_ist:
                for e in unique_pool:
                    e_ist = _email_ist(e)
                    if e_ist and e_ist.year == a_ist.year and e_ist.month == a_ist.month:
                        month_pool.append(e)

            pool = month_pool if len(month_pool) >= len(group_sorted) else unique_pool
            group_list_indexes = {
                s.get("list_index")
                for s in group_sorted
                if s.get("list_index") is not None
            }
            used_outside_group = set()
            for other_state in row_states:
                other_li = other_state.get("list_index")
                if other_li is None or other_li in group_list_indexes or other_li >= len(automation_rows):
                    continue
                other_occ_key = other_state.get("occurrence_key") or _occurrence_group_key(
                    (other_state.get("subject_norm") or "").lower(),
                    other_state.get("requester") or "",
                    other_state.get("service_no") or "",
                )
                if other_occ_key != current_occ_key:
                    continue
                other_a_dt = _parse_time_str(automation_rows[other_li].get("Actual Response Date & Time"))
                other_a_ist = _to_ist(other_a_dt) if other_a_dt else None
                if other_a_ist:
                    used_outside_group.add(other_a_ist.replace(second=0, microsecond=0))

            pool_without_used = []
            for e in pool:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                tkey = e_ist.replace(second=0, microsecond=0)
                if tkey in used_outside_group:
                    continue
                pool_without_used.append(e)
            if len(pool_without_used) >= len(group_sorted):
                pool = pool_without_used

            if len(pool) < len(group_sorted):
                return None

            pick = pool[min(slot_index, len(pool) - 1)]
            pick_ist = _email_ist(pick)
            if not pick_ist:
                return None
            return {
                "when": pick_ist.replace(second=0, microsecond=0),
                "slot_index": slot_index,
                "group_size": len(group_sorted),
            }

        def _apply_shared_occurrence_triplet(
            state,
            row_vals,
            list_index,
            row_idx,
            note_tag: str,
        ) -> bool:
            shared_decision = (state or {}).get("shared_decision") if state else None
            if not shared_decision:
                return False
            if (shared_decision.get("owner") or "") != "shared_occurrence":
                return False
            if not _is_authoritative_occurrence_lane((shared_decision.get("row_type") or "")):
                return False
            if (shared_decision.get("fill_style") or "") != "all_three_same":
                return False
            triplet = shared_decision.get("triplet")
            if not triplet or len(triplet) != 3:
                return False
            cand_c_ist, cand_a_ist, cand_r_ist = triplet
            if not (cand_c_ist and cand_a_ist and cand_r_ist):
                return False
            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                cand_c_ist,
                cand_a_ist,
                cand_r_ist,
                note_tag,
                "occurrence_ess",
            ):
                return False
            t = _format_time(cand_c_ist)
            if not t:
                return False
            row_vals["Created Date & Time"] = t
            row_vals["Actual Response Date & Time"] = t
            row_vals["Actual Resolved Date & Time"] = t
            if row_idx:
                ws.cell(row_idx, created_col).value = t
                ws.cell(row_idx, response_col).value = t
                ws.cell(row_idx, resolved_col).value = t
            _set_row_fill(row_idx, clear_fill)
            if list_index < len(debug_rows):
                who = (
                    debug_rows[list_index].get("ResolvedSource")
                    or debug_rows[list_index].get("AckSource")
                    or debug_rows[list_index].get("CreatedSource")
                    or "SHARED_OCCURRENCE"
                )
                debug_rows[list_index]["CreatedSource"] = who
                debug_rows[list_index]["AckSource"] = who
                debug_rows[list_index]["ResolvedSource"] = who
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; {note_tag}"
            return True

        def _current_row_triplet_ist(row_vals):
            if not row_vals:
                return None
            c_dt = _parse_time_str(row_vals.get("Created Date & Time") or "")
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time") or "")
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time") or "")
            if not (c_dt and a_dt and r_dt):
                return None
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist and a_ist and r_ist):
                return None
            return (
                c_ist.replace(second=0, microsecond=0),
                a_ist.replace(second=0, microsecond=0),
                r_ist.replace(second=0, microsecond=0),
            )

        def _commit_episode_update(
            state,
            row_vals,
            list_index,
            row_idx,
            cand_c_ist,
            cand_a_ist,
            cand_r_ist,
            *,
            created_src: str = "",
            ack_src: str = "",
            resolved_src: str = "",
            note_suffix: str = "",
            clear_blue: bool = False,
            seed_owner: bool = False,
        ) -> bool:
            t_c = _format_time(cand_c_ist)
            t_a = _format_time(cand_a_ist)
            t_r = _format_time(cand_r_ist)
            if not (t_c and t_a and t_r):
                return False

            row_vals["Created Date & Time"] = t_c
            row_vals["Actual Response Date & Time"] = t_a
            row_vals["Actual Resolved Date & Time"] = t_r
            if row_idx:
                ws.cell(row_idx, created_col).value = t_c
                ws.cell(row_idx, response_col).value = t_a
                ws.cell(row_idx, resolved_col).value = t_r
                if clear_blue:
                    _set_row_fill(row_idx, clear_fill)

            if list_index < len(debug_rows):
                if created_src:
                    debug_rows[list_index]["CreatedSource"] = created_src
                if ack_src:
                    debug_rows[list_index]["AckSource"] = ack_src
                if resolved_src:
                    debug_rows[list_index]["ResolvedSource"] = resolved_src
                if seed_owner and (
                    (created_src == "PARSED_FROM_QUOTED_REQUEST" and cand_c_ist < cand_a_ist)
                    or "all-three-same" in (note_suffix or "").lower()
                ):
                    notes_now = debug_rows[list_index].get("Notes", "") or ""
                    if notes_now:
                        debug_rows[list_index]["Notes"] = _remove_note_tokens(
                            notes_now,
                            ["Created retained (response anchor unreliable)"],
                        )
                if note_suffix:
                    notes_now = debug_rows[list_index].get("Notes", "") or ""
                    if note_suffix not in {part.strip() for part in notes_now.split(";") if part.strip()}:
                        debug_rows[list_index]["Notes"] = f"{notes_now}; {note_suffix}" if notes_now else note_suffix

            if state is not None and seed_owner:
                state["seed_locked"] = True
            return True

        def _seeded_lane_should_replace_base(
            state,
            row_vals,
            list_index,
            cand_c_ist,
            cand_a_ist,
            cand_r_ist,
            *,
            owner_tag: str,
            candidate_kind: str,
            created_src: str = "",
            ack_src: str = "",
            resolved_src: str = "",
        ):
            if not state:
                return False, "missing_state"
            if not (cand_c_ist and cand_a_ist and cand_r_ist):
                return False, "missing_seed_candidate"

            cand_c_min = cand_c_ist.replace(second=0, microsecond=0)
            cand_a_min = cand_a_ist.replace(second=0, microsecond=0)
            cand_r_min = cand_r_ist.replace(second=0, microsecond=0)
            candidate_triplet = (cand_c_min, cand_a_min, cand_r_min)
            if not (cand_c_min <= cand_a_min <= cand_r_min):
                return False, "seed_candidate_unordered"

            shared_decision = state.get("shared_decision") or {}
            if (
                (shared_decision.get("owner") or "") == "shared_occurrence"
                and (shared_decision.get("fill_style") or "") == "all_three_same"
                and _is_authoritative_occurrence_lane(shared_decision.get("row_type") or "")
            ):
                decision_triplet = shared_decision.get("triplet")
                if decision_triplet and tuple(decision_triplet) != candidate_triplet:
                    return False, "shared_occurrence_authoritative"

            shared_occ_plan = _preferred_shared_occurrence_plan(state, require_override_for_all_ack=True)
            if shared_occ_plan and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                return False, "occurrence_authoritative_family"

            if _row_is_special_all_same_risk_family(state, row_vals, list_index):
                return False, "special_all_same_family"

            live_reply_ist = _live_reply_lane_for_row(list_index)
            if live_reply_ist:
                live_reply_min = live_reply_ist.replace(second=0, microsecond=0)
                if cand_r_min != live_reply_min:
                    return False, "seed_not_on_live_reply_lane"

            current_triplet = _current_row_triplet_ist(row_vals)
            if current_triplet == candidate_triplet:
                return True, "candidate_already_current"

            profile = _rewrite_guard_profile(row_vals, list_index)
            if not profile.get("ordered"):
                return True, "current_not_ordered"

            current_notes_l = profile.get("notes_l") or ""
            current_created_src = profile.get("created_src") or ""
            base_candidate = state.get("base_candidate") or {}
            base_created_src = (base_candidate.get("created_src") or "").lower()

            current_score = _episode_selection_score(
                list_index=list_index,
                c_min=profile.get("c_min"),
                a_min=profile.get("a_min"),
                r_min=profile.get("r_min"),
                base_strength=profile.get("rewrite_strength") or 0,
                created_src=profile.get("created_src") or "",
                ack_src=profile.get("ack_src") or "",
                resolved_src=profile.get("resolved_src") or "",
                notes_l=current_notes_l,
                candidate_kind="base",
                owner_tag="BaseResolver",
            )
            candidate_strength = _candidate_rewrite_strength(
                candidate_kind,
                owner_tag,
                cand_c_min,
                cand_a_min,
                cand_r_min,
            )
            candidate_score = _episode_selection_score(
                list_index=list_index,
                c_min=cand_c_min,
                a_min=cand_a_min,
                r_min=cand_r_min,
                base_strength=candidate_strength,
                created_src=created_src,
                ack_src=ack_src,
                resolved_src=resolved_src,
                notes_l=owner_tag,
                candidate_kind=candidate_kind,
                owner_tag=owner_tag,
            )
            score_delta = candidate_score - current_score

            if profile.get("strong_live_ack") and not profile.get("low_confidence"):
                return False, "current_strong_live_ack"

            current_r_min = profile.get("r_min")
            candidate_same_reply_lane = not current_r_min or cand_r_min == current_r_min
            candidate_improves_created = bool(profile.get("c_min") and cand_c_min < profile.get("c_min"))
            candidate_is_direct_reply = cand_c_min < cand_a_min and cand_a_min == cand_r_min
            candidate_is_seed_same_time = (
                candidate_kind == "seed_same_time"
                or "all-three-same" in (owner_tag or "").lower()
            )
            candidate_is_deep_proven_pair = bool(
                candidate_kind == "hybrid"
                and created_src == "PARSED_FROM_QUOTED_REQUEST"
                and ack_src == "PARSED_FROM_QUOTED_ACK"
                and cand_c_min < cand_a_min <= cand_r_min
            )
            current_is_shallow_fallback = bool(
                profile.get("c_min")
                and profile.get("a_min")
                and profile.get("r_min")
                and profile.get("c_min") < profile.get("a_min") == profile.get("r_min")
                and current_created_src == "PARSED_FROM_QUOTED_REQUEST"
                and not (profile.get("ack_src") or "").startswith("PARSED_FROM_QUOTED_ACK")
            )

            weak_base = bool(
                profile.get("low_confidence")
                or profile.get("all_same")
                or "created retained (response anchor unreliable)" in current_notes_l
                or "quotedrequestonly" in current_notes_l
                or "created_clamped_to_first" in current_created_src.lower()
                or "created_clamped_to_first" in base_created_src
            )

            if candidate_is_deep_proven_pair and candidate_same_reply_lane and current_is_shallow_fallback:
                return True, "deep_proven_pair_beats_shallow_fallback"

            if (
                candidate_same_reply_lane
                and candidate_improves_created
                and candidate_is_direct_reply
                and profile.get("all_same")
            ):
                return True, "reanchor_all_same_base"

            if (
                candidate_same_reply_lane
                and candidate_improves_created
                and "created retained (response anchor unreliable)" in current_notes_l
            ):
                return True, "reanchor_created_retained"

            if (
                candidate_is_seed_same_time
                and candidate_same_reply_lane
                and (state.get("group_total") or 0) <= 1
                and cand_c_min == cand_a_min == cand_r_min
            ):
                return True, "seed_unique_ess_over_ess_same_time"

            if weak_base and score_delta >= 8:
                return True, "seed_beats_weak_base"

            if candidate_same_reply_lane and score_delta >= 20:
                return True, "seed_clearly_better"

            return False, "keep_base_candidate"

        def _seeded_lane_should_commit_authoritatively(
            state,
            row_vals,
            list_index,
            *,
            lane_mode: str,
            req_ist,
            ack_ist,
            reply_ist,
            created_src: str,
            ack_src: str,
        ) -> bool:
            if not state or not req_ist or not ack_ist or not reply_ist:
                return False
            if not (req_ist <= ack_ist <= reply_ist):
                return False

            live_reply_ist = _live_reply_lane_for_row(list_index)
            if live_reply_ist and reply_ist.replace(second=0, microsecond=0) != live_reply_ist.replace(second=0, microsecond=0):
                return False

            if _row_is_special_all_same_risk_family(state, row_vals, list_index):
                return False

            shared_decision = state.get("shared_decision") or {}
            if (
                (shared_decision.get("owner") or "") == "shared_occurrence"
                and (shared_decision.get("fill_style") or "") == "all_three_same"
                and _is_authoritative_occurrence_lane(shared_decision.get("row_type") or "")
            ):
                decision_triplet = shared_decision.get("triplet")
                if decision_triplet and tuple(decision_triplet) != (
                    req_ist.replace(second=0, microsecond=0),
                    ack_ist.replace(second=0, microsecond=0),
                    reply_ist.replace(second=0, microsecond=0),
                ):
                    return False

            shared_occ_plan = _preferred_shared_occurrence_plan(state, require_override_for_all_ack=True)
            if shared_occ_plan and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                return False

            if lane_mode == "direct-reply":
                return bool(
                    created_src == "PARSED_FROM_QUOTED_REQUEST"
                    and req_ist < ack_ist == reply_ist
                )

            if lane_mode == "req-ack-reply":
                return bool(
                    created_src == "PARSED_FROM_QUOTED_REQUEST"
                    and ack_src == "PARSED_FROM_QUOTED_ACK"
                    and req_ist < ack_ist <= reply_ist
                )

            return False

        def _lock_occurrence_row(state, row_vals, list_index, note_tag: str, *, triplet=None) -> bool:
            if not state or list_index is None:
                return False
            triplet = triplet or _current_row_triplet_ist(row_vals)
            if not triplet or len(triplet) != 3:
                return False
            state["occurrence_locked"] = True
            state["occurrence_lock_triplet"] = tuple(triplet)
            if list_index < len(debug_rows):
                notes = debug_rows[list_index].get("Notes", "") or ""
                if note_tag not in notes:
                    debug_rows[list_index]["Notes"] = f"{notes}; {note_tag}"
            return True

        def _apply_occurrence_plan_authoritatively(
            state,
            row_vals,
            list_index,
            row_idx,
            shared_occ_plan,
            note_tag: str,
        ) -> bool:
            if not state or not shared_occ_plan:
                return False
            lane_kind = shared_occ_plan.get("lane_kind") or ""
            pick_when = shared_occ_plan.get("pick_when")
            if not _is_authoritative_occurrence_lane(lane_kind) or not pick_when:
                return False

            cand_triplet = (pick_when, pick_when, pick_when)
            state["shared_decision"] = {
                "owner": "shared_occurrence",
                "row_type": lane_kind,
                "occurrence_slot": shared_occ_plan.get("slot_index", 0),
                "lane_time": pick_when,
                "fill_style": "all_three_same",
                "confidence": "strong",
                "triplet": cand_triplet,
            }

            t = _format_time(pick_when)
            if not t:
                return False
            row_vals["Created Date & Time"] = t
            row_vals["Actual Response Date & Time"] = t
            row_vals["Actual Resolved Date & Time"] = t
            if row_idx:
                ws.cell(row_idx, created_col).value = t
                ws.cell(row_idx, response_col).value = t
                ws.cell(row_idx, resolved_col).value = t
                _set_row_fill(row_idx, clear_fill)
            if list_index < len(debug_rows):
                who = (
                    debug_rows[list_index].get("ResolvedSource")
                    or debug_rows[list_index].get("AckSource")
                    or debug_rows[list_index].get("CreatedSource")
                    or "SHARED_OCCURRENCE"
                )
                debug_rows[list_index]["CreatedSource"] = who
                debug_rows[list_index]["AckSource"] = who
                debug_rows[list_index]["ResolvedSource"] = who
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; {note_tag}"
            return True

        def _should_preserve_occurrence_same_time(state, row_vals, shared_occ_plan) -> bool:
            if not state or not row_vals or not shared_occ_plan:
                return False
            if not _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                return False
            target_when = shared_occ_plan.get("pick_when")
            current_triplet = _current_row_triplet_ist(row_vals)
            if not target_when or not current_triplet:
                return False
            target_triplet = (
                target_when.replace(second=0, microsecond=0),
                target_when.replace(second=0, microsecond=0),
                target_when.replace(second=0, microsecond=0),
            )
            if current_triplet != target_triplet:
                return False
            list_index = state.get("list_index")
            notes_l = (debug_rows[list_index].get("Notes", "") or "").lower() if list_index is not None and list_index < len(debug_rows) else ""
            shape_hints = _occurrence_slot_shape_hints(state)
            row_anchor = current_triplet[2] or current_triplet[1] or current_triplet[0]
            target_minute = target_triplet[2]
            if row_anchor != target_minute:
                return False
            if shape_hints["has_local_triplet"] and not shape_hints["has_local_same_time"]:
                return False
            return bool(
                "requester span(all-ack->ess)" in notes_l
                or "occurrenceslotshape[allthreesame]" in notes_l
                or "esscontinuationguard[allthreestrictessonly]" in notes_l
            )

        _subject_family_slot_cache = {}

        def _subject_family_slot(state):
            list_index = state.get("list_index")
            if list_index in _subject_family_slot_cache:
                return _subject_family_slot_cache[list_index]
            if not state.get("multi_service_subject"):
                out = (0, 1)
                _subject_family_slot_cache[list_index] = out
                return out
            subject_norm_value = (state.get("subject_norm") or "").lower()
            if not subject_norm_value:
                out = (0, 1)
                _subject_family_slot_cache[list_index] = out
                return out
            family = []
            for s in row_states:
                other_li = s.get("list_index")
                if other_li is None or other_li >= len(automation_rows):
                    continue
                if not s.get("multi_service_subject"):
                    continue
                if (s.get("subject_norm") or "").lower() != subject_norm_value:
                    continue
                family.append(s)
            family.sort(key=lambda s: s.get("row_index") or 10**9)
            slot_index = 0
            for idx, s in enumerate(family):
                if s.get("list_index") == list_index:
                    slot_index = idx
                    break
            out = (slot_index, len(family) or 1)
            _subject_family_slot_cache[list_index] = out
            return out

        def _shared_occurrence_pick(
            state,
            *,
            subject_norm_value: str,
            requester_value: str,
            current_created_ist=None,
            current_ack_ist=None,
            current_resolved_ist=None,
            default_idx: int = 0,
        ):
            occ_key = state.get("occurrence_key") or _occurrence_group_key(
                subject_norm_value,
                requester_value,
                state.get("service_no") or "",
            )
            shared_plan = _preferred_shared_occurrence_plan(state)
            family_idx, family_total = _subject_family_slot(state)
            multi_service = bool(state.get("multi_service_subject")) and family_total >= 2
            if shared_plan:
                pick_idx = shared_plan["slot_index"]
                total_rows = shared_plan["group_size"]
            else:
                pick_idx = family_idx if multi_service else max(0, default_idx)
                total_rows = family_total if multi_service else (
                    state.get("group_total") or _occurrence_group_total(
                        subject_norm_value,
                        requester_value,
                        state.get("service_no") or "",
                    )
                )
            # For multi-service repeated-subject families, current row times are
            # often exactly the stale values we are trying to escape. In that
            # case, route by shared family slot and avoid reusing the row's
            # current timestamps as an occurrence anchor.
            if shared_plan:
                target_ist = shared_plan["pick_when"]
                target_reply_ist = shared_plan["pick_when"]
            elif multi_service:
                target_ist = None
                target_reply_ist = None
            else:
                target_ist = current_ack_ist or current_resolved_ist or current_created_ist
                target_reply_ist = current_resolved_ist or current_ack_ist or current_created_ist
            return {
                "occ_key": occ_key,
                "pick_idx": pick_idx,
                "total_rows": total_rows,
                "multi_service": multi_service,
                "target_ist": target_ist,
                "target_reply_ist": target_reply_ist,
            }

        def _dedupe_multi_service_lanes(items, lane_key_fn):
            out = []
            seen = set()
            for item in items:
                lane_key = lane_key_fn(item)
                if lane_key is None:
                    out.append(item)
                    continue
                if lane_key in seen:
                    continue
                seen.add(lane_key)
                out.append(item)
            return out

        def _episode_candidate_for_row(
            state,
            row_vals,
            list_index,
            *,
            require_risky_notes: bool,
            require_suspicious: bool,
        ):
            requester = state.get("requester") or ""
            subject_norm_value = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            if not requester or not subject_norm_value or not base_thread:
                return None

            precheck = _risk_guard_precheck(state, row_vals, list_index)
            cache_key = (
                list_index,
                require_risky_notes,
                require_suspicious,
                precheck.get("notes_l", ""),
                row_vals.get("Created Date & Time") or "",
                row_vals.get("Actual Response Date & Time") or "",
                row_vals.get("Actual Resolved Date & Time") or "",
            )
            if cache_key in _episode_candidate_cache:
                return _episode_candidate_cache[cache_key]

            c_ist = precheck.get("c_ist")
            a_ist = precheck.get("a_ist")
            r_ist = precheck.get("r_ist")
            if not (c_ist and a_ist and r_ist):
                _episode_candidate_cache[cache_key] = None
                return None
            suspicious_all_same = precheck.get("suspicious_all_same", False)
            suspicious_created_ack = precheck.get("suspicious_created_ack", False)
            if require_suspicious and not (suspicious_all_same or suspicious_created_ack):
                _episode_candidate_cache[cache_key] = None
                return None

            notes_l = precheck.get("notes_l", "")
            if require_risky_notes and not precheck.get("has_risky_notes"):
                _episode_candidate_cache[cache_key] = None
                return None
            row_group_total = state.get("group_total") or 0
            if row_group_total >= 2:
                _occurrence_expected_reply_ist(state)
                # Repeated rows are occurrence problems first. Validate against the
                # same occurrence pool, then let the later distinct mapper own the
                # actual repair so we don't split the triplet here.
                _episode_candidate_cache[cache_key] = None
                return None

            row_tokens = _match_tokens(subject_norm_value)
            row_id_tokens = _id_like_tokens(subject_norm_value)
            if not row_id_tokens:
                desc_text = row_vals.get("Description") or state.get("description") or ""
                row_id_tokens = _id_like_tokens(desc_text)

            anchor_ist = r_ist
            parsed_anchor = False
            if list_index < len(debug_rows):
                created_src_now = (debug_rows[list_index].get("CreatedSource") or "").lower()
                ack_src_now = (debug_rows[list_index].get("AckSource") or "").lower()
                parsed_anchor = ("parsed_from_quoted" in created_src_now) or ("parsed_from_quoted" in ack_src_now)

            thread = _expanded_thread(
                subject_norm_value,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=anchor_ist,
            )
            thread = thread or []
            req_pool = _requester_pool(subject_norm_value, requester, anchor_ist, day_window=14) or []
            any_pool = _requester_pool(subject_norm_value, "", anchor_ist, day_window=14) or []

            merged_sources = []
            seen_keys = set()
            for src in list(thread) + list(req_pool) + list(any_pool):
                if not getattr(src, "sent_time", None):
                    continue
                src_key = (
                    getattr(src, "subject", "") or "",
                    getattr(src, "sender_email", "") or "",
                    getattr(src, "sender_name", "") or "",
                    getattr(src, "sent_time", None),
                )
                if src_key in seen_keys:
                    continue
                seen_keys.add(src_key)
                merged_sources.append(src)

            if not merged_sources:
                _episode_candidate_cache[cache_key] = None
                return None

            anchor_minute = anchor_ist.replace(second=0, microsecond=0)

            def _best_request_before(upper_ist, max_gap):
                live_req_candidates = []
                for cand in merged_sources:
                    cand_ist = _email_ist(cand)
                    if not cand_ist or cand_ist >= upper_ist:
                        continue
                    if max_gap and (upper_ist - cand_ist) > max_gap:
                        continue
                    if _ess_sender(cand) or _system_like_sender(cand):
                        continue
                    if not _row_subject_match_email_quoted(cand, subject_norm_value, row_tokens, row_id_tokens):
                        continue
                    live_req_candidates.append((cand_ist, cand))

                quoted_req_candidates = []
                for src in merged_sources:
                    q_req = _extract_quoted_request_before_ist(src, subject_norm_value, upper_ist)
                    if not q_req or q_req >= upper_ist:
                        continue
                    if max_gap and (upper_ist - q_req) > max_gap:
                        continue
                    quoted_req_candidates.append(q_req)

                chosen_live_req_ist = _best_req_before_ack([t for t, _ in live_req_candidates], upper_ist) if live_req_candidates else None
                chosen_live_req = None
                if chosen_live_req_ist:
                    for cand_ist, cand in live_req_candidates:
                        if cand_ist == chosen_live_req_ist:
                            chosen_live_req = cand
                            break

                chosen_quoted_req = _best_req_before_ack(quoted_req_candidates, upper_ist) if quoted_req_candidates else None

                if chosen_live_req and _email_ist(chosen_live_req):
                    return {
                        "when": _email_ist(chosen_live_req),
                        "src": chosen_live_req.sender_email or chosen_live_req.sender_name,
                        "kind": "live",
                    }
                if chosen_quoted_req:
                    return {
                        "when": chosen_quoted_req,
                        "src": "PARSED_FROM_QUOTED_REQUEST",
                        "kind": "quoted",
                    }
                return None

            resolved_candidates = []
            for cand in merged_sources:
                cand_ist = _email_ist(cand)
                if not cand_ist:
                    continue
                if not _req_match(cand, requester):
                    continue
                if not _row_subject_match_email_quoted(cand, subject_norm_value, row_tokens, row_id_tokens):
                    continue
                if _ack_like(cand) or _ack_like_text_fallback(cand) or _ess_only_short_ack(cand):
                    continue
                minute_match = cand_ist.replace(second=0, microsecond=0) == anchor_minute
                near_later = parsed_anchor and cand_ist > anchor_ist and (cand_ist - anchor_ist) <= timedelta(hours=48)
                if not minute_match and not near_later:
                    continue
                resolved_rank = 0 if near_later else 1
                resolved_candidates.append(
                    (
                        resolved_rank,
                        abs((cand_ist - anchor_ist).total_seconds()),
                        cand,
                    )
                )

            resolved_candidates.sort(key=lambda x: (x[0], x[1], _email_ist(x[2]) or anchor_ist))
            if not resolved_candidates:
                _episode_candidate_cache[cache_key] = None
                return None

            def _episode_locality_bucket(created_when, ack_when, resolved_when):
                backward_created = max(timedelta(0), c_ist - created_when)
                backward_ack = max(timedelta(0), a_ist - ack_when)
                total_backward = backward_created + backward_ack
                if total_backward <= timedelta(hours=2):
                    return 0
                if total_backward <= timedelta(hours=8):
                    return 1
                if total_backward <= timedelta(hours=24):
                    return 2
                return 3

            def _episode_span_bucket(created_when, ack_when, resolved_when):
                req_ack_gap = ack_when - created_when
                ack_res_gap = resolved_when - ack_when
                total_span = resolved_when - created_when
                if (
                    req_ack_gap <= timedelta(minutes=16)
                    and ack_res_gap <= timedelta(hours=6)
                    and total_span <= timedelta(hours=6)
                ):
                    return 0
                if (
                    req_ack_gap <= timedelta(minutes=30)
                    and ack_res_gap <= timedelta(hours=24)
                    and total_span <= timedelta(hours=24)
                ):
                    return 1
                if total_span <= timedelta(hours=48):
                    return 2
                return 3

            def _episode_score(episode, resolved_priority, req_kind_rank, ack_like_rank, mode_rank):
                created_when = episode["created"]
                ack_when = episode["response"]
                resolved_when = episode["resolved"]
                locality_bucket = _episode_locality_bucket(created_when, ack_when, resolved_when)
                span_bucket = _episode_span_bucket(created_when, ack_when, resolved_when)
                backward_created = max(timedelta(0), c_ist - created_when)
                backward_ack = max(timedelta(0), a_ist - ack_when)
                resolved_delta = abs((resolved_when - anchor_ist).total_seconds())
                return (
                    locality_bucket,
                    span_bucket,
                    resolved_priority,
                    mode_rank,
                    0 if req_kind_rank == 0 else 1,
                    ack_like_rank,
                    abs(int(backward_created.total_seconds())),
                    abs(int(backward_ack.total_seconds())),
                    int(resolved_delta),
                )

            best_episode = None
            best_score = None
            for resolved_priority, _delta, resolved_msg in resolved_candidates[:6]:
                resolved_ist = _email_ist(resolved_msg)
                if not resolved_ist:
                    continue

                ack_candidates = []
                for ack_msg in merged_sources:
                    ack_ist = _email_ist(ack_msg)
                    if not ack_ist or ack_ist >= resolved_ist:
                        continue
                    if (resolved_ist - ack_ist) > timedelta(hours=48):
                        continue
                    if not _ess_sender(ack_msg):
                        continue
                    if not _row_subject_match_email_quoted(ack_msg, subject_norm_value, row_tokens, row_id_tokens):
                        continue
                    ack_like_score = 0 if (_ack_like(ack_msg) or _ack_like_text_fallback(ack_msg) or _ess_only_short_ack(ack_msg)) else 1
                    ack_candidates.append((ack_like_score, resolved_ist - ack_ist, ack_msg))

                ack_candidates.sort(key=lambda item: (item[0], item[1]))
                for ack_like_score, _ack_delta, ack_msg in ack_candidates[:8]:
                    ack_ist = _email_ist(ack_msg)
                    if not ack_ist:
                        continue
                    later_real_replies = [
                        e for e in merged_sources
                        if _req_match(e, requester)
                        and _email_ist(e)
                        and _email_ist(e) > ack_ist
                        and _email_ist(e) <= (ack_ist + timedelta(hours=48))
                        and _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens)
                        and not (_ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e))
                    ]
                    req_info = _best_request_before(ack_ist, timedelta(minutes=16))
                    if not req_info:
                        continue
                    req_ist = req_info["when"]
                    if not req_ist or not (req_ist < ack_ist < resolved_ist):
                        continue
                    if resolved_ist == ack_ist.replace(second=resolved_ist.second, microsecond=resolved_ist.microsecond) and later_real_replies:
                        continue
                    if resolved_ist == ack_ist and later_real_replies:
                        continue
                    episode = {
                        "created": req_ist,
                        "response": ack_ist,
                        "resolved": resolved_ist,
                        "created_src": req_info["src"],
                        "ack_src": ack_msg.sender_email or ack_msg.sender_name,
                        "resolved_src": resolved_msg.sender_email or resolved_msg.sender_name,
                        "mode": "req-ack-reply",
                    }
                    score = _episode_score(
                        episode,
                        resolved_priority=resolved_priority,
                        req_kind_rank=0 if req_info["kind"] == "live" else 1,
                        ack_like_rank=ack_like_score,
                        mode_rank=1,
                    )
                    if best_score is None or score < best_score:
                        best_score = score
                        best_episode = episode

                req_info = _best_request_before(resolved_ist, timedelta(minutes=16))
                if req_info and req_info["when"] and req_info["when"] < resolved_ist:
                    episode = {
                        "created": req_info["when"],
                        "response": resolved_ist,
                        "resolved": resolved_ist,
                        "created_src": req_info["src"],
                        "ack_src": resolved_msg.sender_email or resolved_msg.sender_name,
                        "resolved_src": resolved_msg.sender_email or resolved_msg.sender_name,
                        "mode": "direct-reply",
                    }
                    score = _episode_score(
                        episode,
                        resolved_priority=resolved_priority,
                        req_kind_rank=0 if req_info["kind"] == "live" else 1,
                        ack_like_rank=0,
                        mode_rank=0,
                    )
                    if best_score is None or score < best_score:
                        best_score = score
                        best_episode = episode

            if not best_episode:
                _episode_candidate_cache[cache_key] = None
                return None
            if best_episode["created"] == best_episode["response"] == best_episode["resolved"]:
                _episode_candidate_cache[cache_key] = None
                return None
            if best_episode["created"] >= best_episode["response"]:
                _episode_candidate_cache[cache_key] = None
                return None
            if best_episode["response"] > best_episode["resolved"]:
                _episode_candidate_cache[cache_key] = None
                return None
            _episode_candidate_cache[cache_key] = best_episode
            return best_episode

        def _risk_guard_episode_for_row(state, row_vals, list_index):
            return _episode_candidate_for_row(
                state,
                row_vals,
                list_index,
                require_risky_notes=True,
                require_suspicious=True,
            )

        def _best_req_before_ack(reqs, ack_ist):
            if not reqs:
                return None
            strict_reqs = [
                r for r in reqs
                if r.replace(second=0, microsecond=0) != ack_ist.replace(second=0, microsecond=0)
            ]
            chosen = strict_reqs[-1] if strict_reqs else reqs[-1]
            return chosen

        def _best_request_anchor_from_sources(merged_sources, subject_norm_value, row_tokens, row_id_tokens, upper_ist, max_gap):
            if not merged_sources or not upper_ist:
                return None

            live_req_candidates = []
            for cand in merged_sources:
                cand_ist = _email_ist(cand)
                if not cand_ist or cand_ist >= upper_ist:
                    continue
                if max_gap and (upper_ist - cand_ist) > max_gap:
                    continue
                if _ess_sender(cand) or _system_like_sender(cand):
                    continue
                if not _row_subject_match_email_quoted(cand, subject_norm_value, row_tokens, row_id_tokens):
                    continue
                live_req_candidates.append((cand_ist, cand))

            quoted_req_candidates = []
            for src in merged_sources:
                q_req = _extract_quoted_request_before_ist(src, subject_norm_value, upper_ist)
                if not q_req or q_req >= upper_ist:
                    continue
                if max_gap and (upper_ist - q_req) > max_gap:
                    continue
                quoted_req_candidates.append(q_req)

            chosen_live_req_ist = _best_req_before_ack([t for t, _ in live_req_candidates], upper_ist) if live_req_candidates else None
            chosen_live_req = None
            if chosen_live_req_ist:
                for cand_ist, cand in live_req_candidates:
                    if cand_ist == chosen_live_req_ist:
                        chosen_live_req = cand
                        break

            chosen_quoted_req = _best_req_before_ack(quoted_req_candidates, upper_ist) if quoted_req_candidates else None

            if chosen_live_req and _email_ist(chosen_live_req):
                return {
                    "when": _email_ist(chosen_live_req),
                    "src": chosen_live_req.sender_email or chosen_live_req.sender_name,
                    "kind": "live",
                }
            if chosen_quoted_req:
                return {
                    "when": chosen_quoted_req,
                    "src": "PARSED_FROM_QUOTED_REQUEST",
                    "kind": "quoted",
                }
            return None

        def _subject_variant_score(subject_norm_value: str, candidate_subject: str) -> int:
            base_norm = normalize_subject(subject_norm_value or "")
            cand_norm = normalize_subject(candidate_subject or "")
            base_match = normalize_subject_for_match(base_norm)
            cand_match = normalize_subject_for_match(cand_norm)
            if not base_match or not cand_match:
                return 0
            if base_match == cand_match:
                return 100
            if base_match in cand_match or cand_match in base_match:
                return 85
            base_tokens = _match_tokens(base_match)
            cand_tokens = _match_tokens(cand_match)
            if not base_tokens or not cand_tokens:
                return 0
            inter = len(base_tokens & cand_tokens)
            if inter <= 0:
                return 0
            extra = len(cand_tokens - base_tokens)
            missing = len(base_tokens - cand_tokens)
            return max(0, inter * 14 - extra * 8 - missing * 6)

        def _preferred_reply_anchored_quoted_episode(
            reply_pool_real,
            merged_sources,
            subject_norm_value,
            row_tokens,
            row_id_tokens,
            requester_name,
            target_reply_ist,
        ):
            if not reply_pool_real or not merged_sources:
                return None
            best_episode = None
            best_score = None
            for reply_msg in reply_pool_real:
                reply_ist = _email_ist(reply_msg)
                if not reply_ist:
                    continue
                if requester_name and not _req_match(reply_msg, requester_name):
                    continue
                if not _row_subject_match_email_quoted(reply_msg, subject_norm_value, row_tokens, row_id_tokens):
                    continue
                req_anchor_info = _best_request_anchor_from_sources(
                    merged_sources,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens,
                    reply_ist,
                    timedelta(hours=48),
                )
                req_anchor_ist = req_anchor_info.get("when") if req_anchor_info else None
                if not req_anchor_ist or reply_ist <= req_anchor_ist:
                    continue
                gap = reply_ist - req_anchor_ist
                if gap > timedelta(hours=48):
                    continue
                subject_score = _subject_variant_score(subject_norm_value, getattr(reply_msg, "subject", "") or "")
                if subject_score < 60:
                    continue
                score = (
                    1,
                    subject_score,
                    -abs((reply_ist - target_reply_ist).total_seconds()) if target_reply_ist else 0,
                    -abs(gap.total_seconds()),
                    int(reply_ist.timestamp()),
                )
                candidate = {
                    "request": req_anchor_ist,
                    "reply_ist": reply_ist,
                    "reply_msg": reply_msg,
                    "req_src": req_anchor_info.get("src") if req_anchor_info else "PARSED_FROM_QUOTED_REQUEST",
                    "gap": gap,
                }
                if best_score is None or score > best_score:
                    best_score = score
                    best_episode = candidate
            return best_episode

        # Build created-time list in row order.
        created_list = []
        for state in nonblue_row_states:
            idx = state["list_index"]
            created_str = automation_rows[idx].get("Created Date & Time")
            created_dt = _parse_time_str(created_str)
            created_list.append(_to_ist(created_dt) if created_dt else None)

        def _find_prev(i):
            for j in range(i - 1, -1, -1):
                if created_list[j]:
                    return j, created_list[j]
            return None, None

        def _find_next(i):
            for j in range(i + 1, len(created_list)):
                if created_list[j]:
                    return j, created_list[j]
            return None, None

        grace = timedelta(hours=6)
        max_forward = timedelta(days=1)

        for i, state in enumerate(nonblue_row_states):
            thread = state.get("thread")
            requester = state.get("requester")
            if not thread or not requester:
                continue

            description = state.get("description") or ""
            if "maintenance" in description.lower():
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            date_tokens = state.get("date_tokens") or []
            explicit_marker = state.get("explicit_marker")
            group_total = state.get("group_total") or 0
            row_has_ids = state.get("row_has_ids")
            debug = state.get("debug")
            debug_notes = (debug.notes if debug else "") or ""
            date_anchor_missing = state.get("date_anchor_missing") or ("DateAnchorMissing" in debug_notes)
            date_anchor_after = state.get("date_anchor_after") or ("DateAnchorAfter" in debug_notes)
            date_tokens_match_thread = state.get("date_tokens_match_thread")

            # Keep per-occurrence duplicate handling stable for repeated explicit-date rows.
            if group_total >= 2 and explicit_marker:
                continue

            # Skip INC/DR rows unless anchor was missing/after OR the row has an
            # explicit marker (safe path for ordered repeated ID rows).
            if row_has_ids and not (date_anchor_missing or date_anchor_after or explicit_marker):
                continue

            prev_idx, prev_time = _find_prev(i)
            next_idx, next_time = _find_next(i)
            if not prev_time and not next_time:
                continue

            created_dt = created_list[i]
            if prev_time and next_time and prev_time <= next_time:
                window_start = prev_time - grace
                window_end = next_time + grace
                expected_center = prev_time + (next_time - prev_time) / 2
            elif prev_time:
                window_start = prev_time - grace
                window_end = prev_time + max_forward
                expected_center = prev_time
            else:
                window_start = next_time - max_forward
                window_end = next_time + grace
                expected_center = next_time

            if created_dt and window_start <= created_dt <= window_end:
                continue

            # If date tokens exist and match the thread, treat them as strong only
            # when the current created time already falls within the expected window.
            # Otherwise allow a sequence-based correction.
            if date_tokens and date_tokens_match_thread and not (date_anchor_missing or date_anchor_after):
                if created_dt and window_start <= created_dt <= window_end:
                    continue

            # Collect candidate replies in the expected window.
            window_thread = []
            for e in thread:
                sent_ist = _email_ist(e)
                if not sent_ist:
                    continue
                if window_start <= sent_ist <= window_end:
                    window_thread.append(e)

            consultant_in_window = [
                e for e in window_thread
                if _req_match(e, requester)
            ]
            candidate_in_window = consultant_in_window

            ess_only_no_request = "ESS-only; no non-ESS request" in debug_notes
            # Safety guard: do not let ESS-window fallback pick another consultant's
            # thread slice when this requester already has replies elsewhere in thread.
            if not candidate_in_window and ess_only_no_request and not row_has_ids:
                requester_exists_in_thread = any(_req_match(e, requester) for e in thread)
                if not requester_exists_in_thread:
                    candidate_in_window = [
                        e for e in window_thread
                        if _ess_sender(e)
                    ]

            if not candidate_in_window:
                continue

            pick = min(
                candidate_in_window,
                key=lambda e: abs(_email_ist(e) - expected_center),
            )
            sliced = [e for e in window_thread if e.sent_time <= pick.sent_time]
            if not sliced:
                continue

            new_times, new_debug, _episode = _resolve_times_seeded_first(
                thread=sliced,
                requester_name=requester,
                ess_team=ess_team,
                subject_norm=state.get("subject_norm"),
                description=description,
            )
            _seeded_won = _episode is not None

            new_created_dt = _parse_time_str(new_times.created)
            new_created_dt_ist = _to_ist(new_created_dt) if new_created_dt else None
            new_response_dt = _parse_time_str(new_times.response)
            new_response_dt_ist = _to_ist(new_response_dt) if new_response_dt else None
            new_resolved_dt = _parse_time_str(new_times.resolved)
            new_resolved_dt_ist = _to_ist(new_resolved_dt) if new_resolved_dt else None
            if not new_created_dt_ist:
                continue
            if not (new_response_dt_ist and new_resolved_dt_ist):
                continue
            if not (window_start <= new_created_dt_ist <= window_end):
                continue

            orig_times = state.get("times")
            orig_all_same = (
                orig_times.created
                and orig_times.response
                and orig_times.resolved
                and orig_times.created == orig_times.response == orig_times.resolved
            )
            new_all_same = (
                new_times.created
                and new_times.response
                and new_times.resolved
                and new_times.created == new_times.response == new_times.resolved
            )
            if (not orig_all_same and new_all_same):
                # Avoid collapsing timelines that were already distinct.
                continue
            if not _allow_guard_rewrite(
                automation_rows[state["list_index"]],
                state["list_index"],
                new_created_dt_ist,
                new_response_dt_ist,
                new_resolved_dt_ist,
                "OrderSequenceAdjusted",
                "sequence",
            ):
                continue

            # Apply updates to in-memory rows
            list_index = state["list_index"]
            automation_rows[list_index]["Created Date & Time"] = new_times.created
            automation_rows[list_index]["Actual Response Date & Time"] = new_times.response
            automation_rows[list_index]["Actual Resolved Date & Time"] = new_times.resolved

            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = new_debug.created_src
                debug_rows[list_index]["AckSource"] = new_debug.ack_src
                debug_rows[list_index]["ResolvedSource"] = new_debug.resolved_src
                debug_rows[list_index]["Notes"] = f"{new_debug.notes}; OrderSequenceAdjusted; Match={state.get('match_note')}"

            # Apply updates to sheet
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = new_times.created
                ws.cell(row_idx, response_col).value = new_times.response
                ws.cell(row_idx, resolved_col).value = new_times.resolved

            # Update local created list to keep subsequent ordering aligned
            created_list[i] = new_created_dt_ist

        # After sequence adjustments, re-apply "resolved-after-ack" using the full thread.
        # This protects cases where the ordering slice stopped before the true resolution reply
        # (e.g., ack just before midnight and resolution shortly after).
        for state in nonblue_row_states:
            thread = state.get("thread")
            requester = state.get("requester")
            if not thread or not requester:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            description = state.get("description") or ""
            if "maintenance" in description.lower():
                continue

            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if list_index < len(debug_rows):
                notes_now = (debug_rows[list_index].get("Notes", "") or "").lower()
                if "requester follow-up (no in-between request)" in notes_now:
                    continue
                if "failed subject; no ack phrase" in notes_now:
                    continue
                if "no ess or requester replies" in notes_now:
                    continue
                if _row_has_force_same_time_lock(state, debug_rows[list_index].get("Notes", "")):
                    continue
                row_group_total = state.get("group_total") or 0
                ess_like_repeated = (
                    row_group_total >= 2
                    and (
                        "ess-only; no non-ess request" in notes_now
                        or "requester follow-up" in notes_now
                        or "esscontinuationguard[allthree]" in notes_now
                    )
                )
            else:
                notes_now = ""
                ess_like_repeated = False

            row_vals = automation_rows[list_index]
            ack_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            res_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            created_raw = row_vals.get("Created Date & Time")
            response_raw = row_vals.get("Actual Response Date & Time")
            resolved_raw = row_vals.get("Actual Resolved Date & Time")
            if ess_like_repeated and created_raw and created_raw == response_raw == resolved_raw:
                # For repeated ESS-like rows, later distinct-occurrence mapping keeps
                # each occurrence aligned as a full triplet. Do not let this pass
                # move only resolved to a later reply.
                continue
            if not ack_dt:
                continue

            try:
                ack_ist = _to_ist(ack_dt)
            except Exception:
                continue
            res_ist = _to_ist(res_dt) if res_dt else None

            consultant_replies = [
                e for e in thread
                if _match_requester(e.sender_name, e.sender_email, requester)
            ]

            window_end = ack_ist + timedelta(hours=48)
            date_tokens = state.get("date_tokens") or []
            explicit_marker = state.get("explicit_marker")
            enable_postack_fallback_30m = os.getenv("POSTACK_FALLBACK_30M", "0") == "1"
            baseline_date = state.get("baseline_created_date")
            created_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            created_ist = _to_ist(created_dt) if created_dt else None
            anchor_date = _anchor_date(date_tokens)
            if anchor_date:
                anchor_start = _to_ist(datetime(anchor_date.year, anchor_date.month, anchor_date.day))
                window_end = max(window_end, anchor_start + timedelta(hours=36))
            if res_ist and res_ist > (ack_ist + timedelta(minutes=20)):
                continue
            consultant_after = [
                e for e in consultant_replies
                if _email_ist(e) and _email_ist(e) > ack_ist and _email_ist(e) <= window_end
                and _is_real_reply_candidate(e)
            ]
            # Augment with global scan across all emails to catch the next
            # consultant reply even if the thread grouping missed it.
            subj_norm = state.get("subject_norm") or ""
            subj_tokens = _match_tokens(subj_norm)
            subj_inc_set = _inc_tokens(subj_norm)
            min_score = 0.72
            if explicit_marker and anchor_date:
                min_score = 0.60
            allow_global_postack_scan = bool(explicit_marker or anchor_date)
            if allow_global_postack_scan and len(subj_tokens) >= 3 and len(subj_norm) >= 10:
                global_candidates = []
                for e in _emails_for_requester(requester):
                    if not _is_real_reply_candidate(e):
                        continue
                    key_norm = normalize_subject(e.subject or "")
                    key_tokens = _match_tokens(key_norm)
                    if not key_tokens:
                        continue
                    if subj_inc_set:
                        key_inc_set = _inc_tokens(key_norm)
                        if not key_inc_set or subj_inc_set.isdisjoint(key_inc_set):
                            continue
                    score = _token_overlap_score(subj_tokens, key_tokens)
                    contains = (subj_norm in key_norm or key_norm in subj_norm)
                    if score < min_score and not contains:
                        continue
                    sent_ist = _email_ist(e)
                    if not sent_ist:
                        continue
                    if baseline_date and abs((sent_ist.date() - baseline_date).days) > 2:
                        continue
                    if created_ist and sent_ist > (created_ist + timedelta(hours=72)):
                        continue
                    if sent_ist > ack_ist and sent_ist <= window_end:
                        global_candidates.append(e)

                if global_candidates:
                    # Merge global candidates with thread candidates (dedup)
                    seen = set()
                    merged = []
                    for e in consultant_after + global_candidates:
                        dedup_key = (e.subject, e.sender_email, e.sent_time)
                        if dedup_key in seen:
                            continue
                        seen.add(dedup_key)
                        merged.append(e)
                    consultant_after = merged

            if anchor_date and consultant_after:
                on_anchor = [
                    e for e in consultant_after
                    if _email_ist(e) and _email_ist(e).date() == anchor_date
                ]
                if on_anchor:
                    consultant_after = on_anchor

            consultant_after.sort(key=lambda e: e.sent_time)
            fallback_30m_used = False

            if not consultant_after and enable_postack_fallback_30m:
                fallback_after = [
                    e for e in consultant_replies
                    if _email_ist(e) and _email_ist(e) >= (ack_ist + timedelta(minutes=30))
                    and _email_ist(e) <= window_end
                    and _is_real_reply_candidate(e)
                ]
                if anchor_date:
                    on_anchor = [
                        e for e in fallback_after
                        if _email_ist(e) and _email_ist(e).date() == anchor_date
                    ]
                    if on_anchor:
                        fallback_after = on_anchor
                fallback_after.sort(key=lambda e: e.sent_time)
                if fallback_after:
                    consultant_after = fallback_after
                    fallback_30m_used = True

            if not consultant_after:
                continue

            next_reply = _pick_reply_after_ack(consultant_after, ack_ist, requester)
            next_reply_ist = _email_ist(next_reply) if next_reply else None
            if (
                owner_reply_in_window
                and owner_reply is not None
                and (
                    next_reply is None
                    or (next_reply_ist and owner_reply_ist > (next_reply_ist + timedelta(minutes=5)))
                )
            ):
                next_reply = owner_reply
            if not next_reply:
                continue
            next_reply_ist = _email_ist(next_reply)
            next_str = _format_time(next_reply.sent_time)
            if not (next_reply_ist and next_str):
                continue

            row_tokens = _match_tokens(subj_norm)
            row_id_tokens = _id_like_tokens(subj_norm)
            if not row_id_tokens:
                row_id_tokens = _id_like_tokens(description)

            lane_thread = _expanded_thread(
                subj_norm,
                thread,
                requester,
                include_non_ess=True,
                reference_ist=next_reply_ist,
            ) or thread
            lane_episode = _lane_local_episode_from_reply(
                next_reply,
                next_reply_ist,
                lane_thread,
                requester,
                subj_norm,
                row_tokens,
                row_id_tokens,
            )

            note_suffix = "ResolvedAfterAckPost"
            if not consultant_replies:
                note_suffix = "ResolvedAfterAckPostGlobal"
            if fallback_30m_used:
                note_suffix = "ResolvedAfterAckPostFallback30m"

            rebased = False
            if lane_episode and lane_episode.get("request"):
                lane_req_ist = lane_episode.get("request")
                lane_ack_ist = lane_episode.get("ack") or next_reply_ist
                lane_ack_msg = lane_episode.get("ack_msg")
                if (
                    lane_req_ist
                    and lane_req_ist <= lane_ack_ist <= next_reply_ist
                    and (next_reply_ist - lane_req_ist) <= timedelta(hours=48)
                ):
                    current_created_ist = _to_ist(created_dt) if created_dt else None
                    should_rebase_lane = False
                    if current_created_ist and lane_req_ist > (current_created_ist + timedelta(minutes=30)):
                        should_rebase_lane = True
                    if lane_ack_ist > (ack_ist + timedelta(minutes=5)):
                        should_rebase_lane = True
                    if should_rebase_lane:
                        rebased = _apply_guarded_episode_update(
                            state,
                            row_vals,
                            list_index,
                            state.get("row_index"),
                            lane_req_ist,
                            lane_ack_ist,
                            next_reply_ist,
                            owner_tag=note_suffix,
                            candidate_kind="requester_ack",
                            created_src="PARSED_FROM_QUOTED_REQUEST",
                            ack_src=(
                                lane_ack_msg.sender_email or lane_ack_msg.sender_name
                                if lane_ack_msg else (next_reply.sender_email or next_reply.sender_name or "PARSED_FROM_QUOTED_REPLY")
                            ),
                            resolved_src=next_reply.sender_email or next_reply.sender_name,
                            note_suffix=note_suffix,
                        )

            if rebased:
                continue

            row_vals["Actual Resolved Date & Time"] = next_str
            if list_index < len(debug_rows):
                debug_rows[list_index]["ResolvedSource"] = next_reply.sender_email or next_reply.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; {note_suffix}"

            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, resolved_col).value = next_str

        # Cross-subject duplicate triplet guard (narrow):
        # when same consultant has identical created/ack/resolved across
        # different subjects, try to move only resolved to the next valid
        # consultant non-ack reply in that row's own thread.
        seen_by_requester_triplet = {}
        state_by_list_index = {
            s.get("list_index"): s
            for s in row_states
            if s.get("list_index") is not None
        }
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if list_index < len(debug_rows):
                notes_now = (debug_rows[list_index].get("Notes", "") or "").lower()
                if "requester follow-up (no in-between request)" in notes_now:
                    continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            if not (c and a and r):
                continue

            requester = state.get("requester") or ""
            req_key = _requester_key(requester)
            triplet_key = (req_key, c, a, r)
            prev_idx = seen_by_requester_triplet.get(triplet_key)
            if prev_idx is None:
                seen_by_requester_triplet[triplet_key] = list_index
                continue

            prev_state = state_by_list_index.get(prev_idx)
            prev_subject = (prev_state or {}).get("subject_norm") if prev_state else None
            curr_subject = state.get("subject_norm")
            if not prev_subject or not curr_subject or prev_subject == curr_subject:
                continue

            thread = state.get("thread") or []
            if not thread:
                continue

            res_dt = _parse_time_str(r) or _parse_time_str(a)
            if not res_dt:
                continue
            try:
                res_ist = _to_ist(res_dt)
            except Exception:
                continue

            date_tokens = state.get("date_tokens") or []
            anchor_date = _anchor_date(date_tokens)
            candidates = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _is_real_reply_candidate(e)
            ]
            if anchor_date and candidates:
                on_anchor = [e for e in candidates if _email_ist(e) and _email_ist(e).date() == anchor_date]
                if on_anchor:
                    candidates = on_anchor
            if not candidates:
                continue

            candidates.sort(key=lambda e: e.sent_time)
            next_candidates = [
                e for e in candidates
                if _email_ist(e) and _email_ist(e) > res_ist
                and _email_ist(e) <= (res_ist + timedelta(hours=48))
            ]
            if not next_candidates:
                continue

            pick = next_candidates[0]
            new_res = _format_time(pick.sent_time)
            if not new_res or new_res == r:
                continue

            row_idx = state.get("row_index")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            pick_ist = _email_ist(pick)
            if not (c_ist and a_ist and pick_ist and c_ist <= a_ist <= pick_ist):
                continue
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                c_ist,
                a_ist,
                pick_ist,
                owner_tag="CrossSubjectDuplicateGuard",
                candidate_kind="cleanup",
                resolved_src=pick.sender_email or pick.sender_name,
                note_suffix="CrossSubjectDuplicateGuard",
            )

        # Episode consistency guard (conservative):
        # If resolved is far after ack, re-anchor Created/Ack to a newer strong
        # request->ack episode in the same matched thread.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue

            try:
                c_ist = _to_ist(c_dt)
                a_ist = _to_ist(a_dt)
                r_ist = _to_ist(r_dt)
            except Exception:
                continue
            if not (c_ist <= a_ist <= r_ist):
                continue

            old_gap = r_ist - a_ist
            # Only touch rows with clearly stale ack vs resolved.
            if old_gap < timedelta(hours=18):
                continue

            subj_norm = state.get("subject_norm") or ""
            subj_tokens = _match_tokens(subj_norm)
            subj_inc_set = _inc_tokens(subj_norm)
            subj_num_set = _sig_num_tokens(subj_norm)

            episode_candidates = []
            for ack_mail in thread:
                if not ack_mail.sent_time:
                    continue
                if not _req_match(ack_mail, requester):
                    continue
                ack_mail_ist = _email_ist(ack_mail)
                if not ack_mail_ist:
                    continue
                if ack_mail_ist <= a_ist + timedelta(hours=4):
                    continue
                if ack_mail_ist > r_ist:
                    continue
                if not _is_ack_like_reply(ack_mail):
                    continue

                key_norm = normalize_subject(ack_mail.subject or "")
                key_tokens = _match_tokens(key_norm)
                if subj_tokens and key_tokens:
                    score = _token_overlap_score(subj_tokens, key_tokens)
                    contains = (subj_norm in key_norm or key_norm in subj_norm)
                    if score < 0.45 and not contains:
                        continue
                key_inc_set = _inc_tokens(key_norm)
                if subj_inc_set and key_inc_set and subj_inc_set.isdisjoint(key_inc_set):
                    continue
                key_num_set = _sig_num_tokens(key_norm)
                if subj_num_set and key_num_set and subj_num_set.isdisjoint(key_num_set):
                    continue

                req_candidates = [
                    e for e in thread
                    if e.sent_time
                    and not _ess_sender(e)
                    and _email_ist(e) and _email_ist(e) <= ack_mail_ist
                ]
                if not req_candidates:
                    continue
                req_mail = max(req_candidates, key=lambda e: e.sent_time)
                req_ist = _email_ist(req_mail)
                if not req_ist:
                    continue
                if req_ist <= c_ist + timedelta(hours=2):
                    continue
                if req_ist > ack_mail_ist:
                    continue
                req_ack_gap = ack_mail_ist - req_ist
                if req_ack_gap > timedelta(minutes=45):
                    continue
                if req_ack_gap < timedelta(minutes=1):
                    continue

                new_gap = r_ist - ack_mail_ist
                episode_candidates.append((new_gap, -ack_mail_ist.timestamp(), req_mail, ack_mail))

            if not episode_candidates:
                continue

            episode_candidates.sort(key=lambda x: (x[0], x[1]))
            _new_gap, _neg_ack_ts, req_mail, ack_mail = episode_candidates[0]
            req_ist = _to_ist(req_mail.sent_time)
            ack_ist = _to_ist(ack_mail.sent_time)
            if not (req_ist <= ack_ist <= r_ist):
                continue
            # Apply only on clear improvement.
            if (r_ist - ack_ist) >= old_gap - timedelta(hours=2):
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            if new_created == c and new_ack == a:
                continue
            row_idx = state.get("row_index")
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                req_ist,
                ack_ist,
                r_ist,
                owner_tag="EpisodeAckRefreshGuard",
                candidate_kind="requester_ack",
                created_src=req_mail.sender_email or req_mail.sender_name,
                ack_src=ack_mail.sender_email or ack_mail.sender_name,
                note_suffix="EpisodeAckRefreshGuard",
            )

        # Episode ack-refresh guard (strict):
        # Re-anchor Created/Ack only when there is a clearly newer request->ack
        # pair in the same thread and it materially improves Ack->Resolved gap.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue

            old_gap = r_ist - a_ist
            if old_gap < timedelta(hours=12):
                continue

            subj_norm = state.get("subject_norm") or ""
            subj_tokens = _match_tokens(subj_norm)

            requester_candidates = []
            for e in thread:
                if not e.sent_time:
                    continue
                if not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist <= a_ist + timedelta(hours=2) or e_ist > r_ist:
                    continue
                if not _ack_like(e):
                    continue
                key_norm = normalize_subject(e.subject or "")
                key_tokens = _match_tokens(key_norm)
                if subj_tokens and key_tokens:
                    score = _token_overlap_score(subj_tokens, key_tokens)
                    contains = (subj_norm in key_norm or key_norm in subj_norm)
                    if score < 0.45 and not contains:
                        continue
                requester_candidates.append(e)
            if not requester_candidates:
                continue
            requester_candidates.sort(key=lambda e: e.sent_time, reverse=True)

            applied = False
            for ack_mail in requester_candidates:
                ack_ist = _to_ist(ack_mail.sent_time)
                req_before = [
                    e for e in thread
                    if e.sent_time
                    and not _ess_sender(e)
                    and _email_ist(e) and _email_ist(e) <= ack_ist
                ]
                guard_tag = "EpisodeAckRefreshGuard"
                req_mail = None
                req_ist = None
                new_gap = None

                if req_before:
                    req_mail = max(req_before, key=lambda e: e.sent_time)
                    req_ist = _email_ist(req_mail)
                    if not req_ist:
                        continue
                    if req_ist <= c_ist + timedelta(hours=2):
                        continue
                    if not (timedelta(minutes=1) <= (ack_ist - req_ist) <= timedelta(minutes=45)):
                        continue
                    new_gap = r_ist - ack_ist
                    if new_gap >= old_gap - timedelta(hours=4):
                        continue
                else:
                    # Safe fallback for ESS-only rows:
                    # when no non-ESS request exists, allow requester self-episode
                    # re-anchor only on a tight request->ack pair and strong gap improvement.
                    has_non_ess = any(
                        e.sent_time and not _ess_sender(e)
                        and _email_ist(e)
                        for e in thread
                    )
                    if has_non_ess:
                        continue
                    req_before_self = [
                        e for e in thread
                        if e.sent_time
                        and _req_match(e, requester)
                        and _email_ist(e) and _email_ist(e) <= ack_ist
                        and not _ack_like(e)
                    ]
                    if not req_before_self:
                        continue
                    req_mail = max(req_before_self, key=lambda e: e.sent_time)
                    req_ist = _email_ist(req_mail)
                    if not req_ist:
                        continue
                    if req_ist <= c_ist + timedelta(hours=8):
                        continue
                    if not (timedelta(minutes=1) <= (ack_ist - req_ist) <= timedelta(minutes=30)):
                        continue
                    new_gap = r_ist - ack_ist
                    if new_gap >= old_gap - timedelta(hours=8):
                        continue
                    guard_tag = "EpisodeAckRefreshSelfGuard"

                new_created = _format_time(req_mail.sent_time)
                new_ack = _format_time(ack_mail.sent_time)
                if not new_created or not new_ack:
                    continue
                row_idx = state.get("row_index")
                applied = _apply_guarded_episode_update(
                    state,
                    row_vals,
                    list_index,
                    row_idx,
                    req_ist,
                    ack_ist,
                    r_ist,
                    owner_tag=guard_tag,
                    candidate_kind="requester_ack",
                    created_src=req_mail.sender_email or req_mail.sender_name,
                    ack_src=ack_mail.sender_email or ack_mail.sender_name,
                    note_suffix=guard_tag,
                )
                if not applied:
                    continue
                break
            if applied:
                continue

        # ESS-only span rebase guard (strict):
        # For ESS-only span rows with very old Ack, allow rebasing Created/Ack to
        # a clearly later requester episode when no non-ESS request exists.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue
            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if not any(
                tag in notes_l
                for tag in (
                    "ess-only; no non-ess request; span",
                    "ess-only; no non-ess request; requester span",
                    "ess-only; no non-ess request; requester span(ack-like)",
                    "ess-only; no non-ess request; requester span(all-ack->ess)",
                )
            ):
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue

            old_gap = r_ist - a_ist
            if old_gap < timedelta(hours=12):
                continue

            requester_after_all = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _email_ist(e) and _email_ist(e) >= a_ist + timedelta(hours=2)
                and _email_ist(e) <= r_ist
            ]
            requester_after_all.sort(key=lambda e: e.sent_time)
            requester_after_non_ack = _shared_resolution_candidates(requester_after_all)
            if not requester_after_non_ack:
                continue

            req_mail = requester_after_non_ack[0]
            req_ist = _to_ist(req_mail.sent_time)
            ack_mail = None
            # Prefer explicit ack-like requester follow-up within 60m.
            for e in requester_after_all:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist < req_ist:
                    continue
                if e_ist > req_ist + timedelta(minutes=60):
                    break
                if _ack_like(e):
                    ack_mail = e
                    break
            if ack_mail is None:
                # Conservative fallback: treat requester episode start as ack only
                # when it still materially improves stale Ack->Resolved gap.
                ack_mail = req_mail

            ack_ist = _to_ist(ack_mail.sent_time)
            if not (req_ist <= ack_ist <= r_ist):
                continue
            new_gap = r_ist - ack_ist
            if new_gap >= old_gap - timedelta(hours=6):
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            row_idx = state.get("row_index")
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                req_ist,
                ack_ist,
                r_ist,
                owner_tag="ESSSpanRebaseGuard",
                candidate_kind="requester_ack",
                created_src=req_mail.sender_email or req_mail.sender_name,
                ack_src=ack_mail.sender_email or ack_mail.sender_name,
                note_suffix="ESSSpanRebaseGuard",
            )

        # Requester episode rebase guard (very narrow):
        # For stale rows where resolved is already from requester but created/ack
        # stayed on an older non-requester episode, rebase Created/Ack to the
        # first requester non-ack episode in the same thread.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue
            old_gap = r_ist - a_ist
            if old_gap < timedelta(hours=10):
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""

            # Keep this guard narrow to avoid disturbing normal request/ack rows.
            created_is_parsed = isinstance(created_src_now, str) and created_src_now.startswith("PARSED_FROM_")
            span_row = isinstance(notes_now, str) and ("span" in notes_now.lower())
            if not (created_is_parsed or span_row):
                continue
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue
            if _match_requester(ack_src_now, ack_src_now, requester):
                continue

            requester_after = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and not _ack_like(e)
                and _email_ist(e) and _email_ist(e) >= a_ist + timedelta(hours=2)
                and _email_ist(e) <= r_ist
            ]
            requester_after.sort(key=lambda e: e.sent_time)
            if not requester_after:
                continue

            req_mail = requester_after[0]
            req_ist = _to_ist(req_mail.sent_time)
            ack_mail = req_mail
            for e in thread:
                if not e.sent_time:
                    continue
                if not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist < req_ist:
                    continue
                if e_ist > req_ist + timedelta(minutes=45):
                    break
                if _ack_like(e):
                    ack_mail = e
                    break

            ack_ist = _to_ist(ack_mail.sent_time)
            if not (req_ist <= ack_ist <= r_ist):
                continue
            new_gap = r_ist - ack_ist
            if new_gap >= old_gap - timedelta(hours=2):
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            row_idx = state.get("row_index")
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                req_ist,
                ack_ist,
                r_ist,
                owner_tag="RequesterEpisodeRebaseGuard",
                candidate_kind="requester_ack",
                created_src=req_mail.sender_email or req_mail.sender_name,
                ack_src=ack_mail.sender_email or ack_mail.sender_name,
                note_suffix="RequesterEpisodeRebaseGuard",
            )

        # Final ack-delay guard (narrow):
        # In ESS-only rows, prevent very-late ack timestamps from stale carry-over.
        # Do not run this on mixed/requester rows.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue
            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if "ess-only; no non-ess request" not in notes_l:
                continue

            has_non_ess = any(
                e.sent_time
                and not _ess_sender(e)
                and not _system_like_sender(e)
                for e in thread
            )
            if has_non_ess:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            if a_ist <= c_ist:
                continue
            if (a_ist - c_ist) <= timedelta(hours=12):
                continue

            window_end = c_ist + timedelta(hours=12)
            requester_window = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _email_ist(e) and _email_ist(e) > c_ist
                and _email_ist(e) <= window_end
            ]
            requester_window.sort(key=lambda e: e.sent_time)
            ess_window = [
                e for e in thread
                if e.sent_time
                and _ess_sender(e)
                and _email_ist(e) and _email_ist(e) > c_ist
                and _email_ist(e) <= window_end
            ]
            ess_window.sort(key=lambda e: e.sent_time)

            ack_pick = None
            for e in requester_window:
                if _ack_like(e):
                    ack_pick = e
                    break
            if ack_pick is None:
                for e in ess_window:
                    if _ack_like(e):
                        ack_pick = e
                        break
            if ack_pick is None and requester_window:
                ack_pick = requester_window[0]
            if ack_pick is None and ess_window:
                ack_pick = ess_window[0]

            new_ack = _format_time(ack_pick.sent_time) if ack_pick else c
            if not new_ack or new_ack == a:
                continue
            row_idx = state.get("row_index")
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            r_ist = _to_ist(r_dt) if r_dt else None
            ack_pick_ist = _email_ist(ack_pick) if ack_pick else c_ist
            cand_r_ist = r_ist if (r_ist and r_ist >= ack_pick_ist) else ack_pick_ist
            if not (c_ist and ack_pick_ist and cand_r_ist and c_ist <= ack_pick_ist <= cand_r_ist):
                continue
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                c_ist,
                ack_pick_ist,
                cand_r_ist,
                owner_tag="AckDelayWindowGuard",
                candidate_kind="cleanup",
                ack_src=(ack_pick.sender_email or ack_pick.sender_name) if ack_pick else "ACK NOT FOUND",
                note_suffix="AckDelayWindowGuard",
            )

        # Quoted request rebase guard (safe):
        # When Created is stale vs Ack/Resolved, mine requester mails in the same
        # episode for quoted non-ESS request times and re-anchor Created.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            if not requester or not subject_norm:
                continue
            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and (a_dt or r_dt)):
                continue
            c_ist = _to_ist(c_dt)
            upper_ist = _to_ist(a_dt) if a_dt else _to_ist(r_dt)
            if not upper_ist:
                continue
            if (upper_ist - c_ist) < timedelta(hours=18):
                continue

            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=upper_ist,
            )
            if not thread:
                continue

            parsed_candidates = []
            for e in thread:
                if not getattr(e, "sent_time", None):
                    continue
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist > (upper_ist + timedelta(minutes=5)):
                    continue
                if e_ist < (upper_ist - timedelta(days=10)):
                    continue
                if not _req_match(e, requester):
                    continue
                if _ack_like(e) or _ack_like_text_fallback(e):
                    continue
                try:
                    parsed_dt = _extract_request_time_from_email(
                        e,
                        ess_team,
                        max_dt=e.sent_time,
                        subject_norm=subject_norm,
                    )
                except Exception:
                    parsed_dt = None
                if not parsed_dt:
                    continue
                parsed_ist = _to_ist(parsed_dt)
                if parsed_ist > (upper_ist + timedelta(minutes=5)):
                    continue
                if parsed_ist < (upper_ist - timedelta(days=10)):
                    continue
                if parsed_ist <= (c_ist + timedelta(minutes=30)):
                    continue
                parsed_candidates.append(parsed_ist)
            if not parsed_candidates:
                continue

            parsed_candidates.sort()
            pick_ist = parsed_candidates[-1]
            t = _format_time(pick_ist)
            if not t:
                continue

            row_idx = state.get("row_index")
            cand_a_ist = _to_ist(a_dt) if a_dt else pick_ist
            cand_r_ist = _to_ist(r_dt) if r_dt else cand_a_ist
            if cand_a_ist < pick_ist:
                cand_a_ist = pick_ist
            if cand_r_ist < cand_a_ist:
                cand_r_ist = cand_a_ist
            if not (pick_ist <= cand_a_ist <= cand_r_ist):
                continue
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                pick_ist,
                cand_a_ist,
                cand_r_ist,
                owner_tag="QuotedRequestRebaseGuard",
                candidate_kind="quoted",
                created_src="PARSED_FROM_QUOTED_REQUEST",
                ack_src="PARSED_FROM_QUOTED_REQUEST" if cand_a_ist == pick_ist else "",
                resolved_src="PARSED_FROM_QUOTED_REQUEST" if cand_r_ist == cand_a_ist == pick_ist else "",
                note_suffix="QuotedRequestRebaseGuard",
            )

        # Final non-ack resolved guard (global, safe):
        # If resolved lands on an ack-like requester reminder, rebase resolved
        # to latest requester non-ack reply in the same thread.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not r_dt:
                continue
            floor_dt = None
            if c_dt and a_dt:
                floor_dt = max(_to_ist(c_dt), _to_ist(a_dt))
            elif a_dt:
                floor_dt = _to_ist(a_dt)
            elif c_dt:
                floor_dt = _to_ist(c_dt)
            r_ist = _to_ist(r_dt)

            resolved_emails = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _email_ist(e) and abs((_email_ist(e) - r_ist).total_seconds()) <= 61
            ]
            resolved_mail = None
            if resolved_emails:
                resolved_mail = min(
                    resolved_emails,
                    key=lambda e: abs((_email_ist(e) - r_ist).total_seconds()),
                )
            else:
                # Sheet timestamps are minute-precision; tolerate small drift.
                near_hits = [
                    e for e in thread
                    if e.sent_time
                    and _req_match(e, requester)
                    and _email_ist(e) and abs((_email_ist(e) - r_ist).total_seconds()) <= 300
                ]
                if near_hits:
                    resolved_mail = min(
                        near_hits,
                        key=lambda e: abs((_email_ist(e) - r_ist).total_seconds()),
                    )
            if not resolved_mail or not _ack_like(resolved_mail):
                continue

            non_ack_pool = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _is_real_reply_candidate(e)
                and _email_ist(e) and _email_ist(e) <= r_ist
                and (floor_dt is None or _email_ist(e) >= floor_dt)
            ]
            if not non_ack_pool:
                continue

            pick = max(non_ack_pool, key=lambda e: e.sent_time)

            new_res = _format_time(pick.sent_time)
            if not new_res or new_res == r:
                continue
            row_idx = state.get("row_index")
            c_ist = _to_ist(c_dt) if c_dt else None
            a_ist = _to_ist(a_dt) if a_dt else None
            pick_ist = _email_ist(pick)
            if not (pick_ist and floor_dt and pick_ist >= floor_dt):
                continue
            if c_ist and a_ist:
                cand_c_ist = c_ist
                cand_a_ist = a_ist
            elif a_ist:
                cand_c_ist = a_ist
                cand_a_ist = a_ist
            elif c_ist:
                cand_c_ist = c_ist
                cand_a_ist = c_ist
            else:
                continue
            if not (cand_c_ist <= cand_a_ist <= pick_ist):
                continue
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                cand_c_ist,
                cand_a_ist,
                pick_ist,
                owner_tag="ResolvedNonAckGuard",
                candidate_kind="cleanup",
                resolved_src=pick.sender_email or pick.sender_name,
                note_suffix="ResolvedNonAckGuard",
            )

        # Late-episode rebase guard (narrow):
        # If resolved is much later than ack and resolved belongs to requester,
        # re-anchor Created/Ack to the latest requester episode near resolved.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue
            if (r_ist - a_ist) < timedelta(hours=48):
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue
            if _match_requester(ack_src_now, ack_src_now, requester) and not str(created_src_now).startswith("PARSED_FROM_"):
                continue

            requester_timeline = [
                e for e in thread
                if e.sent_time and _req_match(e, requester) and _email_ist(e) and _email_ist(e) <= r_ist
            ]
            requester_timeline.sort(key=lambda e: e.sent_time)
            if not requester_timeline:
                continue

            # Use latest requester episode ending at resolved.
            end_idx = len(requester_timeline) - 1
            episode_start_idx = end_idx
            for j in range(end_idx, 0, -1):
                curr_t = _email_ist(requester_timeline[j])
                prev_t = _email_ist(requester_timeline[j - 1])
                if not curr_t or not prev_t:
                    continue
                if (curr_t - prev_t) > timedelta(hours=6):
                    break
                episode_start_idx = j - 1

            episode_slice = requester_timeline[episode_start_idx : end_idx + 1]
            non_ack_episode = _shared_resolution_candidates(episode_slice)
            if not non_ack_episode:
                continue
            req_mail = non_ack_episode[0]
            req_ist = _email_ist(req_mail)
            if not req_ist:
                continue
            if req_ist <= a_ist + timedelta(hours=2):
                continue

            ack_mail = req_mail
            for e in episode_slice:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist < req_ist:
                    continue
                if e_ist > req_ist + timedelta(minutes=60):
                    break
                if _ack_like(e):
                    ack_mail = e
                    break
            ack_ist = _email_ist(ack_mail)
            if not ack_ist or not (req_ist <= ack_ist <= r_ist):
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            if new_created == c and new_ack == a:
                continue

            row_idx = state.get("row_index")
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                req_ist,
                ack_ist,
                r_ist,
                owner_tag="LateEpisodeRebaseGuard",
                candidate_kind="requester_ack",
                created_src=req_mail.sender_email or req_mail.sender_name,
                ack_src=ack_mail.sender_email or ack_mail.sender_name,
                note_suffix="LateEpisodeRebaseGuard",
            )

        # Requester-ack ownership guard (narrow):
        # If resolved is from requester but ack is not, align ack to first
        # requester reply after created within the row thread.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            thread = state.get("thread") or []
            requester = state.get("requester") or ""
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue
            if _match_requester(ack_src_now, ack_src_now, requester):
                continue

            requester_after_created = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _email_ist(e) and _email_ist(e) >= c_ist
                and _email_ist(e) <= r_ist
            ]
            requester_after_created.sort(key=lambda e: e.sent_time)
            if not requester_after_created:
                continue

            ack_pick = requester_after_created[0]
            ack_pick_ist = _email_ist(ack_pick)
            if not ack_pick_ist:
                continue

            new_ack = _format_time(ack_pick.sent_time)
            if not new_ack:
                continue

            new_created = c
            # Only rebase created in stale parsed-anchor cases.
            if (
                isinstance(created_src_now, str)
                and created_src_now.startswith("PARSED_FROM_")
                and (ack_pick_ist - c_ist) >= timedelta(hours=24)
            ):
                new_created = new_ack

            changed = (new_ack != a) or (new_created != c)
            if not changed:
                continue

            row_vals["Actual Response Date & Time"] = new_ack
            row_vals["Created Date & Time"] = new_created
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = new_created
                ws.cell(row_idx, response_col).value = new_ack
            if list_index < len(debug_rows):
                if new_created != c:
                    debug_rows[list_index]["CreatedSource"] = ack_pick.sender_email or ack_pick.sender_name
                debug_rows[list_index]["AckSource"] = ack_pick.sender_email or ack_pick.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterAckOwnershipGuard"

        # Generic stale-created guard for ESS-only rows:
        # if Created stayed on an older non-requester episode while Ack/Resolved
        # are requester-owned, rebase Created to the requester episode near Ack.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester or not subject_norm:
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            if "ESS-only; no non-ESS request" not in notes_now:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue
            if (a_ist - c_ist) < timedelta(hours=12):
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            if _match_requester(created_src_now, created_src_now, requester):
                continue
            if not _match_requester(ack_src_now, ack_src_now, requester):
                continue
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue

            row_tokens = _match_tokens(subject_norm)
            if not row_tokens:
                continue

            requester_before_ack = []
            for e in thread:
                if not e.sent_time or not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist or e_ist > a_ist:
                    continue
                # Keep this near the active episode only.
                if e_ist < (a_ist - timedelta(days=2)):
                    continue
                mail_subj = normalize_subject(e.subject or "")
                mail_tokens = _match_tokens(mail_subj)
                sim = _token_overlap_score(row_tokens, mail_tokens) if mail_tokens else 0.0
                contains = (subject_norm in mail_subj or mail_subj in subject_norm) if mail_subj else False
                if sim >= 0.40 or contains:
                    requester_before_ack.append(e)

            requester_before_ack.sort(key=lambda e: e.sent_time)
            if not requester_before_ack:
                continue

            # Prefer a non-ack requester message as Created anchor.
            non_ack_before_ack = _shared_resolution_candidates(requester_before_ack)
            if not non_ack_before_ack:
                continue
            req_mail = non_ack_before_ack[0]

            req_ist = _email_ist(req_mail)
            if not req_ist or req_ist > a_ist:
                continue

            new_created = _format_time(req_mail.sent_time)
            if not new_created or new_created == c:
                continue
            row_vals["Created Date & Time"] = new_created
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = new_created
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = req_mail.sender_email or req_mail.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; StaleCreatedRequesterEpisodeGuard"

        # Generic requester subject-episode ownership guard:
        # when resolved belongs to requester but ack source does not, re-anchor
        # Created/Ack to the first requester episode that matches this row subject.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester or not subject_norm:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue
            if (r_ist - a_ist) < timedelta(hours=8):
                continue

            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue
            if _match_requester(ack_src_now, ack_src_now, requester):
                continue

            row_tokens = _match_tokens(subject_norm)
            if not row_tokens:
                continue

            window_end = min(r_ist + timedelta(minutes=5), c_ist + timedelta(days=10))
            requester_timeline = []
            for e in thread:
                if not e.sent_time or not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist or e_ist < c_ist or e_ist > window_end:
                    continue
                mail_subj = normalize_subject(e.subject or "")
                mail_tokens = _match_tokens(mail_subj)
                sim = _token_overlap_score(row_tokens, mail_tokens) if mail_tokens else 0.0
                contains = (subject_norm in mail_subj or mail_subj in subject_norm) if mail_subj else False
                if sim >= 0.35 or contains:
                    requester_timeline.append(e)

            requester_timeline.sort(key=lambda e: e.sent_time)
            if not requester_timeline:
                continue

            requester_non_ack = _shared_resolution_candidates(requester_timeline)
            if not requester_non_ack:
                continue
            req_mail = requester_non_ack[0]
            req_ist = _email_ist(req_mail)
            if not req_ist:
                continue
            if req_ist <= a_ist + timedelta(hours=2):
                continue
            if (r_ist - req_ist) > timedelta(days=2):
                continue

            ack_mail = req_mail
            for e in requester_timeline:
                e_ist = _email_ist(e)
                if not e_ist or e_ist < req_ist:
                    continue
                if e_ist > req_ist + timedelta(minutes=60):
                    break
                if _ack_like(e):
                    ack_mail = e
                    break

            ack_ist = _email_ist(ack_mail)
            if not ack_ist or ack_ist > (r_ist + timedelta(minutes=5)):
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            if new_created == c and new_ack == a:
                continue

            row_vals["Created Date & Time"] = new_created
            row_vals["Actual Response Date & Time"] = new_ack
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = new_created
                ws.cell(row_idx, response_col).value = new_ack
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = req_mail.sender_email or req_mail.sender_name
                debug_rows[list_index]["AckSource"] = ack_mail.sender_email or ack_mail.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterSubjectEpisodeGuard"

        # Mixed requester-ownership guard (global):
        # If Created/Ack were taken from the same non-requester source but Resolved
        # belongs to requester, re-anchor Created/Ack to requester episode near Resolved.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester or not subject_norm:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue
            if (r_ist - a_ist) < timedelta(hours=4):
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue
            if _match_requester(ack_src_now, ack_src_now, requester):
                continue
            if not created_src_now or not ack_src_now:
                continue
            if _requester_key(created_src_now) != _requester_key(ack_src_now):
                continue

            row_tokens = _match_tokens(subject_norm)
            if not row_tokens:
                continue

            r_soft = r_ist + timedelta(minutes=5)
            requester_timeline = []
            for e in thread:
                if not e.sent_time or not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist or e_ist > r_soft:
                    continue
                mail_subj = normalize_subject(e.subject or "")
                mail_tokens = _match_tokens(mail_subj)
                sim = _token_overlap_score(row_tokens, mail_tokens) if mail_tokens else 0.0
                contains = (subject_norm in mail_subj or mail_subj in subject_norm) if mail_subj else False
                if sim >= 0.30 or contains:
                    requester_timeline.append(e)

            requester_timeline.sort(key=lambda e: e.sent_time)
            if not requester_timeline:
                continue

            end_idx = len(requester_timeline) - 1
            start_idx = end_idx
            for j in range(end_idx, 0, -1):
                curr_t = _email_ist(requester_timeline[j])
                prev_t = _email_ist(requester_timeline[j - 1])
                if not curr_t or not prev_t:
                    continue
                if (curr_t - prev_t) > timedelta(hours=8):
                    break
                start_idx = j - 1
            episode = requester_timeline[start_idx : end_idx + 1]
            if not episode:
                continue

            non_ack_episode = _shared_resolution_candidates(episode)
            if not non_ack_episode:
                continue
            req_mail = non_ack_episode[0]
            req_ist = _email_ist(req_mail)
            if not req_ist:
                continue
            if req_ist <= c_ist + timedelta(minutes=30):
                continue
            if req_ist < (a_ist - timedelta(minutes=5)):
                continue
            if (r_ist - req_ist) > timedelta(days=2):
                continue

            ack_mail = req_mail
            for e in episode:
                e_ist = _email_ist(e)
                if not e_ist or e_ist < req_ist:
                    continue
                if e_ist > req_ist + timedelta(minutes=60):
                    break
                if _ack_like(e):
                    ack_mail = e
                    break
            ack_ist = _email_ist(ack_mail)
            if not ack_ist or ack_ist > r_soft:
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            if new_created == c and new_ack == a:
                continue

            row_vals["Created Date & Time"] = new_created
            row_vals["Actual Response Date & Time"] = new_ack
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = new_created
                ws.cell(row_idx, response_col).value = new_ack
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = req_mail.sender_email or req_mail.sender_name
                debug_rows[list_index]["AckSource"] = ack_mail.sender_email or ack_mail.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; MixedRequesterEpisodeGuard"

        # Resolved-window reanchor guard (global):
        # If Created/Ack are from same non-requester source but Resolved is requester,
        # anchor Created/Ack from requester activity close to Resolved timestamp.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester or not subject_norm:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            if not created_src_now or not ack_src_now:
                continue
            if _requester_key(created_src_now) != _requester_key(ack_src_now):
                continue
            if _match_requester(ack_src_now, ack_src_now, requester):
                continue
            if not _match_requester(resolved_src_now, resolved_src_now, requester):
                continue

            row_tokens = _match_tokens(subject_norm)
            if not row_tokens:
                continue

            win_start = r_ist - timedelta(hours=48)
            win_end = r_ist + timedelta(minutes=5)
            requester_window = []
            for e in thread:
                if not e.sent_time or not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist or e_ist < win_start or e_ist > win_end:
                    continue
                mail_subj = normalize_subject(e.subject or "")
                mail_tokens = _match_tokens(mail_subj)
                sim = _token_overlap_score(row_tokens, mail_tokens) if mail_tokens else 0.0
                contains = (subject_norm in mail_subj or mail_subj in subject_norm) if mail_subj else False
                if sim >= 0.25 or contains:
                    requester_window.append(e)

            requester_window.sort(key=lambda e: e.sent_time)
            if not requester_window:
                continue

            requester_non_ack = _shared_resolution_candidates(requester_window)
            if not requester_non_ack:
                continue
            req_mail = requester_non_ack[0]
            req_ist = _email_ist(req_mail)
            if not req_ist:
                continue
            if req_ist < (c_ist - timedelta(minutes=5)):
                continue

            ack_mail = req_mail
            for e in requester_window:
                e_ist = _email_ist(e)
                if not e_ist or e_ist < req_ist:
                    continue
                if e_ist > req_ist + timedelta(minutes=60):
                    break
                if _ack_like(e):
                    ack_mail = e
                    break
            ack_ist = _email_ist(ack_mail)
            if not ack_ist or ack_ist > win_end:
                continue

            new_created = _format_time(req_mail.sent_time)
            new_ack = _format_time(ack_mail.sent_time)
            if not new_created or not new_ack:
                continue
            if new_created == c and new_ack == a:
                continue

            row_vals["Created Date & Time"] = new_created
            row_vals["Actual Response Date & Time"] = new_ack
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = new_created
                ws.cell(row_idx, response_col).value = new_ack
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = req_mail.sender_email or req_mail.sender_name
                debug_rows[list_index]["AckSource"] = ack_mail.sender_email or ack_mail.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ResolvedWindowReanchorGuard"

        # Requester-span ack-like guard (global):
        # For ACK NOT FOUND + requester span(ack-like), prefer latest requester
        # non-ack in thread so reminder mails do not anchor old timestamps.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if list_index < len(debug_rows):
                notes_now = (debug_rows[list_index].get("Notes", "") or "").lower()
                if "requester follow-up (no in-between request)" in notes_now:
                    continue
            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester:
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if (
                "requester span(ack-like)" not in notes_l
                and "requester span(all-ack->ess)" not in notes_l
            ):
                continue
            if str(ack_src_now).strip().upper() != "ACK NOT FOUND":
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            r_dt = _parse_time_str(r)
            if not (c_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            r_ist = _to_ist(r_dt)

            requester_msgs = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _email_ist(e)
            ]
            requester_msgs.sort(key=lambda e: e.sent_time)
            if not requester_msgs:
                continue

            episodes = []
            current = [requester_msgs[0]]
            for e in requester_msgs[1:]:
                prev_t = _email_ist(current[-1])
                cur_t = _email_ist(e)
                if not prev_t or not cur_t:
                    continue
                if (cur_t - prev_t) > timedelta(hours=8):
                    episodes.append(current)
                    current = [e]
                else:
                    current.append(e)
            episodes.append(current)
            if not episodes:
                continue

            latest_episode = episodes[-1]
            latest_non_ack = [e for e in latest_episode if _is_real_reply_candidate(e)]
            if not latest_non_ack:
                continue
            latest_req = latest_non_ack[-1]
            latest_ist = _email_ist(latest_req)
            if not latest_ist:
                continue
            if latest_ist <= c_ist:
                continue
            if latest_ist < (r_ist - timedelta(days=2)):
                continue

            baseline_date = state.get("baseline_created_date")
            if baseline_date:
                if abs((latest_ist.date() - baseline_date).days) > 2:
                    continue
            else:
                # Keep conservative bound when baseline isn't available.
                if (latest_ist - c_ist) > timedelta(days=7):
                    continue

            t = _format_time(latest_req.sent_time)
            if not t:
                continue
            if latest_ist <= r_ist:
                continue
            row_vals["Actual Resolved Date & Time"] = t
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, resolved_col).value = t
            if list_index < len(debug_rows):
                who = latest_req.sender_email or latest_req.sender_name
                debug_rows[list_index]["ResolvedSource"] = who
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterSpanAckLikeGuard; ResolvedOnly"

        # Episode consistency guard (global):
        # For mixed-source/problem rows, pick one requester episode and derive all
        # three timestamps from that episode to avoid cross-episode mixing.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester or not subject_norm:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist <= a_ist <= r_ist):
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""

            mixed_sources = (
                created_src_now
                and ack_src_now
                and (_requester_key(created_src_now) == _requester_key(ack_src_now))
                and (not _match_requester(ack_src_now, ack_src_now, requester))
                and _match_requester(resolved_src_now, resolved_src_now, requester)
            )
            span_ack_missing = (
                "requester span" in (notes_now or "").lower()
                and str(ack_src_now).strip().upper() == "ACK NOT FOUND"
            )
            if not (mixed_sources or span_ack_missing):
                continue

            row_tokens = _match_tokens(subject_norm)
            if not row_tokens:
                continue

            requester_msgs = []
            for e in thread:
                if not e.sent_time or not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                mail_subj = normalize_subject(e.subject or "")
                mail_tokens = _match_tokens(mail_subj)
                sim = _token_overlap_score(row_tokens, mail_tokens) if mail_tokens else 0.0
                contains = (subject_norm in mail_subj or mail_subj in subject_norm) if mail_subj else False
                if sim >= 0.20 or contains:
                    requester_msgs.append(e)

            requester_msgs.sort(key=lambda e: e.sent_time)
            if not requester_msgs:
                continue

            episodes = []
            current = [requester_msgs[0]]
            for e in requester_msgs[1:]:
                prev_t = _email_ist(current[-1])
                cur_t = _email_ist(e)
                if not prev_t or not cur_t:
                    continue
                if (cur_t - prev_t) > timedelta(hours=8):
                    episodes.append(current)
                    current = [e]
                else:
                    current.append(e)
            episodes.append(current)
            if not episodes:
                continue

            baseline_date = state.get("baseline_created_date")
            chosen = None
            chosen_mode = "fallback"

            # For requester-span(ack-like) rows, always take the latest requester episode.
            # This prevents stale old episodes from winning via baseline.
            if span_ack_missing:
                chosen = episodes[-1]
                chosen_mode = "latest_span"

            # For mixed-source rows, prefer the episode nearest to resolved timestamp.
            if chosen is None and mixed_sources:
                near_ranked = []
                for ep in episodes:
                    ep_times = [(_email_ist(m), m) for m in ep if _email_ist(m)]
                    if not ep_times:
                        continue
                    min_abs = min(abs((t - r_ist).total_seconds()) for t, _ in ep_times)
                    before_flag = 0 if any(t <= r_ist for t, _ in ep_times) else 1
                    near_ranked.append((before_flag, min_abs, -_email_ist(ep[-1]).timestamp(), ep))
                if near_ranked:
                    near_ranked.sort()
                    chosen = near_ranked[0][3]
                    chosen_mode = "near_resolved"

            # Baseline-aware fallback for non-span/non-mixed rows.
            if chosen is None and baseline_date:
                def _episode_anchor(ep):
                    non_ack = [m for m in ep if not _ack_like(m)]
                    return non_ack[0] if non_ack else None
                ranked = []
                for ep in episodes:
                    anchor = _episode_anchor(ep)
                    if not anchor:
                        continue
                    anchor_t = _email_ist(anchor)
                    if not anchor_t:
                        continue
                    day_delta = abs((anchor_t.date() - baseline_date).days)
                    has_r = any(abs((_email_ist(m) - r_ist).total_seconds()) <= 21600 for m in ep if _email_ist(m))
                    ranked.append((day_delta, 0 if has_r else 1, -_email_ist(ep[-1]).timestamp(), ep))
                if ranked:
                    ranked.sort()
                    chosen = ranked[0][3]
                    chosen_mode = "baseline"

            if chosen is None:
                by_resolved = [
                    ep for ep in episodes
                    if any(abs((_email_ist(m) - r_ist).total_seconds()) <= 21600 for m in ep if _email_ist(m))
                ]
                chosen = by_resolved[-1] if by_resolved else episodes[-1]
                chosen_mode = "resolved_or_latest"

            if not chosen:
                continue

            chosen_non_ack = [m for m in chosen if not _ack_like(m)]
            if chosen_non_ack:
                ep_created = chosen_non_ack[0]
                ep_created_ist = _email_ist(ep_created)
                if not ep_created_ist:
                    continue

                ep_ack = ep_created
                for m in chosen:
                    m_ist = _email_ist(m)
                    if not m_ist or m_ist < ep_created_ist:
                        continue
                    if m_ist > ep_created_ist + timedelta(minutes=60):
                        break
                    if _ack_like(m):
                        ep_ack = m
                        break
                ep_ack_ist = _email_ist(ep_ack)
                if not ep_ack_ist:
                    continue

                non_ack_after_ack = [
                    m for m in chosen
                    if _email_ist(m)
                    and _email_ist(m) >= ep_ack_ist
                    and _is_real_reply_candidate(m)
                ]
                if not non_ack_after_ack:
                    continue
                ep_resolved = non_ack_after_ack[-1]
                ep_resolved_ist = _email_ist(ep_resolved)
                if not ep_resolved_ist:
                    continue

                if not (ep_created_ist <= ep_ack_ist <= ep_resolved_ist):
                    continue
            else:
                # All requester mails in this episode are ack-like. Keep this isolated
                # to requester-span rows with missing ACK source; collapse to latest
                # requester mail in the latest episode to avoid stale cross-episode picks.
                if not span_ack_missing:
                    continue
                ep_resolved = chosen[-1]
                ep_resolved_ist = _email_ist(ep_resolved)
                if not ep_resolved_ist:
                    continue
                ep_created = ep_resolved
                ep_ack = ep_resolved
                ep_created_ist = ep_resolved_ist
                ep_ack_ist = ep_resolved_ist
                chosen_mode = "latest_span_all_ack"

            # Safe bound: do not jump to a very distant day unexpectedly.
            # Do not apply this cap for latest-span rows.
            if baseline_date and chosen_mode not in ("latest_span", "latest_span_all_ack"):
                old_delta = abs((c_ist.date() - baseline_date).days)
                new_delta = abs((ep_created_ist.date() - baseline_date).days)
                if new_delta > 5 and new_delta > old_delta:
                    continue

            new_created = _format_time(ep_created.sent_time)
            new_ack = _format_time(ep_ack.sent_time)
            new_resolved = _format_time(ep_resolved.sent_time)
            if not new_created or not new_ack or not new_resolved:
                continue
            if new_created == c and new_ack == a and new_resolved == r:
                continue

            row_idx = state.get("row_index")
            _apply_guarded_episode_update(
                state,
                row_vals,
                list_index,
                row_idx,
                ep_created_ist,
                ep_ack_ist,
                ep_resolved_ist,
                owner_tag=f"EpisodeConsistencyGuard[{chosen_mode}]",
                candidate_kind="requester_ack",
                created_src=ep_created.sender_email or ep_created.sender_name,
                ack_src=ep_ack.sender_email or ep_ack.sender_name,
                resolved_src=ep_resolved.sender_email or ep_resolved.sender_name,
                note_suffix=f"EpisodeConsistencyGuard[{chosen_mode}]",
            )

        # Final requester-span(ack-like) fallback (global):
        # If a row still sits on an old timestamp with ACK NOT FOUND, move it to
        # latest requester mail in the same expanded thread.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester:
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            # Keep this guard generic but isolated to ESS-only span-like rows.
            if "ess-only; no non-ess request" not in notes_l:
                continue

            row_vals = automation_rows[list_index]
            c = row_vals.get("Created Date & Time")
            c_dt = _parse_time_str(c)
            c_ist = _to_ist(c_dt) if c_dt else None
            if not c_ist:
                continue
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            r_ist = _to_ist(r_dt) if r_dt else None

            row_tokens = _match_tokens(subject_norm)
            requester_msgs = [
                e for e in thread
                if e.sent_time
                and _req_match(e, requester)
                and _email_ist(e)
                and (
                    not row_tokens
                    or (
                        (lambda s_norm, s_tokens: (
                            (_token_overlap_score(row_tokens, s_tokens) >= 0.45)
                            or (subject_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        ))(
                            normalize_subject(e.subject or ""),
                            _match_tokens(normalize_subject(e.subject or "")),
                        )
                    )
                )
            ]
            if not requester_msgs:
                continue
            requester_non_ack = [e for e in requester_msgs if _is_real_reply_candidate(e)]
            requester_non_ack_exists = bool(requester_non_ack)
            used_ess_non_ack = False
            latest_req = max(requester_non_ack, key=lambda e: e.sent_time) if requester_non_ack else max(requester_msgs, key=lambda e: e.sent_time)
            latest_ist = _email_ist(latest_req)
            if not latest_ist:
                continue
            # Only move when there is a clear stale gap from either created/resolved.
            ref_ist = c_ist if not r_ist else max(c_ist, r_ist)
            if (latest_ist - ref_ist) < timedelta(hours=6):
                continue

            t = _format_time(latest_req.sent_time)
            if not t:
                continue
            if requester_non_ack_exists:
                if r_ist and latest_ist <= r_ist:
                    continue
                row_vals["Actual Resolved Date & Time"] = t
            else:
                # All requester mails are ack-like in this span row. Keep timestamps
                # consistent without promoting update-chasing requester mails.
                # Prefer latest ESS non-ack reply before that requester tail.
                ess_non_ack = []
                for e in thread:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if not _ess_sender(e):
                        continue
                    if not _is_real_reply_candidate(e):
                        continue
                    if e_ist < c_ist:
                        continue
                    if e_ist > latest_ist:
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(e.subject or "")
                        s_tokens = _match_tokens(s_norm)
                        if s_tokens:
                            score = _token_overlap_score(row_tokens, s_tokens)
                            contains = subject_norm and (subject_norm in s_norm or s_norm in subject_norm)
                            if score < 0.45 and not contains:
                                continue
                    ess_non_ack.append(e)
                ess_non_ack.sort(key=lambda e: e.sent_time)

                if ess_non_ack:
                    best = ess_non_ack[-1]
                    best_t = _format_time(best.sent_time)
                    if not best_t:
                        continue
                    if r_ist and _email_ist(best) and _email_ist(best) <= r_ist:
                        continue
                    row_vals["Actual Resolved Date & Time"] = best_t
                    t = best_t
                    latest_req = best
                    used_ess_non_ack = True
                else:
                    a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                    a_ist = _to_ist(a_dt) if a_dt else None
                    if c_ist == latest_ist and a_ist == latest_ist and r_ist == latest_ist:
                        continue
                    row_vals["Created Date & Time"] = t
                    row_vals["Actual Response Date & Time"] = t
                    row_vals["Actual Resolved Date & Time"] = t
            row_idx = state.get("row_index")
            if row_idx:
                if requester_non_ack_exists or used_ess_non_ack:
                    ws.cell(row_idx, resolved_col).value = t
                else:
                    ws.cell(row_idx, created_col).value = t
                    ws.cell(row_idx, response_col).value = t
                    ws.cell(row_idx, resolved_col).value = t
            if list_index < len(debug_rows):
                who = latest_req.sender_email or latest_req.sender_name
                if used_ess_non_ack:
                    debug_rows[list_index]["ResolvedSource"] = who
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterSpanAllAckGuard; ResolvedFromEssNonAck"
                elif requester_non_ack_exists:
                    debug_rows[list_index]["ResolvedSource"] = who
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterSpanAckLikeFinalGuard; ResolvedOnly"
                else:
                    debug_rows[list_index]["CreatedSource"] = who
                    debug_rows[list_index]["AckSource"] = who
                    debug_rows[list_index]["ResolvedSource"] = who
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterSpanAllAckGuard"

        # Final resolved non-ack guard (global):
        # If resolved currently points to an ack-like/update-like requester reply
        # (e.g., "thank you for the update/information"), move resolved to the latest
        # prior non-ack requester reply in the same subject episode.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if list_index < len(debug_rows):
                notes_now = (debug_rows[list_index].get("Notes", "") or "").lower()
                if "requester follow-up (no in-between request)" in notes_now:
                    continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester:
                continue

            row_vals = automation_rows[list_index]
            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt) if a_dt else None
            r_ist = _to_ist(r_dt)

            row_tokens = _match_tokens(subject_norm)
            requester_msgs = []
            for e in thread:
                if not e.sent_time or not _req_match(e, requester):
                    continue
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if row_tokens:
                    s_norm = normalize_subject(e.subject or "")
                    s_tokens = _match_tokens(s_norm)
                    if s_tokens:
                        score = _token_overlap_score(row_tokens, s_tokens)
                        contains = subject_norm and (subject_norm in s_norm or s_norm in subject_norm)
                        if score < 0.45 and not contains:
                            continue
                requester_msgs.append(e)

            if not requester_msgs:
                continue

            # Find the message that likely produced current resolved timestamp.
            resolved_hits = []
            for e in requester_msgs:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if abs((e_ist - r_ist).total_seconds()) <= 180:
                    resolved_hits.append(e)
            if resolved_hits:
                resolved_mail = max(resolved_hits, key=lambda e: e.sent_time)
            else:
                # Minute-level sheet timestamps can miss seconds; use nearest requester
                # mail at/before resolved (soft +5m tolerance).
                prior_msgs = [
                    e for e in requester_msgs
                    if _email_ist(e) and _email_ist(e) <= (r_ist + timedelta(minutes=5))
                ]
                if not prior_msgs:
                    continue
                resolved_mail = max(prior_msgs, key=lambda e: e.sent_time)
            if _is_real_reply_candidate(resolved_mail):
                continue

            fallback = [
                e for e in requester_msgs
                if _is_real_reply_candidate(e)
                and _email_ist(e)
                and _email_ist(e) <= (r_ist + timedelta(minutes=5))
            ]
            pick = max(fallback, key=lambda e: e.sent_time) if fallback else None
            pick_ist = _email_ist(pick) if pick else None
            pick_src = (pick.sender_email or pick.sender_name) if pick else ""
            used_quoted = False
            if not pick_ist:
                quoted_ist = _extract_quoted_requester_reply_ist(
                    resolved_mail,
                    requester,
                    subject_norm,
                    c_ist,
                    r_ist,
                )
                if not quoted_ist:
                    continue
                pick_ist = quoted_ist
                pick_src = "PARSED_FROM_QUOTED_REPLY"
                used_quoted = True
            # Safety: do not jump too far back in time on ambiguous long chains.
            if (r_ist - pick_ist) > timedelta(days=14):
                continue

            t = _format_time(pick_ist)
            if not t:
                continue
            reanchored = False
            if a_ist and pick_ist < a_ist:
                row_vals["Actual Response Date & Time"] = t
                if c_ist > pick_ist:
                    row_vals["Created Date & Time"] = t
                reanchored = True
            row_vals["Actual Resolved Date & Time"] = t
            row_idx = state.get("row_index")
            if row_idx:
                if reanchored:
                    ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                    ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                ws.cell(row_idx, resolved_col).value = t
            if list_index < len(debug_rows):
                if reanchored:
                    debug_rows[list_index]["CreatedSource"] = pick_src
                    debug_rows[list_index]["AckSource"] = pick_src
                debug_rows[list_index]["ResolvedSource"] = pick_src
                extra = "; ResolvedAckLikeQuotedGuard" if used_quoted else ""
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ResolvedAckLikeGuard{extra}{'; ResolvedAckLikeReanchor' if reanchored else ''}"

        # Quoted request-anchor guard (isolated):
        # If created was retained due unreliable response anchor and response exists,
        # allow a strictly-bounded quoted request timestamp from the same chain to
        # re-anchor Created only.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if list_index < len(debug_rows):
                notes_now = (debug_rows[list_index].get("Notes", "") or "").lower()
                if "requester follow-up (no in-between request)" in notes_now:
                    continue

            base_thread = state.get("thread") or []
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            thread = _expanded_thread(subject_norm, base_thread, requester)
            if not thread or not requester:
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            row_vals = automation_rows[list_index]
            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt) if r_dt else None
            if a_ist < c_ist:
                continue
            row_tokens = _match_tokens(subject_norm)

            source_mismatch = (
                bool(created_src_now)
                and bool(requester)
                and not _match_requester(created_src_now, created_src_now, requester)
            )
            ack_from_requester = bool(requester) and bool(ack_src_now) and _match_requester(ack_src_now, ack_src_now, requester)
            span_missing_non_ess = "ess-only; no non-ess request" in notes_l
            large_created_gap = (a_ist - c_ist) >= timedelta(hours=8)
            needs_quoted_anchor = (
                "created retained (response anchor unreliable)" in notes_l
                or "created_clamped_to_first" in (created_src_now.lower() if isinstance(created_src_now, str) else "")
                or (span_missing_non_ess and source_mismatch and large_created_gap)
                or (source_mismatch and ack_from_requester and large_created_gap)
            )
            if (
                not needs_quoted_anchor
            ):
                continue

            live_reply_ist = _live_reply_lane_for_row(list_index)
            live_reply_msg = _live_reply_message_for_row(list_index)
            row_id_tokens = _id_like_tokens(subject_norm)
            if not row_id_tokens:
                desc_text = row_vals.get("Description") or state.get("description") or ""
                row_id_tokens = _id_like_tokens(desc_text)
            if live_reply_ist and live_reply_msg:
                lane_thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=live_reply_ist,
                ) or thread
                lane_episode = _lane_local_episode_from_reply(
                    live_reply_msg,
                    live_reply_ist,
                    lane_thread,
                    requester,
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                )
                if lane_episode and lane_episode.get("request"):
                    lane_req_ist = lane_episode.get("request")
                    lane_ack_ist = lane_episode.get("ack") or live_reply_ist
                    lane_ack_msg = lane_episode.get("ack_msg")
                    direct_reply_mode = lane_ack_msg is None or lane_ack_ist == live_reply_ist
                    if (
                        lane_req_ist
                        and lane_req_ist > c_ist
                        and lane_req_ist < live_reply_ist
                        and (live_reply_ist - lane_req_ist) <= timedelta(hours=48)
                        and lane_req_ist <= lane_ack_ist <= live_reply_ist
                    ):
                        applied = _apply_guarded_episode_update(
                            state,
                            row_vals,
                            list_index,
                            state.get("row_index"),
                            lane_req_ist,
                            lane_ack_ist,
                            live_reply_ist,
                            owner_tag="QuotedRequestAnchorGuard",
                            candidate_kind="quoted" if direct_reply_mode else "hybrid",
                            created_src="PARSED_FROM_QUOTED_REQUEST",
                            ack_src=(
                                live_reply_msg.sender_email or live_reply_msg.sender_name or "PARSED_FROM_QUOTED_REPLY"
                                if direct_reply_mode else
                                (lane_ack_msg.sender_email or lane_ack_msg.sender_name or "ESS_LANE_LOCAL_ACK")
                            ),
                            resolved_src=live_reply_msg.sender_email or live_reply_msg.sender_name or "LANE_LOCAL_REPLY",
                            note_suffix="QuotedRequestAnchorGuard",
                        )
                        if applied:
                            continue

            quoted_candidates = []
            for e in thread:
                if not e.sent_time:
                    continue
                if not _req_match(e, requester):
                    continue
                q_ist = _extract_quoted_request_before_ist(e, subject_norm, a_ist)
                if not q_ist:
                    continue
                quoted_candidates.append(q_ist)

            if not quoted_candidates:
                continue

            q_pick = max(quoted_candidates)
            # Safety bounds: keep within same practical episode.
            if q_pick <= c_ist:
                continue
            if q_pick >= a_ist:
                continue
            if (a_ist - q_pick) > timedelta(hours=24):
                continue

            t = _format_time(q_pick)
            if not t:
                continue
            direct_reply_candidates = []
            for e in thread:
                if not e.sent_time:
                    continue
                if not _req_match(e, requester):
                    continue
                if not _is_real_reply_candidate(e):
                    continue
                e_ist = _email_ist(e)
                if not e_ist or e_ist <= q_pick:
                    continue
                if (e_ist - q_pick) > timedelta(hours=48):
                    continue
                if row_tokens:
                    s_norm = normalize_subject(e.subject or "")
                    s_tokens = _match_tokens(s_norm)
                    if s_tokens:
                        score = _token_overlap_score(row_tokens, s_tokens)
                        contains = subject_norm and (subject_norm in s_norm or s_norm in subject_norm)
                        if score < 0.45 and not contains:
                            continue
                direct_reply_candidates.append((e_ist, e))
            direct_reply_candidates.sort(key=lambda item: item[0])

            sync_response_reply = None
            if direct_reply_candidates:
                if r_ist:
                    resolved_near = [
                        item for item in direct_reply_candidates
                        if abs((item[0] - r_ist).total_seconds()) <= 300
                    ]
                    if resolved_near:
                        sync_response_reply = resolved_near[0]
                if sync_response_reply is None and (
                    (r_ist and a_ist > (r_ist + timedelta(minutes=5)))
                    or ((a_ist - q_pick) > timedelta(hours=24))
                ):
                    sync_response_reply = direct_reply_candidates[0]

            row_vals["Created Date & Time"] = t
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = t
            if sync_response_reply:
                reply_ist, reply_msg = sync_response_reply
                reply_t = _format_time(reply_ist)
                if reply_t:
                    row_vals["Actual Response Date & Time"] = reply_t
                    if row_idx:
                        ws.cell(row_idx, response_col).value = reply_t
                    if (not r_ist) or (r_ist < reply_ist) or (a_ist > (r_ist + timedelta(minutes=5))):
                        row_vals["Actual Resolved Date & Time"] = reply_t
                        if row_idx:
                            ws.cell(row_idx, resolved_col).value = reply_t
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REQUEST"
                if sync_response_reply:
                    debug_rows[list_index]["AckSource"] = sync_response_reply[1].sender_email or sync_response_reply[1].sender_name or "PARSED_FROM_QUOTED_REPLY"
                    if (not r_ist) or (r_ist < sync_response_reply[0]) or (a_ist > ((r_ist or sync_response_reply[0]) + timedelta(minutes=5))):
                        debug_rows[list_index]["ResolvedSource"] = sync_response_reply[1].sender_email or sync_response_reply[1].sender_name or "PARSED_FROM_QUOTED_REPLY"
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedRequestAnchorGuard; QuotedDirectReplySync"
                else:
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedRequestAnchorGuard"

        if True:
            # Live request anchor guard (global):
            # If a row drifted into ESS-only/parsed anchoring, but a live non-ESS
            # requester mail exists in the same subject episode before ack/resolved,
            # re-anchor Created to that live request.
            live_request_states = nonblue_row_states
            for state in live_request_states:
                list_index = state.get("list_index")
                if list_index is None or list_index >= len(automation_rows):
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                requester = state.get("requester") or ""
                subject_norm = (state.get("subject_norm") or "").lower()
                if not subject_norm:
                    continue

                notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
                notes_l = (notes_now or "").lower()
                created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
                ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
                resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
                baseline_date = state.get("baseline_created_date")
                baseline_mid = None
                if baseline_date:
                    try:
                        baseline_mid = _to_ist(datetime(baseline_date.year, baseline_date.month, baseline_date.day))
                    except Exception:
                        baseline_mid = None
                no_requester_recovery = (
                    "norequesterthreadrecovery" in notes_l
                    or "no ess or requester replies" in notes_l
                )
                parsed_created = isinstance(created_src_now, str) and created_src_now.startswith("PARSED_FROM_")
                trigger = (
                    parsed_created
                    or "dateanchorafter" in notes_l
                    or "dateanchormissing" in notes_l
                    or no_requester_recovery
                    or "ambiguousresolvedbyrequester" in notes_l
                )
                if not trigger:
                    continue

                c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                c_ist = _to_ist(c_dt) if c_dt else None
                a_ist = _to_ist(a_dt) if a_dt else None
                r_ist = _to_ist(r_dt) if r_dt else None
                baseline_drift = 0
                if baseline_date and c_ist:
                    try:
                        baseline_drift = abs((c_ist.date() - baseline_date).days)
                    except Exception:
                        baseline_drift = 0
                mixed_owner_drift = False
                if requester and c_ist and a_ist and (a_ist - c_ist) >= timedelta(hours=2):
                    try:
                        mixed_owner_drift = (
                            bool(ack_src_now)
                            and bool(resolved_src_now)
                            and (not _match_requester(str(ack_src_now), str(ack_src_now), requester))
                            and _match_requester(str(resolved_src_now), str(resolved_src_now), requester)
                        )
                    except Exception:
                        mixed_owner_drift = False
                # Keep mixed-owner recovery narrow: it should only run when drift is
                # clearly large, otherwise it can pull from a wrong late episode.
                if mixed_owner_drift and (baseline_drift < 3) and (not parsed_created):
                    mixed_owner_drift = False
                # Keep this guard isolated: do not rewrite stable rows.
                if (
                    not parsed_created
                    and not no_requester_recovery
                    and not mixed_owner_drift
                    and baseline_drift <= 2
                ):
                    continue
                # Recovery mode keeps existing early-window behavior for the original
                # no-requester case only. Mixed-owner has its own tighter path below.
                recovery_mode = no_requester_recovery
                upper_ist = a_ist or r_ist
                if not c_ist:
                    continue
                if not upper_ist:
                    upper_ist = c_ist + timedelta(days=7)
                if recovery_mode:
                    upper_ist = max(upper_ist, c_ist + timedelta(days=7))
                if (upper_ist - c_ist) < timedelta(hours=18) and not recovery_mode:
                    continue

                base_thread = state.get("thread") or []
                thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=upper_ist,
                )
                row_tokens = _match_tokens(subject_norm)
                candidates = []
                for e in (thread or []):
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist > (upper_ist + timedelta(minutes=5)):
                        continue
                    if e_ist < (upper_ist - timedelta(days=14)):
                        continue
                    if baseline_date and abs((e_ist.date() - baseline_date).days) > 5:
                        continue
                    if e_ist < (c_ist - timedelta(hours=2)):
                        continue
                    if _ess_sender(e):
                        continue
                    if _system_like_sender(e):
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(e.subject or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    candidates.append(e)

                pick = None
                pick_ist = None
                pick_src = ""
                allow_global_fallback = not mixed_owner_drift
                def _baseline_rank_mail(e):
                    e_ist = _email_ist(e)
                    if not e_ist or not baseline_date:
                        return (9999, 999999999)
                    day_delta = abs((e_ist.date() - baseline_date).days)
                    prox = abs((e_ist - (baseline_mid or e_ist)).total_seconds())
                    return (day_delta, prox)

                def _baseline_rank_dt(dt_ist):
                    if not dt_ist or not baseline_date:
                        return (9999, 999999999)
                    day_delta = abs((dt_ist.date() - baseline_date).days)
                    prox = abs((dt_ist - (baseline_mid or dt_ist)).total_seconds())
                    return (day_delta, prox)

                if candidates:
                    candidates.sort(key=lambda e: e.sent_time)
                    if baseline_date:
                        pick = min(candidates, key=lambda e: (_baseline_rank_mail(e), -_email_ist(e).timestamp()))
                    elif mixed_owner_drift:
                        local_end = c_ist + timedelta(hours=72)
                        local_candidates = [
                            e for e in candidates
                            if _email_ist(e) and _email_ist(e) >= (c_ist - timedelta(minutes=15)) and _email_ist(e) <= local_end
                        ]
                        if local_candidates:
                            pick = min(local_candidates, key=lambda e: abs((_email_ist(e) - c_ist).total_seconds()))
                    elif recovery_mode:
                        early_window_end = c_ist + timedelta(hours=72)
                        early_candidates = [
                            e for e in candidates
                            if _email_ist(e) and _email_ist(e) >= c_ist and _email_ist(e) <= early_window_end
                        ]
                        pick = early_candidates[0] if early_candidates else candidates[-1]
                    else:
                        pick = candidates[-1]
                    pick_ist = _email_ist(pick)
                    pick_src = pick.sender_email or pick.sender_name

                # Fallback-1: use requester pool across all emails for this subject episode.
                if (not pick_ist) and allow_global_fallback:
                    requester_pool = _requester_pool(subject_norm, requester, upper_ist, day_window=21)
                    fallback_pool = []
                    for e in requester_pool:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if e_ist > (upper_ist + timedelta(minutes=5)):
                            continue
                        if e_ist < (upper_ist - timedelta(days=14)):
                            continue
                        if baseline_date and abs((e_ist.date() - baseline_date).days) > 5:
                            continue
                        if _ack_like(e) or _ack_like_text_fallback(e):
                            continue
                        fallback_pool.append(e)
                    if fallback_pool:
                        if baseline_date:
                            pick = min(fallback_pool, key=lambda e: (_baseline_rank_mail(e), -_email_ist(e).timestamp()))
                        else:
                            pick = max(fallback_pool, key=lambda e: e.sent_time)
                        pick_ist = _email_ist(pick)
                        pick_src = pick.sender_email or pick.sender_name

                # Fallback-2: parse quoted non-ESS request from requester emails.
                if (not pick_ist) and allow_global_fallback:
                    parsed_candidates = []
                    for e in (thread or []):
                        if not getattr(e, "sent_time", None):
                            continue
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if e_ist > (upper_ist + timedelta(minutes=5)):
                            continue
                        if not _req_match(e, requester):
                            continue
                        try:
                            parsed_dt = _extract_request_time_from_email(
                                e,
                                ess_team,
                                max_dt=e.sent_time,
                                subject_norm=subject_norm,
                            )
                        except Exception:
                            parsed_dt = None
                        if not parsed_dt:
                            continue
                        parsed_ist = _to_ist(parsed_dt)
                        if parsed_ist > (upper_ist + timedelta(minutes=5)):
                            continue
                        if parsed_ist < (upper_ist - timedelta(days=14)):
                            continue
                        if baseline_date and abs((parsed_ist.date() - baseline_date).days) > 5:
                            continue
                        parsed_candidates.append(parsed_ist)
                    if parsed_candidates:
                        if baseline_date:
                            pick_ist = min(parsed_candidates, key=lambda dt: (_baseline_rank_dt(dt), -dt.timestamp()))
                        else:
                            parsed_candidates.sort()
                            pick_ist = parsed_candidates[-1]
                        pick_src = "PARSED_FROM_QUOTED_REQUEST"

                if not pick_ist:
                    continue
                if mixed_owner_drift:
                    if pick_ist < (c_ist - timedelta(minutes=15)) or pick_ist > (c_ist + timedelta(hours=72)):
                        continue
                if baseline_date:
                    try:
                        current_delta = abs((c_ist.date() - baseline_date).days)
                        pick_delta = abs((pick_ist.date() - baseline_date).days)
                    except Exception:
                        current_delta = 9999
                        pick_delta = 9999
                    # Do not worsen day drift when a baseline is known.
                    if pick_delta > current_delta and pick_delta > 2:
                        continue
                if abs((pick_ist - c_ist).total_seconds()) <= 120:
                    continue
                if pick_ist <= (c_ist + timedelta(minutes=15)):
                    continue

                t = _format_time(pick.sent_time) if pick else _format_time(pick_ist)
                if not t:
                    continue
                row_vals["Created Date & Time"] = t
                if a_ist and a_ist < pick_ist:
                    row_vals["Actual Response Date & Time"] = t
                if r_ist and r_ist < pick_ist:
                    row_vals["Actual Resolved Date & Time"] = t

                row_idx = state.get("row_index")
                if row_idx:
                    ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                    ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                    ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
                if list_index < len(debug_rows):
                    who = pick_src
                    debug_rows[list_index]["CreatedSource"] = who
                    if a_ist and a_ist < pick_ist:
                        debug_rows[list_index]["AckSource"] = who
                    if r_ist and r_ist < pick_ist:
                        debug_rows[list_index]["ResolvedSource"] = who
                    if no_requester_recovery:
                        suffix = "; LiveRequestAnchorGuardNoRequesterRecovery"
                    elif mixed_owner_drift:
                        suffix = "; LiveRequestAnchorGuardMixedOwner"
                    else:
                        suffix = "; LiveRequestAnchorGuard"
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}{suffix}"

        if True:
            # ESS continuation guard (global):
            # If the row is ESS-initiated/span-like and the same consultant keeps
            # replying with no external request in-between, align to consultant tail.
            used_ess_continuation = set()
            for state in nonblue_row_states:
                list_index = state.get("list_index")
                if list_index is None or list_index >= len(automation_rows):
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                requester = state.get("requester") or ""
                subject_norm = (state.get("subject_norm") or "").lower()
                if not requester or not subject_norm:
                    continue

                notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
                notes_l = (notes_now or "").lower()
                created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
                ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
                resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
                mixed_source_pattern = (
                    isinstance(created_src_now, str)
                    and created_src_now.startswith("PARSED_FROM_")
                    and bool(ack_src_now)
                    and bool(requester)
                    and not _match_requester(ack_src_now, ack_src_now, requester)
                    and _match_requester(resolved_src_now, resolved_src_now, requester)
                )
                if (
                    "ess-only; no non-ess request" not in notes_l
                    and "requester follow-up (no in-between request)" not in notes_l
                    and not mixed_source_pattern
                ):
                    continue

                c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                if not (c_dt and a_dt and r_dt):
                    continue
                c_ist = _to_ist(c_dt)
                a_ist = _to_ist(a_dt)
                r_ist = _to_ist(r_dt)
                # For ESS-only continuation rows, always allow collapse to a single
                # requester reply (unique per subject/requester), even if times
                # currently differ.

                base_thread = state.get("thread") or []
                thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=r_ist,
                )
                requester_pool = _requester_pool(subject_norm, requester, r_ist, day_window=21)
                merged_msgs = []
                for e in (thread or []):
                    merged_msgs.append(e)
                if not merged_msgs:
                    continue
                dedup = {}
                for e in merged_msgs:
                    dedup[(getattr(e, "subject", ""), getattr(e, "sender_email", ""), getattr(e, "sender_name", ""), getattr(e, "sent_time", None))] = e
                merged_msgs = list(dedup.values())
                merged_msgs.sort(key=lambda e: e.sent_time if getattr(e, "sent_time", None) else datetime.max)

                consultant_msgs = []
                for e in merged_msgs:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if not _req_match(e, requester):
                        continue
                    if not _is_real_reply_candidate(e):
                        continue
                    consultant_msgs.append(e)
                consultant_msgs.sort(key=lambda e: e.sent_time)
                if len(consultant_msgs) < (1 if mixed_source_pattern else 2):
                    continue

                first_ist = _email_ist(consultant_msgs[0])
                latest = None
                latest_ist = None
                for cand in consultant_msgs:
                    cand_ist = _email_ist(cand)
                    if not cand_ist:
                        continue
                    key = (
                        subject_norm,
                        requester.strip().lower(),
                        state.get("service_bucket") or "",
                        cand_ist.replace(second=0, microsecond=0),
                    )
                    if key in used_ess_continuation:
                        continue
                    latest = cand
                    latest_ist = cand_ist
                    used_ess_continuation.add(key)
                    break
                if not first_ist or not latest_ist:
                    continue

                non_ess_between = False
                row_tokens = _match_tokens(subject_norm)
                for e in emails:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist <= first_ist or e_ist > latest_ist:
                        continue
                    if row_tokens:
                        s_norm = normalize_subject(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    if _ess_sender(e):
                        continue
                    if _system_like_sender(e):
                        continue
                    if _req_match(e, requester):
                        continue
                    non_ess_between = True
                    break
                if non_ess_between:
                    continue

                # Strict ESS-only collapse: always use latest requester reply
                # when there is no non-ESS request in-between.
                if "ess-only; no non-ess request" in notes_l and latest_ist and latest_ist >= c_ist:
                    key = (
                        subject_norm,
                        requester.strip().lower(),
                        state.get("service_bucket") or "",
                        latest_ist.replace(second=0, microsecond=0),
                    )
                    if key in used_ess_continuation:
                        if list_index < len(debug_rows):
                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationDupSkipped"
                        continue
                    t = _format_time(latest.sent_time)
                    if not t:
                        continue
                    if not _allow_guard_rewrite(
                        row_vals,
                        list_index,
                        latest_ist,
                        latest_ist,
                        latest_ist,
                        "ESSContinuationGuard[AllThreeStrict]",
                        "continuation",
                    ):
                        continue
                    used_ess_continuation.add(key)
                    row_vals["Created Date & Time"] = t
                    row_vals["Actual Response Date & Time"] = t
                    row_vals["Actual Resolved Date & Time"] = t
                    row_idx = state.get("row_index")
                    if row_idx:
                        ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                        ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                        ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
                    if list_index < len(debug_rows):
                        who = latest.sender_email or latest.sender_name
                        debug_rows[list_index]["CreatedSource"] = who
                        debug_rows[list_index]["AckSource"] = who
                        debug_rows[list_index]["ResolvedSource"] = who
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationGuard[AllThreeStrict]"
                    continue

                if (latest_ist - c_ist) > timedelta(days=5):
                    continue

                t = _format_time(latest.sent_time)
                if not t:
                    continue

                if c_ist == a_ist and a_ist == r_ist:
                    if not _allow_guard_rewrite(
                        row_vals,
                        list_index,
                        latest_ist,
                        latest_ist,
                        latest_ist,
                        "ESSContinuationGuard[AllThree]",
                        "continuation",
                    ):
                        continue
                    row_vals["Created Date & Time"] = t
                    row_vals["Actual Response Date & Time"] = t
                    row_vals["Actual Resolved Date & Time"] = t
                    mode = "AllThree"
                else:
                    if latest_ist <= (r_ist + timedelta(minutes=3)):
                        continue
                    if latest_ist >= c_ist:
                        if not _allow_guard_rewrite(
                            row_vals,
                            list_index,
                            c_ist,
                            latest_ist,
                            latest_ist,
                            "ESSContinuationGuard[ResponseResolved]",
                            "continuation",
                        ):
                            continue
                        row_vals["Actual Response Date & Time"] = t
                        row_vals["Actual Resolved Date & Time"] = t
                        mode = "ResponseResolved"
                    else:
                        continue

                row_idx = state.get("row_index")
                if row_idx:
                    ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                    ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                    ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
                if list_index < len(debug_rows):
                    who = latest.sender_email or latest.sender_name
                    if mode == "AllThree":
                        debug_rows[list_index]["CreatedSource"] = who
                        debug_rows[list_index]["AckSource"] = who
                    else:
                        debug_rows[list_index]["AckSource"] = who
                    debug_rows[list_index]["ResolvedSource"] = who
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationGuard[{mode}]"

        # Consultant-window environment guard:
        # Derive environment from consultant replies only, bounded by the selected
        # Created..Resolved timeframe, to avoid cross-request contamination in
        # long mixed threads.
        if env_col:
            subj_explicit_env_re = re.compile(
                r"\b(prod|prd|production|fcp|bip|uat|fct|biu|qa|fcq|biq|dev|development|fcd|bid)\b",
                flags=re.IGNORECASE,
            )
            for state in nonblue_row_states:
                list_index = state.get("list_index")
                if list_index is None or list_index >= len(automation_rows):
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                requester = state.get("requester") or ""
                if not requester:
                    continue

                description = state.get("description") or ""
                subject_text = _subject_for_description(description)
                # Keep subject-explicit env untouched.
                if subject_text and subj_explicit_env_re.search(subject_text):
                    continue

                c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                if not (c_dt and r_dt):
                    continue
                c_ist = _to_ist(c_dt)
                r_ist = _to_ist(r_dt)
                if r_ist < c_ist:
                    continue

                base_thread = state.get("thread") or []
                subject_norm = (state.get("subject_norm") or "").lower()
                thread = _expanded_thread(subject_norm, base_thread, requester)
                if not thread:
                    continue

                window_start = c_ist - timedelta(minutes=5)
                window_end = r_ist + timedelta(minutes=5)
                consultant_parts = []
                for e in thread:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist < window_start or e_ist > window_end:
                        continue
                    if not _req_match(e, requester):
                        continue
                    if getattr(e, "body", None):
                        consultant_parts.append(e.body)
                    if getattr(e, "body_html", None):
                        consultant_parts.append(e.body_html)
                if not consultant_parts:
                    continue

                env_window = resolve_environment(subject_text, "\n".join(consultant_parts))
                if not env_window:
                    continue
                env_current = row_vals.get("Environment") or ""
                if env_window == env_current:
                    continue

                row_vals["Environment"] = env_window
                row_idx = state.get("row_index")
                if row_idx:
                    ws.cell(row_idx, env_col).value = env_window
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; EnvConsultantWindowGuard"

        # Continuation all-three lock:
        # For ESS-only continuation/requester-span rows, if created is already chosen
        # correctly but response/resolved were later moved by post-ack passes, keep all
        # three equal to created.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if "ess-only; no non-ess request" not in notes_l:
                continue
            if not (
                "requester span" in notes_l
                or "requester follow-up (no in-between request)" in notes_l
                or "requester follow-up(top-only)" in notes_l
                or "ess-only continuation(top requester)" in notes_l
                or "esscontinuationguard[allthree]" in notes_l
            ):
                continue
            if not (
                "resolvedafterackpost" in notes_l
                or "resolvedafterackrelated" in notes_l
                or "resolvedafterack" in notes_l
                or "resolvedwithin48hafterack" in notes_l
            ):
                continue

            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not c_dt:
                continue
            if a_dt and r_dt and _to_ist(a_dt) == _to_ist(c_dt) and _to_ist(r_dt) == _to_ist(c_dt):
                continue

            # Allow same-consultant and teammate continuation tails:
            # in ESS-only continuation rows, latest actionable reply can be by
            # requester or ESS teammate. Keep this generic and do not require
            # requester-only source matching here.
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            res_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            # Lock only when created is requester-owned; otherwise this can clobber
            # legitimate request->ack->resolve sequences.
            if requester and created_src_now:
                try:
                    if not _match_requester(str(created_src_now), str(created_src_now), requester):
                        continue
                except Exception:
                    continue
            elif requester:
                continue

            # Extra safety: do not force all-three if a non-ESS request exists in the
            # active episode or if requester has actionable (non-ack) follow-ups.
            subject_norm = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=_to_ist(r_dt or a_dt or c_dt),
            )
            if thread:
                upper_ist = _to_ist(r_dt or a_dt or c_dt)
                lower_ist = _to_ist(c_dt) - timedelta(minutes=2)
                has_external_between = False
                requester_actionable_after_created = False
                for e in thread:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist < lower_ist or e_ist > (upper_ist + timedelta(minutes=5)):
                        continue
                    if (not _ess_sender(e)) and (not _system_like_sender(e)):
                        has_external_between = True
                        break
                    if _req_match(e, requester) and e_ist > (_to_ist(c_dt) + timedelta(minutes=2)):
                        if (not _ack_like(e)) and (not _ack_like_text_fallback(e)):
                            requester_actionable_after_created = True
                            break
                if has_external_between or requester_actionable_after_created:
                    continue

            row_vals["Actual Response Date & Time"] = c
            row_vals["Actual Resolved Date & Time"] = c

            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, response_col).value = c
                ws.cell(row_idx, resolved_col).value = c

            if list_index < len(debug_rows):
                src = debug_rows[list_index].get("CreatedSource", "") or ack_src_now or res_src_now
                if src:
                    debug_rows[list_index]["AckSource"] = src
                    debug_rows[list_index]["ResolvedSource"] = src
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ContinuationAllThreeFinalLock"

        # Parsed-gap reanchor guard:
        # If Created came from parsed quote and response/resolved are much later,
        # re-anchor Created to the latest real non-ESS request before response.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            if not subject_norm:
                continue

            created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if not (isinstance(created_src_now, str) and created_src_now.startswith("PARSED_")):
                continue
            if not (
                "monotonicguard" in notes_l
                or "resolvedafterack" in notes_l
                or "resolvedwithin48hafterack" in notes_l
            ):
                continue

            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            cutoff_dt = a_dt or r_dt
            if not (c_dt and cutoff_dt):
                continue
            c_ist = _to_ist(c_dt)
            cutoff_ist = _to_ist(cutoff_dt)
            if cutoff_ist <= c_ist:
                continue
            old_gap = cutoff_ist - c_ist
            if old_gap < timedelta(hours=12):
                continue

            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=cutoff_ist,
            )
            if not thread:
                continue

            row_tokens = _match_tokens(subject_norm)
            candidates = []
            for e in thread:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist > (cutoff_ist + timedelta(minutes=3)):
                    continue
                if _ess_sender(e) or _system_like_sender(e):
                    continue
                if row_tokens:
                    s_norm = normalize_subject(getattr(e, "subject", "") or "")
                    s_tokens = _match_tokens(s_norm)
                    score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                    contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                    if score < 0.45 and not contains:
                        continue
                candidates.append(e)

            if not candidates:
                continue
            pick = max(candidates, key=lambda e: e.sent_time)
            pick_ist = _email_ist(pick)
            if not pick_ist:
                continue
            if pick_ist <= (c_ist + timedelta(minutes=15)):
                continue
            new_gap = cutoff_ist - pick_ist
            if new_gap < timedelta(0) or new_gap > timedelta(days=7):
                continue
            if new_gap >= (old_gap - timedelta(minutes=5)):
                continue

            t = _format_time(pick.sent_time)
            if not t:
                continue
            row_vals["Created Date & Time"] = t
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = t
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = pick.sender_email or pick.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ParsedGapReanchorGuard"

        # Baseline stale-created guard (safe + narrow):
        # When Created drifts >1 day away from ServiceNow baseline on risky/ambiguous
        # rows, re-anchor Created to the nearest valid request episode around baseline.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            baseline_date = state.get("baseline_created_date")
            if not baseline_date:
                continue

            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            if not requester or not subject_norm:
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            risky = (
                "norequesterthreadrecovery" in notes_l
                or "no ess or requester replies" in notes_l
                or "ambiguousresolvedbyrequester" in notes_l
                or "score:" in notes_l
                or "altunion:" in notes_l
                or "dateanchormissing" in notes_l
                or "dateanchorafter" in notes_l
            )
            if not risky:
                continue

            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            if not c_dt:
                continue
            c_ist = _to_ist(c_dt)
            if abs((c_ist.date() - baseline_date).days) <= 1:
                continue

            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            a_ist = _to_ist(a_dt) if a_dt else None
            r_ist = _to_ist(r_dt) if r_dt else None

            baseline_mid = _to_ist(datetime(baseline_date.year, baseline_date.month, baseline_date.day))
            window_start = baseline_mid - timedelta(hours=18)
            window_end = baseline_mid + timedelta(hours=72)

            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=baseline_mid,
            )
            if not thread:
                continue

            row_tokens = _match_tokens(subject_norm)
            req_candidates = []
            for e in thread:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist < window_start or e_ist > window_end:
                    continue
                if _ess_sender(e) or _system_like_sender(e):
                    continue
                if row_tokens:
                    s_norm = normalize_subject(getattr(e, "subject", "") or "")
                    s_tokens = _match_tokens(s_norm)
                    score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                    contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                    if score < 0.45 and not contains:
                        continue
                req_candidates.append(e)

            pick_ist = None
            pick_src = ""
            if req_candidates:
                req_candidates.sort(key=lambda e: e.sent_time)
                pick = req_candidates[0]
                pick_ist = _email_ist(pick)
                pick_src = pick.sender_email or pick.sender_name
            else:
                # Quoted fallback from requester messages in the same episode window.
                quoted_candidates = []
                for e in thread:
                    if not e.sent_time or not _req_match(e, requester):
                        continue
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist < (window_start - timedelta(hours=6)) or e_ist > (window_end + timedelta(hours=6)):
                        continue
                    q_ist = _extract_quoted_request_before_ist(e, subject_norm, a_ist or r_ist or e_ist)
                    if not q_ist:
                        continue
                    if q_ist < window_start or q_ist > window_end:
                        continue
                    quoted_candidates.append(q_ist)
                if quoted_candidates:
                    quoted_candidates.sort()
                    pick_ist = quoted_candidates[-1]
                    pick_src = "PARSED_FROM_QUOTED_REQUEST"

            if not pick_ist:
                continue
            if abs((pick_ist - c_ist).total_seconds()) <= 120:
                continue

            t = _format_time(pick_ist)
            if not t:
                continue

            row_vals["Created Date & Time"] = t
            if a_ist and a_ist < pick_ist:
                row_vals["Actual Response Date & Time"] = t
            if r_ist and r_ist < pick_ist:
                row_vals["Actual Resolved Date & Time"] = t

            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = pick_src
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BaselineStaleCreatedGuard"

        # Risk-row episode lock (isolated, opt-in):
        # For clearly problematic rows, freeze all three timestamps from one
        # consultant/request episode to avoid cross-episode mixing.
        # Disabled by default because it can over-correct stable rows.
        locked_list_indexes = set()
        enable_risk_episode_lock = os.getenv("ENABLE_RISK_EPISODE_LOCK", "0") == "1"
        if enable_risk_episode_lock:
            for state in nonblue_row_states:
                list_index = state.get("list_index")
                if list_index is None or list_index >= len(automation_rows):
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                requester = state.get("requester") or ""
                subject_norm = (state.get("subject_norm") or "").lower()
                if not requester or not subject_norm:
                    continue

                notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
                notes_l = (notes_now or "").lower()
                baseline_date = state.get("baseline_created_date")
                ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
                resolved_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""

                c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                if not (c_dt and a_dt and r_dt):
                    continue
                c_ist = _to_ist(c_dt)
                a_ist = _to_ist(a_dt)
                r_ist = _to_ist(r_dt)

                baseline_drift = 0
                if baseline_date:
                    try:
                        baseline_drift = abs((c_ist.date() - baseline_date).days)
                    except Exception:
                        baseline_drift = 0
                source_mixed = False
                if requester and ack_src_now and resolved_src_now:
                    try:
                        source_mixed = (
                            (not _match_requester(str(ack_src_now), str(ack_src_now), requester))
                            and _match_requester(str(resolved_src_now), str(resolved_src_now), requester)
                        )
                    except Exception:
                        source_mixed = False
                # Keep this pass on clearly divergent rows only; 12h was too sensitive
                # and pulled many stable rows into expensive episode locking.
                large_gap = (a_ist - c_ist) > timedelta(hours=24)

                strong_risk_markers = (
                    "norequesterthreadrecovery",
                    "no ess or requester replies",
                    "ambiguousresolvedbyrequester",
                )
                weak_risk_markers = (
                    "dateanchormissing",
                    "dateanchorafter",
                )
                strong_risky = any(m in notes_l for m in strong_risk_markers)
                weak_risky = any(m in notes_l for m in weak_risk_markers)
                # Only allow weak marker rows when there is a strong drift signal.
                if not strong_risky and not (weak_risky and (baseline_drift >= 3 or source_mixed or large_gap)):
                    continue
                # Keep lock pass narrow for performance and safety.
                if baseline_drift <= 2 and (not source_mixed) and (not large_gap):
                    continue

                base_thread = state.get("thread") or []
                ref_ist = c_ist
                if baseline_date:
                    try:
                        ref_ist = _to_ist(datetime(baseline_date.year, baseline_date.month, baseline_date.day))
                    except Exception:
                        ref_ist = c_ist
                thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=ref_ist,
                )
                if not thread:
                    continue

                row_tokens = _match_tokens(subject_norm)
                subject_fit_cache = {}

                def _subject_fit(email_obj):
                    k = id(email_obj)
                    if k in subject_fit_cache:
                        return subject_fit_cache[k]
                    if not row_tokens:
                        subject_fit_cache[k] = True
                        return True
                    s_norm = normalize_subject(getattr(email_obj, "subject", "") or "")
                    s_tokens = _match_tokens(s_norm)
                    score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                    contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                    ok = score >= 0.45 or contains
                    subject_fit_cache[k] = ok
                    return ok

                consultant_msgs = [
                    e for e in thread
                    if getattr(e, "sent_time", None)
                    and _req_match(e, requester)
                    and _email_ist(e)
                    and _subject_fit(e)
                ]
                consultant_msgs.sort(key=lambda e: e.sent_time)
                if not consultant_msgs:
                    continue

                episodes = []
                current = [consultant_msgs[0]]
                for e in consultant_msgs[1:]:
                    prev_ist = _email_ist(current[-1])
                    now_ist = _email_ist(e)
                    if prev_ist and now_ist and (now_ist - prev_ist) > timedelta(hours=24):
                        episodes.append(current)
                        current = [e]
                    else:
                        current.append(e)
                episodes.append(current)

                anchor_ist = a_ist or r_ist or c_ist
                if baseline_date:
                    try:
                        anchor_ist = _to_ist(datetime(baseline_date.year, baseline_date.month, baseline_date.day))
                    except Exception:
                        pass

                best_episode = None
                best_rank = None
                for ep in episodes:
                    non_ack_ep = [e for e in ep if not _ack_like(e) and not _ack_like_text_fallback(e)]
                    seed = non_ack_ep[0] if non_ack_ep else ep[0]
                    seed_ist = _email_ist(seed)
                    if not seed_ist:
                        continue
                    if baseline_date:
                        day_delta = abs((seed_ist.date() - baseline_date).days)
                    else:
                        day_delta = abs((seed_ist - anchor_ist).total_seconds()) / 86400.0
                    prox = abs((seed_ist - anchor_ist).total_seconds())
                    ack_penalty = 1 if not non_ack_ep else 0
                    rank = (day_delta, prox, ack_penalty)
                    if best_rank is None or rank < best_rank:
                        best_rank = rank
                        best_episode = ep

                if not best_episode:
                    continue

                non_ack_best = [e for e in best_episode if not _ack_like(e) and not _ack_like_text_fallback(e)]
                response_mail = non_ack_best[0] if non_ack_best else best_episode[0]
                resolved_mail = non_ack_best[-1] if non_ack_best else best_episode[-1]
                response_ist = _email_ist(response_mail)
                if not response_ist:
                    continue

                non_ess_reqs = [
                    e for e in thread
                    if getattr(e, "sent_time", None)
                    and _email_ist(e)
                    and _subject_fit(e)
                    and not _ess_sender(e)
                    and not _system_like_sender(e)
                    and _email_ist(e) <= response_ist
                    and _email_ist(e) >= (response_ist - timedelta(days=7))
                ]
                if baseline_date:
                    by_day = [e for e in non_ess_reqs if _email_ist(e).date() == baseline_date]
                    if by_day:
                        non_ess_reqs = by_day
                created_mail = max(non_ess_reqs, key=lambda e: e.sent_time) if non_ess_reqs else response_mail

                t_c = _format_time(created_mail.sent_time)
                t_a = _format_time(response_mail.sent_time)
                t_r = _format_time(resolved_mail.sent_time)
                if not (t_c and t_a and t_r):
                    continue

                row_vals["Created Date & Time"] = t_c
                row_vals["Actual Response Date & Time"] = t_a
                row_vals["Actual Resolved Date & Time"] = t_r
                row_idx = state.get("row_index")
                if row_idx:
                    ws.cell(row_idx, created_col).value = t_c
                    ws.cell(row_idx, response_col).value = t_a
                    ws.cell(row_idx, resolved_col).value = t_r
                if list_index < len(debug_rows):
                    debug_rows[list_index]["CreatedSource"] = created_mail.sender_email or created_mail.sender_name
                    debug_rows[list_index]["AckSource"] = response_mail.sender_email or response_mail.sender_name
                    debug_rows[list_index]["ResolvedSource"] = resolved_mail.sender_email or resolved_mail.sender_name
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RiskEpisodeLock"
                locked_list_indexes.add(list_index)

        # Final monotonic safety:
        # ensure Created <= Ack <= Resolved after all guards.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if list_index in locked_list_indexes:
                continue
            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            orig_c = row_vals.get("Created Date & Time")
            orig_a = row_vals.get("Actual Response Date & Time")
            orig_r = row_vals.get("Actual Resolved Date & Time")
            orig_created_src = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
            orig_ack_src = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            orig_resolved_src = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            orig_notes = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            c = row_vals.get("Created Date & Time")
            a = row_vals.get("Actual Response Date & Time")
            r = row_vals.get("Actual Resolved Date & Time")
            c_dt = _parse_time_str(c)
            a_dt = _parse_time_str(a)
            r_dt = _parse_time_str(r)
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)

            changed = False
            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            res_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""

            if a_ist < c_ist:
                repaired_created = False
                # Try safe requester/non-ESS re-anchor before collapsing ack to created.
                base_thread = state.get("thread") or []
                subject_norm = (state.get("subject_norm") or "").lower()
                thread = _expanded_thread(subject_norm, base_thread, requester, include_non_ess=True, reference_ist=a_ist)
                if thread:
                    candidate_reqs = []
                    for e in thread:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if e_ist > (a_ist + timedelta(minutes=3)):
                            continue
                        if e_ist < (a_ist - timedelta(days=14)):
                            continue
                        if _ess_sender(e) or _system_like_sender(e):
                            continue
                        candidate_reqs.append(e)
                    if candidate_reqs:
                        pick_req = max(candidate_reqs, key=lambda e: e.sent_time)
                        pick_req_ist = _email_ist(pick_req)
                        if pick_req_ist and pick_req_ist <= a_ist:
                            new_c = _format_time(pick_req.sent_time)
                            if new_c:
                                row_vals["Created Date & Time"] = new_c
                                c = new_c
                                c_dt = _parse_time_str(c)
                                c_ist = _to_ist(c_dt) if c_dt else c_ist
                                repaired_created = True
                                changed = True
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["CreatedSource"] = pick_req.sender_email or pick_req.sender_name
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; MonotonicCreatedRepair"
                if a_ist < c_ist:
                    a = c
                    a_dt = _parse_time_str(a)
                    a_ist = _to_ist(a_dt) if a_dt else c_ist
                    row_vals["Actual Response Date & Time"] = a
                    changed = True
            if r_ist < a_ist:
                prefer_back_anchor = (
                    "requester span(all-ack->ess)" in (notes_now or "").lower()
                    or "requester span(all-ack->requester-fallback)" in (notes_now or "").lower()
                    or (
                        str(ack_src_now).strip().upper() == "ACK NOT FOUND"
                        and c == a
                        and res_src_now
                        and requester
                        and not _match_requester(res_src_now, res_src_now, requester)
                    )
                )

                if prefer_back_anchor:
                    # Keep older non-ack resolved episode and re-anchor created/ack back.
                    a = r
                    row_vals["Actual Response Date & Time"] = a
                    a_dt = _parse_time_str(a)
                    a_ist = _to_ist(a_dt) if a_dt else r_ist
                    if c_ist > a_ist:
                        c = a
                        row_vals["Created Date & Time"] = c
                        c_dt = _parse_time_str(c)
                        c_ist = _to_ist(c_dt) if c_dt else a_ist
                    changed = True
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; MonotonicBackAnchorGuard"
                else:
                    # Before collapsing resolved to ack, try a safe requester-based
                    # repair in the same thread episode.
                    base_thread = state.get("thread") or []
                    subject_norm = (state.get("subject_norm") or "").lower()
                    thread = _expanded_thread(subject_norm, base_thread, requester)
                    repaired = False
                    if thread and requester:
                        repair_candidates = [
                            e for e in thread
                            if e.sent_time
                            and _req_match(e, requester)
                            and _email_ist(e)
                            and _email_ist(e) >= a_ist
                            and _email_ist(e) <= (a_ist + timedelta(hours=48))
                            and not _ack_like(e)
                            and not _ack_like_text_fallback(e)
                        ]
                        repair_candidates.sort(key=lambda e: e.sent_time)
                        if repair_candidates:
                            pick = repair_candidates[0]
                            repaired_r = _format_time(pick.sent_time)
                            if repaired_r:
                                r = repaired_r
                                row_vals["Actual Resolved Date & Time"] = r
                                repaired = True
                                changed = True
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["ResolvedSource"] = pick.sender_email or pick.sender_name
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; MonotonicResolvedRepair"
                    if not repaired:
                        r = a
                        row_vals["Actual Resolved Date & Time"] = r
                        changed = True

            if not changed:
                continue
            final_c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            final_a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            final_r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            final_c_ist = _to_ist(final_c_dt) if final_c_dt else None
            final_a_ist = _to_ist(final_a_dt) if final_a_dt else None
            final_r_ist = _to_ist(final_r_dt) if final_r_dt else None
            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                final_c_ist,
                final_a_ist,
                final_r_ist,
                "MonotonicGuard",
                "monotonic",
            ):
                row_vals["Created Date & Time"] = orig_c
                row_vals["Actual Response Date & Time"] = orig_a
                row_vals["Actual Resolved Date & Time"] = orig_r
                if list_index < len(debug_rows):
                    debug_rows[list_index]["CreatedSource"] = orig_created_src
                    debug_rows[list_index]["AckSource"] = orig_ack_src
                    debug_rows[list_index]["ResolvedSource"] = orig_resolved_src
                    debug_rows[list_index]["Notes"] = orig_notes
                continue
            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
            if list_index < len(debug_rows):
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; MonotonicGuard"

        def _is_strict_all_live_builder_row(state, row_vals, list_index) -> bool:
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            created_src = str(debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else "")
            ack_src = str(debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else "")
            resolved_src = str(debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else "")
            cache_key = (
                list_index,
                notes_l,
                created_src,
                ack_src,
                resolved_src,
                str(row_vals.get("Created Date & Time") or ""),
                str(row_vals.get("Actual Response Date & Time") or ""),
                str(row_vals.get("Actual Resolved Date & Time") or ""),
                state.get("requester") or "",
                (state.get("subject_norm") or "").lower(),
            )
            cached = strict_all_live_builder_cache.get(cache_key)
            if cached is not None:
                return cached
            risky_markers = (
                "quotedrequestonly",
                "quotedpairgap>16m",
                "created retained (response anchor unreliable)",
                "dateanchorafter",
                "dateanchormissing",
                "dateanchorignoredstale",
                "norequesterthreadrecovery",
                "ambiguousresolvedbyrequester",
                "requester follow-up",
                "requester span(",
                "esscontinuationguard[",
                "ess-only; no non-ess request",
                "hybrid",
            )
            if any(marker in notes_l for marker in risky_markers):
                strict_all_live_builder_cache[cache_key] = False
                return False

            profile = _rewrite_guard_profile(row_vals, list_index)
            if not profile.get("ordered"):
                strict_all_live_builder_cache[cache_key] = False
                return False
            if profile.get("low_confidence", False):
                strict_all_live_builder_cache[cache_key] = False
                return False

            srcs = (created_src, ack_src, resolved_src)
            if not all(srcs):
                strict_all_live_builder_cache[cache_key] = False
                return False
            if any(src.strip().lower().startswith("parsed_") for src in srcs):
                strict_all_live_builder_cache[cache_key] = False
                return False
            if any(src.strip().lower() == "ack not found" for src in srcs):
                strict_all_live_builder_cache[cache_key] = False
                return False

            live_reply_ist = _live_reply_lane_for_row(list_index)
            if not live_reply_ist:
                strict_all_live_builder_cache[cache_key] = False
                return False
            r_min = profile.get("r_min")
            if not r_min or r_min != live_reply_ist.replace(second=0, microsecond=0):
                strict_all_live_builder_cache[cache_key] = False
                return False

            requester = state.get("requester") or ""
            if requester and not _match_requester(resolved_src, resolved_src, requester):
                strict_all_live_builder_cache[cache_key] = False
                return False

            strict_all_live_builder_cache[cache_key] = True
            return True

        def _row_is_special_all_same_risk_family(state, row_vals, list_index) -> bool:
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            if _row_has_force_same_time_lock(state, notes_l):
                return True
            if "ess-only; no non-ess request" in notes_l:
                return True
            if "requester span(all-ack->ess)" in notes_l:
                return True
            shared_decision = state.get("shared_decision") or {}
            if (
                (shared_decision.get("owner") or "") == "shared_occurrence"
                and (shared_decision.get("fill_style") or "") == "all_three_same"
                and _is_authoritative_occurrence_lane(shared_decision.get("row_type") or "")
            ):
                return True
            shared_occ_plan = _preferred_shared_occurrence_plan(state)
            if shared_occ_plan and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                return True
            return False

        def _authoritative_seed_short_circuit_ok(state, row_vals, list_index) -> bool:
            initial_lane_episode = state.get("initial_lane_episode") or {}
            if not (state.get("seed_locked") and initial_lane_episode.get("authoritative")):
                return False
            base_c = initial_lane_episode.get("created")
            base_a = initial_lane_episode.get("response")
            base_r = initial_lane_episode.get("resolved")
            if not (base_c and base_a and base_r):
                return False
            seed_triplet = (
                base_c.replace(second=0, microsecond=0),
                base_a.replace(second=0, microsecond=0),
                base_r.replace(second=0, microsecond=0),
            )
            current_triplet = _current_row_triplet_ist(row_vals)
            if not current_triplet or current_triplet != seed_triplet:
                return False
            if _row_is_special_all_same_risk_family(state, row_vals, list_index):
                return False

            shared_decision = state.get("shared_decision") or {}
            decision_triplet = shared_decision.get("triplet")
            if decision_triplet and tuple(decision_triplet) != seed_triplet:
                return False

            shared_occ_plan = _preferred_shared_occurrence_plan(state)
            if shared_occ_plan:
                pick_when = shared_occ_plan.get("pick_when")
                if pick_when and seed_triplet[2] != pick_when and seed_triplet[1] != pick_when:
                    return False
                if _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                    return False

            return True

        # Mixed-owner episode clamp (isolated):
        # When Created is early, Response/Resolved are very late, and ownership is
        # split (Ack non-requester, Resolved requester), clamp to requester's local
        # episode near Created instead of a later drifted episode.
        for state in nonblue_row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            row_vals = automation_rows[list_index]
            if _is_strict_all_live_builder_row(state, row_vals, list_index):
                continue
            requester = state.get("requester") or ""
            if not requester:
                continue

            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if (a_ist - c_ist) <= timedelta(hours=18):
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if not (
                "monotonicguard" in notes_l
                or "liverequestanchorguardmixedowner" in notes_l
            ):
                continue

            ack_src_now = debug_rows[list_index].get("AckSource", "") if list_index < len(debug_rows) else ""
            res_src_now = debug_rows[list_index].get("ResolvedSource", "") if list_index < len(debug_rows) else ""
            if not (ack_src_now and res_src_now):
                continue
            try:
                ack_is_req = _match_requester(str(ack_src_now), str(ack_src_now), requester)
                res_is_req = _match_requester(str(res_src_now), str(res_src_now), requester)
            except Exception:
                continue
            if ack_is_req or (not res_is_req):
                continue

            subject_norm = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=r_ist,
            )
            if not thread:
                continue

            win_start = c_ist - timedelta(minutes=5)
            win_end = min(c_ist + timedelta(hours=72), r_ist + timedelta(minutes=5))
            req_local = []
            for e in thread:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist < win_start or e_ist > win_end:
                    continue
                if not _req_match(e, requester):
                    continue
                if _ack_like(e) or _ack_like_text_fallback(e):
                    continue
                req_local.append(e)
            if not req_local:
                continue

            req_local.sort(key=lambda e: e.sent_time)
            response_pick = req_local[0]
            response_pick_ist = _email_ist(response_pick)
            if not response_pick_ist:
                continue
            # Apply only when materially earlier than current response.
            if response_pick_ist >= (a_ist - timedelta(minutes=5)):
                continue

            resolved_local = [
                e for e in req_local
                if _email_ist(e) >= response_pick_ist
                and _email_ist(e) <= (response_pick_ist + timedelta(hours=72))
            ]
            resolved_pick = resolved_local[-1] if resolved_local else response_pick

            t_a = _format_time(response_pick.sent_time)
            t_r = _format_time(resolved_pick.sent_time)
            if not (t_a and t_r):
                continue
            row_vals["Actual Response Date & Time"] = t_a
            row_vals["Actual Resolved Date & Time"] = t_r

            row_idx = state.get("row_index")
            if row_idx:
                ws.cell(row_idx, response_col).value = t_a
                ws.cell(row_idx, resolved_col).value = t_r
            if list_index < len(debug_rows):
                debug_rows[list_index]["AckSource"] = response_pick.sender_email or response_pick.sender_name
                debug_rows[list_index]["ResolvedSource"] = resolved_pick.sender_email or resolved_pick.sender_name
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; MixedOwnerEpisodeClamp"

        # Blue-gap final audit:
        # - Mark rows blue when Created->Response gap > 16 minutes.
        # - Before marking, try one safe local re-anchor for missed ack.
        # - If re-anchor fixes the gap, clear blue.
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        clear_fill = PatternFill(fill_type=None)

        def _is_blue_cell_fill(cell):
            try:
                rgb = (cell.fill.start_color.rgb or "").upper()
                return rgb.endswith("BDD7EE")
            except Exception:
                return False

        def _row_has_blue_fill(row_idx):
            for col in range(1, ws.max_column + 1):
                if _is_blue_cell_fill(ws.cell(row_idx, col)):
                    return True
            return False

        def _set_row_fill(row_idx, fill_obj):
            for col in range(1, ws.max_column + 1):
                ws.cell(row_idx, col).fill = fill_obj

        blue_gap_audit_started_at = _stage_timer_start()
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue

            row_vals = automation_rows[list_index]
            if _is_strict_all_live_builder_row(state, row_vals, list_index):
                continue
            requester = state.get("requester") or ""
            if not requester:
                continue

            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt) if r_dt else None

            # Safe local re-anchor only for rows currently >16m.
            if (a_ist - c_ist) > timedelta(minutes=16):
                subject_norm = (state.get("subject_norm") or "").lower()
                base_thread = state.get("thread") or []
                thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=a_ist,
                )
                if thread:
                    local_req = []
                    for e in thread:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if not _req_match(e, requester):
                            continue
                        if e_ist <= c_ist:
                            continue
                        if e_ist > (c_ist + timedelta(minutes=16)):
                            continue
                        local_req.append(e)
                    if local_req:
                        local_req.sort(key=lambda e: e.sent_time)
                        ack_pick = local_req[0]
                        ack_pick_ist = _email_ist(ack_pick)
                        if ack_pick_ist and ack_pick_ist < a_ist:
                            t_a = _format_time(ack_pick.sent_time)
                            if t_a:
                                row_vals["Actual Response Date & Time"] = t_a
                                a_dt = _parse_time_str(t_a)
                                a_ist = _to_ist(a_dt) if a_dt else a_ist
                                if r_ist and r_ist < a_ist:
                                    row_vals["Actual Resolved Date & Time"] = t_a
                                    r_ist = a_ist
                                ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                                ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["AckSource"] = ack_pick.sender_email or ack_pick.sender_name
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueGapReanchor"

            # Final blue decision after re-anchor.
            a_dt2 = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            if not a_dt2:
                continue
            a_ist2 = _to_ist(a_dt2)
            is_blue_gap = (a_ist2 - c_ist) > timedelta(minutes=16)
            has_blue = _row_has_blue_fill(row_idx)
            if is_blue_gap:
                if not has_blue:
                    _set_row_fill(row_idx, blue_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueGap>16m"
            else:
                notes_now = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
                if "quotedpairgap>16m" in notes_now:
                    # Keep blue marking for invalid quoted-pair gaps unless we
                    # already collapsed to an ESS-only all-three-same row.
                    c_dt2 = _parse_time_str(row_vals.get("Created Date & Time"))
                    r_dt2 = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                    all_same = bool(c_dt2 and a_dt2 and r_dt2 and c_dt2 == a_dt2 == r_dt2)
                    if all_same and ("ess-only; no non-ess request" in notes_now):
                        if has_blue:
                            _set_row_fill(row_idx, clear_fill)
                        if list_index < len(debug_rows):
                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedEssOnly"
                    else:
                        if not has_blue:
                            _set_row_fill(row_idx, blue_fill)
                    continue
                if has_blue:
                    _set_row_fill(row_idx, clear_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueCleared"

        # Blue-only strict realign pass (isolated):
        # Process only rows already marked blue, so stricter logic cannot disturb
        # correctly filled rows.
        used_quoted_pair_keys = set()
        used_ess_continuation_blue = set()

        # Quoted-request-only tagging pass (isolated):
        # Identify rows where no live non-ESS request exists, but a non-ESS
        # request is present in quoted history. These should be processed
        # separately from ESS-only continuation.
        quoted_request_only = set()
        quoted_only_hybrid_pairs = {}
        quoted_only_hybrid_req_sources = {}
        quoted_only_hybrid_req_debug = {}
        ess_email_set = {e.strip().lower() for e in ess_team or []}
        ess_domain_set = {"invenio-solutions.com", "inveniolsi.com"}
        ess_name_tokens = set()
        for em in ess_email_set:
            if "@" in em:
                local = em.split("@", 1)[0]
                for tok in re.split(r"[._\\-]+", local):
                    if len(tok) >= 3:
                        ess_name_tokens.add(tok.lower())
        # Expand ESS name tokens with observed sender names in this thread.
        ess_name_tokens_all = set(ess_name_tokens)
        for st in row_states:
            thread_msgs = st.get("thread") or []
            for msg in thread_msgs:
                # Only add tokens for ESS senders to avoid misclassifying non-ESS names.
                sender_email = (getattr(msg, "sender_email", "") or "").lower()
                sender_domain = sender_email.split("@", 1)[-1] if "@" in sender_email else ""
                if (sender_email in ess_email_set) or (sender_domain in ess_domain_set):
                    name_raw = (getattr(msg, "sender_name", "") or "").lower()
                    if name_raw:
                        for tok in re.split(r"[^a-z0-9]+", name_raw):
                            if len(tok) >= 3:
                                ess_name_tokens_all.add(tok)
        _ess_name_stop = {
            "admin", "support", "team", "service", "services", "ops", "operations",
            "enterprise", "ess", "es", "help", "helpdesk", "desk", "mailbox",
            "noreply", "no", "reply", "system", "group", "global",
        }
        def _ess_name_only(from_line: str) -> bool:
            name_blob = re.sub(r"[^a-z0-9]+", " ", (from_line or "").lower()).strip()
            if not name_blob:
                return False
            tokens = {t for t in name_blob.split() if len(t) >= 3 and t not in _ess_name_stop}
            if not tokens:
                return False
            ess_tokens = {t for t in ess_name_tokens_all if t not in _ess_name_stop}
            matches = tokens & ess_tokens
            if not matches:
                return False
            if len(matches) >= 2:
                return True
            # Single-token match: require a long token to avoid false ESS hits like "admin".
            tok = next(iter(matches))
            return len(tok) >= 6

        def _quoted_from_line_is_ess_shared(from_line: str) -> bool:
            key = (from_line or "").strip().lower()
            cached = _quoted_from_line_is_ess_cache.get(key)
            if cached is not None:
                return cached
            addr_hits = re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", from_line or "", flags=re.I)
            if not addr_hits:
                out = _ess_name_only(from_line)
                _quoted_from_line_is_ess_cache[key] = out
                return out
            emails_l = [em.lower() for em in addr_hits]
            # Keep quoted sender classification narrow and consistent across the
            # quoted-family flows. Domain-only matching is too broad because
            # requester/business users can share company domains with ESS.
            out = any(em in ess_email_set for em in emails_l) or _ess_name_only(from_line)
            _quoted_from_line_is_ess_cache[key] = out
            return out
        _quoted_from_line_is_ess_cache = {}
        _quoted_message_lines_cache = {}
        _bounded_quoted_header_cache = {}

        def _clean_quoted_message_lines(email_obj):
            cache_key = _email_stable_key(email_obj)
            if cache_key in _quoted_message_lines_cache:
                return _quoted_message_lines_cache[cache_key]
            _quoted_message_lines_cache[cache_key] = _extract_canonical_message_lines(email_obj)
            return _quoted_message_lines_cache[cache_key]
        def _extract_bounded_quoted_header_candidates(email_obj, *, allow_relaxed: bool = False):
            cache_key = (_email_stable_key(email_obj), bool(allow_relaxed))
            cached = _bounded_quoted_header_cache.get(cache_key)
            if cached is not None:
                return list(cached)
            blocks = []
            for candidate in _extract_canonical_quoted_header_candidates(email_obj, allow_relaxed=allow_relaxed):
                sent_dt = candidate.sent_dt
                if not sent_dt:
                    continue
                from_line = re.sub(r"(?i)^from\b\s*:?\s*", "", candidate.from_line or "").strip()
                subj_text = re.sub(r"(?i)^(subject|objet)\b\s*:?\s*", "", candidate.subject_line or "").strip()
                blocks.append((from_line, _to_ist(sent_dt), subj_text))
            _bounded_quoted_header_cache[cache_key] = tuple(blocks)
            return list(_bounded_quoted_header_cache[cache_key])
        def _extract_quoted_blocks_with_subject(email_obj):
            return _extract_bounded_quoted_header_candidates(email_obj, allow_relaxed=False)
        def _extract_quoted_blocks_relaxed(email_obj):
            return _extract_bounded_quoted_header_candidates(email_obj, allow_relaxed=True)
        quoted_block_cache = {}
        raw_eml_quoted_cache = {}
        raw_eml_quoted_summary_cache = {}
        raw_eml_header_summary_cache = dict(_eml_header_prewarm)
        raw_eml_episode_fallback_cache = {}
        email_by_path_cache = {}
        raw_id_path_cache = {}
        _quoted_block_matches_row_cache = {}
        # Pre-warm raw_eml_quoted_cache from already-loaded EmailRecord objects.
        # Prevents _get_quoted_blocks_from_eml_path from re-reading files from disk.
        for _pw_e in emails:
            _pw_path = str(getattr(_pw_e, "path", "") or "")
            if not _pw_path or _pw_path in raw_eml_quoted_cache:
                continue
            try:
                _pw_blocks = _extract_quoted_blocks_with_subject(_pw_e)
                if not _pw_blocks:
                    _pw_blocks = _extract_quoted_blocks_relaxed(_pw_e)
                raw_eml_quoted_cache[_pw_path] = _pw_blocks
            except Exception:
                raw_eml_quoted_cache[_pw_path] = []

        # Pre-warm email_by_path_cache - turns O(n) linear scan per call into O(1) dict lookup.
        email_by_path_cache.update({
            str(getattr(_pw_e, "path", "") or ""): _pw_e
            for _pw_e in emails
            if getattr(_pw_e, "path", None)
        })
        eml_id_index = None
        def _get_quoted_blocks_with_subject_cached(msg):
            key = _email_stable_key(msg)
            if key in quoted_block_cache:
                return quoted_block_cache[key]
            blocks = _extract_quoted_blocks_with_subject(msg)
            quoted_block_cache[key] = blocks
            return blocks
        def _get_quoted_blocks_from_eml_path(path: str):
            if not path:
                return []
            if path in raw_eml_quoted_cache:
                return raw_eml_quoted_cache[path]
            try:
                with open(path, "rb") as f:
                    msg = BytesParser(policy=policy.default).parse(f)
            except Exception:
                raw_eml_quoted_cache[path] = []
                return []
            plain_parts = []
            html_parts = []
            try:
                if msg.is_multipart():
                    for part in msg.walk():
                        ctype = part.get_content_type()
                        try:
                            content = part.get_content()
                        except Exception:
                            content = None
                        if ctype == "text/plain" and isinstance(content, str):
                            plain_parts.append(content.strip())
                        elif ctype == "text/html" and isinstance(content, str):
                            html_parts.append(content)
                else:
                    content = msg.get_content()
                    if isinstance(content, str):
                        plain_parts.append(content.strip())
            except Exception:
                pass
            raw_plain = "\n".join(p for p in plain_parts if p)
            raw_html = "\n".join(h for h in html_parts if isinstance(h, str))
            temp = SimpleNamespace(body=raw_plain, body_html="", body_html_raw=raw_html)
            blocks = _extract_quoted_blocks_with_subject(temp)
            if not blocks:
                blocks = _extract_quoted_blocks_relaxed(temp)
            raw_eml_quoted_cache[path] = blocks
            return blocks
        raw_id_token_cache = {}
        def _canonical_raw_text(msg):
            key = ("canonical_raw_text", _email_stable_key(msg))
            cached = _raw_match_token_cache.get(key)
            if cached is not None:
                return cached
            lines = _extract_canonical_message_lines(msg)
            text = " ".join(lines)
            if getattr(msg, "subject", ""):
                text = f"{getattr(msg, 'subject', '')} {text}".strip()
            _raw_match_token_cache[key] = text
            return text
        def _raw_match_tokens(msg):
            key = _email_stable_key(msg)
            cached = _raw_match_token_cache.get(key)
            if cached is not None:
                return cached
            raw = _canonical_raw_text(msg).lower()
            raw = raw.replace("\u2013", "-").replace("\u2014", "-").replace("\u2011", "-")
            raw = raw.replace("&ndash;", "-").replace("&mdash;", "-").replace("&#8209;", "-")
            out = _match_tokens(raw)
            _raw_match_token_cache[key] = out
            return out
        def _row_id_in_raw(msg, row_id_tokens_set):
            if not row_id_tokens_set:
                return True
            # Fast path: check subject only.
            subj_norm = _subject_norm_cached(getattr(msg, "subject", "") or "")
            subj_ids = _id_like_tokens(subj_norm)
            if subj_ids and (not subj_ids.isdisjoint(row_id_tokens_set)):
                return True

            key = _email_stable_key(msg)
            if key in raw_id_token_cache:
                return bool(raw_id_token_cache[key] & row_id_tokens_set)

            raw = _canonical_raw_text(msg).lower()
            raw = raw.replace("\u2013", "-").replace("\u2014", "-").replace("\u2011", "-")
            raw = raw.replace("&ndash;", "-").replace("&mdash;", "-").replace("&#8209;", "-")
            raw_tokens = {
                m.group(0)
                for m in re.finditer(r"(?<![a-z0-9])[a-z]{2,}[a-z0-9\\-]*\\d[a-z0-9\\-]*(?![a-z0-9])", raw)
            }
            raw_id_token_cache[key] = raw_tokens
            return bool(raw_tokens & row_id_tokens_set)
        def _quoted_subject_confirms_row(
            q_norm: str,
            q_ids: set,
            q_tokens: set,
            subject_norm_value: str,
            row_tokens: set,
            row_id_tokens_set: set,
        ) -> bool:
            if q_norm and not _strict_subject_identity_gate(
                subject_norm_value or "",
                q_norm,
                iface_tokens=_interface_tokens(subject_norm_value or ""),
            ):
                return False
            if row_id_tokens_set:
                has_id_overlap = bool(q_ids and not row_id_tokens_set.isdisjoint(q_ids))
                if not has_id_overlap:
                    return False
            if not _quoted_subject_short_variant_ok(subject_norm_value, q_norm):
                return False
            if not row_tokens:
                return True
            score = _token_overlap_score(row_tokens, q_tokens) if q_tokens else 0.0
            contains = bool(
                subject_norm_value
                and q_norm
                and (subject_norm_value in q_norm or q_norm in subject_norm_value)
            )
            if q_norm and not _quoted_subject_variant_owns_row(subject_norm_value, q_norm):
                return False
            if score >= 0.45 or contains:
                return True
            return _fresh_picker_subject_safe(
                subject_norm_value,
                q_norm,
                iface_tokens=_interface_tokens(subject_norm_value),
                allow_added_inc=True,
            )

        def _quoted_block_matches_row(
            msg,
            q_subj: str,
            subject_norm_value: str,
            row_tokens: set,
            row_id_tokens_set: set,
            *,
            allow_raw_id_fallback: bool = False,
        ) -> bool:
            cache_key = (
                id(msg),
                q_subj or "",
                subject_norm_value or "",
                tuple(sorted(row_tokens or ())),
                tuple(sorted(row_id_tokens_set or ())),
                bool(allow_raw_id_fallback),
            )
            cached = _quoted_block_matches_row_cache.get(cache_key)
            if cached is not None:
                return cached
            if row_id_tokens_set:
                if not q_subj:
                    out = bool(allow_raw_id_fallback and _row_id_in_raw(msg, row_id_tokens_set))
                    if out:
                        row_short_variants = _subject_short_variant_tokens(subject_norm_value or "")
                        if row_short_variants and row_short_variants.isdisjoint(_raw_match_tokens(msg)):
                            out = False
                    _quoted_block_matches_row_cache[cache_key] = out
                    return out
                q_norm = normalize_subject(q_subj or "")
                q_ids = _id_like_tokens(q_norm)
                q_tokens = _match_tokens(q_norm)
                out = _quoted_subject_confirms_row(
                    q_norm,
                    q_ids,
                    q_tokens,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens_set,
                )
                _quoted_block_matches_row_cache[cache_key] = out
                return out
            if row_tokens:
                if not q_subj:
                    _quoted_block_matches_row_cache[cache_key] = False
                    return False
                q_norm = normalize_subject(q_subj or "")
                q_ids = _id_like_tokens(q_norm)
                q_tokens = _match_tokens(q_norm)
                out = _quoted_subject_confirms_row(
                    q_norm,
                    q_ids,
                    q_tokens,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens_set,
                )
                _quoted_block_matches_row_cache[cache_key] = out
                return out
            _quoted_block_matches_row_cache[cache_key] = True
            return True

        def _get_quoted_summaries_from_eml_path(path: str):
            if not path:
                return []
            if path in raw_eml_quoted_summary_cache:
                return raw_eml_quoted_summary_cache[path]
            blocks = _get_quoted_blocks_from_eml_path(path)
            if not blocks:
                raw_eml_quoted_summary_cache[path] = []
                return []
            summaries = []
            for from_line, sent_ist, q_subj in blocks:
                if not q_subj:
                    continue
                q_norm = normalize_subject(q_subj or "")
                q_ids = _id_like_tokens(q_norm)
                if not q_ids:
                    continue
                q_tokens = _match_tokens(q_norm)
                is_ess = _quoted_from_line_is_ess_shared(from_line)
                summaries.append((sent_ist, q_norm, q_ids, q_tokens, is_ess))
            raw_eml_quoted_summary_cache[path] = summaries
            return summaries

        _lane_local_episode_cache = {}
        _quoted_sender_matches_live_cache = {}
        _seed_locked_lane_context_cache = {}
        _seed_locked_lane_episode_cache = {}

        def _quoted_sender_matches_live_shared(from_line: str, email_obj) -> bool:
            cache_key = ((from_line or "").strip().lower(), id(email_obj))
            cached = _quoted_sender_matches_live_cache.get(cache_key)
            if cached is not None:
                return cached
            sender_email = (getattr(email_obj, "sender_email", "") or "").strip().lower()
            sender_name = (getattr(email_obj, "sender_name", "") or "").strip().lower()
            blob = (from_line or "").strip().lower()
            if sender_email and sender_email in blob:
                _quoted_sender_matches_live_cache[cache_key] = True
                return True
            if not sender_name:
                _quoted_sender_matches_live_cache[cache_key] = False
                return False
            quoted_tokens = {
                tok for tok in re.split(r"[^a-z0-9]+", blob)
                if len(tok) >= 3 and tok not in _ess_name_stop
            }
            sender_tokens = {
                tok for tok in re.split(r"[^a-z0-9]+", sender_name)
                if len(tok) >= 3 and tok not in _ess_name_stop
            }
            overlap = quoted_tokens & sender_tokens
            out = len(overlap) >= 2 or any(len(tok) >= 6 for tok in overlap)
            _quoted_sender_matches_live_cache[cache_key] = out
            return out

        def _lane_local_episode_from_reply(
            reply_msg,
            reply_ist,
            thread,
            requester_name: str,
            subject_norm_value: str,
            row_tokens: set,
            row_id_tokens_set: set,
        ):
            thread_sig = tuple(id(e) for e in (thread or ()))
            cache_key = (
                id(reply_msg) if reply_msg else None,
                reply_ist.replace(second=0, microsecond=0) if reply_ist else None,
                requester_name or "",
                subject_norm_value or "",
                thread_sig,
            )
            cached = _lane_local_episode_cache.get(cache_key)
            if cached is not None:
                return cached
            if not (reply_msg and reply_ist and thread):
                _lane_local_episode_cache[cache_key] = None
                return None

            quoted_blocks = _get_quoted_blocks_with_subject_cached(reply_msg)
            if not quoted_blocks:
                quoted_blocks = _get_quoted_blocks_from_eml_path(getattr(reply_msg, "path", ""))
            if not quoted_blocks:
                _lane_local_episode_cache[cache_key] = None
                return None

            lane_blocks = []
            for idx, (from_line, sent_ist, q_subj) in enumerate(quoted_blocks):
                if not sent_ist or sent_ist >= reply_ist:
                    continue
                subject_matches = _quoted_block_matches_row(
                    reply_msg,
                    q_subj,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens_set,
                    allow_raw_id_fallback=True,
                )
                allow_blank_subject_anchor = (
                    not q_subj
                    and idx <= 2
                    and sent_ist < reply_ist
                    and (reply_ist - sent_ist) <= timedelta(hours=48)
                )
                if not subject_matches and not allow_blank_subject_anchor:
                    continue
                lane_blocks.append((idx, from_line or "", sent_ist, q_subj or ""))
                if len(lane_blocks) >= 10:
                    break
            if not lane_blocks:
                _lane_local_episode_cache[cache_key] = None
                return None

            def _quoted_sender_is_ess(from_line: str):
                return _quoted_from_line_is_ess_shared(from_line)

            ack_idx = None
            ack_ist = None
            ack_msg = None
            live_ack_candidates = []
            for e in thread:
                e_ist = _email_ist(e)
                if not e_ist or e_ist >= reply_ist:
                    continue
                if not _ess_sender(e):
                    continue
                if not _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens_set):
                    continue
                if not _is_shared_ess_ack_candidate(e):
                    continue
                live_ack_candidates.append((e_ist, e))
            for idx, from_line, sent_ist, _q_subj in lane_blocks:
                if _quoted_sender_is_ess(from_line) is not True:
                    continue

                live_ack_matches = []
                for e_ist, e in live_ack_candidates:
                    delta = abs((e_ist - sent_ist).total_seconds())
                    if delta > 300:
                        continue
                    sender_score = 1 if _quoted_sender_matches_live_shared(from_line, e) else 0
                    live_ack_matches.append((sender_score, -delta, e_ist, id(e), e))
                if not live_ack_matches:
                    continue
                live_ack_matches.sort(reverse=True)
                ack_idx = idx
                ack_ist = live_ack_matches[0][2]
                ack_msg = live_ack_matches[0][4]
                break

            req_ist = None
            if ack_idx is not None and ack_ist:
                for next_idx, next_from_line, next_sent_ist, _next_q_subj in lane_blocks:
                    if next_idx <= ack_idx:
                        continue
                    if next_sent_ist >= ack_ist:
                        continue
                    if _quoted_sender_is_ess(next_from_line) is not False:
                        continue
                    if (ack_ist - next_sent_ist) > timedelta(minutes=16):
                        continue
                    req_ist = next_sent_ist
                    break

            if ack_ist and not req_ist:
                ack_idx = None
                ack_ist = None
                ack_msg = None

            if not req_ist:
                for _idx, from_line, sent_ist, _q_subj in lane_blocks:
                    if _quoted_sender_is_ess(from_line) is False:
                        req_ist = sent_ist
                        break

            if not req_ist:
                _lane_local_episode_cache[cache_key] = None
                return None

            out = {
                "request": req_ist,
                "ack": ack_ist,
                "ack_msg": ack_msg,
                "reply": reply_ist,
                "reply_msg": reply_msg,
            }
            _lane_local_episode_cache[cache_key] = out
            return out

        def _seed_locked_lane_context(state, row_vals, list_index):
            if not state:
                return None
            requester = state.get("requester") or ""
            subject_norm_value = (state.get("subject_norm") or "").lower()
            if not requester or not subject_norm_value:
                return None

            live_reply_ist = _live_reply_lane_for_row(list_index)
            live_reply_msg = _live_reply_message_for_row(list_index)
            if not (live_reply_ist and live_reply_msg):
                return None

            row_tokens = _match_tokens(subject_norm_value)
            row_id_tokens = _id_like_tokens(subject_norm_value)
            if not row_id_tokens:
                desc_text = row_vals.get("Description") or state.get("description") or ""
                row_id_tokens = _id_like_tokens(desc_text)

            base_thread = state.get("thread") or []
            cache_key = (
                id(base_thread),
                id(live_reply_msg),
                live_reply_ist.replace(second=0, microsecond=0),
                requester,
                subject_norm_value,
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
            )
            cached = _seed_locked_lane_context_cache.get(cache_key)
            if cached is not None:
                return cached

            lane_msgs = []
            seen = set()
            for e in list(base_thread) + [live_reply_msg]:
                if id(e) in seen:
                    continue
                seen.add(id(e))
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if e_ist > live_reply_ist:
                    continue
                if (live_reply_ist - e_ist) > timedelta(hours=48):
                    continue
                lane_msgs.append(e)
            lane_msgs.sort(key=lambda e: e.sent_time)
            out = {
                "requester": requester,
                "subject_norm": subject_norm_value,
                "live_reply_ist": live_reply_ist,
                "live_reply_msg": live_reply_msg,
                "row_tokens": row_tokens,
                "row_id_tokens": row_id_tokens,
                "thread": lane_msgs,
            }
            _seed_locked_lane_context_cache[cache_key] = out
            return out

        def _seed_lane_episode_from_locked_reply(
            reply_msg,
            reply_ist,
            thread,
            requester_name: str,
            subject_norm_value: str,
            row_tokens: set,
            row_id_tokens_set: set,
            *,
            allow_strict_same_time: bool = False,
        ):
            thread_sig = tuple(id(e) for e in (thread or ()))
            cache_key = (
                id(reply_msg) if reply_msg else None,
                reply_ist.replace(second=0, microsecond=0) if reply_ist else None,
                requester_name or "",
                subject_norm_value or "",
                thread_sig,
                bool(allow_strict_same_time),
            )
            cached = _seed_locked_lane_episode_cache.get(cache_key)
            if cached is not None:
                return cached
            if not (reply_msg and reply_ist and thread):
                _seed_locked_lane_episode_cache[cache_key] = None
                return None
            reply_is_dl_only_reroute = (
                workbook_kind == "incident_business"
                and _is_ess_dl_only_reroute(reply_msg, ess_team)
            )

            quoted_blocks = _get_quoted_blocks_with_subject_cached(reply_msg)
            if not quoted_blocks:
                quoted_blocks = _get_quoted_blocks_from_eml_path(getattr(reply_msg, "path", ""))
            if not quoted_blocks:
                _seed_locked_lane_episode_cache[cache_key] = None
                return None

            primary_blocks = []
            for idx, (from_line, sent_ist, q_subj) in enumerate(quoted_blocks):
                if not sent_ist or sent_ist >= reply_ist:
                    continue
                if (reply_ist - sent_ist) > timedelta(hours=48):
                    continue
                subject_matches = _quoted_block_matches_row(
                    reply_msg,
                    q_subj,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens_set,
                    allow_raw_id_fallback=True,
                )
                allow_blank_subject_anchor = (
                    not q_subj
                    and not primary_blocks
                    and idx <= 2
                    and sent_ist < reply_ist
                    and (reply_ist - sent_ist) <= timedelta(hours=48)
                )
                if not subject_matches and not allow_blank_subject_anchor:
                    if primary_blocks:
                        break
                    continue
                primary_blocks.append((idx, from_line or "", sent_ist, q_subj or ""))
                if len(primary_blocks) >= 6:
                    break
            if not primary_blocks:
                _seed_locked_lane_episode_cache[cache_key] = None
                return None

            first_idx, first_from_line, first_sent_ist, _first_q_subj = primary_blocks[0]
            first_is_ess = _quoted_from_line_is_ess_shared(first_from_line)
            reply_flags = _shared_reply_flags(email_record=reply_msg)
            html_probe_text = _seed_current_html_text(reply_msg)
            html_probe = (
                SimpleNamespace(
                    subject=getattr(reply_msg, "subject", "") or "",
                    sender_name=getattr(reply_msg, "sender_name", "") or "",
                    sender_email=getattr(reply_msg, "sender_email", "") or "",
                    sent_time=getattr(reply_msg, "sent_time", None),
                    body=html_probe_text,
                    body_html="",
                )
                if html_probe_text
                else None
            )
            reply_ackish = bool(
                reply_flags["ack_candidate"]
                or reply_flags["thanks_info"]
                or reply_flags["nonfinal_followup"]
                or _is_shared_ess_ack_candidate(reply_msg)
                or _is_ack_like_reply(reply_msg)
                or _is_thanks_info_reply(reply_msg)
                or _email_has_explicit_ack_signal(reply_msg)
                or (
                    html_probe is not None
                    and (
                        _is_shared_ess_ack_candidate(html_probe)
                        or _is_ack_like_reply(html_probe)
                        or _is_thanks_info_reply(html_probe)
                        or _email_has_explicit_ack_signal(html_probe)
                    )
                )
            )

            def _lower_non_ess_below(reference_idx, reference_ist, reference_msg=None):
                lower_non_ess = None
                for next_idx, next_from_line, next_sent_ist, _next_q_subj in primary_blocks:
                    if next_idx <= reference_idx:
                        continue
                    if (
                        reference_msg is not None
                        and next_sent_ist
                        and abs((next_sent_ist - reference_ist).total_seconds()) <= 300
                        and _quoted_sender_matches_live_shared(next_from_line, reference_msg)
                    ):
                        continue
                    if next_sent_ist >= reference_ist:
                        continue
                    if _quoted_from_line_is_ess_shared(next_from_line) is False:
                        lower_non_ess = next_sent_ist
                        break
                return lower_non_ess

            if first_is_ess is False:
                out = {
                    "request": first_sent_ist,
                    "ack": None,
                    "ack_msg": None,
                    "reply": reply_ist,
                    "reply_msg": reply_msg,
                    "mode": "direct-reply",
                }
                _seed_locked_lane_episode_cache[cache_key] = out
                return out

            if first_is_ess is True:
                if (
                    allow_strict_same_time
                    and _ess_sender(reply_msg)
                    and not reply_is_dl_only_reroute
                    and reply_flags["substantive_reply"]
                    and not reply_flags["ack_candidate"]
                    and not reply_flags["ignore_reply"]
                ):
                    live_ess_same_time_candidates = []
                    for e in thread:
                        e_ist = _email_ist(e)
                        if not e_ist or e_ist >= reply_ist:
                            continue
                        if not _ess_sender(e):
                            continue
                        if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                            continue
                        if not _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens_set):
                            continue
                        flags = _shared_reply_flags(email_record=e)
                        if flags["ignore_reply"] or flags["ack_candidate"] or not flags["substantive_reply"]:
                            continue
                        live_ess_same_time_candidates.append((e_ist, e))

                    same_time_match = None
                    for e_ist, e in live_ess_same_time_candidates:
                        delta = abs((e_ist - first_sent_ist).total_seconds())
                        if delta > 300:
                            continue
                        if not _quoted_sender_matches_live_shared(first_from_line, e):
                            continue
                        cand = (-delta, e_ist, id(e), e)
                        if same_time_match is None or cand > same_time_match:
                            same_time_match = cand

                    if same_time_match is not None:
                        matched_ess_ist = same_time_match[1]
                        matched_ess_msg = same_time_match[3]
                        lower_non_ess = _lower_non_ess_below(first_idx, matched_ess_ist, matched_ess_msg)
                        if lower_non_ess is None:
                            out = {
                                "request": reply_ist,
                                "ack": reply_ist,
                                "ack_msg": reply_msg,
                                "reply": reply_ist,
                                "reply_msg": reply_msg,
                                "mode": "all-three-same",
                            }
                            _seed_locked_lane_episode_cache[cache_key] = out
                            return out
                        out = {
                            "request": lower_non_ess,
                            "ack": matched_ess_ist,
                            "ack_msg": None,
                            "reply": reply_ist,
                            "reply_msg": reply_msg,
                            "mode": "req-ack-reply",
                        }
                        _seed_locked_lane_episode_cache[cache_key] = out
                        return out

                if allow_strict_same_time and _ess_sender(reply_msg) and not reply_is_dl_only_reroute and reply_ackish:
                    lower_non_ess = _lower_non_ess_below(first_idx, first_sent_ist)
                    if lower_non_ess is None:
                        out = {
                            "request": reply_ist,
                            "ack": reply_ist,
                            "ack_msg": reply_msg,
                            "reply": reply_ist,
                            "reply_msg": reply_msg,
                            "mode": "all-three-same",
                        }
                    else:
                        out = {
                            "request": lower_non_ess,
                            "ack": first_sent_ist,
                            "ack_msg": None,
                            "reply": reply_ist,
                            "reply_msg": reply_msg,
                            "mode": "req-ack-reply",
                        }
                    _seed_locked_lane_episode_cache[cache_key] = out
                    return out

                live_ack_candidates = []
                for e in thread:
                    e_ist = _email_ist(e)
                    if not e_ist or e_ist >= reply_ist:
                        continue
                    if not _ess_sender(e):
                        continue
                    if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                        continue
                    if not _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens_set):
                        continue
                    if not _is_shared_ess_ack_candidate(e):
                        continue
                    live_ack_candidates.append((e_ist, e))

                ack_match = None
                for e_ist, e in live_ack_candidates:
                    delta = abs((e_ist - first_sent_ist).total_seconds())
                    if delta > 300:
                        continue
                    sender_score = 1 if _quoted_sender_matches_live_shared(first_from_line, e) else 0
                    cand = (sender_score, -delta, e_ist, id(e), e)
                    if ack_match is None or cand > ack_match:
                        ack_match = cand

                if ack_match is not None:
                    ack_ist = ack_match[2]
                    ack_msg = ack_match[4]
                    ack_echo_skips = 0
                    for next_idx, next_from_line, next_sent_ist, _next_q_subj in primary_blocks[1:]:
                        if next_idx <= first_idx:
                            continue
                        next_is_ess = _quoted_from_line_is_ess_shared(next_from_line)
                        echoed_ack_block = bool(
                            ack_msg
                            and next_sent_ist
                            and abs((next_sent_ist - ack_ist).total_seconds()) <= 300
                            and _quoted_sender_matches_live_shared(next_from_line, ack_msg)
                        )
                        if echoed_ack_block and ack_echo_skips < 1:
                            ack_echo_skips += 1
                            continue
                        if next_sent_ist >= ack_ist:
                            break
                        if next_is_ess is not False:
                            break
                        if (ack_ist - next_sent_ist) > timedelta(minutes=16):
                            break
                        out = {
                            "request": next_sent_ist,
                            "ack": ack_ist,
                            "ack_msg": ack_msg,
                            "reply": reply_ist,
                            "reply_msg": reply_msg,
                        }
                        _seed_locked_lane_episode_cache[cache_key] = out
                        return out

                _seed_locked_lane_episode_cache[cache_key] = None
                return None

            _seed_locked_lane_episode_cache[cache_key] = None
            return None

        def _seed_lane_local_initial_episode(
            state,
            row_vals,
            list_index,
            row_idx,
        ) -> bool:
            lane_ctx = _seed_locked_lane_context(state, row_vals, list_index)
            if not lane_ctx:
                return False

            requester = lane_ctx["requester"]
            subject_norm_value = lane_ctx["subject_norm"]
            live_reply_ist = lane_ctx["live_reply_ist"]
            live_reply_msg = lane_ctx["live_reply_msg"]
            row_tokens = lane_ctx["row_tokens"]
            row_id_tokens = lane_ctx["row_id_tokens"]
            thread = lane_ctx["thread"]

            lane_episode = _seed_lane_episode_from_locked_reply(
                live_reply_msg,
                live_reply_ist,
                thread,
                requester,
                subject_norm_value,
                row_tokens,
                row_id_tokens,
                allow_strict_same_time=bool((state.get("group_total") or 0) <= 1),
            )
            if not lane_episode or not lane_episode.get("request"):
                fallback_pair = quoted_only_hybrid_pairs.get(list_index)
                if fallback_pair:
                    fb_req_ist, fb_ack_ist = fallback_pair
                    if (
                        fb_req_ist
                        and live_reply_ist
                        and fb_req_ist < live_reply_ist
                        and (live_reply_ist - fb_req_ist) <= timedelta(hours=48)
                    ):
                        ack_pick_ist = live_reply_ist
                        ack_pick_msg = None
                        if fb_ack_ist and fb_req_ist <= fb_ack_ist <= live_reply_ist:
                            ack_pick_ist = fb_ack_ist
                            if fb_ack_ist != live_reply_ist:
                                for e in thread:
                                    e_ist = _email_ist(e)
                                    if not e_ist:
                                        continue
                                    if e_ist.replace(second=0, microsecond=0) != fb_ack_ist.replace(second=0, microsecond=0):
                                        continue
                                    if not _ess_sender(e):
                                        continue
                                    if workbook_kind == "incident_business" and _is_ess_dl_only_reroute(e, ess_team):
                                        continue
                                    if not _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens):
                                        continue
                                    ack_pick_msg = e
                                    break
                        lane_episode = {
                            "request": fb_req_ist,
                            "ack": ack_pick_ist,
                            "ack_msg": ack_pick_msg,
                            "reply": live_reply_ist,
                            "reply_msg": live_reply_msg,
                        }
            if not lane_episode or not lane_episode.get("request"):
                return False

            req_ist = lane_episode.get("request")
            ack_ist = lane_episode.get("ack") or live_reply_ist
            reply_ist = lane_episode.get("reply") or live_reply_ist
            if not (req_ist and ack_ist and reply_ist):
                return False
            if not (req_ist <= ack_ist <= reply_ist):
                return False

            ack_msg = lane_episode.get("ack_msg")
            reply_msg = lane_episode.get("reply_msg") or live_reply_msg
            lane_mode = lane_episode.get("mode") or ("direct-reply" if (ack_msg is None or ack_ist == reply_ist) else "req-ack-reply")
            direct_reply_mode = lane_mode == "direct-reply"
            all_three_same_mode = lane_mode == "all-three-same"

            if all_three_same_mode:
                owner_tag = "LaneLocalInitialEpisode[all-three-same]"
                candidate_kind = "seed_same_time"
                same_time_src = reply_msg.sender_email or reply_msg.sender_name or "ESS_LOCKED_LANE_REPLY"
                created_src = same_time_src
                ack_src = same_time_src
                resolved_src = same_time_src
            else:
                owner_tag = (
                    "LaneLocalInitialEpisode[direct-reply]"
                    if direct_reply_mode
                    else "LaneLocalInitialEpisode[req-ack-reply]"
                )
                candidate_kind = "quoted" if direct_reply_mode else "hybrid"
                created_src = "PARSED_FROM_QUOTED_REQUEST"
                ack_src = (
                    (reply_msg.sender_email or reply_msg.sender_name or "PARSED_FROM_QUOTED_REPLY")
                    if direct_reply_mode
                    else (
                        (ack_msg.sender_email or ack_msg.sender_name)
                        if ack_msg is not None
                        else "PARSED_FROM_QUOTED_ACK"
                    )
                )
                resolved_src = reply_msg.sender_email or reply_msg.sender_name or "LANE_LOCAL_REPLY"

            _trace_focus_row(
                "seed_lane:candidate",
                state=state,
                row_vals=row_vals,
                list_index=list_index,
                mode=lane_mode,
                candidate=f"{_format_time(req_ist)} / {_format_time(ack_ist)} / {_format_time(reply_ist)}",
            )

            if _seeded_lane_should_commit_authoritatively(
                state,
                row_vals,
                list_index,
                lane_mode=lane_mode,
                req_ist=req_ist,
                ack_ist=ack_ist,
                reply_ist=reply_ist,
                created_src=created_src,
                ack_src=ack_src,
            ):
                should_take_seed, seed_reason = True, "seeded_authoritative_local_pair"
            else:
                should_take_seed, seed_reason = _seeded_lane_should_replace_base(
                    state,
                    row_vals,
                    list_index,
                    req_ist,
                    ack_ist,
                    reply_ist,
                    owner_tag=owner_tag,
                    candidate_kind=candidate_kind,
                    created_src=created_src,
                    ack_src=ack_src,
                    resolved_src=resolved_src,
                )
            _trace_focus_row(
                "seed_lane:winner_decision",
                state=state,
                row_vals=row_vals,
                list_index=list_index,
                decision=seed_reason,
                take_seed=should_take_seed,
            )
            applied = False
            if should_take_seed:
                applied = _commit_episode_update(
                    state,
                    row_vals,
                    list_index,
                    row_idx,
                    req_ist,
                    ack_ist,
                    reply_ist,
                    created_src=created_src,
                    ack_src=ack_src,
                    resolved_src=resolved_src,
                    note_suffix=owner_tag,
                    seed_owner=True,
                )

            if applied or _current_row_triplet_ist(row_vals) == (
                req_ist.replace(second=0, microsecond=0),
                ack_ist.replace(second=0, microsecond=0),
                reply_ist.replace(second=0, microsecond=0),
            ):
                state["initial_lane_episode"] = {
                    "created": req_ist,
                    "response": ack_ist,
                    "resolved": reply_ist,
                    "mode": lane_mode,
                    "created_src": created_src,
                    "ack_src": ack_src,
                    "resolved_src": resolved_src,
                    "authoritative": True,
                    "reply_minute": reply_ist.replace(second=0, microsecond=0),
                    "occurrence_key": state.get("occurrence_key"),
                    "service_bucket": state.get("service_bucket"),
                }
                state["seed_locked"] = True

            _trace_focus_row(
                "seed_lane:result",
                state=state,
                row_vals=row_vals,
                list_index=list_index,
                applied=applied,
            )
            return applied

        def _should_promote_lane_seed_primary(state, row_vals, list_index) -> bool:
            initial_lane_episode = state.get("initial_lane_episode") or {}
            if state.get("seed_locked") and initial_lane_episode.get("authoritative"):
                return False
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            if _has_source_locked_same_time(notes_l):
                return False

            if _is_strict_all_live_builder_row(state, row_vals, list_index):
                return False
            if _row_is_special_all_same_risk_family(state, row_vals, list_index):
                return False

            requester = state.get("requester") or ""
            subject_norm_value = (state.get("subject_norm") or "").lower()
            if not requester or not subject_norm_value:
                return False

            live_reply_ist = _live_reply_lane_for_row(list_index)
            live_reply_msg = _live_reply_message_for_row(list_index)
            if not (live_reply_ist and live_reply_msg):
                return False

            profile = _rewrite_guard_profile(row_vals, list_index)
            weak_base = bool(
                profile.get("low_confidence")
                or profile.get("all_same")
                or "created retained (response anchor unreliable)" in notes_l
                or "created_clamped_to_first" in (profile.get("created_src") or "").lower()
            )
            if not weak_base:
                return False
            current_r_min = profile.get("r_min")
            live_reply_min = live_reply_ist.replace(second=0, microsecond=0)
            if current_r_min and current_r_min != live_reply_min:
                return False

            lane_ctx = _seed_locked_lane_context(state, row_vals, list_index)
            if not lane_ctx:
                return False

            lane_episode = _seed_lane_episode_from_locked_reply(
                lane_ctx["live_reply_msg"],
                lane_ctx["live_reply_ist"],
                lane_ctx["thread"],
                lane_ctx["requester"],
                lane_ctx["subject_norm"],
                lane_ctx["row_tokens"],
                lane_ctx["row_id_tokens"],
                allow_strict_same_time=bool((state.get("group_total") or 0) <= 1),
            )
            if not lane_episode or not lane_episode.get("request"):
                return False

            req_ist = lane_episode.get("request")
            ack_ist = lane_episode.get("ack") or lane_ctx["live_reply_ist"]
            reply_ist = lane_episode.get("reply") or lane_ctx["live_reply_ist"]
            if not (req_ist and ack_ist and reply_ist):
                return False
            if not (req_ist <= ack_ist <= reply_ist):
                return False
            if reply_ist.replace(second=0, microsecond=0) != live_reply_min:
                return False

            current_triplet = _current_row_triplet_ist(row_vals)
            candidate_triplet = (
                req_ist.replace(second=0, microsecond=0),
                ack_ist.replace(second=0, microsecond=0),
                reply_ist.replace(second=0, microsecond=0),
            )
            if current_triplet and current_triplet == candidate_triplet:
                return False
            if current_triplet and candidate_triplet[0] >= current_triplet[0]:
                return False
            return True

        def _email_for_path(path: str):
            if not path:
                return None
            if path in email_by_path_cache:
                return email_by_path_cache[path]
            found = None
            for e in emails:
                e_path = str(getattr(e, "path", "") or "")
                if e_path and e_path == path:
                    found = e
                    break
            email_by_path_cache[path] = found
            return found

        def _get_eml_header_summary(path: str):
            if not path:
                return None
            if path in raw_eml_header_summary_cache:
                return raw_eml_header_summary_cache[path]
            try:
                with open(path, "rb") as f:
                    msg = BytesParser(policy=policy.default).parse(f)
            except Exception:
                raw_eml_header_summary_cache[path] = None
                return None
            try:
                sent_dt = parsedate_to_datetime(msg.get("Date")) if msg.get("Date") else None
            except Exception:
                sent_dt = None
            subject_raw = str(msg.get("Subject") or "")
            subject_norm = normalize_subject(subject_raw)
            subject_ids = _id_like_tokens(subject_norm)
            sender_email = ""
            sender_name = ""
            try:
                from_header = msg.get("From")
                if getattr(from_header, "addresses", None):
                    sender_email = (from_header.addresses[0].addr_spec or "").lower()
                    sender_name = from_header.addresses[0].display_name or ""
                else:
                    sender_name = str(from_header or "")
            except Exception:
                sender_name = str(msg.get("From") or "")
            summary = {
                "sent_dt": sent_dt,
                "subject_raw": subject_raw,
                "subject_norm": subject_norm,
                "subject_ids": subject_ids,
                "sender_email": sender_email,
                "sender_name": sender_name,
            }
            raw_eml_header_summary_cache[path] = summary
            return summary

        def _final_confident_eml_episode(
            subject_norm_value: str,
            row_tokens: set,
            row_id_tokens_set: set,
            requester_name: str,
            merged_sources: list,
            quoted_pick_idx: int,
            total_rows: int,
            target_reply_ist,
        ):
            path_candidates = []
            seen_paths = set()
            for src in merged_sources or []:
                src_path = str(getattr(src, "path", "") or "")
                if src_path and src_path not in seen_paths:
                    seen_paths.add(src_path)
                    path_candidates.append(src_path)
            for path in _find_eml_paths_by_id(row_id_tokens_set):
                if path and path not in seen_paths:
                    seen_paths.add(path)
                    path_candidates.append(path)
            cache_key = (
                subject_norm_value or "",
                requester_name or "",
                tuple(path_candidates),
                tuple(sorted(row_id_tokens_set)) if row_id_tokens_set else (),
                int(quoted_pick_idx or 0),
                int(total_rows or 0),
                (
                    target_reply_ist.replace(second=0, microsecond=0).isoformat()
                    if target_reply_ist else ""
                ),
            )
            cached = raw_eml_episode_fallback_cache.get(cache_key)
            if cached is not None:
                return cached
            if not path_candidates:
                raw_eml_episode_fallback_cache[cache_key] = None
                return None

            candidate_episodes = []
            for path in path_candidates:
                header_summary = _get_eml_header_summary(path)
                if not header_summary or not header_summary.get("sent_dt"):
                    continue
                header_ist = _to_ist(header_summary.get("sent_dt"))
                if not header_ist:
                    continue
                header_subject_norm = header_summary.get("subject_norm") or normalize_subject(header_summary.get("subject_raw") or "")
                header_ids = header_summary.get("subject_ids") or _id_like_tokens(header_subject_norm)
                header_tokens = _match_tokens(header_subject_norm)
                if not _quoted_subject_confirms_row(
                    header_subject_norm,
                    header_ids,
                    header_tokens,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens_set,
                ):
                    continue
                header_email = _email_for_path(path)
                sender_email = (header_summary.get("sender_email") or "").lower()
                sender_domain = sender_email.split("@", 1)[-1] if "@" in sender_email else ""
                is_ess_header = False
                if header_email is not None:
                    is_ess_header = _ess_sender(header_email)
                else:
                    is_ess_header = bool(
                        (sender_email and sender_email in ess_email_set)
                        or (sender_domain and sender_domain in ess_domain_set)
                    )
                if not is_ess_header:
                    continue
                if header_email is not None:
                    ack_like_or_nonrequester = (
                        _ack_like(header_email)
                        or _ack_like_text_fallback(header_email)
                        or _ess_only_short_ack(header_email)
                        or (not _req_match(header_email, requester_name))
                    )
                    if not ack_like_or_nonrequester:
                        continue

                quoted_req_candidates = []
                for from_line, sent_ist, q_subj in _get_quoted_blocks_from_eml_path(path):
                    q_norm = normalize_subject(q_subj or "")
                    q_ids = _id_like_tokens(q_norm)
                    q_tokens = _match_tokens(q_norm)
                    if q_subj and not _quoted_subject_confirms_row(
                        q_norm,
                        q_ids,
                        q_tokens,
                        subject_norm_value,
                        row_tokens,
                        row_id_tokens_set,
                    ):
                        continue
                    is_ess = _quoted_from_line_is_ess_shared(from_line)
                    if is_ess:
                        continue
                    if sent_ist >= header_ist:
                        continue
                    if (header_ist - sent_ist) > timedelta(minutes=16):
                        continue
                    quoted_req_candidates.append(sent_ist)
                if not quoted_req_candidates:
                    continue
                req_ist = _best_req_before_ack(sorted(quoted_req_candidates), header_ist)
                if not req_ist:
                    continue

                reply_matches = []
                for e in merged_sources or []:
                    e_ist = _email_ist(e)
                    if not e_ist or e_ist <= header_ist:
                        continue
                    if (e_ist - header_ist) > timedelta(hours=48):
                        continue
                    if not _ess_sender(e):
                        continue
                    flags = _shared_reply_flags(email_record=e)
                    if flags["ack_candidate"]:
                        continue
                    if flags["ignore_reply"]:
                        continue
                    parent_ess, has_non_ess = _parent_sender_info(e)
                    if (parent_ess is True) and (not has_non_ess):
                        continue
                    if not _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens_set):
                        continue
                    reply_matches.append(e)
                if reply_matches:
                    if target_reply_ist:
                        reply_matches.sort(
                            key=lambda e: (
                                abs((_email_ist(e) - target_reply_ist).total_seconds()),
                                _email_ist(e),
                            )
                        )
                    else:
                        reply_matches.sort(key=lambda e: _email_ist(e))
                    reply_pick = reply_matches[0]
                else:
                    reply_pick = header_email

                if reply_pick is None:
                    continue
                candidate_episodes.append(
                    {
                        "req_ist": req_ist,
                        "ack_ist": header_ist,
                        "reply": reply_pick,
                        "req_src": "PARSED_FROM_RAW_EML_REQUEST",
                    }
                )

            if not candidate_episodes:
                raw_eml_episode_fallback_cache[cache_key] = None
                return None

            deduped = []
            seen_episode_minutes = set()
            for episode in sorted(candidate_episodes, key=lambda ep: (ep["ack_ist"], ep["req_ist"])):
                minute_key = (
                    episode["req_ist"].replace(second=0, microsecond=0),
                    episode["ack_ist"].replace(second=0, microsecond=0),
                )
                if minute_key in seen_episode_minutes:
                    continue
                seen_episode_minutes.add(minute_key)
                deduped.append(episode)

            if total_rows and len(deduped) <= quoted_pick_idx:
                raw_eml_episode_fallback_cache[cache_key] = None
                return None

            if target_reply_ist:
                deduped.sort(
                    key=lambda ep: (
                        abs((_email_ist(ep["reply"]) - target_reply_ist).total_seconds()) if _email_ist(ep["reply"]) else float("inf"),
                        ep["ack_ist"],
                    )
                )
            else:
                deduped.sort(key=lambda ep: ep["ack_ist"])

            chosen = deduped[quoted_pick_idx] if total_rows and quoted_pick_idx < len(deduped) else deduped[0]
            raw_eml_episode_fallback_cache[cache_key] = chosen
            return chosen

        def _build_eml_id_index():
            nonlocal eml_id_index
            if eml_id_index is not None:
                return eml_id_index
            eml_id_index = {}
            all_row_id_tokens = set()
            for st in row_states:
                subj_norm = (st.get("subject_norm") or "").lower()
                toks = _id_like_tokens(subj_norm)
                if not toks:
                    li = st.get("list_index")
                    desc = ""
                    if li is not None and li < len(automation_rows):
                        desc = automation_rows[li].get("Description") or ""
                    if not desc:
                        desc = st.get("description") or ""
                    toks = _id_like_tokens(desc or "")
                all_row_id_tokens.update(toks)
            if not all_row_id_tokens:
                return eml_id_index
            try:
                for p in eml_root.rglob("*.eml"):
                    path = str(p)
                    header_summary = _get_eml_header_summary(path)
                    path_tokens = set()
                    if header_summary:
                        path_tokens.update((header_summary.get("subject_ids") or set()) & all_row_id_tokens)
                    summaries = _get_quoted_summaries_from_eml_path(path)
                    if not summaries:
                        if not path_tokens:
                            continue
                    for _sent_ist, _q_norm, q_ids, _q_tokens, _is_ess in summaries:
                        path_tokens.update(q_ids & all_row_id_tokens)
                    if not path_tokens:
                        continue
                    for tok in path_tokens:
                        eml_id_index.setdefault(tok, set()).add(path)
            except Exception:
                eml_id_index = {}
            if eml_id_index:
                eml_id_index = {k: list(v) for k, v in eml_id_index.items()}
            return eml_id_index

        def _find_eml_paths_by_id(row_id_tokens_set: set):
            if not row_id_tokens_set:
                return []
            key = tuple(sorted(row_id_tokens_set))
            if key in raw_id_path_cache:
                return raw_id_path_cache[key]
            hits = []
            idx = _build_eml_id_index()
            if idx:
                seen = set()
                for tok in row_id_tokens_set:
                    for path in idx.get(tok, []):
                        if path not in seen:
                            seen.add(path)
                            hits.append(path)
            raw_id_path_cache[key] = hits
            return hits
        _stage_timer_stop("blue_gap_final_audit", blue_gap_audit_started_at, items=len(row_states))
        quoted_only_tag_started_at = _stage_timer_start()
        relaxed_quoted_block_cache = {}
        quoted_request_context_cache = {}
        quoted_pair_analysis_cache = {}
        quoted_hybrid_analysis_cache = {}
        quoted_row_partition_cache = {}

        def _get_relaxed_quoted_blocks_cached(msg):
            key = _email_stable_key(msg)
            if key in relaxed_quoted_block_cache:
                return relaxed_quoted_block_cache[key]
            blocks = _extract_quoted_blocks_relaxed(msg)
            relaxed_quoted_block_cache[key] = blocks
            return blocks

        def _row_match_context(state, row_vals, list_index):
            subject_norm_value = (state.get("subject_norm") or "").lower()
            requester_name = state.get("requester") or ""
            desc_text = ""
            if row_vals:
                desc_text = row_vals.get("Description") or ""
            if not desc_text:
                desc_text = state.get("description") or ""
            c_dt = _parse_time_str(row_vals.get("Created Date & Time")) if row_vals else None
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time")) if row_vals else None
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time")) if row_vals else None
            cache_key = (
                list_index,
                subject_norm_value,
                requester_name,
                desc_text,
                str(row_vals.get("Created Date & Time") or "") if row_vals else "",
                str(row_vals.get("Actual Response Date & Time") or "") if row_vals else "",
                str(row_vals.get("Actual Resolved Date & Time") or "") if row_vals else "",
            )
            cached = row_match_context_cache.get(cache_key)
            if cached is not None:
                return cached

            row_tokens = _match_tokens(subject_norm_value)
            row_id_tokens = _id_like_tokens(subject_norm_value)
            if not row_id_tokens:
                row_id_tokens = _id_like_tokens(desc_text or "")
            ref_ist = _to_ist(r_dt) if r_dt else (_to_ist(a_dt) if a_dt else _to_ist(c_dt) if c_dt else None)
            out = {
                "subject_norm": subject_norm_value,
                "requester": requester_name,
                "row_tokens": row_tokens,
                "row_id_tokens": row_id_tokens,
                "ref_ist": ref_ist,
                "created_dt": c_dt,
                "ack_dt": a_dt,
                "resolved_dt": r_dt,
            }
            row_match_context_cache[cache_key] = out
            return out

        def _get_quoted_row_partitions_cached(
            msg,
            subject_norm_value,
            row_tokens,
            row_id_tokens,
            *,
            allow_raw_id_fallback=False,
            use_path_fallback=False,
            use_relaxed_fallback=False,
        ):
            cache_key = (
                id(msg),
                subject_norm_value or "",
                tuple(sorted(row_tokens)) if row_tokens else (),
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
                bool(allow_raw_id_fallback),
                bool(use_path_fallback),
                bool(use_relaxed_fallback),
            )
            cached = quoted_row_partition_cache.get(cache_key)
            if cached is not None:
                return cached

            quoted_blocks = _get_quoted_blocks_with_subject_cached(msg)
            if (not quoted_blocks) and use_path_fallback:
                quoted_blocks = _get_quoted_blocks_from_eml_path(getattr(msg, "path", ""))
            if (not quoted_blocks) and use_relaxed_fallback:
                quoted_blocks = _get_relaxed_quoted_blocks_cached(msg)

            q_non_ess = []
            q_ess = []
            for from_line, sent_ist, q_subj in quoted_blocks or ():
                if not _quoted_block_matches_row(
                    msg,
                    q_subj,
                    subject_norm_value,
                    row_tokens,
                    row_id_tokens,
                    allow_raw_id_fallback=allow_raw_id_fallback,
                ):
                    continue
                if _quoted_from_line_is_ess_shared(from_line):
                    q_ess.append(sent_ist)
                else:
                    q_non_ess.append(sent_ist)

            out = (tuple(q_non_ess), tuple(q_ess))
            quoted_row_partition_cache[cache_key] = out
            return out

        def _merge_subject_filtered_msgs(
            existing_msgs,
            extra_msgs,
            subject_norm_value,
            row_tokens,
            row_id_tokens,
            *,
            strict_quoted_variant=False,
        ):
            merged_msgs = list(existing_msgs or [])
            merged_ids = {id(e) for e in merged_msgs}
            for e in extra_msgs or ():
                if id(e) in merged_ids:
                    continue
                match_ok = (
                    _row_subject_match_email_quoted(e, subject_norm_value, row_tokens, row_id_tokens)
                    if strict_quoted_variant
                    else _row_subject_match_email(e, subject_norm_value, row_tokens, row_id_tokens)
                )
                if not match_ok:
                    continue
                merged_msgs.append(e)
                merged_ids.add(id(e))
            return merged_msgs

        email_subject_id_index = None

        def _build_email_subject_id_index():
            nonlocal email_subject_id_index
            if email_subject_id_index is not None:
                return email_subject_id_index
            email_subject_id_index = {}
            all_row_id_tokens = set()
            for st in row_states:
                subj_norm = (st.get("subject_norm") or "").lower()
                toks = _id_like_tokens(subj_norm)
                if not toks:
                    li = st.get("list_index")
                    desc = ""
                    if li is not None and li < len(automation_rows):
                        desc = automation_rows[li].get("Description") or ""
                    if not desc:
                        desc = st.get("description") or ""
                    toks = _id_like_tokens(desc or "")
                all_row_id_tokens.update(toks)
            if not all_row_id_tokens:
                return email_subject_id_index

            for e in emails:
                subj_raw = getattr(e, "subject", "") or ""
                subj_norm = _subject_norm_cached(subj_raw)
                email_tokens = set(_id_like_tokens(subj_norm) & all_row_id_tokens)
                if not email_tokens:
                    email_tokens.update(_id_like_tokens(subj_raw) & all_row_id_tokens)
                if not email_tokens:
                    continue
                for tok in email_tokens:
                    email_subject_id_index.setdefault(tok, []).append(e)
            return email_subject_id_index

        def _find_emails_by_id(row_id_tokens_set: set):
            if not row_id_tokens_set:
                return []
            idx = _build_email_subject_id_index()
            if not idx:
                return []
            out = []
            seen = set()
            for tok in row_id_tokens_set:
                for e in idx.get(tok, []):
                    key = id(e)
                    if key in seen:
                        continue
                    seen.add(key)
                    out.append(e)
            return out

        def _quoted_request_row_context(
            subject_norm_value,
            base_thread,
            requester_name,
            ref_ist,
            row_tokens,
            row_id_tokens,
            *,
            allow_broad_subject_fallback=True,
        ):
            ref_day = ""
            if ref_ist:
                try:
                    ref_day = ref_ist.date().isoformat()
                except Exception:
                    ref_day = ""
            ctx_key = (
                subject_norm_value or "",
                requester_name or "",
                ref_day,
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
                bool(allow_broad_subject_fallback),
            )
            cached = quoted_request_context_cache.get(ctx_key)
            if cached is not None:
                return cached

            thread = _expanded_thread(
                subject_norm_value,
                base_thread,
                requester_name,
                include_non_ess=True,
                reference_ist=ref_ist,
            )
            thread = thread or []
            merged_msgs = list(thread)
            merged_ids = {id(e) for e in merged_msgs}
            thread_ids = set(merged_ids)

            if row_id_tokens:
                id_matches = _find_emails_by_id(row_id_tokens)
                if id_matches:
                    merged_msgs = _merge_subject_filtered_msgs(
                        merged_msgs,
                        id_matches,
                        subject_norm_value,
                        row_tokens,
                        row_id_tokens,
                        strict_quoted_variant=True,
                    )
                    merged_ids = {id(e) for e in merged_msgs}
                    thread_ids = set(id(e) for e in thread)

            has_live_non_ess = False
            for e in thread:
                if not _req_match(e, requester_name):
                    continue
                if not _ess_sender(e):
                    has_live_non_ess = True
                    break

            # Broad subject-only mailbox expansion is expensive. When the row has
            # concrete ID tokens, later stages already use targeted ID lookups; and
            # when the current thread already contains a live non-ESS request, the
            # quoted-only stage will skip anyway.
            if allow_broad_subject_fallback and (not row_id_tokens) and (not has_live_non_ess):
                extra_any = _requester_pool(subject_norm_value, "", ref_ist, day_window=30) or []
                if extra_any:
                    merged_msgs = _merge_subject_filtered_msgs(
                        merged_msgs,
                        extra_any,
                        subject_norm_value,
                        row_tokens,
                        row_id_tokens,
                        strict_quoted_variant=True,
                    )
                    merged_ids = {id(e) for e in merged_msgs}

            if not has_live_non_ess:
                for e in merged_msgs:
                    if id(e) in thread_ids:
                        continue
                    if not _req_match(e, requester_name):
                        continue
                    if not _ess_sender(e):
                        has_live_non_ess = True
                        break

            out = (merged_msgs, has_live_non_ess)
            quoted_request_context_cache[ctx_key] = out
            return out

        def _quoted_request_pair_analysis(merged_msgs, subject_norm, row_tokens, row_id_tokens, *, quoted_gap_mins=16, quoted_day_slack=0):
            quoted_pair_key = (
                tuple(id(e) for e in merged_msgs),
                subject_norm,
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
                int(quoted_gap_mins),
                int(quoted_day_slack),
            )
            cached_pair_analysis = quoted_pair_analysis_cache.get(quoted_pair_key)
            if cached_pair_analysis is not None:
                has_pair, has_wide_pair, latest_quoted_req, quoted_non_ess_all = cached_pair_analysis
                return has_pair, has_wide_pair, latest_quoted_req, list(quoted_non_ess_all)

            has_pair = False
            has_wide_pair = False
            latest_quoted_req = None
            quoted_non_ess_all = []
            for msg in merged_msgs:
                q_non_ess, q_ess = _get_quoted_row_partitions_cached(
                    msg,
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                    allow_raw_id_fallback=True,
                    use_path_fallback=True,
                    use_relaxed_fallback=True,
                )
                if q_non_ess:
                    quoted_non_ess_all.extend(q_non_ess)
                    latest_q = max(q_non_ess)
                    if (latest_quoted_req is None) or (latest_q > latest_quoted_req):
                        latest_quoted_req = latest_q
                if q_non_ess and q_ess:
                    sorted_non_ess = sorted(q_non_ess)
                    sorted_ess = sorted(q_ess)
                    for ack_ist in sorted_ess:
                        reqs = [
                            r for r in sorted_non_ess
                            if r < ack_ist and abs((ack_ist.date() - r.date()).days) <= quoted_day_slack
                        ]
                        if not reqs:
                            continue
                        req_ist = _best_req_before_ack(reqs, ack_ist)
                        gap = ack_ist - req_ist
                        if gap <= timedelta(minutes=quoted_gap_mins):
                            has_pair = True
                            break
                        has_wide_pair = True
                if has_pair:
                    break

            if (not has_pair) and row_id_tokens:
                for msg in merged_msgs:
                    if not _row_id_in_raw(msg, row_id_tokens):
                        continue
                    q_non_ess = []
                    q_ess = []
                    for from_line, sent_ist, _q_subj in _get_relaxed_quoted_blocks_cached(msg):
                        if _quoted_from_line_is_ess_shared(from_line):
                            q_ess.append(sent_ist)
                        else:
                            q_non_ess.append(sent_ist)
                    if q_non_ess:
                        quoted_non_ess_all.extend(q_non_ess)
                        latest_q = max(q_non_ess)
                        if (latest_quoted_req is None) or (latest_q > latest_quoted_req):
                            latest_quoted_req = latest_q
                    if q_non_ess and q_ess:
                        sorted_non_ess = sorted(q_non_ess)
                        sorted_ess = sorted(q_ess)
                        for ack_ist in sorted_ess:
                            reqs = [r for r in sorted_non_ess if r < ack_ist and abs((ack_ist.date() - r.date()).days) <= quoted_day_slack]
                            if not reqs:
                                continue
                            req_ist = _best_req_before_ack(reqs, ack_ist)
                            gap = ack_ist - req_ist
                            if gap <= timedelta(minutes=quoted_gap_mins):
                                has_pair = True
                                break
                            has_wide_pair = True
                    if has_pair:
                        break

            if (not has_pair) and row_id_tokens:
                raw_paths = _find_eml_paths_by_id(row_id_tokens)
                for path in raw_paths:
                    quoted_summaries = _get_quoted_summaries_from_eml_path(path)
                    if not quoted_summaries:
                        continue
                    q_non_ess = []
                    q_ess = []
                    for sent_ist, q_norm, q_ids, q_tokens, is_ess in quoted_summaries:
                        if not _quoted_subject_confirms_row(
                            q_norm,
                            q_ids,
                            q_tokens,
                            subject_norm,
                            row_tokens,
                            row_id_tokens,
                        ):
                            continue
                        if is_ess:
                            q_ess.append(sent_ist)
                        else:
                            q_non_ess.append(sent_ist)
                            quoted_non_ess_all.append(sent_ist)
                    if q_non_ess:
                        latest_q = max(q_non_ess)
                        if (latest_quoted_req is None) or (latest_q > latest_quoted_req):
                            latest_quoted_req = latest_q
                    if q_non_ess and q_ess:
                        sorted_non_ess = sorted(q_non_ess)
                        sorted_ess = sorted(q_ess)
                        for ack_ist in sorted_ess:
                            reqs = [r for r in sorted_non_ess if r < ack_ist and abs((ack_ist.date() - r.date()).days) <= quoted_day_slack]
                            if not reqs:
                                continue
                            req_ist = _best_req_before_ack(reqs, ack_ist)
                            gap = ack_ist - req_ist
                            if gap <= timedelta(minutes=quoted_gap_mins):
                                has_pair = True
                                break
                            has_wide_pair = True
                    if has_pair:
                        break

            cached_pair_analysis = (has_pair, has_wide_pair, latest_quoted_req, tuple(quoted_non_ess_all))
            quoted_pair_analysis_cache[quoted_pair_key] = cached_pair_analysis
            return has_pair, has_wide_pair, latest_quoted_req, list(quoted_non_ess_all)

        def _ess_only_msg_partition(
            merged_msgs,
            *,
            subject_norm,
            row_tokens,
            row_id_tokens,
            requester,
            quoted_only,
            ess_email_set,
            ess_domain_set,
        ):
            cache_key = (
                tuple(id(e) for e in merged_msgs),
                subject_norm or "",
                tuple(sorted(row_tokens)) if row_tokens else (),
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
                requester or "",
                bool(quoted_only),
            )
            cached = ess_only_msg_partition_cache.get(cache_key)
            if cached is not None:
                return cached

            consultant_msgs = []
            ess_reply_msgs = []
            observed_ess_name_tokens = set()
            for e in merged_msgs:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                if row_id_tokens:
                    s_ids = _id_like_tokens(s_norm)
                    if (not s_ids) or row_id_tokens.isdisjoint(s_ids):
                        continue
                elif row_tokens:
                    s_tokens = _match_tokens(s_norm)
                    score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                    contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                    if score < 0.45 and not contains:
                        continue

                is_acky = (not quoted_only) and (
                    _ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e)
                )
                if _ess_sender(e) and not is_acky:
                    ess_reply_msgs.append(e)
                if _req_match(e, requester) and not is_acky:
                    consultant_msgs.append(e)

                sender_email = (getattr(e, "sender_email", "") or "").lower()
                sender_domain = sender_email.split("@", 1)[-1] if "@" in sender_email else ""
                if (sender_email in ess_email_set) or (sender_domain in ess_domain_set):
                    name_raw = (getattr(e, "sender_name", "") or "").lower()
                    if name_raw:
                        for tok in re.split(r"[^a-z0-9]+", name_raw):
                            if len(tok) >= 3:
                                observed_ess_name_tokens.add(tok)

            consultant_msgs.sort(key=lambda e: e.sent_time)
            ess_reply_msgs.sort(key=lambda e: e.sent_time)
            out = (consultant_msgs, ess_reply_msgs, tuple(sorted(observed_ess_name_tokens)))
            ess_only_msg_partition_cache[cache_key] = out
            return out

        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            row_vals = automation_rows[list_index]
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            if "quotedrequestonly" in notes_l:
                quoted_request_only.add(list_index)
                continue
            if _is_strict_all_live_builder_row(state, row_vals, list_index):
                continue
            row_ctx = _row_match_context(state, row_vals, list_index)
            requester = row_ctx["requester"]
            if not requester:
                continue
            subject_norm = row_ctx["subject_norm"]
            base_thread = state.get("thread") or []
            c_dt = row_ctx["created_dt"]
            a_dt = row_ctx["ack_dt"]
            r_dt = row_ctx["resolved_dt"]
            ref_ist = row_ctx["ref_ist"]

            # Tag only when a real quoted non-ESS request + ESS ack pair exists.
            # Keep strict ack gap for validity; if a wider gap exists, mark blue.
            quoted_gap_mins = 16
            quoted_day_slack = 0
            row_tokens = row_ctx["row_tokens"]
            row_id_tokens = row_ctx["row_id_tokens"]
            merged_msgs, has_live_non_ess = _quoted_request_row_context(
                subject_norm,
                base_thread,
                requester,
                ref_ist,
                row_tokens,
                row_id_tokens,
            )
            if has_live_non_ess:
                continue
            quoted_pair_key = (
                tuple(id(e) for e in merged_msgs),
                subject_norm,
                tuple(sorted(row_id_tokens)) if row_id_tokens else (),
            )
            has_pair, has_wide_pair, latest_quoted_req, quoted_non_ess_all = _quoted_request_pair_analysis(
                merged_msgs,
                subject_norm,
                row_tokens,
                row_id_tokens,
                quoted_gap_mins=quoted_gap_mins,
                quoted_day_slack=quoted_day_slack,
            )
            if (
                (not has_pair)
                and (not has_wide_pair)
                and (not quoted_non_ess_all)
                and (not row_id_tokens)
                and (not has_live_non_ess)
            ):
                widened_msgs = _merge_subject_filtered_msgs(
                    merged_msgs,
                    _requester_pool(subject_norm, "", ref_ist, day_window=30) or [],
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                    strict_quoted_variant=True,
                )
                if len(widened_msgs) != len(merged_msgs):
                    merged_msgs = widened_msgs
                    quoted_pair_key = (
                        tuple(id(e) for e in merged_msgs),
                        subject_norm,
                        tuple(sorted(row_id_tokens)) if row_id_tokens else (),
                    )
                    has_pair, has_wide_pair, latest_quoted_req, quoted_non_ess_all = _quoted_request_pair_analysis(
                        merged_msgs,
                        subject_norm,
                        row_tokens,
                        row_id_tokens,
                        quoted_gap_mins=quoted_gap_mins,
                        quoted_day_slack=quoted_day_slack,
                    )
            # Hybrid: quoted request + live ESS reply within ack window.
            hybrid_pair = None
            if (not has_pair) and quoted_non_ess_all:
                occ_meta = _shared_occurrence_pick(
                    state,
                    subject_norm_value=subject_norm,
                    requester_value=requester,
                    current_created_ist=None,
                    current_ack_ist=None,
                    current_resolved_ist=None,
                    default_idx=0,
                )
                quoted_hybrid_key = quoted_pair_key + (
                    requester,
                    int(occ_meta.get("total_rows") or 0),
                )
                cached_hybrid_analysis = quoted_hybrid_analysis_cache.get(quoted_hybrid_key)
                if cached_hybrid_analysis is None:
                    live_ack = None
                    live_req = latest_quoted_req
                    live_req_src = "PARSED_FROM_QUOTED_REQUEST"
                    live_req_debug = "quoted"
                    live_candidates = []
                    quoted_candidates_all = sorted({q for q in quoted_non_ess_all if q is not None})
                    total_rows_for_subject = occ_meta["total_rows"]
                    reply_led_candidate = None
                    if total_rows_for_subject <= 1:
                        resolved_candidates = []
                        for cand in merged_msgs:
                            cand_ist = _email_ist(cand)
                            if not cand_ist:
                                continue
                            if not _req_match(cand, requester):
                                continue
                            if _ack_like(cand) or _ack_like_text_fallback(cand) or _ess_only_short_ack(cand):
                                continue
                            if not _row_subject_match_email_quoted(cand, subject_norm, row_tokens, row_id_tokens):
                                continue
                            resolved_candidates.append(cand)
                        resolved_candidates.sort(key=lambda e: e.sent_time)
                        if resolved_candidates:
                            target_reply = resolved_candidates[-1]
                            target_reply_ist = _email_ist(target_reply)
                            best_reply_led = None
                            best_reply_led_score = None
                            if target_reply_ist:
                                for ack_msg in merged_msgs:
                                    ack_ist = _email_ist(ack_msg)
                                    if not ack_ist or ack_ist >= target_reply_ist:
                                        continue
                                    if (target_reply_ist - ack_ist) > timedelta(hours=48):
                                        continue
                                    if not _ess_sender(ack_msg):
                                        continue
                                    if not _row_subject_match_email_quoted(ack_msg, subject_norm, row_tokens, row_id_tokens):
                                        continue

                                    quoted_req_candidates = [
                                        q for q in quoted_candidates_all
                                        if q < ack_ist and (ack_ist - q) <= timedelta(minutes=quoted_gap_mins)
                                    ]
                                    quoted_req = _best_req_before_ack(quoted_req_candidates, ack_ist) if quoted_req_candidates else None

                                    live_non_ess_candidates = []
                                    for cand in merged_msgs:
                                        cand_ist = _email_ist(cand)
                                        if not cand_ist or cand_ist >= ack_ist:
                                            continue
                                        if (ack_ist - cand_ist) > timedelta(minutes=quoted_gap_mins):
                                            continue
                                        if _ess_sender(cand) or _system_like_sender(cand):
                                            continue
                                        if not _row_subject_match_email_quoted(cand, subject_norm, row_tokens, row_id_tokens):
                                            continue
                                        live_non_ess_candidates.append((cand_ist, cand))
                                    live_non_ess_req = None
                                    if live_non_ess_candidates:
                                        live_req_times = [t for t, _ in live_non_ess_candidates]
                                        chosen_live_req_ist = _best_req_before_ack(live_req_times, ack_ist)
                                        if chosen_live_req_ist:
                                            for cand_ist, cand in live_non_ess_candidates:
                                                if cand_ist == chosen_live_req_ist:
                                                    live_non_ess_req = cand
                                                    break

                                    chosen_req = None
                                    chosen_req_src = None
                                    chosen_mode = None
                                    if live_non_ess_req and _email_ist(live_non_ess_req):
                                        chosen_req = _email_ist(live_non_ess_req)
                                        chosen_req_src = live_non_ess_req.sender_email or live_non_ess_req.sender_name
                                        chosen_mode = "reply-led-live-non-ess"
                                    elif quoted_req:
                                        chosen_req = quoted_req
                                        chosen_req_src = "PARSED_FROM_QUOTED_REQUEST"
                                        chosen_mode = "reply-led-quoted"

                                    if not chosen_req:
                                        continue

                                    score = (
                                        abs((target_reply_ist - ack_ist).total_seconds()),
                                        0 if chosen_mode == "reply-led-live-non-ess" else 1,
                                        abs((ack_ist - chosen_req).total_seconds()),
                                    )
                                    if best_reply_led_score is None or score < best_reply_led_score:
                                        best_reply_led_score = score
                                        best_reply_led = (chosen_req, ack_ist, chosen_req_src, chosen_mode)

                            if best_reply_led:
                                live_req, live_ack, live_req_src, live_req_debug = best_reply_led
                                reply_led_candidate = (live_ack, target_reply, live_req)
                                live_candidates.append(reply_led_candidate)
                    if reply_led_candidate is None:
                        for e in merged_msgs:
                            if not _req_match(e, requester):
                                continue
                            if not _ess_sender(e):
                                continue
                            e_ist = _email_ist(e)
                            if not e_ist:
                                continue
                            req_candidates = [
                                q for q in quoted_candidates_all
                                if q < e_ist and (e_ist - q) <= timedelta(minutes=quoted_gap_mins)
                            ]
                            if not req_candidates:
                                continue
                            strict_req_candidates = [
                                q for q in req_candidates
                                if q.replace(second=0, microsecond=0) != e_ist.replace(second=0, microsecond=0)
                            ]
                            chosen_req = max(strict_req_candidates or req_candidates)
                            live_candidates.append((e_ist, e, chosen_req))
                    if live_candidates:
                        if reply_led_candidate is not None:
                            live_candidates = [reply_led_candidate]
                        else:
                            live_candidates.sort(key=lambda x: x[0])
                        live_ack, live_ack_msg, chosen_req = live_candidates[0]
                        live_req = chosen_req
                        if not live_req_debug.startswith("reply-led"):
                            live_req_src = "PARSED_FROM_QUOTED_REQUEST"
                            live_req_debug = "quoted-window-before-ack"
                        live_non_ess_req = None
                        for cand in merged_msgs:
                            cand_ist = _email_ist(cand)
                            if not cand_ist:
                                continue
                            if cand_ist >= live_ack:
                                continue
                            if (live_ack - cand_ist) > timedelta(minutes=quoted_gap_mins):
                                continue
                            if _ess_sender(cand) or _system_like_sender(cand):
                                continue
                            if not _row_subject_match_email_quoted(cand, subject_norm, row_tokens, row_id_tokens):
                                continue
                            if (live_non_ess_req is None) or (cand_ist > _email_ist(live_non_ess_req)):
                                live_non_ess_req = cand
                        if (not live_req_debug.startswith("reply-led")) and live_non_ess_req and _email_ist(live_non_ess_req):
                            live_req = _email_ist(live_non_ess_req)
                            live_req_src = live_non_ess_req.sender_email or live_non_ess_req.sender_name
                            live_req_debug = "live-non-ess"
                    cached_hybrid_analysis = (
                        bool(live_ack),
                        live_req,
                        live_ack,
                        live_req_src,
                        live_req_debug,
                    )
                    quoted_hybrid_analysis_cache[quoted_hybrid_key] = cached_hybrid_analysis
                else:
                    has_hybrid_pair, live_req, live_ack, live_req_src, live_req_debug = cached_hybrid_analysis
                if live_ack:
                    has_pair = True
                    hybrid_pair = (live_req, live_ack)
                    quoted_only_hybrid_pairs[list_index] = hybrid_pair
                    quoted_only_hybrid_req_sources[list_index] = live_req_src
                    quoted_only_hybrid_req_debug[list_index] = live_req_debug
            # Note: full-mailbox quoted scans disabled for performance.
            if has_pair:
                quoted_request_only.add(list_index)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedRequestOnly"
            elif has_wide_pair:
                # Found a quoted pair but gap is too large: mark blue for audit.
                _set_row_fill(row_idx, blue_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedPairGap>16m"

        _stage_timer_stop("quoted_request_only_tag", quoted_only_tag_started_at, items=len(row_states))
        lane_local_initial_seed_started_at = _stage_timer_start()
        lane_local_initial_seeded = 0
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue

            row_vals = automation_rows[list_index]
            if _row_is_special_all_same_risk_family(state, row_vals, list_index):
                continue

            if _seed_lane_local_initial_episode(state, row_vals, list_index, row_idx):
                lane_local_initial_seeded += 1
        _stage_timer_stop("lane_local_initial_episode_seed", lane_local_initial_seed_started_at, items=lane_local_initial_seeded)
        occurrence_lock_started_at = _stage_timer_start()
        occurrence_locked_rows = 0
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue

            row_vals = automation_rows[list_index]
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            slot_shape_plan = _occurrence_slot_authoritative_same_time_plan(state)
            shared_occ_plan = slot_shape_plan or _preferred_shared_occurrence_plan(state, require_override_for_all_ack=True)
            if not shared_occ_plan:
                continue

            shared_decision = state.get("shared_decision") or {}
            locked = False
            if (
                _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or "")
                and (
                    slot_shape_plan is not None
                    or "requester span(all-ack->ess)" in notes_l
                )
            ):
                locked = _apply_occurrence_plan_authoritatively(
                    state,
                    row_vals,
                    list_index,
                    row_idx,
                    shared_occ_plan,
                    (
                        "OccurrenceSlotShape[AllThreeSame]"
                        if slot_shape_plan is not None
                        else "ESSContinuationGuard[AllThreeStrictEssOnly]"
                    ),
                )
                if locked:
                    decision_triplet = (state.get("shared_decision") or {}).get("triplet")
                    locked = _lock_occurrence_row(
                        state,
                        row_vals,
                        list_index,
                        "OccurrenceLocked",
                        triplet=decision_triplet,
                        )
            else:
                lane_when = shared_occ_plan.get("pick_when") or shared_decision.get("lane_time")
                current_triplet = _current_row_triplet_ist(row_vals)
                if current_triplet and lane_when:
                    if (
                        current_triplet[1] == lane_when
                        or current_triplet[2] == lane_when
                    ):
                        locked = _lock_occurrence_row(
                            state,
                            row_vals,
                            list_index,
                            "OccurrenceLocked",
                            triplet=current_triplet,
                        )

            if locked:
                occurrence_locked_rows += 1

        _stage_timer_stop("occurrence_lock_enforcement", occurrence_lock_started_at, items=occurrence_locked_rows)
        # ESS-only strict pass (isolated from blue):
        # Run for ESS-only rows regardless of blue, but only if times are not already equal.
        used_ess_continuation_ess_only = set()
        ess_only_reply_index = {}
        ess_only_reply_month_gate = {}
        ess_email_set = {e.strip().lower() for e in ess_team or []}
        ess_domain_set = {"invenio-solutions.com", "inveniolsi.com"}
        ess_name_tokens = set()
        for em in ess_email_set:
            if "@" in em:
                local = em.split("@", 1)[0]
                for tok in re.split(r"[._\\-]+", local):
                    if len(tok) >= 3:
                        ess_name_tokens.add(tok.lower())

        ess_only_strict_started_at = _stage_timer_start()
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            quoted_only = (list_index in quoted_request_only) or ("quotedrequestonly" in notes_l)
            if "quotedpairgap>16m" in notes_l and ("ess-only; no non-ess request" not in notes_l):
                # Invalid quoted pair gap should remain blue; skip ESS-only handling.
                continue
            if (not quoted_only) and ("ess-only; no non-ess request" not in notes_l):
                continue
            row_vals = automation_rows[list_index]
            if _is_strict_all_live_builder_row(state, row_vals, list_index):
                continue
            row_ctx = _row_match_context(state, row_vals, list_index)
            requester = row_ctx["requester"]
            if not requester:
                continue

            c_dt = row_ctx["created_dt"]
            a_dt = row_ctx["ack_dt"]
            r_dt = row_ctx["resolved_dt"]
            if c_dt and a_dt and r_dt and (c_dt == a_dt == r_dt):
                continue

            subject_norm = row_ctx["subject_norm"]
            base_thread = state.get("thread") or []
            ref_ist = row_ctx["ref_ist"]
            row_tokens = row_ctx["row_tokens"]
            row_id_tokens = row_ctx["row_id_tokens"]
            merged_msgs, _ = _quoted_request_row_context(
                subject_norm,
                base_thread,
                requester,
                ref_ist,
                row_tokens,
                row_id_tokens,
            )

            # Try quoted non-ESS request + ESS ack (same-day, <=16m).
            quoted_gap_mins = 16
            quoted_day_slack = 0
            parent_ess_cache = {}
            def _parent_sender_info(msg):
                key = _email_stable_key(msg)
                if key in parent_ess_cache:
                    return parent_ess_cache[key]
                parent_is_ess = None
                has_non_ess = False
                q_non_ess, q_ess = _get_quoted_row_partitions_cached(
                    msg,
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                )
                if q_non_ess:
                    has_non_ess = True
                    parent_is_ess = False
                elif q_ess:
                    parent_is_ess = True
                parent_ess_cache[key] = (parent_is_ess, has_non_ess)
                return parent_ess_cache[key]
            occ_key = state.get("occurrence_key") or _occurrence_group_key(subject_norm, requester, state.get("service_no") or "")
            idx = ess_only_reply_index.get(occ_key, 0)
            consultant_msgs, ess_reply_msgs, observed_ess_name_tokens = _ess_only_msg_partition(
                merged_msgs,
                subject_norm=subject_norm,
                row_tokens=row_tokens,
                row_id_tokens=row_id_tokens,
                requester=requester,
                quoted_only=quoted_only,
                ess_email_set=ess_email_set,
                ess_domain_set=ess_domain_set,
            )
            if (not consultant_msgs) and (not ess_reply_msgs) and (not row_id_tokens):
                widened_msgs = _merge_subject_filtered_msgs(
                    merged_msgs,
                    _requester_pool(subject_norm, "", ref_ist, day_window=30) or [],
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                    strict_quoted_variant=True,
                )
                if len(widened_msgs) != len(merged_msgs):
                    merged_msgs = widened_msgs
                    consultant_msgs, ess_reply_msgs, observed_ess_name_tokens = _ess_only_msg_partition(
                        merged_msgs,
                        subject_norm=subject_norm,
                        row_tokens=row_tokens,
                        row_id_tokens=row_id_tokens,
                        requester=requester,
                        quoted_only=quoted_only,
                        ess_email_set=ess_email_set,
                        ess_domain_set=ess_domain_set,
                    )
            ess_name_tokens_all = set(ess_name_tokens)
            ess_name_tokens_all.update(observed_ess_name_tokens)
            occ_meta = _shared_occurrence_pick(
                state,
                subject_norm_value=subject_norm,
                requester_value=requester,
                current_created_ist=_to_ist(c_dt) if c_dt else None,
                current_ack_ist=_to_ist(a_dt) if a_dt else None,
                current_resolved_ist=_to_ist(r_dt) if r_dt else None,
                default_idx=idx,
            )
            multi_service_subject = occ_meta["multi_service"]
            quoted_pick_idx = occ_meta["pick_idx"]
            quoted_pick_total = occ_meta["total_rows"]
            target_ist = occ_meta["target_ist"]
            if (not multi_service_subject) and idx < len(consultant_msgs):
                target_ist = _email_ist(consultant_msgs[idx])

            anchor_day = None
            date_tokens = state.get("date_tokens") or []
            for t in date_tokens:
                dt = _parse_date_token(t)
                if dt:
                    anchor_day = dt.date() if hasattr(dt, "date") else dt
                    break
            if not anchor_day:
                anchor_day = state.get("baseline_created_date")
            candidate_pairs_all = []
            # If row has ID tokens, add only ID-matching emails (targeted, not full scan).
            id_match_msgs = []
            if row_id_tokens:
                merged_ids = {id(e) for e in merged_msgs}
                for e in _find_emails_by_id(row_id_tokens):
                    if not _row_subject_match_email_quoted(e, subject_norm, row_tokens, row_id_tokens):
                        continue
                    id_match_msgs.append(e)
                    if id(e) not in merged_ids:
                        merged_msgs.append(e)
                        merged_ids.add(id(e))
            for msg in sorted(merged_msgs, key=lambda e: e.sent_time if getattr(e, "sent_time", None) else datetime.max, reverse=True):
                q_non_ess, q_ess = _get_quoted_row_partitions_cached(
                    msg,
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                )
                if q_non_ess and q_ess:
                    sorted_non_ess = sorted(q_non_ess)
                    sorted_ess = sorted(q_ess)
                    for ack_ist in sorted_ess:
                        reqs = [
                            r
                            for r in sorted_non_ess
                            if r < ack_ist
                            and abs((ack_ist.date() - r.date()).days) <= quoted_day_slack
                        ]
                        if not reqs:
                            continue
                        req_ist = _best_req_before_ack(reqs, ack_ist)
                        if (ack_ist - req_ist) <= timedelta(minutes=quoted_gap_mins):
                            candidate_pairs_all.append((req_ist, ack_ist))

            # Fallback: if nothing found, scan only ID-matching messages (targeted).
            if (not candidate_pairs_all) and id_match_msgs:
                for msg in sorted(id_match_msgs, key=lambda e: e.sent_time if getattr(e, "sent_time", None) else datetime.max, reverse=True):
                    q_non_ess, q_ess = _get_quoted_row_partitions_cached(
                        msg,
                        subject_norm,
                        row_tokens,
                        row_id_tokens,
                    )
                    if q_non_ess and q_ess:
                        sorted_non_ess = sorted(q_non_ess)
                        sorted_ess = sorted(q_ess)
                        for ack_ist in sorted_ess:
                            reqs = [
                                r
                                for r in sorted_non_ess
                                if r < ack_ist
                                and abs((ack_ist.date() - r.date()).days) <= quoted_day_slack
                            ]
                            if not reqs:
                                continue
                            req_ist = _best_req_before_ack(reqs, ack_ist)
                            if (ack_ist - req_ist) <= timedelta(minutes=quoted_gap_mins):
                                candidate_pairs_all.append((req_ist, ack_ist))
                    if candidate_pairs_all:
                        break
            if not candidate_pairs_all:
                for ack_msg in sorted(
                    merged_msgs,
                    key=lambda e: e.sent_time if getattr(e, "sent_time", None) else datetime.max,
                ):
                    ack_ist = _email_ist(ack_msg)
                    if not ack_ist:
                        continue
                    if not _ess_sender(ack_msg):
                        continue
                    if not _row_subject_match_email_quoted(ack_msg, subject_norm, row_tokens, row_id_tokens):
                        continue
                    ack_like_or_nonrequester = (
                        _ack_like(ack_msg)
                        or _ack_like_text_fallback(ack_msg)
                        or _ess_only_short_ack(ack_msg)
                        or (not _req_match(ack_msg, requester))
                    )
                    if not ack_like_or_nonrequester:
                        continue
                    req_info = _best_request_anchor_from_sources(
                        merged_msgs,
                        subject_norm,
                        row_tokens,
                        row_id_tokens,
                        ack_ist,
                        timedelta(minutes=16),
                    )
                    if not req_info:
                        continue
                    candidate_pairs_all.append((req_info["when"], ack_ist))
            hybrid_pair_used = False

            if not candidate_pairs_all and row_id_tokens:
                # Last-resort: ID-only quoted scan (no quoted subject required).
                for msg in emails:
                    if not _row_id_in_raw(msg, row_id_tokens):
                        continue
                    q_non_ess = []
                    q_ess = []
                    quoted_blocks = _get_quoted_blocks_with_subject_cached(msg)
                    for from_line, sent_ist, _q_subj in quoted_blocks:
                        if _quoted_from_line_is_ess_shared(from_line):
                            q_ess.append(sent_ist)
                        else:
                            q_non_ess.append(sent_ist)
                    if q_non_ess and q_ess:
                        q_non_ess.sort()
                        q_ess.sort()
                        for ack_ist in q_ess:
                            reqs = [
                                r
                                for r in q_non_ess
                                if r < ack_ist
                                and abs((ack_ist.date() - r.date()).days) <= quoted_day_slack
                            ]
                            if not reqs:
                                continue
                            req_ist = _best_req_before_ack(reqs, ack_ist)
                            if (ack_ist - req_ist) <= timedelta(minutes=quoted_gap_mins):
                                candidate_pairs_all.append((req_ist, ack_ist))
                    if candidate_pairs_all:
                        break

            if quoted_only:
                hybrid_pair = quoted_only_hybrid_pairs.get(list_index)
                if hybrid_pair:
                    candidate_pairs_all.append(hybrid_pair)

            pair_req = None
            pair_ack = None
            pair_reply = None
            pair_req_src = "PARSED_FROM_QUOTED_REQUEST"
            blue_direct_episode = None
            selected_pair_is_hybrid = False
            direct_reply_gap_blue = False
            # Do not block quoted reanchor by span notes; we require a real
            # consultant reply after the ack, which is the safer filter.
            if candidate_pairs_all:
                # De-duplicate exact pairs (same req+ack).
                uniq = {}
                for req_ist, ack_ist in candidate_pairs_all:
                    uniq[(req_ist, ack_ist)] = (req_ist, ack_ist)
                candidate_pairs = list(uniq.values())
                # Note: full-mailbox pair widening disabled for performance.
                # If baseline anchor looks stale and there are no explicit date tokens,
                # ignore the baseline day.
                has_date_tokens = bool(date_tokens)
                baseline_stale = False
                if anchor_day and (not has_date_tokens) and candidate_pairs:
                    earliest_day = min(p[0].date() for p in candidate_pairs)
                    latest_day = max(p[0].date() for p in candidate_pairs)
                    if min(abs((anchor_day - earliest_day).days), abs((anchor_day - latest_day).days)) > 7:
                        baseline_stale = True
                if baseline_stale:
                    anchor_day = None

                anchor_filtered = False
                if anchor_day:
                    # If there is no consultant reply on the anchor day, don't force that day.
                    has_consultant_on_anchor = any(
                        _email_ist(e) and _email_ist(e).date() == anchor_day for e in consultant_msgs
                    )
                    if not has_consultant_on_anchor:
                        anchor_day = None
                if anchor_day:
                    filtered = [
                        p for p in candidate_pairs
                        if p[0].date() == anchor_day and p[1].date() == anchor_day
                    ]
                    if filtered:
                        candidate_pairs = filtered
                        anchor_filtered = True

                if target_ist and not anchor_filtered and (not quoted_only):
                    near_pairs = [
                        p for p in candidate_pairs
                        if abs((p[1] - target_ist).total_seconds()) <= (7 * 24 * 3600)
                    ]
                    if near_pairs:
                        candidate_pairs = near_pairs
                    candidate_pairs.sort(key=lambda p: (abs((p[1] - target_ist).total_seconds()), p[1]))
                else:
                    candidate_pairs.sort(key=lambda p: p[1])  # by ack time

                # For quoted-only rows, prefer the day of the consultant reply occurrence
                # if a matching pair exists on that day (avoids cross-episode bleed).
                if quoted_only:
                    c_day = None
                    if multi_service_subject and target_ist:
                        c_day = target_ist.date()
                    elif consultant_msgs:
                        pick_msg = consultant_msgs[quoted_pick_idx] if quoted_pick_idx < len(consultant_msgs) else consultant_msgs[0]
                        pick_ist = _email_ist(pick_msg)
                        if pick_ist:
                            c_day = pick_ist.date()
                    if c_day and any(p[1].date() == c_day for p in candidate_pairs):
                        anchor_day = c_day
                        filtered = [
                            p for p in candidate_pairs
                            if p[0].date() == anchor_day and p[1].date() == anchor_day
                        ]
                        if filtered:
                            candidate_pairs = filtered
                # Require a real consultant reply shortly after the ack,
                # otherwise the pair is likely from a different episode.
                lane_pair_override = None
                candidate_pairs_with_reply = []
                # Prefer requester-matched replies, but fall back to any ESS reply.
                reply_pool = consultant_msgs if consultant_msgs else ess_reply_msgs
                # If still empty, scan full mailbox for ESS replies matching this subject/ID.
                if not reply_pool:
                    reply_pool = []
                    for e in emails:
                        if not _ess_sender(e):
                            continue
                        if _ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e):
                            continue
                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                        s_ids = _id_like_tokens(s_norm)
                        if row_id_tokens:
                            if (not s_ids) or row_id_tokens.isdisjoint(s_ids):
                                continue
                        if row_tokens:
                            s_tokens = _match_tokens(s_norm)
                            score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                            contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                            if score < 0.45 and not contains:
                                continue
                        reply_pool.append(e)
                # Only treat real replies as valid for quoted pairs. If the direct
                # parent is ESS, allow it only when the quoted chain also contains
                # a non-ESS request for this subject/ID.
                reply_pool_real = []
                for e in reply_pool:
                    if _ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e):
                        continue
                    parent_ess, has_non_ess = _parent_sender_info(e)
                    if (parent_ess is True) and (not has_non_ess):
                        continue
                    reply_pool_real.append(e)
                reply_pool_real.sort(key=lambda e: e.sent_time)
                reply_window_hours = 48
                direct_reply_candidate = None
                total_rows = quoted_pick_total
                target_reply_ist = occ_meta["target_reply_ist"]
                if (not multi_service_subject) and consultant_msgs:
                    if total_rows <= 1:
                        target_reply_ist = _email_ist(consultant_msgs[-1])
                    else:
                        target_reply_ist = _email_ist(
                            consultant_msgs[min(quoted_pick_idx, len(consultant_msgs) - 1)]
                        )
                elif reply_pool_real:
                    target_reply_ist = _email_ist(reply_pool_real[-1])

                preferred_reply_episode = None
                lane_local_episode = None
                if quoted_only:
                    preferred_reply_episode = _preferred_reply_anchored_quoted_episode(
                        reply_pool_real,
                        merged_msgs,
                        subject_norm,
                        row_tokens,
                        row_id_tokens,
                        requester,
                        target_reply_ist,
                    )
                    target_reply_msg = None
                    if target_reply_ist:
                        for e in reply_pool_real:
                            e_ist = _email_ist(e)
                            if e_ist and abs((e_ist - target_reply_ist).total_seconds()) <= 60:
                                target_reply_msg = e
                                break
                    if target_reply_msg is None and reply_pool_real:
                        target_reply_msg = preferred_reply_episode.get("reply_msg") if preferred_reply_episode else reply_pool_real[-1]
                    target_reply_when = _email_ist(target_reply_msg) if target_reply_msg else target_reply_ist
                    if target_reply_msg and target_reply_when:
                        lane_local_episode = _lane_local_episode_from_reply(
                            target_reply_msg,
                            target_reply_when,
                            merged_msgs,
                            requester,
                            subject_norm,
                            row_tokens,
                            row_id_tokens,
                        )
                        if lane_local_episode and lane_local_episode.get("ack"):
                            lane_pair_override = (
                                lane_local_episode["request"],
                                lane_local_episode["ack"],
                                target_reply_msg,
                            )
                            candidate_pairs_with_reply = [lane_pair_override]

                for req_ist, ack_ist in candidate_pairs:
                    if lane_pair_override and req_ist == lane_pair_override[0] and ack_ist == lane_pair_override[1]:
                        continue
                    reply_matches = []
                    for e in reply_pool_real:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if e_ist <= ack_ist:
                            continue
                        if (e_ist - ack_ist) > timedelta(hours=reply_window_hours):
                            continue
                        reply_matches.append(e)
                    reply_pick = None
                    if reply_matches:
                        if quoted_only:
                            if total_rows <= 1:
                                if target_reply_ist:
                                    reply_matches.sort(
                                        key=lambda e: (
                                            abs((_email_ist(e) - target_reply_ist).total_seconds()),
                                            abs((_email_ist(e) - ack_ist).total_seconds()),
                                            -_email_ist(e).timestamp(),
                                        )
                                    )
                                    reply_pick = reply_matches[0]
                                else:
                                    reply_pick = reply_matches[-1]
                            else:
                                if target_reply_ist:
                                    reply_matches.sort(
                                        key=lambda e: (
                                            abs((_email_ist(e) - target_reply_ist).total_seconds()),
                                            _email_ist(e),
                                        )
                                    )
                                else:
                                    reply_matches.sort(key=lambda e: _email_ist(e))
                                if quoted_pick_idx < len(reply_matches):
                                    reply_pick = reply_matches[quoted_pick_idx]
                        else:
                            reply_pick = reply_matches[0]
                    if reply_pick:
                        candidate_pairs_with_reply.append((req_ist, ack_ist, reply_pick))

                if candidate_pairs_with_reply:
                    if quoted_only:
                        if target_reply_ist:
                            candidate_pairs_with_reply.sort(
                                key=lambda p: (
                                    abs((_email_ist(p[2]) - target_reply_ist).total_seconds()),
                                    abs((p[1] - _email_ist(p[2])).total_seconds()),
                                    abs((p[1] - target_reply_ist).total_seconds()),
                                    -_email_ist(p[2]).timestamp(),
                                )
                            )
                        else:
                            candidate_pairs_with_reply.sort(
                                key=lambda p: (-_email_ist(p[2]).timestamp(), p[1])
                            )
                        if multi_service_subject:
                            candidate_pairs_with_reply = _dedupe_multi_service_lanes(
                                candidate_pairs_with_reply,
                                lambda p: (
                                    _email_ist(p[2]).replace(second=0, microsecond=0)
                                    if _email_ist(p[2]) else None
                                ),
                            )
                        if total_rows <= 1:
                            pick_idx = 0
                        elif quoted_pick_idx < len(candidate_pairs_with_reply):
                            pick_idx = quoted_pick_idx
                        else:
                            pick_idx = None
                        if pick_idx is None:
                            pair_req, pair_ack, pair_reply = None, None, None
                        else:
                            pair_req, pair_ack, pair_reply = candidate_pairs_with_reply[pick_idx]
                        if quoted_only and preferred_reply_episode and target_reply_ist:
                            chosen_reply_ist = _email_ist(pair_reply) if pair_reply else None
                            preferred_reply_ist = preferred_reply_episode.get("reply_ist")
                            preferred_hits_target = bool(
                                preferred_reply_ist
                                and abs((preferred_reply_ist - target_reply_ist).total_seconds()) <= 60
                            )
                            chosen_hits_target = bool(
                                chosen_reply_ist
                                and abs((chosen_reply_ist - target_reply_ist).total_seconds()) <= 60
                            )
                            if preferred_hits_target and not chosen_hits_target:
                                preferred_gap = preferred_reply_episode.get("gap")
                                if preferred_gap and preferred_gap <= timedelta(minutes=16):
                                    pair_req = preferred_reply_episode["request"]
                                    pair_ack = preferred_reply_ist
                                    pair_reply = preferred_reply_episode["reply_msg"]
                                    pair_req_src = preferred_reply_episode["req_src"]
                                else:
                                    blue_direct_episode = (
                                        preferred_reply_episode["request"],
                                        preferred_reply_ist,
                                        preferred_reply_episode["reply_msg"],
                                        preferred_reply_episode["req_src"],
                                    )
                                    direct_reply_gap_blue = True
                                    pair_req, pair_ack, pair_reply = None, None, None
                    else:
                        if quoted_pick_idx < len(consultant_msgs):
                            target_ist = _email_ist(consultant_msgs[quoted_pick_idx])
                            if target_ist:
                                candidate_pairs_with_reply.sort(
                                    key=lambda p: (
                                        abs((_email_ist(p[2]) - target_ist).total_seconds()),
                                        _email_ist(p[2]),
                                    )
                                )
                            else:
                                candidate_pairs_with_reply.sort(key=lambda p: _email_ist(p[2]))
                        else:
                            candidate_pairs_with_reply.sort(key=lambda p: _email_ist(p[2]))
                        pair_req, pair_ack, pair_reply = candidate_pairs_with_reply[0]
                    if quoted_only and hybrid_pair and pair_req and pair_ack:
                        selected_pair_is_hybrid = (pair_req, pair_ack) == hybrid_pair
                else:
                    # Direct-reply fallback for quoted-request-only rows:
                    # if no ack pair is usable, allow a real consultant reply
                    # directly after the quoted request, but keep SLA blue when
                    # the direct reply misses the 16-minute window.
                    if quoted_only:
                        if lane_local_episode and lane_local_episode.get("request") and lane_local_episode.get("reply_msg"):
                            if lane_local_episode.get("ack"):
                                pair_req = lane_local_episode["request"]
                                pair_ack = lane_local_episode["ack"]
                                pair_reply = lane_local_episode["reply_msg"]
                                pair_req_src = "PARSED_FROM_QUOTED_REQUEST"
                            else:
                                blue_direct_episode = (
                                    lane_local_episode["request"],
                                    lane_local_episode["reply"],
                                    lane_local_episode["reply_msg"],
                                    "PARSED_FROM_QUOTED_REQUEST",
                                )
                                direct_reply_gap_blue = True
                        if preferred_reply_episode:
                            preferred_gap = preferred_reply_episode.get("gap")
                            if preferred_gap and preferred_gap <= timedelta(minutes=16):
                                pair_req = preferred_reply_episode["request"]
                                pair_ack = preferred_reply_episode["reply_ist"]
                                pair_reply = preferred_reply_episode["reply_msg"]
                                pair_req_src = preferred_reply_episode["req_src"]
                            else:
                                blue_direct_episode = (
                                    preferred_reply_episode["request"],
                                    preferred_reply_episode["reply_ist"],
                                    preferred_reply_episode["reply_msg"],
                                    preferred_reply_episode["req_src"],
                                )
                                direct_reply_gap_blue = True
                                pair_req, pair_ack, pair_reply = None, None, None
                            continue
                        direct_pair_candidates = []
                        blue_direct_candidates = []
                        saw_direct_reply = False
                        for e in reply_pool_real:
                            e_ist = _email_ist(e)
                            if not e_ist:
                                continue
                            saw_direct_reply = True
                            req_anchor_info = (
                                _best_request_anchor_from_sources(
                                    merged_msgs,
                                    subject_norm,
                                    row_tokens,
                                    row_id_tokens,
                                    e_ist,
                                    timedelta(hours=48),
                                )
                                if e else None
                            )
                            req_anchor_ist = (req_anchor_info.get("when") if req_anchor_info else None) or latest_quoted_req
                            if not req_anchor_ist or e_ist <= req_anchor_ist:
                                continue
                            direct_gap = e_ist - req_anchor_ist
                            if direct_gap and direct_gap <= timedelta(minutes=16):
                                direct_pair_candidates.append(
                                    (
                                        req_anchor_ist,
                                        e_ist,
                                        e,
                                        req_anchor_info.get("src") if req_anchor_info else "PARSED_FROM_QUOTED_REQUEST",
                                    )
                                )
                            else:
                                blue_direct_candidates.append(
                                    (
                                        req_anchor_ist,
                                        e_ist,
                                        e,
                                        req_anchor_info.get("src") if req_anchor_info else "PARSED_FROM_QUOTED_REQUEST",
                                    )
                                )
                        if direct_pair_candidates:
                            if multi_service_subject:
                                direct_pair_candidates = _dedupe_multi_service_lanes(
                                    direct_pair_candidates,
                                    lambda p: p[1].replace(second=0, microsecond=0) if p[1] else None,
                                )
                            if multi_service_subject and target_reply_ist:
                                direct_pair_candidates.sort(
                                    key=lambda p: (
                                        abs((p[1] - target_reply_ist).total_seconds()),
                                        p[1],
                                    )
                                )
                                pair_req, pair_ack, pair_reply, pair_req_src = direct_pair_candidates[0]
                            else:
                                chosen_direct = (
                                    direct_pair_candidates[quoted_pick_idx]
                                    if quoted_pick_idx < len(direct_pair_candidates)
                                    else None
                                )
                                if chosen_direct:
                                    pair_req, pair_ack, pair_reply, pair_req_src = chosen_direct
                                else:
                                    pair_req, pair_ack, pair_reply = None, None, None
                        elif blue_direct_candidates:
                            if multi_service_subject:
                                blue_direct_candidates = _dedupe_multi_service_lanes(
                                    blue_direct_candidates,
                                    lambda p: p[1].replace(second=0, microsecond=0) if p[1] else None,
                                )
                            if multi_service_subject and target_reply_ist:
                                blue_direct_candidates.sort(
                                    key=lambda p: (
                                        abs((p[1] - target_reply_ist).total_seconds()),
                                        p[1],
                                    )
                                )
                                blue_direct_episode = blue_direct_candidates[0]
                            else:
                                blue_direct_episode = (
                                    blue_direct_candidates[quoted_pick_idx]
                                    if quoted_pick_idx < len(blue_direct_candidates)
                                    else blue_direct_candidates[0]
                                )
                            direct_reply_gap_blue = True
                            pair_req, pair_ack, pair_reply = None, None, None
                        elif saw_direct_reply:
                            direct_reply_gap_blue = True
                            pair_req, pair_ack, pair_reply = None, None, None
                        else:
                            pair_req, pair_ack, pair_reply = None, None, None
                    else:
                        # No usable pair for this occurrence; do not reanchor.
                        pair_req, pair_ack, pair_reply = None, None, None

            if quoted_only and (not pair_req) and latest_quoted_req and (not direct_reply_gap_blue):
                direct_reply_pool = consultant_msgs if consultant_msgs else ess_reply_msgs
                direct_replies = []
                for e in direct_reply_pool:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if _ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e):
                        continue
                    parent_ess, has_non_ess = _parent_sender_info(e)
                    if (parent_ess is True) and (not has_non_ess):
                        continue
                    if e_ist <= latest_quoted_req:
                        continue
                    if (e_ist - latest_quoted_req) > timedelta(hours=48):
                        continue
                    direct_replies.append(e)
                if direct_replies:
                    if len(direct_replies) > 1 and multi_service_subject and target_reply_ist:
                        direct_replies.sort(
                            key=lambda e: (
                                abs((_email_ist(e) - target_reply_ist).total_seconds()),
                                _email_ist(e),
                            )
                        )
                        direct_reply_candidate = direct_replies[0]
                    else:
                        if len(direct_replies) == 1:
                            direct_reply_candidate = direct_replies[0] if quoted_pick_idx == 0 else None
                        else:
                            direct_reply_candidate = direct_replies[quoted_pick_idx] if quoted_pick_idx < len(direct_replies) else None
                    req_anchor_info = (
                        _best_request_anchor_from_sources(
                            merged_msgs,
                            subject_norm,
                            row_tokens,
                            row_id_tokens,
                            _email_ist(direct_reply_candidate),
                            timedelta(hours=48),
                        )
                        if direct_reply_candidate else None
                    )
                    req_anchor_ist = latest_quoted_req or (req_anchor_info.get("when") if req_anchor_info else None)
                    direct_gap = (_email_ist(direct_reply_candidate) - req_anchor_ist) if (direct_reply_candidate and req_anchor_ist) else None
                    if direct_reply_candidate and direct_gap <= timedelta(minutes=16):
                        pair_req = req_anchor_ist
                        pair_ack = _email_ist(direct_reply_candidate)
                        pair_reply = direct_reply_candidate
                        pair_req_src = (req_anchor_info.get("src") if req_anchor_info else "PARSED_FROM_QUOTED_REQUEST")
                    else:
                        direct_reply_gap_blue = True

            current_notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            allow_raw_eml_episode_fallback = (
                quoted_only
                and "ess-only; no non-ess request" in current_notes_l
                and "requester span(all-ack->ess)" not in current_notes_l
            )

            if allow_raw_eml_episode_fallback and (not pair_req) and (not pair_ack) and (not direct_reply_gap_blue):
                raw_eml_episode = _final_confident_eml_episode(
                    subject_norm,
                    row_tokens,
                    row_id_tokens,
                    requester,
                    merged_msgs,
                    quoted_pick_idx,
                    quoted_pick_total,
                    target_reply_ist,
                )
                if raw_eml_episode:
                    pair_req = raw_eml_episode.get("req_ist")
                    pair_ack = raw_eml_episode.get("ack_ist")
                    pair_reply = raw_eml_episode.get("reply")
                    pair_req_src = raw_eml_episode.get("req_src") or "PARSED_FROM_RAW_EML_REQUEST"

            if pair_req and pair_ack:
                if quoted_only:
                    cur_c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                    cur_a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                    cur_r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                    cur_c_ist = _to_ist(cur_c_dt) if cur_c_dt else None
                    cur_a_ist = _to_ist(cur_a_dt) if cur_a_dt else None
                    cur_r_ist = _to_ist(cur_r_dt) if cur_r_dt else None
                    if cur_c_ist and cur_a_ist:
                        cur_c_min = cur_c_ist.replace(second=0, microsecond=0)
                        cur_a_min = cur_a_ist.replace(second=0, microsecond=0)
                        pair_req_min = pair_req.replace(second=0, microsecond=0)
                        pair_ack_min = pair_ack.replace(second=0, microsecond=0)
                        if (
                            cur_a_min >= cur_c_min
                            and (cur_a_min - cur_c_min) <= timedelta(minutes=16)
                            and cur_c_min >= pair_req_min
                            and cur_a_min <= pair_ack_min
                            and (cur_r_ist is None or cur_r_ist >= cur_a_ist)
                        ):
                            _set_row_fill(row_idx, clear_fill)
                            if list_index < len(debug_rows):
                                notes_now = debug_rows[list_index].get("Notes", "")
                                if "QuotedRequestOnlyPreservedLiveAck" not in notes_now:
                                    debug_rows[list_index]["Notes"] = (
                                        f"{notes_now}; QuotedRequestOnlyPreservedLiveAck; BlueClearedStrict"
                                    )
                            continue
                        # Do not let quoted-only reanchor move a row backward to an
                        # older episode when the current row already has a valid
                        # newer local request->ack window. This keeps quoted repair
                        # from overwriting a stronger live/local episode.
                        if (
                            cur_a_min >= cur_c_min
                            and (cur_a_min - cur_c_min) <= timedelta(minutes=16)
                            and cur_r_ist is not None
                            and cur_r_ist >= cur_a_ist
                            and pair_ack_min < cur_a_min
                        ):
                            continue

                t_c = _format_time(pair_req)
                t_a = _format_time(pair_ack)

                # Pick resolved reply for quoted-only:
                # - If only one row for this subject+requester: pick latest reply after ack (within 48h).
                # - If multiple rows: pick by occurrence order (earliest for row1, next for row2, etc.).
                occ_key = state.get("occurrence_key") or _occurrence_group_key(subject_norm, requester, state.get("service_no") or "")
                total_rows = state.get("group_total") or _occurrence_group_total(subject_norm, requester, state.get("service_no") or "")
                resolved_pick = None
                replies_after_ack = []
                reply_source_pool = reply_pool_real if "reply_pool_real" in locals() and reply_pool_real else consultant_msgs
                for e in reply_source_pool:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if _ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e):
                        continue
                    parent_ess, has_non_ess = _parent_sender_info(e)
                    if (parent_ess is True) and (not has_non_ess):
                        continue
                    if e_ist <= pair_ack:
                        continue
                    if (e_ist - pair_ack) > timedelta(hours=48):
                        continue
                    replies_after_ack.append(e)
                if replies_after_ack:
                    if multi_service_subject:
                        replies_after_ack = _dedupe_multi_service_lanes(
                            replies_after_ack,
                            lambda e: (
                                _email_ist(e).replace(second=0, microsecond=0)
                                if _email_ist(e) else None
                            ),
                        )
                    if total_rows <= 1:
                        resolved_pick = replies_after_ack[-1]
                    else:
                        resolved_pick = replies_after_ack[min(quoted_pick_idx, len(replies_after_ack) - 1)]
                elif "pair_reply" in locals():
                    resolved_pick = pair_reply

                if resolved_pick and resolved_pick.sent_time:
                    t_r = _format_time(resolved_pick.sent_time)
                    resolved_pick_ist = _email_ist(resolved_pick)
                else:
                    t_r = row_vals.get("Actual Resolved Date & Time") or t_a
                    r_dt_now = _parse_time_str(t_r)
                    r_ist_now = _to_ist(r_dt_now) if r_dt_now else None
                    if not r_ist_now or r_ist_now < pair_ack:
                        t_r = t_a
                    resolved_pick_ist = _to_ist(_parse_time_str(t_r)) if t_r else None

                if t_c and t_a and t_r:
                    candidate_kind = "hybrid" if (quoted_only and selected_pair_is_hybrid) else "quoted"
                    if not _allow_guard_rewrite(
                        row_vals,
                        list_index,
                        pair_req,
                        pair_ack,
                        resolved_pick_ist or pair_ack,
                        "QuotedRequestOnlyCandidate",
                        candidate_kind,
                    ):
                        continue
                    row_vals["Created Date & Time"] = t_c
                    row_vals["Actual Response Date & Time"] = t_a
                    row_vals["Actual Resolved Date & Time"] = t_r
                    ws.cell(row_idx, created_col).value = t_c
                    ws.cell(row_idx, response_col).value = t_a
                    ws.cell(row_idx, resolved_col).value = t_r
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REQUEST"
                        debug_rows[list_index]["AckSource"] = "PARSED_FROM_QUOTED_REPLY"
                        if resolved_pick:
                            debug_rows[list_index]["ResolvedSource"] = resolved_pick.sender_email or resolved_pick.sender_name
                        elif t_r == t_a:
                            debug_rows[list_index]["ResolvedSource"] = "PARSED_FROM_QUOTED_REPLY"
                        if quoted_only and selected_pair_is_hybrid:
                            debug_rows[list_index]["CreatedSource"] = quoted_only_hybrid_req_sources.get(list_index) or pair_req_src
                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; HybridReqSrc={quoted_only_hybrid_req_debug.get(list_index, 'unknown')}"
                        else:
                            debug_rows[list_index]["CreatedSource"] = pair_req_src
                        if quoted_only and selected_pair_is_hybrid:
                            note_tag = "QuotedRequestOnlyHybridLiveAck"
                        elif quoted_only and resolved_pick and _email_ist(resolved_pick) == pair_ack:
                            note_tag = "QuotedRequestOnlyDirectReply"
                        elif quoted_only and pair_req_src == "PARSED_FROM_RAW_EML_REQUEST":
                            note_tag = "QuotedRequestOnlyRawEmlFallback"
                        else:
                            note_tag = "QuotedRequestOnlyReanchor" if quoted_only else "ESSOnlyQuotedPairReanchor"
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; {note_tag}"
                    # Advance occurrence index when a quoted pair is used.
                    occ_key = state.get("occurrence_key") or _occurrence_group_key(subject_norm, requester, state.get("service_no") or "")
                    idx = ess_only_reply_index.get(occ_key, 0)
                    ess_only_reply_index[occ_key] = idx + 1
                    continue
            if quoted_only:
                if direct_reply_gap_blue:
                    if blue_direct_episode:
                        blue_req, blue_reply_ist, blue_reply_msg, blue_req_src = blue_direct_episode
                        if _allow_guard_rewrite(
                            row_vals,
                            list_index,
                            blue_req,
                            blue_reply_ist,
                            blue_reply_ist,
                            "QuotedRequestOnlyCandidate",
                            "quoted",
                        ):
                            t_c = _format_time(blue_req)
                            t_a = _format_time(blue_reply_ist)
                            row_vals["Created Date & Time"] = t_c
                            row_vals["Actual Response Date & Time"] = t_a
                            row_vals["Actual Resolved Date & Time"] = t_a
                            ws.cell(row_idx, created_col).value = t_c
                            ws.cell(row_idx, response_col).value = t_a
                            ws.cell(row_idx, resolved_col).value = t_a
                            if list_index < len(debug_rows):
                                debug_rows[list_index]["CreatedSource"] = blue_req_src
                                debug_rows[list_index]["AckSource"] = blue_reply_msg.sender_email or blue_reply_msg.sender_name
                                debug_rows[list_index]["ResolvedSource"] = blue_reply_msg.sender_email or blue_reply_msg.sender_name
                                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedDirectReplyPreservedBlue"
                    _set_row_fill(row_idx, blue_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedDirectReplyGap>16m"
                    continue
                # If this is a quoted-request-only row and no valid pair was found,
                # allow ESS-only fallback instead of forcing blue.
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedRequestOnlyNoPair"

            # Fallback: collapse to consultant reply by occurrence (ESS-only)
            if merged_msgs:
                notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
                shared_occ_plan = _preferred_shared_occurrence_plan(state, require_override_for_all_ack=True)
                if (
                    "requester span(all-ack->ess)" in notes_l
                    and _apply_occurrence_plan_authoritatively(
                        state,
                        row_vals,
                        list_index,
                        row_idx,
                        shared_occ_plan,
                        "ESSContinuationGuard[AllThreeStrictEssOnly]",
                    )
                ):
                    occ_key = (shared_occ_plan or {}).get("occ_key")
                    if occ_key is not None:
                        ess_only_reply_index[occ_key] = (shared_occ_plan or {}).get("slot_index", 0) + 1
                    continue
                if consultant_msgs:
                    occ_meta = _shared_occurrence_pick(
                        state,
                        subject_norm_value=subject_norm,
                        requester_value=requester,
                        current_created_ist=c_ist if 'c_ist' in locals() else None,
                        current_ack_ist=a_ist if 'a_ist' in locals() else None,
                        current_resolved_ist=r_ist if 'r_ist' in locals() else None,
                        default_idx=idx,
                    )
                    pick_idx = occ_meta["pick_idx"]
                    # Pick reply by occurrence for this subject+requester.
                    key_count = occ_meta["occ_key"]
                    total_rows = occ_meta["total_rows"]
                    if occ_meta["multi_service"]:
                        consultant_msgs = _dedupe_multi_service_lanes(
                            consultant_msgs,
                            lambda e: (
                                _email_ist(e).replace(second=0, microsecond=0)
                                if _email_ist(e) else None
                            ),
                        )
                    if shared_occ_plan and shared_occ_plan.get("pick") is not None:
                        pick = shared_occ_plan["pick"]
                    elif total_rows <= 1:
                        pick = consultant_msgs[-1]
                    else:
                        # Occurrence-based pick, but keep replies within the same month
                        # as the first valid reply for this subject/requester.
                        month_gate_used = False
                        month_key = ess_only_reply_month_gate.get(occ_key)
                        if month_key is None:
                            base_pick = consultant_msgs[min(pick_idx, len(consultant_msgs) - 1)]
                            base_ist = _email_ist(base_pick)
                            if base_ist:
                                month_key = (base_ist.year, base_ist.month)
                                ess_only_reply_month_gate[occ_key] = month_key
                        if month_key:
                            same_month = []
                            for e in consultant_msgs:
                                e_ist = _email_ist(e)
                                if not e_ist:
                                    continue
                                if (e_ist.year, e_ist.month) == month_key:
                                    same_month.append(e)
                            if same_month:
                                pick = same_month[min(pick_idx, len(same_month) - 1)]
                                month_gate_used = True
                            else:
                                pick = consultant_msgs[min(pick_idx, len(consultant_msgs) - 1)]
                        else:
                            pick = consultant_msgs[min(pick_idx, len(consultant_msgs) - 1)]
                    allow_acky = "requester span(all-ack->ess)" in notes_l
                    # Validator only: if this looks like ack, do not collapse
                    # unless all requester spans were ack-like.
                    if not pick or (_ess_only_short_ack(pick) and not allow_acky):
                        continue
                    cand_ist = _email_ist(pick)
                    if not cand_ist:
                        continue
                    key = (
                        subject_norm,
                        requester.strip().lower(),
                        state.get("service_bucket") or "",
                        cand_ist.replace(second=0, microsecond=0),
                    )
                    if key in used_ess_continuation_ess_only:
                        continue
                    if pick.sent_time:
                        candidate_kind = (
                            "occurrence_ess"
                            if shared_occ_plan and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or "")
                            else "continuation"
                        )
                        if not _allow_guard_rewrite(
                            row_vals,
                            list_index,
                            cand_ist,
                            cand_ist,
                            cand_ist,
                            "ESSContinuationGuard[AllThreeStrictEssOnly]",
                            candidate_kind,
                        ):
                            continue
                        t = _format_time(pick.sent_time)
                        if t:
                            row_vals["Created Date & Time"] = t
                            row_vals["Actual Response Date & Time"] = t
                            row_vals["Actual Resolved Date & Time"] = t
                            ws.cell(row_idx, created_col).value = t
                            ws.cell(row_idx, response_col).value = t
                            ws.cell(row_idx, resolved_col).value = t
                            _set_row_fill(row_idx, clear_fill)
                            if list_index < len(debug_rows):
                                who = pick.sender_email or pick.sender_name
                                debug_rows[list_index]["CreatedSource"] = who
                                debug_rows[list_index]["AckSource"] = who
                                debug_rows[list_index]["ResolvedSource"] = who
                                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationGuard[AllThreeStrictEssOnly]"
                            # Advance occurrence index only when we collapse
                            ess_only_reply_index[occ_key] = idx + 1
                            used_ess_continuation_ess_only.add(key)
                elif shared_occ_plan and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                    pick = shared_occ_plan.get("pick")
                    cand_ist = _email_ist(pick) if pick else None
                    if not cand_ist:
                        continue
                    if not _allow_guard_rewrite(
                        row_vals,
                        list_index,
                        cand_ist,
                        cand_ist,
                        cand_ist,
                        "ESSContinuationGuard[AllThreeStrictEssOnly]",
                        "occurrence_ess",
                    ):
                        continue
                    t = _format_time(pick.sent_time)
                    if not t:
                        continue
                    row_vals["Created Date & Time"] = t
                    row_vals["Actual Response Date & Time"] = t
                    row_vals["Actual Resolved Date & Time"] = t
                    ws.cell(row_idx, created_col).value = t
                    ws.cell(row_idx, response_col).value = t
                    ws.cell(row_idx, resolved_col).value = t
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        who = pick.sender_email or pick.sender_name
                        debug_rows[list_index]["CreatedSource"] = who
                        debug_rows[list_index]["AckSource"] = who
                        debug_rows[list_index]["ResolvedSource"] = who
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationGuard[AllThreeStrictEssOnly]"
                    occ_key = shared_occ_plan.get("occ_key")
                    if occ_key is not None:
                        ess_only_reply_index[occ_key] = shared_occ_plan.get("slot_index", 0) + 1
        _stage_timer_stop("ess_only_strict_pass", ess_only_strict_started_at, items=len(row_states))
        blue_quoted_started_at = _stage_timer_start()
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if not _row_has_blue_fill(row_idx):
                continue
            # If ESS-only quoted reanchor already fixed this row, do not override it in blue pass.
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            if (
                ("essonlyquotedpairreanchor" in notes_l)
                or ("quotedrequestonlyreanchor" in notes_l)
                or ("quotedpairgap>16m" in notes_l)
            ):
                continue

            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            if not requester:
                continue

            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt) if r_dt else a_ist
            old_gap = a_ist - c_ist
            if old_gap <= timedelta(minutes=16):
                _set_row_fill(row_idx, clear_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueCleared"
                continue

            subject_norm = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=a_ist,
            )
            thread = thread or []

            row_tokens = _match_tokens(subject_norm)
            baseline_date = state.get("baseline_created_date")
            owner_hint = ""
            if list_index < len(debug_rows):
                owner_hint = (
                    debug_rows[list_index].get("ResolvedSource")
                    or debug_rows[list_index].get("AckSource")
                    or ""
                )

            # Blue pre-fix: quoted pair re-anchor.
            # Recover (created, response) from quoted history when current created
            # is stale and causes a large blue gap.
            quoted_sources = []
            for e in (thread or []):
                quoted_sources.append(e)
            requester_pool_any_q = _requester_pool(subject_norm, "", c_ist, day_window=30) or []
            for e in requester_pool_any_q:
                quoted_sources.append(e)
            # Add anchor-day (baseline or created day) subject-matched sources to improve quoted recovery.
            anchor_day = baseline_date or c_ist.date()
            if anchor_day and row_tokens:
                for e in emails:
                    e_ist = _email_ist(e)
                    if not e_ist or e_ist.date() != anchor_day:
                        continue
                    s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                    s_tokens = _match_tokens(s_norm)
                    score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                    contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                    if score < 0.45 and not contains:
                        continue
                    quoted_sources.append(e)
            if not quoted_sources:
                continue
            row_id_tokens = _id_like_tokens(subject_norm)
            if not row_id_tokens:
                desc_text = row_vals.get("Description") or ""
                if not desc_text:
                    desc_text = state.get("description") or ""
                row_id_tokens = _id_like_tokens(desc_text)

            def _iter_live_row_replies():
                seen = set()
                reply_sources = []
                for e in (thread or []):
                    if id(e) in seen:
                        continue
                    seen.add(id(e))
                    reply_sources.append(e)
                for e in requester_pool_any_q or []:
                    if id(e) in seen:
                        continue
                    seen.add(id(e))
                    reply_sources.append(e)
                for e in reply_sources:
                    if not _ess_sender(e):
                        continue
                    if not _req_match(e, requester):
                        continue
                    if _ack_like(e) or _ack_like_text_fallback(e) or _ess_only_short_ack(e):
                        continue
                    s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                    if row_id_tokens:
                        s_ids = _id_like_tokens(s_norm)
                        if (not s_ids) or row_id_tokens.isdisjoint(s_ids):
                            continue
                    elif row_tokens:
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    yield e_ist, e

            def _locked_live_reply_ist():
                target_values = [
                    row_vals.get("Actual Resolved Date & Time"),
                    row_vals.get("Actual Response Date & Time"),
                ]
                live_replies = list(_iter_live_row_replies())
                if not live_replies:
                    return None
                for raw_value in target_values:
                    target_dt = _parse_time_str(raw_value)
                    target_ist = _to_ist(target_dt) if target_dt else None
                    if not target_ist:
                        continue
                    for reply_ist, _msg in live_replies:
                        if abs((reply_ist - target_ist).total_seconds()) <= 60:
                            return reply_ist
                return None

            locked_live_reply_ist = _locked_live_reply_ist()

            quoted_pairs = []
            quoted_pair_applied = False
            seen_pairs = set()
            for e in quoted_sources:
                if not getattr(e, "sent_time", None):
                    continue
                q_ack_ist = _extract_quoted_requester_reply_ist(
                    e,
                    requester,
                    subject_norm,
                    c_ist - timedelta(days=5),
                    c_ist + timedelta(minutes=1),
                )
                if not q_ack_ist:
                    continue
                q_req_ist = _extract_quoted_request_before_ist(
                    e,
                    subject_norm,
                    q_ack_ist,
                )
                if not q_req_ist or q_req_ist >= q_ack_ist:
                    continue
                if q_req_ist <= (c_ist - timedelta(days=5)):
                    continue
                if q_ack_ist >= a_ist:
                    continue
                pair_key = (
                    q_req_ist.replace(second=0, microsecond=0),
                    q_ack_ist.replace(second=0, microsecond=0),
                )
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                quoted_pairs.append((q_req_ist, q_ack_ist))

            blue_cleared_strict = False
            if quoted_pairs:
                # Prefer baseline-day pair when available to avoid later-episode drift.
                q_req_ist, q_ack_ist = None, None
                if baseline_date:
                    same_day_pairs = [
                        p for p in quoted_pairs
                        if p[0].date() == baseline_date
                    ]
                    if same_day_pairs:
                        same_day_pairs.sort(key=lambda p: ((p[1] - p[0]), p[1]))
                        q_req_ist, q_ack_ist = same_day_pairs[0]
                if not q_req_ist:
                    # Fallback: latest ack before current created/response.
                    quoted_pairs.sort(key=lambda p: p[1])
                    q_req_ist, q_ack_ist = quoted_pairs[-1]
                q_gap = q_ack_ist - q_req_ist
                if (
                    q_gap > timedelta(seconds=0)
                    and q_gap <= timedelta(minutes=16)
                    and q_gap < old_gap
                ):
                    pair_key = (
                        subject_norm,
                        requester.strip().lower(),
                        state.get("service_bucket") or "",
                        q_req_ist.replace(second=0, microsecond=0),
                        q_ack_ist.replace(second=0, microsecond=0),
                    )
                    if pair_key in used_quoted_pair_keys:
                        if list_index < len(debug_rows):
                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairDupSkipped"
                    else:
                        t_c = _format_time(q_req_ist)
                        t_a = _format_time(q_ack_ist)
                        t_r = row_vals.get("Actual Resolved Date & Time") or t_a
                        r_dt_now = _parse_time_str(t_r)
                        r_ist_now = _to_ist(r_dt_now) if r_dt_now else None
                        if locked_live_reply_ist and q_ack_ist > (locked_live_reply_ist + timedelta(minutes=1)):
                            if list_index < len(debug_rows):
                                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairBlockedLiveReplyLane"
                            continue
                        if not r_ist_now or r_ist_now < q_ack_ist:
                            t_r = t_a
                        elif locked_live_reply_ist and abs((r_ist_now - locked_live_reply_ist).total_seconds()) <= 60:
                            t_r = _format_time(locked_live_reply_ist)
                        if t_c and t_a and t_r:
                            cand_r_ist = _to_ist(_parse_time_str(t_r)) if t_r else None
                            if not _allow_guard_rewrite(
                                row_vals,
                                list_index,
                                q_req_ist,
                                q_ack_ist,
                                cand_r_ist or q_ack_ist,
                                "BlueQuotedPairReanchor",
                                "quoted",
                            ):
                                continue
                            used_quoted_pair_keys.add(pair_key)
                            row_vals["Created Date & Time"] = t_c
                            row_vals["Actual Response Date & Time"] = t_a
                            row_vals["Actual Resolved Date & Time"] = t_r
                            ws.cell(row_idx, created_col).value = t_c
                            ws.cell(row_idx, response_col).value = t_a
                            ws.cell(row_idx, resolved_col).value = t_r
                            c_ist = q_req_ist
                            a_ist = q_ack_ist
                            old_gap = a_ist - c_ist
                            if list_index < len(debug_rows):
                                debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REQUEST"
                                debug_rows[list_index]["AckSource"] = "PARSED_FROM_QUOTED_REPLY"
                                if t_r == t_a:
                                    debug_rows[list_index]["ResolvedSource"] = "PARSED_FROM_QUOTED_REPLY"
                                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairReanchor"
                            if old_gap <= timedelta(minutes=16):
                                _set_row_fill(row_idx, clear_fill)
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"
                                blue_cleared_strict = True
                            quoted_pair_applied = True

            # Blue pre-fix (non-ESS quoted pair fallback):
            # If no requester-matched quoted pair was found, allow a quoted
            # non-ESS request + ESS ack pair to re-anchor (safe, blue-only).
            # Additionally, for ESS-only blue rows, allow same-day quoted
            # non-ESS pair even when requester-match path failed.
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            if (not quoted_pair_applied) and ("ess-only; no non-ess request" not in notes_l):
                def _parent_sender_is_ess_msg(msg):
                    blocks = _get_quoted_blocks_with_subject_cached(msg)
                    for from_line, _sent_ist, q_subj in blocks:
                        if not _quoted_block_matches_row(
                            msg,
                            q_subj,
                            subject_norm,
                            row_tokens,
                            row_id_tokens,
                        ):
                            continue
                        return _quoted_from_line_is_ess_shared(from_line)
                    return None
                def _has_real_reply_after_ack(ack_ist):
                    for e_ist, e in _iter_live_row_replies():
                        parent_ess = _parent_sender_is_ess_msg(e)
                        if parent_ess is True:
                            continue
                        if e_ist < ack_ist:
                            continue
                        if (e_ist - ack_ist) > timedelta(hours=48):
                            continue
                        if locked_live_reply_ist and abs((e_ist - locked_live_reply_ist).total_seconds()) > 60:
                            continue
                        return True
                    return False
                def _collect_non_ess_pairs(subject_filter: str, sources, win_start=None, win_end=None, center_ist=None, require_improve: bool = True):
                    ess_email_set = {e.strip().lower() for e in ess_team or []}
                    ess_name_tokens = set()
                    for em in ess_email_set:
                        if "@" in em:
                            local = em.split("@", 1)[0]
                            for tok in re.split(r"[._\\-]+", local):
                                if len(tok) >= 3:
                                    ess_name_tokens.add(tok.lower())
                    non_ess_times = []
                    ess_times = []
                    for e in sources:
                        for from_line, sent_ist in _extract_quoted_blocks(e, subject_filter):
                            if win_start and sent_ist < win_start:
                                continue
                            if win_end and sent_ist > win_end:
                                continue
                            if sent_ist <= (c_ist - timedelta(days=5)):
                                continue
                            if sent_ist >= a_ist:
                                continue
                            if _quoted_from_line_is_ess_shared(from_line):
                                ess_times.append(sent_ist)
                            else:
                                non_ess_times.append(sent_ist)
                    if not (ess_times and non_ess_times):
                        return []
                    ess_times.sort()
                    non_ess_times.sort()
                    pairs = []
                    for ack_ist in ess_times:
                        reqs = [r for r in non_ess_times if r < ack_ist]
                        if not reqs:
                            continue
                        req_ist = reqs[-1]
                        gap = ack_ist - req_ist
                        if gap <= timedelta(minutes=16) and (
                            (not require_improve) or gap < old_gap or old_gap <= timedelta(seconds=0)
                        ):
                            score = abs((req_ist - center_ist).total_seconds()) if center_ist else 0
                            pairs.append((score, req_ist, ack_ist))
                    pairs.sort(key=lambda p: p[0])
                    return [(req, ack) for _, req, ack in pairs]

                def _choose_pair(best_a, best_b):
                    if not best_a:
                        return best_b
                    if not best_b:
                        return best_a
                    da = abs((best_a[0] - c_ist).total_seconds())
                    db = abs((best_b[0] - c_ist).total_seconds())
                    return best_a if da <= db else best_b

                # Prefer same-day quoted pair first (tightest window).
                if baseline_date:
                    day_start = datetime(baseline_date.year, baseline_date.month, baseline_date.day, tzinfo=c_ist.tzinfo)
                else:
                    day_start = datetime(c_ist.year, c_ist.month, c_ist.day, tzinfo=c_ist.tzinfo)
                day_end = day_start + timedelta(days=1)
                extra_sources = _requester_pool("", "", c_ist, day_window=30)
                if extra_sources and row_tokens:
                    filtered = []
                    for e in extra_sources:
                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score >= 0.45 or contains:
                            filtered.append(e)
                    extra_sources = filtered
                if extra_sources:
                    merged_sources = list(quoted_sources) + list(extra_sources)
                else:
                    merged_sources = quoted_sources
                pairs_subj = _collect_non_ess_pairs(subject_norm, merged_sources, day_start, day_end, c_ist, require_improve=True)
                pairs_relaxed = _collect_non_ess_pairs("", merged_sources, day_start, day_end, c_ist, require_improve=True)
                candidate_pairs = pairs_subj + pairs_relaxed
                pick_pair = None
                for cand in candidate_pairs:
                    key = (
                        subject_norm,
                        requester.strip().lower(),
                        state.get("service_bucket") or "",
                        cand[0].replace(second=0, microsecond=0),
                        cand[1].replace(second=0, microsecond=0),
                    )
                    if key in used_quoted_pair_keys:
                        continue
                    pick_pair = cand
                    break
                if not pick_pair:
                    # Fallback: allow quoted pairs within ±48h of Created (closest wins).
                    win_start = c_ist - timedelta(hours=48)
                    win_end = c_ist + timedelta(hours=48)
                    pairs_subj = _collect_non_ess_pairs(subject_norm, quoted_sources, win_start, win_end, c_ist, require_improve=True)
                    pairs_relaxed = _collect_non_ess_pairs("", quoted_sources, win_start, win_end, c_ist, require_improve=True)
                    candidate_pairs = pairs_subj + pairs_relaxed
                    pick_pair = None
                    for cand in candidate_pairs:
                        key = (
                            subject_norm,
                            requester.strip().lower(),
                            state.get("service_bucket") or "",
                            cand[0].replace(second=0, microsecond=0),
                            cand[1].replace(second=0, microsecond=0),
                        )
                        if key in used_quoted_pair_keys:
                            continue
                        pick_pair = cand
                        break

                if pick_pair:
                    q_req_ist, q_ack_ist = pick_pair
                    if not _has_real_reply_after_ack(q_ack_ist):
                        if list_index < len(debug_rows):
                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; QuotedPairNoReplyAfterAck"
                        pick_pair = None
                if pick_pair:
                    pair_key = (
                        subject_norm,
                        requester.strip().lower(),
                        state.get("service_bucket") or "",
                        q_req_ist.replace(second=0, microsecond=0),
                        q_ack_ist.replace(second=0, microsecond=0),
                    )
                    if pair_key in used_quoted_pair_keys:
                        if list_index < len(debug_rows):
                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairDupSkipped"
                        pick_pair = None
                    else:
                        t_c = _format_time(q_req_ist)
                        t_a = _format_time(q_ack_ist)
                        t_r = row_vals.get("Actual Resolved Date & Time") or t_a
                        r_dt_now = _parse_time_str(t_r)
                        r_ist_now = _to_ist(r_dt_now) if r_dt_now else None
                        if locked_live_reply_ist and q_ack_ist > (locked_live_reply_ist + timedelta(minutes=1)):
                            if list_index < len(debug_rows):
                                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairBlockedLiveReplyLane"
                            continue
                        if not r_ist_now or r_ist_now < q_ack_ist:
                            t_r = t_a
                        elif locked_live_reply_ist and abs((r_ist_now - locked_live_reply_ist).total_seconds()) <= 60:
                            t_r = _format_time(locked_live_reply_ist)
                        if t_c and t_a and t_r:
                            cand_r_ist = _to_ist(_parse_time_str(t_r)) if t_r else None
                            if not _allow_guard_rewrite(
                                row_vals,
                                list_index,
                                q_req_ist,
                                q_ack_ist,
                                cand_r_ist or q_ack_ist,
                                "BlueQuotedPairReanchorNonESS",
                                "quoted",
                            ):
                                continue
                            used_quoted_pair_keys.add(pair_key)
                            row_vals["Created Date & Time"] = t_c
                            row_vals["Actual Response Date & Time"] = t_a
                            row_vals["Actual Resolved Date & Time"] = t_r
                            ws.cell(row_idx, created_col).value = t_c
                            ws.cell(row_idx, response_col).value = t_a
                            ws.cell(row_idx, resolved_col).value = t_r
                            c_ist = q_req_ist
                            a_ist = q_ack_ist
                            old_gap = a_ist - c_ist
                            if list_index < len(debug_rows):
                                debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REQUEST"
                                debug_rows[list_index]["AckSource"] = "PARSED_FROM_QUOTED_REPLY"
                                if t_r == t_a:
                                    debug_rows[list_index]["ResolvedSource"] = "PARSED_FROM_QUOTED_REPLY"
                                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairReanchorNonESS"
                            quoted_pair_applied = True
                            if old_gap <= timedelta(minutes=16):
                                _set_row_fill(row_idx, clear_fill)
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"
                                blue_cleared_strict = True

            # Blue-only ESS continuation: collapse to unique consultant reply
            # when ESS-only and no non-ESS request exists between first/last replies.
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            shared_occ_plan = _preferred_shared_occurrence_plan(
                state,
                quoted_sources=quoted_sources,
                c_ist=c_ist,
                require_override_for_all_ack=True,
            )
            if (
                "requester span(all-ack->ess)" in notes_l
                and _apply_occurrence_plan_authoritatively(
                    state,
                    row_vals,
                    list_index,
                    row_idx,
                    shared_occ_plan,
                    "ESSContinuationGuard[AllThreeStrictEssOnly]",
                )
            ):
                occ_key = (shared_occ_plan or {}).get("occ_key")
                if occ_key is not None:
                    ess_only_reply_index[occ_key] = (shared_occ_plan or {}).get("slot_index", 0) + 1
                continue
            allow_occurrence_ess_quoted = bool(
                shared_occ_plan
                and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or "")
                and "ess-only; no non-ess request" in notes_l
                and "requester span(all-ack->ess)" in notes_l
            )
            if (
                ("quotedrequestonly" in notes_l)
                or ("quotedpairgap>16m" in notes_l)
                or ("quotedrequestonlynopair" in notes_l)
            ) and not allow_occurrence_ess_quoted:
                # Quoted-request-only rows (or invalid quoted gap) are handled separately; do not collapse here.
                continue
            if (not quoted_pair_applied) and "ess-only; no non-ess request" in notes_l:
                base_thread = state.get("thread") or []
                thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=r_ist,
                )
                requester_pool = _requester_pool(subject_norm, requester, r_ist, day_window=21)
                merged_msgs = []
                for e in (thread or []):
                    merged_msgs.append(e)
                for e in (requester_pool or []):
                    merged_msgs.append(e)

                # Only block ESS-only collapse if a valid same-day non-ESS request + ESS ack pair exists.
                has_non_ess_quoted = False
                ess_email_set = {e.strip().lower() for e in ess_team or []}
                day_start = datetime(c_ist.year, c_ist.month, c_ist.day, tzinfo=c_ist.tzinfo)
                day_end = day_start + timedelta(days=1)
                quoted_non_ess = []
                quoted_ess = []
                for e in (quoted_sources or []):
                    for from_line, sent_ist in _extract_quoted_blocks(e, subject_norm):
                        if sent_ist < day_start or sent_ist >= day_end:
                            continue
                        if _quoted_from_line_is_ess_shared(from_line):
                            quoted_ess.append(sent_ist)
                        else:
                            quoted_non_ess.append(sent_ist)
                if quoted_non_ess and quoted_ess:
                    quoted_non_ess.sort()
                    quoted_ess.sort()
                    for ack_ist in quoted_ess:
                        reqs = [r for r in quoted_non_ess if r < ack_ist]
                        if not reqs:
                            continue
                        req_ist = reqs[-1]
                        if (ack_ist - req_ist) <= timedelta(minutes=16):
                            has_non_ess_quoted = True
                            break

                # Only collapse if we see no non-ESS sender in live thread or quoted blocks.
                has_non_ess_live = False
                row_tokens = _match_tokens(subject_norm)
                for e in (merged_msgs or []):
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if row_tokens:
                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    if _ess_sender(e):
                        continue
                    if _system_like_sender(e):
                        continue
                    has_non_ess_live = True
                    break

                if (not has_non_ess_quoted) and (not has_non_ess_live):
                    if merged_msgs:
                        dedup = {}
                        for e in merged_msgs:
                            dedup[(getattr(e, "subject", ""), getattr(e, "sender_email", ""), getattr(e, "sender_name", ""), getattr(e, "sent_time", None))] = e
                        merged_msgs = list(dedup.values())
                        merged_msgs.sort(key=lambda e: e.sent_time if getattr(e, "sent_time", None) else datetime.max)

                        # ESS-only enhanced continuation: run even if row wasn't blue,
                        # but only when times are not already equal.
                        force_ess_blue = False
                        if "ess-only; no non-ess request" in notes_l:
                            if c_dt and a_dt and r_dt:
                                force_ess_blue = not (c_dt == a_dt == r_dt)

                        if (
                            shared_occ_plan
                            and _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or "")
                            and ("ess-only; no non-ess request" in notes_l)
                            and (force_ess_blue or _row_has_blue_fill(row_idx))
                        ):
                            occ_pick = shared_occ_plan.get("pick")
                            occ_ist = _email_ist(occ_pick) if occ_pick else None
                            if occ_ist and _allow_guard_rewrite(
                                row_vals,
                                list_index,
                                occ_ist,
                                occ_ist,
                                occ_ist,
                                "ESSContinuationGuard[AllThreeStrictEssOnly]",
                                "occurrence_ess",
                            ):
                                t_occ = _format_time(occ_pick.sent_time)
                                if t_occ:
                                    row_vals["Created Date & Time"] = t_occ
                                    row_vals["Actual Response Date & Time"] = t_occ
                                    row_vals["Actual Resolved Date & Time"] = t_occ
                                    ws.cell(row_idx, created_col).value = t_occ
                                    ws.cell(row_idx, response_col).value = t_occ
                                    ws.cell(row_idx, resolved_col).value = t_occ
                                    _set_row_fill(row_idx, clear_fill)
                                    if list_index < len(debug_rows):
                                        who = occ_pick.sender_email or occ_pick.sender_name
                                        debug_rows[list_index]["CreatedSource"] = who
                                        debug_rows[list_index]["AckSource"] = who
                                        debug_rows[list_index]["ResolvedSource"] = who
                                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationGuard[AllThreeStrictEssOnly]"
                                    continue

                        consultant_msgs = []
                        for e in merged_msgs:
                            e_ist = _email_ist(e)
                            if not e_ist:
                                continue
                            if not _req_match(e, requester):
                                continue
                            if not _is_real_reply_candidate(e):
                                continue
                            consultant_msgs.append(e)
                        consultant_msgs.sort(key=lambda e: e.sent_time)

                        if consultant_msgs:
                            first_ist = _email_ist(consultant_msgs[0])
                            latest = None
                            latest_ist = None
                            # Prefer latest consultant reply (not earliest).
                            for cand in reversed(consultant_msgs):
                                cand_ist = _email_ist(cand)
                                if not cand_ist:
                                    continue
                                key = (
                                    subject_norm,
                                    requester.strip().lower(),
                                    state.get("service_bucket") or "",
                                    cand_ist.replace(second=0, microsecond=0),
                                )
                                if key in used_ess_continuation_blue:
                                    continue
                                latest = cand
                                latest_ist = cand_ist
                                used_ess_continuation_blue.add(key)
                                break
                            if ("ess-only; no non-ess request" in notes_l) and (force_ess_blue or _row_has_blue_fill(row_idx)):
                                # Prefer a same-day non-ESS request + ESS ack quoted pair,
                                # using the latest ack time found in the latest message.
                                ess_email_set = {e.strip().lower() for e in ess_team or []}
                                pair_req = None
                                pair_ack = None
                                # Broaden search to any subject-matched emails (not just requester)
                                # so we don't miss quoted pairs in a teammate's reply.
                                broader_msgs = list(merged_msgs)
                                extra_any = _requester_pool(subject_norm, "", r_ist, day_window=30) or []
                                if extra_any:
                                    for e in extra_any:
                                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                                        s_tokens = _match_tokens(s_norm)
                                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                                        if score < 0.45 and not contains:
                                            continue
                                        broader_msgs.append(e)
                                for msg in sorted(broader_msgs, key=lambda e: e.sent_time if getattr(e, "sent_time", None) else datetime.max, reverse=True):
                                    q_non_ess = []
                                    q_ess = []
                                    for from_line, sent_ist in _extract_quoted_blocks(msg, subject_norm):
                                        if _quoted_from_line_is_ess_shared(from_line):
                                            q_ess.append(sent_ist)
                                        else:
                                            q_non_ess.append(sent_ist)
                                    if q_non_ess and q_ess:
                                        q_non_ess.sort()
                                        q_ess.sort()
                                        best = None
                                        for ack_ist in q_ess:
                                            reqs = [r for r in q_non_ess if r < ack_ist and r.date() == ack_ist.date()]
                                            if not reqs:
                                                continue
                                            req_ist = reqs[-1]
                                            if (ack_ist - req_ist) <= timedelta(minutes=16):
                                                cand = (ack_ist, req_ist)
                                                if not best or cand[0] > best[0]:
                                                    best = cand
                                        if best:
                                            pair_ack, pair_req = best[0], best[1]
                                    if pair_req and pair_ack:
                                        break
                                if pair_req and pair_ack:
                                    t_c = _format_time(pair_req)
                                    t_a = _format_time(pair_ack)
                                    t_r = row_vals.get("Actual Resolved Date & Time") or t_a
                                    r_dt_now = _parse_time_str(t_r)
                                    r_ist_now = _to_ist(r_dt_now) if r_dt_now else None
                                    if locked_live_reply_ist and pair_ack > (locked_live_reply_ist + timedelta(minutes=1)):
                                        if list_index < len(debug_rows):
                                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairBlockedLiveReplyLane"
                                        continue
                                    if not r_ist_now or r_ist_now < pair_ack:
                                        t_r = t_a
                                    elif locked_live_reply_ist and abs((r_ist_now - locked_live_reply_ist).total_seconds()) <= 60:
                                        t_r = _format_time(locked_live_reply_ist)
                                    if t_c and t_a and t_r:
                                        cand_r_ist = _to_ist(_parse_time_str(t_r)) if t_r else None
                                        if not _allow_guard_rewrite(
                                            row_vals,
                                            list_index,
                                            pair_req,
                                            pair_ack,
                                            cand_r_ist or pair_ack,
                                            "BlueQuotedPairReanchorNonESSLatest",
                                            "quoted",
                                        ):
                                            continue
                                        row_vals["Created Date & Time"] = t_c
                                        row_vals["Actual Response Date & Time"] = t_a
                                        row_vals["Actual Resolved Date & Time"] = t_r
                                        ws.cell(row_idx, created_col).value = t_c
                                        ws.cell(row_idx, response_col).value = t_a
                                        ws.cell(row_idx, resolved_col).value = t_r
                                        _set_row_fill(row_idx, clear_fill)
                                        if list_index < len(debug_rows):
                                            debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REQUEST"
                                            debug_rows[list_index]["AckSource"] = "PARSED_FROM_QUOTED_REPLY"
                                            if t_r == t_a:
                                                debug_rows[list_index]["ResolvedSource"] = "PARSED_FROM_QUOTED_REPLY"
                                            debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueQuotedPairReanchorNonESSLatest"
                                        blue_cleared_strict = True
                                        continue
                            if first_ist and latest_ist and latest_ist >= c_ist:
                                non_ess_between = False
                                row_tokens = _match_tokens(subject_norm)
                                for e in merged_msgs:
                                    e_ist = _email_ist(e)
                                    if not e_ist:
                                        continue
                                    if e_ist <= first_ist or e_ist > latest_ist:
                                        continue
                                    if row_tokens:
                                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                                        s_tokens = _match_tokens(s_norm)
                                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                                        if score < 0.45 and not contains:
                                            continue
                                    if _ess_sender(e):
                                        continue
                                    if _system_like_sender(e):
                                        continue
                                    if _req_match(e, requester):
                                        continue
                                    non_ess_between = True
                                    break
                            force_collapse = "requester span(all-ack->ess)" in notes_l
                            if (not non_ess_between) or force_collapse:
                                if latest is None:
                                    continue
                                t = _format_time(latest.sent_time)
                                if t:
                                    row_vals["Created Date & Time"] = t
                                    row_vals["Actual Response Date & Time"] = t
                                    row_vals["Actual Resolved Date & Time"] = t
                                    ws.cell(row_idx, created_col).value = row_vals.get("Created Date & Time")
                                    ws.cell(row_idx, response_col).value = row_vals.get("Actual Response Date & Time")
                                    ws.cell(row_idx, resolved_col).value = row_vals.get("Actual Resolved Date & Time")
                                    _set_row_fill(row_idx, clear_fill)
                                    if list_index < len(debug_rows):
                                        who = latest.sender_email or latest.sender_name
                                        debug_rows[list_index]["CreatedSource"] = who
                                        debug_rows[list_index]["AckSource"] = who
                                        debug_rows[list_index]["ResolvedSource"] = who
                                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; ESSContinuationGuard[AllThreeStrictBlue]"
                                    continue

            if blue_cleared_strict and "ess-only; no non-ess request" not in notes_l:
                continue

            def _owner_hint_match(e):
                if not owner_hint:
                    return False
                hint = str(owner_hint).strip().lower()
                sender_email = (getattr(e, "sender_email", "") or "").strip().lower()
                sender_name = (getattr(e, "sender_name", "") or "").strip().lower()
                sender_blob = f"{sender_name} {sender_email}".strip()
                if not sender_blob:
                    return False
                if "@" in hint:
                    return hint in sender_blob
                return hint in sender_blob

            def _same_sender(a, b):
                a_email = (getattr(a, "sender_email", "") or "").strip().lower()
                b_email = (getattr(b, "sender_email", "") or "").strip().lower()
                if a_email and b_email:
                    return a_email == b_email
                a_name = (getattr(a, "sender_name", "") or "").strip().lower()
                b_name = (getattr(b, "sender_name", "") or "").strip().lower()
                return bool(a_name and b_name and a_name == b_name)

            def _collect_ack_candidates(messages, win_start_ist, win_end_ist, baseline_days):
                out = []
                for e in messages:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist < win_start_ist or e_ist > win_end_ist:
                        continue
                    if not (_req_match(e, requester) or _owner_hint_match(e)):
                        continue
                    if not (_ack_like(e) or _ack_like_text_fallback(e)):
                        continue
                    if (
                        baseline_days is not None
                        and baseline_date
                        and abs((e_ist.date() - baseline_date).days) > baseline_days
                    ):
                        continue
                    if row_tokens:
                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)
                out.sort(key=lambda e: e.sent_time)
                return out

            def _collect_ess_nonack_candidates(messages, win_start_ist, win_end_ist, baseline_days):
                out = []
                for e in messages:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist < win_start_ist or e_ist > win_end_ist:
                        continue
                    if not _ess_sender(e):
                        continue
                    if _system_like_sender(e):
                        continue
                    if not _is_real_reply_candidate(e):
                        continue
                    if (
                        baseline_days is not None
                        and baseline_date
                        and abs((e_ist.date() - baseline_date).days) > baseline_days
                    ):
                        continue
                    if row_tokens:
                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)
                out.sort(key=lambda e: e.sent_time)
                return out

            def _collect_blue_candidates(messages, win_start_ist, win_end_ist, baseline_days):
                out = []
                for e in messages:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    if e_ist < win_start_ist or e_ist > win_end_ist:
                        continue
                    if not (_req_match(e, requester) or _owner_hint_match(e)):
                        continue
                    if not _is_real_reply_candidate(e):
                        continue
                    if (
                        baseline_days is not None
                        and baseline_date
                        and abs((e_ist.date() - baseline_date).days) > baseline_days
                    ):
                        continue
                    if row_tokens:
                        s_norm = _subject_norm_cached(getattr(e, "subject", "") or "")
                        s_tokens = _match_tokens(s_norm)
                        score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                        contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                        if score < 0.45 and not contains:
                            continue
                    out.append(e)
                out.sort(key=lambda e: e.sent_time)
                return out

            # Blue fix #1: recover true ack (first response) when it exists but
            # row drifted due to multi-thread same-subject matching.
            ack_pool_used = False
            ack_candidates = _collect_ack_candidates(
                thread,
                c_ist - timedelta(minutes=5),
                c_ist + timedelta(hours=48),
                baseline_days=7,
            )
            if not ack_candidates:
                requester_pool_for_ack = _requester_pool(subject_norm, requester, c_ist, day_window=30)
                ack_candidates = _collect_ack_candidates(
                    requester_pool_for_ack,
                    c_ist - timedelta(minutes=5),
                    c_ist + timedelta(hours=48),
                    baseline_days=None,
                )
                ack_pool_used = bool(ack_candidates)
            if not ack_candidates and owner_hint:
                requester_pool_for_ack_any = _requester_pool(subject_norm, "", c_ist, day_window=30)
                ack_candidates = _collect_ack_candidates(
                    requester_pool_for_ack_any,
                    c_ist - timedelta(minutes=5),
                    c_ist + timedelta(hours=48),
                    baseline_days=None,
                )
                ack_pool_used = bool(ack_candidates)
            if ack_candidates:
                ack_pick = None
                for e in ack_candidates:
                    if _email_ist(e) >= c_ist:
                        ack_pick = e
                        break
                if not ack_pick:
                    ack_pick = ack_candidates[0]
                ack_pick_ist = _email_ist(ack_pick)
                if ack_pick_ist and (ack_pick_ist - c_ist) < old_gap:
                    requester_pool_for_resolved = _requester_pool(subject_norm, requester, ack_pick_ist, day_window=30)
                    resolved_non_ack = _collect_blue_candidates(
                        thread,
                        ack_pick_ist,
                        ack_pick_ist + timedelta(hours=72),
                        baseline_days=7,
                    )
                    if not resolved_non_ack:
                        resolved_non_ack = _collect_blue_candidates(
                            requester_pool_for_resolved,
                            ack_pick_ist,
                            ack_pick_ist + timedelta(hours=72),
                            baseline_days=None,
                        )
                    resolved_same_sender = [e for e in resolved_non_ack if _same_sender(e, ack_pick)]
                    resolved_pick = resolved_same_sender[-1] if resolved_same_sender else (resolved_non_ack[-1] if resolved_non_ack else ack_pick)
                    t_a = _format_time(ack_pick.sent_time)
                    t_r = _format_time(resolved_pick.sent_time)
                    if t_a and t_r:
                        note_tag = "BlueStrictAckReanchorPool" if ack_pool_used else "BlueStrictAckReanchor"
                        applied = _apply_guarded_episode_update(
                            state,
                            row_vals,
                            list_index,
                            row_idx,
                            c_ist,
                            ack_pick_ist,
                            _email_ist(resolved_pick) or ack_pick_ist,
                            owner_tag=note_tag,
                            candidate_kind="quoted",
                            created_src=debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else "",
                            ack_src=ack_pick.sender_email or ack_pick.sender_name,
                            resolved_src=resolved_pick.sender_email or resolved_pick.sender_name,
                            note_suffix=note_tag,
                        )
                        if applied:
                            old_gap = ack_pick_ist - c_ist
                            if old_gap <= timedelta(minutes=16):
                                _set_row_fill(row_idx, clear_fill)
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"
                                continue

            notes_l_now = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            strict_candidates = []
            used_requester_pool = False
            if "ess-only; no non-ess request" not in notes_l_now:
                strict_candidates = _collect_blue_candidates(
                    thread,
                    c_ist - timedelta(minutes=5),
                    c_ist + timedelta(hours=96),
                    baseline_days=2,
                )
                if not strict_candidates:
                    strict_candidates = _collect_blue_candidates(
                        thread,
                        c_ist - timedelta(minutes=5),
                        c_ist + timedelta(hours=240),
                        baseline_days=7,
                    )
                if not strict_candidates:
                    requester_pool = _requester_pool(subject_norm, requester, c_ist, day_window=30)
                    strict_candidates = _collect_blue_candidates(
                        requester_pool,
                        c_ist - timedelta(minutes=5),
                        c_ist + timedelta(days=30),
                        baseline_days=None,
                    )
                    used_requester_pool = bool(strict_candidates)
                if not strict_candidates and owner_hint:
                    requester_pool_any = _requester_pool(subject_norm, "", c_ist, day_window=30)
                    strict_candidates = _collect_blue_candidates(
                        requester_pool_any,
                        c_ist - timedelta(minutes=5),
                        c_ist + timedelta(days=30),
                        baseline_days=None,
                    )
                    used_requester_pool = bool(strict_candidates)
                if not strict_candidates:
                    strict_candidates = _collect_ess_nonack_candidates(
                        thread,
                        c_ist - timedelta(minutes=5),
                        c_ist + timedelta(hours=240),
                        baseline_days=7,
                    )
                if not strict_candidates:
                    requester_pool_ess_any = _requester_pool(subject_norm, "", c_ist, day_window=30)
                    strict_candidates = _collect_ess_nonack_candidates(
                        requester_pool_ess_any,
                        c_ist - timedelta(minutes=5),
                        c_ist + timedelta(days=30),
                        baseline_days=None,
                    )
                    used_requester_pool = bool(strict_candidates)
                if not strict_candidates:
                    # Blue fallback: quoted-history re-anchor (response only) when
                    # no reliable live candidate is available.
                    quoted_sources = []
                    for e in (thread or []):
                        quoted_sources.append(e)
                    requester_pool_any_for_quoted = _requester_pool(subject_norm, "", a_ist, day_window=30)
                    for e in requester_pool_any_for_quoted:
                        quoted_sources.append(e)
                    quoted_candidates = []
                    seen_q = set()
                    for e in quoted_sources:
                        if not getattr(e, "sent_time", None):
                            continue
                        q_ist = _extract_quoted_request_before_ist(
                            e,
                            subject_norm,
                            a_ist + timedelta(minutes=5),
                        )
                        if not q_ist:
                            continue
                        if q_ist <= c_ist:
                            continue
                        if q_ist >= a_ist:
                            continue
                        if (q_ist - c_ist) > timedelta(hours=72):
                            continue
                        if baseline_date and abs((q_ist.date() - baseline_date).days) > 7:
                            continue
                        k = q_ist.replace(second=0, microsecond=0)
                        if k in seen_q:
                            continue
                        seen_q.add(k)
                        quoted_candidates.append(q_ist)
                    if quoted_candidates:
                        quoted_candidates.sort()
                        # Prefer earliest valid post-created quote as response anchor.
                        q_pick = quoted_candidates[0]
                        new_gap_q = q_pick - c_ist
                        if new_gap_q < old_gap:
                            t_a = _format_time(q_pick)
                            t_r = row_vals.get("Actual Resolved Date & Time") or t_a
                            r_dt_now = _parse_time_str(t_r)
                            r_ist_now = _to_ist(r_dt_now) if r_dt_now else None
                            if not r_ist_now or r_ist_now < q_pick:
                                t_r = t_a
                            if t_a and t_r:
                                row_vals["Actual Response Date & Time"] = t_a
                                row_vals["Actual Resolved Date & Time"] = t_r
                                ws.cell(row_idx, response_col).value = t_a
                                ws.cell(row_idx, resolved_col).value = t_r
                                if list_index < len(debug_rows):
                                    debug_rows[list_index]["AckSource"] = "PARSED_FROM_QUOTED_REPLY"
                                    if t_r == t_a:
                                        debug_rows[list_index]["ResolvedSource"] = "PARSED_FROM_QUOTED_REPLY"
                                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueStrictQuotedReanchor"
                                if new_gap_q <= timedelta(minutes=16):
                                    _set_row_fill(row_idx, clear_fill)
                                    if list_index < len(debug_rows):
                                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"
                                    continue
            if strict_candidates and list_index < len(debug_rows):
                notes_l_now = (debug_rows[list_index].get("Notes") or "").lower()
                if "bluequotedpairdupskipped" in notes_l_now:
                    strict_candidates = []

            if strict_candidates:
                response_pick = None
                for e in strict_candidates:
                    if _email_ist(e) >= c_ist:
                        response_pick = e
                        break
                if not response_pick:
                    response_pick = strict_candidates[0]
                response_pick_ist = _email_ist(response_pick)
                if not response_pick_ist:
                    continue

                resolved_pool = [
                    e for e in strict_candidates
                    if _email_ist(e) >= response_pick_ist
                    and _email_ist(e) <= (response_pick_ist + timedelta(hours=72))
                ]
                resolved_same_sender = [e for e in resolved_pool if _same_sender(e, response_pick)]
                resolved_pick = resolved_same_sender[-1] if resolved_same_sender else (resolved_pool[-1] if resolved_pool else response_pick)
                resolved_pick_ist = _email_ist(resolved_pick)
                if not resolved_pick_ist:
                    continue

                new_gap = response_pick_ist - c_ist
                # Apply only if strictly improves (or fully resolves) the blue gap.
                if new_gap >= old_gap:
                    continue

                t_a = _format_time(response_pick.sent_time)
                t_r = _format_time(resolved_pick.sent_time)
                if not (t_a and t_r):
                    continue
                row_vals["Actual Response Date & Time"] = t_a
                row_vals["Actual Resolved Date & Time"] = t_r
                ws.cell(row_idx, response_col).value = t_a
                ws.cell(row_idx, resolved_col).value = t_r
                if list_index < len(debug_rows):
                    debug_rows[list_index]["AckSource"] = response_pick.sender_email or response_pick.sender_name
                    debug_rows[list_index]["ResolvedSource"] = resolved_pick.sender_email or resolved_pick.sender_name
                    note_suffix = "; BlueStrictEpisodeAlignPool" if used_requester_pool else "; BlueStrictEpisodeAlign"
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}{note_suffix}"

                # Re-evaluate blue after strict align.
                a_dt2 = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                if not a_dt2:
                    continue
                a_ist2 = _to_ist(a_dt2)
                if (a_ist2 - c_ist) <= timedelta(minutes=16):
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

        _stage_timer_stop("blue_quoted_and_continuation_passes", blue_quoted_started_at, items=len(row_states))
        # Duplicate same-time de-collision (isolated):
        # If multiple rows for the same requester ended up with identical
        # created/response/resolved, move only secondary rows to another valid
        # requester episode without touching non-duplicate rows.
        # Requester ack-like quoted episode guard (global, targeted):
        # For ESS-only rows where requester-span fallback landed on an old ESS
        # episode and latest requester mail is an update/ack-like, recover the
        # latest quoted requester reply under that mail and use it as the true
        # episode anchor.
        episode_guard_started_at = _stage_timer_start()
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue
            notes = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes or "").lower()
            if "requester span(all-ack->ess)" not in notes_l:
                continue

            row_vals = automation_rows[list_index]
            requester = state.get("requester") or ""
            if not requester:
                continue
            notes_l = (debug_rows[list_index].get("Notes") or "").lower() if list_index < len(debug_rows) else ""
            if "ess-only; no non-ess request" in notes_l:
                continue
            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)

            subject_norm = (state.get("subject_norm") or "").lower()
            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=a_ist,
            )
            if not thread:
                continue
            requester_live = [e for e in thread if _req_match(e, requester) and _email_ist(e)]
            if not requester_live:
                continue
            requester_live.sort(key=lambda e: e.sent_time)
            latest_top = requester_live[-1]
            latest_top_ist = _email_ist(latest_top)
            if not latest_top_ist:
                continue
            if not (_ack_like(latest_top) or _ack_like_text_fallback(latest_top)):
                continue

            q_ist = _extract_quoted_requester_reply_ist(
                latest_top,
                requester,
                subject_norm,
                latest_top_ist - timedelta(days=14),
                latest_top_ist + timedelta(minutes=1),
            )
            if not q_ist:
                continue
            if q_ist >= latest_top_ist:
                continue
            if (latest_top_ist - q_ist) > timedelta(days=14):
                continue
            # Apply only when this meaningfully improves stale old picks.
            if q_ist <= (a_ist + timedelta(minutes=5)):
                continue

            t = _format_time(q_ist)
            if not t:
                continue
            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                q_ist,
                q_ist,
                q_ist,
                "RequesterAckLikeQuotedEpisodeGuard",
                "requester_ack",
            ):
                continue
            row_vals["Created Date & Time"] = t
            row_vals["Actual Response Date & Time"] = t
            row_vals["Actual Resolved Date & Time"] = t
            ws.cell(row_idx, created_col).value = t
            ws.cell(row_idx, response_col).value = t
            ws.cell(row_idx, resolved_col).value = t
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REPLY"
                debug_rows[list_index]["AckSource"] = "PARSED_FROM_QUOTED_REPLY"
                debug_rows[list_index]["ResolvedSource"] = "PARSED_FROM_QUOTED_REPLY"
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; RequesterAckLikeQuotedEpisodeGuard"

            if _row_has_blue_fill(row_idx):
                _set_row_fill(row_idx, clear_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

        # ESS hybrid ack episode guard (global, targeted):
        # For continuation/requester-follow-up rows that collapsed to one timestamp,
        # recover the actual episode when the thread shows:
        # quoted non-ESS request -> in-between ESS ack/update -> later final reply.
        # Keep this thread-local and cache-backed to avoid slowing the workbook fill.
        def _quoted_line_is_ess_local(from_line: str):
            return _quoted_from_line_is_ess_shared(from_line)

        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue

            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l = (notes_now or "").lower()
            if not (
                "requester follow-up (no in-between request)" in notes_l
                or "requester follow-up(top-only)" in notes_l
                or "ess-only continuation(top requester)" in notes_l
            ):
                continue

            row_vals = automation_rows[list_index]
            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and a_dt and r_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt)
            if not (c_ist and a_ist and r_ist):
                continue

            c_min = c_ist.replace(second=0, microsecond=0)
            a_min = a_ist.replace(second=0, microsecond=0)
            r_min = r_ist.replace(second=0, microsecond=0)
            if not (c_min == a_min == r_min):
                continue

            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            if not requester or not subject_norm:
                continue

            row_tokens = _match_tokens(subject_norm)
            row_id_tokens = _id_like_tokens(subject_norm)
            if not row_id_tokens:
                row_id_tokens = _id_like_tokens(row_vals.get("Description") or state.get("description") or "")

            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=r_ist,
            )
            if not thread:
                continue

            thread_msgs = []
            for e in thread:
                e_ist = _email_ist(e)
                if not e_ist or _system_like_sender(e):
                    continue
                if not _row_subject_match_email_quoted(e, subject_norm, row_tokens, row_id_tokens):
                    continue
                thread_msgs.append((e_ist, e))
            if len(thread_msgs) < 2:
                continue
            thread_msgs.sort(key=lambda item: item[0])

            final_reply_ist, final_reply = thread_msgs[-1]

            lane_episode = _lane_local_episode_from_reply(
                final_reply,
                final_reply_ist,
                [msg for _, msg in thread_msgs],
                requester,
                subject_norm,
                row_tokens,
                row_id_tokens,
            )

            req_pick = lane_episode.get("request") if lane_episode else None
            lane_ack_ist = lane_episode.get("ack") if lane_episode else None
            lane_ack_msg = lane_episode.get("ack_msg") if lane_episode else None

            if not req_pick:
                for _msg_ist, msg in thread_msgs:
                    quoted_blocks = _get_quoted_blocks_with_subject_cached(msg)
                    if not quoted_blocks:
                        quoted_blocks = _get_quoted_blocks_from_eml_path(getattr(msg, "path", ""))
                    for from_line, sent_ist, q_subj in quoted_blocks:
                        if not sent_ist or sent_ist >= final_reply_ist:
                            continue
                        if _quoted_line_is_ess_local(from_line) is not False:
                            continue
                        q_norm = normalize_subject(q_subj or "")
                        q_ids = _id_like_tokens(q_norm)
                        q_tokens = _match_tokens(q_norm)
                        if not _quoted_subject_confirms_row(
                            q_norm,
                            q_ids,
                            q_tokens,
                            subject_norm,
                            row_tokens,
                            row_id_tokens,
                        ):
                            continue
                        if req_pick is None or sent_ist > req_pick:
                            req_pick = sent_ist

            if not req_pick:
                continue

            ack_candidates = []
            if lane_ack_ist and lane_ack_msg and req_pick < lane_ack_ist <= final_reply_ist:
                ack_candidates.append((lane_ack_ist, lane_ack_msg))
            else:
                for msg_ist, msg in thread_msgs:
                    if msg_ist <= req_pick or msg_ist > final_reply_ist:
                        continue
                    if not _ess_sender(msg):
                        continue
                    if not (_ack_like(msg) or _ack_like_text_fallback(msg) or _ess_only_short_ack(msg)):
                        continue
                    if (msg_ist - req_pick) > timedelta(minutes=60):
                        continue
                    ack_candidates.append((msg_ist, msg))
            if not ack_candidates:
                continue

            ack_candidates.sort(key=lambda item: item[0])
            ack_ist, ack_msg = ack_candidates[0]

            resolved_candidates = [
                (msg_ist, msg)
                for msg_ist, msg in thread_msgs
                if msg_ist >= ack_ist and (msg_ist - ack_ist) <= timedelta(hours=48)
            ]
            if not resolved_candidates:
                continue
            resolved_ist, resolved_msg = resolved_candidates[-1]
            if resolved_ist < ack_ist:
                continue

            t_c = _format_time(req_pick)
            t_a = _format_time(ack_ist)
            t_r = _format_time(resolved_ist)
            if not (t_c and t_a and t_r):
                continue
            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                req_pick,
                ack_ist,
                resolved_ist,
                "EssHybridAckEpisodeGuard",
                "hybrid",
            ):
                continue

            row_vals["Created Date & Time"] = t_c
            row_vals["Actual Response Date & Time"] = t_a
            row_vals["Actual Resolved Date & Time"] = t_r
            ws.cell(row_idx, created_col).value = t_c
            ws.cell(row_idx, response_col).value = t_a
            ws.cell(row_idx, resolved_col).value = t_r
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = "PARSED_FROM_QUOTED_REQUEST"
                debug_rows[list_index]["AckSource"] = ack_msg.sender_email or ack_msg.sender_name or "ESS_HYBRID_ACK"
                debug_rows[list_index]["ResolvedSource"] = resolved_msg.sender_email or resolved_msg.sender_name or "ESS_HYBRID_ACK"
                debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; EssHybridAckEpisodeGuard"
            if (ack_ist.replace(second=0, microsecond=0) - req_pick.replace(second=0, microsecond=0)) <= timedelta(minutes=16):
                _set_row_fill(row_idx, clear_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

        _stage_timer_stop("requester_ack_and_hybrid_episode_guards", episode_guard_started_at, items=len(row_states) * 2)
        # Risky fallback collapse guard (narrow):
        # Inspect only risky fallback rows that collapsed all three timestamps to
        # one point, and only re-anchor when a clearly better episode is found.
        risky_and_duplicate_started_at = _stage_timer_start()
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue

            row_vals = automation_rows[list_index]
            precheck = _risk_guard_precheck(state, row_vals, list_index)
            if not precheck.get("has_risky_notes"):
                continue
            if not (precheck.get("suspicious_all_same") or precheck.get("suspicious_created_ack")):
                continue
            episode = _risk_guard_episode_for_row(state, row_vals, list_index)
            if not episode:
                continue

            t_c = _format_time(episode["created"])
            t_a = _format_time(episode["response"])
            t_r = _format_time(episode["resolved"])
            if not (t_c and t_a and t_r):
                continue

            current_c = row_vals.get("Created Date & Time") or ""
            current_a = row_vals.get("Actual Response Date & Time") or ""
            current_r = row_vals.get("Actual Resolved Date & Time") or ""
            if (current_c, current_a, current_r) == (t_c, t_a, t_r):
                continue

            state["risk_guard_episode"] = {
                "created": t_c,
                "response": t_a,
                "resolved": t_r,
                "created_src": episode["created_src"],
                "ack_src": episode["ack_src"],
                "resolved_src": episode["resolved_src"],
                "mode": episode["mode"],
            }
            if list_index < len(debug_rows):
                debug_rows[list_index]["Notes"] = (
                    f"{debug_rows[list_index].get('Notes','')}; "
                    f"RiskyFallbackValidator[{episode['mode']}]"
                )

        def _apply_risk_guard_episode_to_owner(state, owner_match, owner_tag):
            episode = state.get("risk_guard_episode")
            if not episode:
                return
            if (state.get("group_total") or 0) >= 2:
                return

            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                return
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                return

            row_vals = automation_rows[list_index]
            precheck = _risk_guard_precheck(state, row_vals, list_index)
            notes_l = precheck.get("notes_l", "")
            if not owner_match(notes_l):
                return

            c_ist = precheck.get("c_ist")
            a_ist = precheck.get("a_ist")
            r_ist = precheck.get("r_ist")
            if not (c_ist and a_ist and r_ist):
                return

            suspicious_all_same = precheck.get("suspicious_all_same", False)
            suspicious_created_ack = precheck.get("suspicious_created_ack", False)
            if not (suspicious_all_same or suspicious_created_ack):
                return

            t_c = episode.get("created")
            t_a = episode.get("response")
            t_r = episode.get("resolved")
            if not (t_c and t_a and t_r):
                return
            if (row_vals.get("Created Date & Time"), row_vals.get("Actual Response Date & Time"), row_vals.get("Actual Resolved Date & Time")) == (t_c, t_a, t_r):
                return

            e_c_dt = _parse_time_str(t_c)
            e_a_dt = _parse_time_str(t_a)
            e_r_dt = _parse_time_str(t_r)
            if not (e_c_dt and e_a_dt and e_r_dt):
                return
            e_c_ist = _to_ist(e_c_dt)
            e_a_ist = _to_ist(e_a_dt)
            e_r_ist = _to_ist(e_r_dt)
            if not (e_c_ist and e_a_ist and e_r_ist):
                return
            if not (e_c_ist < e_a_ist <= e_r_ist):
                return
            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                e_c_ist,
                e_a_ist,
                e_r_ist,
                owner_tag,
                "risk",
            ):
                return

            row_vals["Created Date & Time"] = t_c
            row_vals["Actual Response Date & Time"] = t_a
            row_vals["Actual Resolved Date & Time"] = t_r
            ws.cell(row_idx, created_col).value = t_c
            ws.cell(row_idx, response_col).value = t_a
            ws.cell(row_idx, resolved_col).value = t_r
            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = episode.get("created_src", "")
                debug_rows[list_index]["AckSource"] = episode.get("ack_src", "")
                debug_rows[list_index]["ResolvedSource"] = episode.get("resolved_src", "")
                debug_rows[list_index]["Notes"] = (
                    f"{debug_rows[list_index].get('Notes','')}; "
                    f"{owner_tag}[{episode.get('mode', '')}]"
                )
            if _row_has_blue_fill(row_idx):
                _set_row_fill(row_idx, clear_fill)
                if list_index < len(debug_rows):
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

        for state in row_states:
            if state.get("occurrence_locked"):
                continue
            _apply_risk_guard_episode_to_owner(
                state,
                lambda notes_l: "quotedrequestonly" in notes_l,
                "QuotedRiskValidatorReanchor",
            )
            _apply_risk_guard_episode_to_owner(
                state,
                lambda notes_l: "requester span(" in notes_l,
                "RequesterSpanRiskValidatorReanchor",
            )
            _apply_risk_guard_episode_to_owner(
                state,
                lambda notes_l: (
                    "ess-only; no non-ess request" in notes_l
                    or "esscontinuationguard[" in notes_l
                ),
                "EssRiskValidatorReanchor",
            )

        def _system_notification_episode_for_row(state, row_vals, list_index):
            requester = state.get("requester") or ""
            subject_norm = (state.get("subject_norm") or "").lower()
            if not requester or not subject_norm:
                return None

            current_c = _to_ist(_parse_time_str(row_vals.get("Created Date & Time")))
            current_a = _to_ist(_parse_time_str(row_vals.get("Actual Response Date & Time")))
            current_r = _to_ist(_parse_time_str(row_vals.get("Actual Resolved Date & Time")))
            reference_ist = current_r or current_a or current_c

            base_thread = state.get("thread") or []
            thread = _expanded_thread(
                subject_norm,
                base_thread,
                requester,
                include_non_ess=True,
                reference_ist=reference_ist,
            ) or []
            if not thread:
                return None

            row_tokens = _match_tokens(subject_norm)
            row_id_tokens = _id_like_tokens(subject_norm)
            if not row_id_tokens:
                desc_text = row_vals.get("Description") or state.get("description") or ""
                row_id_tokens = _id_like_tokens(desc_text)

            matched = []
            seen = set()
            for e in thread:
                if not e or id(e) in seen:
                    continue
                seen.add(id(e))
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                if not _row_subject_match_email_quoted(e, subject_norm, row_tokens, row_id_tokens):
                    continue
                matched.append((e_ist, e))
            if not matched:
                return None
            matched.sort(key=lambda item: item[0])

            system_msgs = [(e_ist, e) for e_ist, e in matched if _system_like_sender(e)]
            requester_replies = [
                (e_ist, e)
                for e_ist, e in matched
                if _req_match(e, requester)
                and _is_real_reply_candidate(e)
                and (not _system_like_sender(e))
            ]
            if not system_msgs or not requester_replies:
                return None

            def _has_stronger_live_before(cap_ist):
                for e_ist, e in matched:
                    if e_ist >= cap_ist:
                        break
                    if _ess_sender(e) or _system_like_sender(e):
                        continue
                    if _req_match(e, requester):
                        continue
                    return True
                return False

            def _has_stronger_quoted_before(cap_ist):
                for _e_ist, e in matched:
                    q_ist = _extract_quoted_request_before_ist(e, subject_norm, cap_ist)
                    if q_ist and q_ist < cap_ist:
                        return True
                return False

            for sys_ist, sys_msg in reversed(system_msgs):
                ack_pick = None
                for e_ist, e in matched:
                    if e_ist <= sys_ist:
                        continue
                    if (not _ess_sender(e)) or _system_like_sender(e):
                        continue
                    if not _is_shared_ack_candidate(e):
                        continue
                    ack_pick = (e_ist, e)
                    break

                if ack_pick:
                    ack_ist, ack_msg = ack_pick
                    if _has_stronger_live_before(ack_ist) or _has_stronger_quoted_before(ack_ist):
                        continue
                    consultant_after = [e for e_ist, e in requester_replies if e_ist > ack_ist]
                    resolved_msg = _pick_reply_after_ack(consultant_after, ack_ist, requester_name=requester) if consultant_after else None
                    resolved_ist = _email_ist(resolved_msg) if resolved_msg else None
                    if not resolved_ist or resolved_ist < ack_ist:
                        continue
                    return {
                        "created": sys_ist,
                        "response": ack_ist,
                        "resolved": resolved_ist,
                        "created_src": sys_msg.sender_email or sys_msg.sender_name or "SYSTEM_NOTIFICATION",
                        "ack_src": ack_msg.sender_email or ack_msg.sender_name or "SYSTEM_ACK",
                        "resolved_src": resolved_msg.sender_email or resolved_msg.sender_name or "SYSTEM_RESOLVED",
                        "mode": "with-ack",
                    }

                consultant_after = [(e_ist, e) for e_ist, e in requester_replies if e_ist > sys_ist]
                if not consultant_after:
                    continue
                consultant_ist, consultant_msg = consultant_after[-1]
                if _has_stronger_live_before(consultant_ist) or _has_stronger_quoted_before(consultant_ist):
                    continue
                return {
                    "created": consultant_ist,
                    "response": consultant_ist,
                    "resolved": consultant_ist,
                    "created_src": consultant_msg.sender_email or consultant_msg.sender_name or "SYSTEM_NOACK_ALL_SAME",
                    "ack_src": "ACK NOT FOUND",
                    "resolved_src": consultant_msg.sender_email or consultant_msg.sender_name or "SYSTEM_NOACK_ALL_SAME",
                    "mode": "no-ack-all-same",
                }

            return None

        system_notif_guard_started_at = _stage_timer_start()
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ") or state.get("occurrence_locked"):
                continue

            row_vals = automation_rows[list_index]
            episode = _system_notification_episode_for_row(state, row_vals, list_index)
            if not episode:
                continue

            t_c = _format_time(episode["created"])
            t_a = _format_time(episode["response"])
            t_r = _format_time(episode["resolved"])
            if not (t_c and t_a and t_r):
                continue
            if (row_vals.get("Created Date & Time"), row_vals.get("Actual Response Date & Time"), row_vals.get("Actual Resolved Date & Time")) == (t_c, t_a, t_r):
                continue

            if not _allow_guard_rewrite(
                row_vals,
                list_index,
                episode["created"],
                episode["response"],
                episode["resolved"],
                "SystemNotificationEpisodeGuard",
                "system_notification",
            ):
                continue

            row_vals["Created Date & Time"] = t_c
            row_vals["Actual Response Date & Time"] = t_a
            row_vals["Actual Resolved Date & Time"] = t_r
            ws.cell(row_idx, created_col).value = t_c
            ws.cell(row_idx, response_col).value = t_a
            ws.cell(row_idx, resolved_col).value = t_r

            if list_index < len(debug_rows):
                debug_rows[list_index]["CreatedSource"] = episode.get("created_src", "")
                debug_rows[list_index]["AckSource"] = episode.get("ack_src", "")
                debug_rows[list_index]["ResolvedSource"] = episode.get("resolved_src", "")
                debug_rows[list_index]["Notes"] = (
                    f"{debug_rows[list_index].get('Notes','')}; "
                    f"SystemNotificationEpisode[{episode.get('mode', '')}]"
                )

            if _row_has_blue_fill(row_idx):
                created_gap = episode["response"].replace(second=0, microsecond=0) - episode["created"].replace(second=0, microsecond=0)
                if created_gap <= timedelta(minutes=16):
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

        _stage_timer_stop("system_notification_episode_guard", system_notif_guard_started_at, items=len(row_states))

        duplicate_groups = {}
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if state.get("is_dep_req") or state.get("is_dep_succ"):
                continue
            if state.get("occurrence_locked"):
                continue
            notes_now = debug_rows[list_index].get("Notes", "") if list_index < len(debug_rows) else ""
            notes_l_now = (notes_now or "").lower()
            allow_nonblue_distinct = (
                "ess-only; no non-ess request" in notes_l_now
                or "requester follow-up" in notes_l_now
                or "esscontinuationguard[allthree]" in notes_l_now
            )
            if blue_only_post_resolver and not _row_has_blue_fill(row_idx) and not allow_nonblue_distinct:
                continue
            row_vals = automation_rows[list_index]
            req = _requester_key(state.get("requester") or "")
            service_bucket = state.get("service_bucket") or _subject_service_bucket(
                (state.get("subject_norm") or "").lower(),
                state.get("service_no") or "",
            )
            c = row_vals.get("Created Date & Time") or ""
            a = row_vals.get("Actual Response Date & Time") or ""
            r = row_vals.get("Actual Resolved Date & Time") or ""
            if not (req and c and a and r):
                continue
            subject_key = (state.get("subject_norm") or "").lower()
            if allow_nonblue_distinct and subject_key and c == a:
                # For ESS-like repeated rows, a prior post-pass can move only
                # resolved while leaving created/response on the same
                # occurrence. Group by requester+subject+created/ack so the
                # later distinct-occurrence repair can still re-align all
                # three timestamps together.
                key = ("ess_like_created_ack", req, subject_key, service_bucket, c, a)
            else:
                key = ("triplet", req, service_bucket, c, a, r)
            duplicate_groups.setdefault(key, []).append(state)

        for dkey, group in duplicate_groups.items():
            if len(group) < 2:
                continue

            # ESS-only distinct occurrence mapping:
            # If all rows in the duplicate group are ESS-only/requester follow-up
            # and we have enough distinct requester replies, map each row to a
            # unique reply by occurrence (chronological).
            ess_like_group = True
            for s in group:
                li = s.get("list_index")
                notes = debug_rows[li].get("Notes", "") if li is not None and li < len(debug_rows) else ""
                notes_l = (notes or "").lower()
                if not (
                    "ess-only; no non-ess request" in notes_l
                    or "requester follow-up" in notes_l
                    or "esscontinuationguard[allthree]" in notes_l
                ):
                    ess_like_group = False
                    break
            if ess_like_group:
                group_sorted = sorted(group, key=lambda x: x.get("row_index") or 10**9)
                shared_group_picks = []
                for s in group_sorted:
                    shared_plan = _preferred_shared_occurrence_plan(s, require_override_for_all_ack=True)
                    if not shared_plan or shared_plan.get("group_size", 0) < 2:
                        shared_group_picks = []
                        break
                    slot_index = shared_plan.get("slot_index", 0)
                    pool = shared_plan.get("pool") or []
                    if slot_index >= len(pool):
                        shared_group_picks = []
                        break
                    shared_group_picks.append((s, pool[slot_index], shared_plan))
                if len(shared_group_picks) == len(group_sorted):
                    for s, pick, _shared_plan in shared_group_picks:
                        li = s.get("list_index")
                        ri = s.get("row_index")
                        if li is None or li >= len(automation_rows) or not ri:
                            continue
                        t = _format_time(getattr(pick, "sent_time", None))
                        if not t:
                            continue
                        row_vals = automation_rows[li]
                        row_vals["Created Date & Time"] = t
                        row_vals["Actual Response Date & Time"] = t
                        row_vals["Actual Resolved Date & Time"] = t
                        ws.cell(ri, created_col).value = t
                        ws.cell(ri, response_col).value = t
                        ws.cell(ri, resolved_col).value = t
                        if li < len(debug_rows):
                            who = pick.sender_email or pick.sender_name
                            debug_rows[li]["CreatedSource"] = who
                            debug_rows[li]["AckSource"] = who
                            debug_rows[li]["ResolvedSource"] = who
                            debug_rows[li]["Notes"] = f"{debug_rows[li].get('Notes','')}; DistinctOccurrenceMap"
                    continue

                anchor_state = group_sorted[0]
                anchor_li = anchor_state.get("list_index")
                requester = anchor_state.get("requester") or ""
                subject_norm = (anchor_state.get("subject_norm") or "").lower()
                base_thread = anchor_state.get("thread") or []
                if requester and subject_norm and base_thread and anchor_li is not None and anchor_li < len(automation_rows):
                    row_tokens = _match_tokens(subject_norm)
                    group_list_indexes = {
                        s.get("list_index")
                        for s in group_sorted
                        if s.get("list_index") is not None
                    }
                    used_outside_group = set()
                    for other_state in row_states:
                        other_li = other_state.get("list_index")
                        if other_li is None or other_li in group_list_indexes or other_li >= len(automation_rows):
                            continue
                        if _requester_key(other_state.get("requester") or "") != _requester_key(requester):
                            continue
                        other_subject = (other_state.get("subject_norm") or "").lower()
                        if other_subject != subject_norm:
                            continue
                        other_vals = automation_rows[other_li]
                        other_a_dt = _parse_time_str(other_vals.get("Actual Response Date & Time"))
                        other_a_ist = _to_ist(other_a_dt) if other_a_dt else None
                        if other_a_ist:
                            used_outside_group.add(other_a_ist.replace(second=0, microsecond=0))
                    reply_pool = []
                    for e in base_thread:
                        if not _req_match(e, requester):
                            continue
                        if _ack_like(e) or _ack_like_text_fallback(e):
                            continue
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if row_tokens:
                            s_norm = normalize_subject(getattr(e, "subject", "") or "")
                            s_tokens = _match_tokens(s_norm)
                            score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                            contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                            if score < 0.45 and not contains:
                                continue
                        reply_pool.append(e)
                    reply_pool.sort(key=lambda e: e.sent_time)

                    # De-duplicate by minute to avoid mapping to identical timestamps.
                    unique_pool = []
                    seen_ts = set()
                    for e in reply_pool:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        key = e_ist.replace(second=0, microsecond=0)
                        if key in seen_ts:
                            continue
                        seen_ts.add(key)
                        unique_pool.append(e)

                    # Prefer replies in the same month as the anchor row.
                    a_dt = _parse_time_str(automation_rows[anchor_li].get("Actual Response Date & Time"))
                    a_ist = _to_ist(a_dt) if a_dt else None
                    month_pool = []
                    if a_ist:
                        for e in unique_pool:
                            e_ist = _email_ist(e)
                            if e_ist and e_ist.year == a_ist.year and e_ist.month == a_ist.month:
                                month_pool.append(e)

                    pool = month_pool if len(month_pool) >= len(group_sorted) else unique_pool
                    pool_without_used = []
                    for e in pool:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        tkey = e_ist.replace(second=0, microsecond=0)
                        if tkey in used_outside_group:
                            continue
                        pool_without_used.append(e)
                    if len(pool_without_used) >= len(group_sorted):
                        pool = pool_without_used
                    if len(pool) >= len(group_sorted):
                        for idx, s in enumerate(group_sorted):
                            li = s.get("list_index")
                            ri = s.get("row_index")
                            if li is None or li >= len(automation_rows) or not ri:
                                continue
                            pick = pool[min(idx, len(pool) - 1)]
                            t = _format_time(pick.sent_time)
                            if not t:
                                continue
                            row_vals = automation_rows[li]
                            row_vals["Created Date & Time"] = t
                            row_vals["Actual Response Date & Time"] = t
                            row_vals["Actual Resolved Date & Time"] = t
                            ws.cell(ri, created_col).value = t
                            ws.cell(ri, response_col).value = t
                            ws.cell(ri, resolved_col).value = t
                            if li < len(debug_rows):
                                who = pick.sender_email or pick.sender_name
                                debug_rows[li]["CreatedSource"] = who
                                debug_rows[li]["AckSource"] = who
                                debug_rows[li]["ResolvedSource"] = who
                                debug_rows[li]["Notes"] = f"{debug_rows[li].get('Notes','')}; DistinctOccurrenceMap"
                        continue

            # Keep one stable anchor row untouched:
            # prefer non-ambiguous match; otherwise earliest row index.
            anchor_state = None
            for s in group:
                li = s.get("list_index")
                notes = debug_rows[li].get("Notes", "") if li is not None and li < len(debug_rows) else ""
                notes_l = (notes or "").lower()
                ambiguous = ("match=score:" in notes_l) or ("ambiguous" in notes_l)
                if not ambiguous:
                    anchor_state = s
                    break
            if anchor_state is None:
                anchor_state = min(group, key=lambda x: x.get("row_index") or 10**9)
            anchor_subject_norm = (anchor_state.get("subject_norm") or "").lower()
            anchor_tokens = _match_tokens(anchor_subject_norm)

            used_response_ists = set()
            anchor_li = anchor_state.get("list_index")
            anchor_a_dt = _parse_time_str(automation_rows[anchor_li].get("Actual Response Date & Time")) if anchor_li is not None else None
            if anchor_a_dt:
                used_response_ists.add(_to_ist(anchor_a_dt).replace(second=0, microsecond=0))

            for s in group:
                li = s.get("list_index")
                ri = s.get("row_index")
                if li is None or li >= len(automation_rows) or not ri:
                    continue
                if s is anchor_state:
                    continue

                row_vals = automation_rows[li]
                requester = s.get("requester") or ""
                subject_norm = (s.get("subject_norm") or "").lower()
                # Same consultant + same-subject family only.
                if anchor_tokens:
                    row_tokens_for_group = _match_tokens(subject_norm)
                    sim_group = _token_overlap_score(anchor_tokens, row_tokens_for_group) if row_tokens_for_group else 0.0
                    contains_group = bool(
                        anchor_subject_norm and subject_norm and (
                            anchor_subject_norm in subject_norm or subject_norm in anchor_subject_norm
                        )
                    )
                    if sim_group < 0.40 and not contains_group:
                        continue
                a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                if not (requester and subject_norm and a_dt):
                    continue
                a_ist = _to_ist(a_dt)
                c_ist = _to_ist(c_dt) if c_dt else None
                baseline_date = s.get("baseline_created_date")
                created_src_now = debug_rows[li].get("CreatedSource", "") if li < len(debug_rows) else ""

                base_thread = s.get("thread") or []
                thread = _expanded_thread(
                    subject_norm,
                    base_thread,
                    requester,
                    include_non_ess=True,
                    reference_ist=a_ist,
                )
                if not thread:
                    continue

                row_tokens = _match_tokens(subject_norm)
                def _collect_candidates(max_days: int, enforce_baseline: bool):
                    out = []
                    for e in thread:
                        e_ist = _email_ist(e)
                        if not e_ist:
                            continue
                        if not _req_match(e, requester):
                            continue
                        if _ack_like(e) or _ack_like_text_fallback(e):
                            continue
                        if abs((e_ist - a_ist).total_seconds()) <= 60:
                            continue
                        if abs((e_ist - a_ist).total_seconds()) > (max_days * 24 * 3600):
                            continue
                        if enforce_baseline and baseline_date and abs((e_ist.date() - baseline_date).days) > 3:
                            continue
                        if row_tokens:
                            s_norm = normalize_subject(getattr(e, "subject", "") or "")
                            s_tokens = _match_tokens(s_norm)
                            score = _token_overlap_score(row_tokens, s_tokens) if s_tokens else 0.0
                            contains = bool(subject_norm and s_norm and (subject_norm in s_norm or s_norm in subject_norm))
                            if score < 0.45 and not contains:
                                continue
                        out.append(e)
                    return out

                candidates = _collect_candidates(max_days=14, enforce_baseline=True)
                if not candidates:
                    candidates = _collect_candidates(max_days=45, enforce_baseline=False)
                if not candidates:
                    continue

                # Pick nearest candidate not already used by anchored duplicates.
                candidates.sort(key=lambda e: abs((_email_ist(e) - a_ist).total_seconds()))
                pick = None
                for e in candidates:
                    tkey = _email_ist(e).replace(second=0, microsecond=0)
                    if tkey in used_response_ists:
                        continue
                    pick = e
                    break
                if not pick:
                    continue

                t = _format_time(pick.sent_time)
                if not t:
                    continue
                row_vals["Actual Response Date & Time"] = t
                row_vals["Actual Resolved Date & Time"] = t
                pick_ist = _email_ist(pick)
                # If created was parsed and sits after the selected episode, align it.
                if (
                    pick_ist
                    and c_ist
                    and pick_ist < c_ist
                    and isinstance(created_src_now, str)
                    and created_src_now.startswith("PARSED_FROM_")
                ):
                    row_vals["Created Date & Time"] = t
                    ws.cell(ri, created_col).value = t
                ws.cell(ri, response_col).value = t
                ws.cell(ri, resolved_col).value = t
                used_response_ists.add(_email_ist(pick).replace(second=0, microsecond=0))
                if li < len(debug_rows):
                    who = pick.sender_email or pick.sender_name
                    if (
                        pick_ist
                        and c_ist
                        and pick_ist < c_ist
                        and isinstance(created_src_now, str)
                        and created_src_now.startswith("PARSED_FROM_")
                    ):
                        debug_rows[li]["CreatedSource"] = who
                    debug_rows[li]["AckSource"] = who
                    debug_rows[li]["ResolvedSource"] = who
                    debug_rows[li]["Notes"] = f"{debug_rows[li].get('Notes','')}; DuplicateDecollision"

        if workbook_kind == "incident_business":
            final_occurrence_lock_started_at = _stage_timer_start()
            final_occurrence_locked_rows = 0
            for state in row_states:
                list_index = state.get("list_index")
                row_idx = state.get("row_index")
                if list_index is None or list_index >= len(automation_rows) or not row_idx:
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                notes_l = (debug_rows[list_index].get("Notes", "") or "").lower() if list_index < len(debug_rows) else ""
                if not _is_all_ack_to_ess_notes(notes_l):
                    continue
                row_vals = automation_rows[list_index]
                shared_occ_plan = _preferred_shared_occurrence_plan(state, require_override_for_all_ack=True)
                if not shared_occ_plan or not _is_authoritative_occurrence_lane(shared_occ_plan.get("lane_kind") or ""):
                    continue

                pick_when = shared_occ_plan.get("pick_when")
                if not pick_when:
                    continue
                target_triplet = (
                    pick_when.replace(second=0, microsecond=0),
                    pick_when.replace(second=0, microsecond=0),
                    pick_when.replace(second=0, microsecond=0),
                )

                current_triplet = _current_row_triplet_ist(row_vals)
                if current_triplet != target_triplet:
                    t = _format_time(pick_when)
                    if not t:
                        continue

                    row_vals["Created Date & Time"] = t
                    row_vals["Actual Response Date & Time"] = t
                    row_vals["Actual Resolved Date & Time"] = t
                    ws.cell(row_idx, created_col).value = t
                    ws.cell(row_idx, response_col).value = t
                    ws.cell(row_idx, resolved_col).value = t
                    _set_row_fill(row_idx, clear_fill)

                    state["shared_decision"] = {
                        "owner": "shared_occurrence",
                        "row_type": shared_occ_plan.get("lane_kind") or "ess_over_ess",
                        "occurrence_slot": shared_occ_plan.get("slot_index", 0),
                        "lane_time": pick_when,
                        "fill_style": "all_three_same",
                        "confidence": "strong",
                        "triplet": target_triplet,
                    }

                    if list_index < len(debug_rows):
                        who = (
                            debug_rows[list_index].get("ResolvedSource")
                            or debug_rows[list_index].get("AckSource")
                            or debug_rows[list_index].get("CreatedSource")
                            or "SHARED_OCCURRENCE"
                        )
                        debug_rows[list_index]["CreatedSource"] = who
                        debug_rows[list_index]["AckSource"] = who
                        debug_rows[list_index]["ResolvedSource"] = who
                        notes_now = debug_rows[list_index].get("Notes", "") or ""
                        if "ESSContinuationGuard[AllThreeStrictEssOnly]" not in notes_now:
                            debug_rows[list_index]["Notes"] = f"{notes_now}; ESSContinuationGuard[AllThreeStrictEssOnly]"
                elif _should_preserve_occurrence_same_time(state, row_vals, shared_occ_plan):
                    state["shared_decision"] = {
                        "owner": "shared_occurrence",
                        "row_type": shared_occ_plan.get("lane_kind") or "ess_over_ess",
                        "occurrence_slot": shared_occ_plan.get("slot_index", 0),
                        "lane_time": pick_when,
                        "fill_style": "all_three_same",
                        "confidence": "strong",
                        "triplet": target_triplet,
                    }
                    if list_index < len(debug_rows):
                        notes_now = debug_rows[list_index].get("Notes", "") or ""
                        if "ESSContinuationGuard[AllThreeStrictEssOnly]" not in notes_now:
                            debug_rows[list_index]["Notes"] = f"{notes_now}; ESSContinuationGuard[AllThreeStrictEssOnly]"

                if _lock_occurrence_row(
                    state,
                    row_vals,
                    list_index,
                    "OccurrenceLocked",
                    triplet=target_triplet,
                ):
                    final_occurrence_locked_rows += 1

            _stage_timer_stop("final_occurrence_lock_apply", final_occurrence_lock_started_at, items=final_occurrence_locked_rows)

        _stage_timer_stop("risky_and_duplicate_repair_passes", risky_and_duplicate_started_at, items=len(row_states))
        if workbook_kind == "task_business":
            raw_task_candidate_cache = {}
            task_request_anchor_cache = {}
            task_msgs_by_core = {}
            task_msgs_by_core_day = {}
            task_non_ess_msgs_by_core = {}
            task_non_ess_msgs_by_core_day = {}

            def _raw_task_candidates_for_ids(row_id_tokens_set: set):
                key = tuple(sorted(row_id_tokens_set))
                if key in raw_task_candidate_cache:
                    return raw_task_candidate_cache[key]
                out = []
                for path in _find_eml_paths_by_id(row_id_tokens_set):
                    header_summary = _get_eml_header_summary(path)
                    if not header_summary or not header_summary.get("sent_dt"):
                        continue
                    out.append((
                        header_summary.get("subject_raw") or "",
                        header_summary.get("sent_dt"),
                        header_summary.get("sender_email") or "",
                        header_summary.get("sender_name") or "",
                    ))
                raw_task_candidate_cache[key] = out
                return out

            for e in emails:
                e_ist = _email_ist(e)
                if not e_ist:
                    continue
                s_core = _task_subject_core(_subject_norm_cached(getattr(e, "subject", "") or ""))
                if not s_core:
                    continue
                key_day = (s_core, e_ist.date())
                task_msgs_by_core.setdefault(s_core, []).append(e)
                task_msgs_by_core_day.setdefault(key_day, []).append(e)
                if not _ess_sender(e) and not _system_like_sender(e):
                    task_non_ess_msgs_by_core.setdefault(s_core, []).append(e)
                    task_non_ess_msgs_by_core_day.setdefault(key_day, []).append(e)

            for bucket in (task_msgs_by_core, task_msgs_by_core_day, task_non_ess_msgs_by_core, task_non_ess_msgs_by_core_day):
                for key in list(bucket.keys()):
                    bucket[key].sort(key=lambda e: _email_ist(e) or datetime.min)

            task_post_started_at = _stage_timer_start()
            for state in row_states:
                list_index = state.get("list_index")
                row_idx = state.get("row_index")
                if list_index is None or list_index >= len(automation_rows) or not row_idx:
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                created_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                ack_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                resolved_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                if not (created_dt and ack_dt):
                    continue
                created_ist = _to_ist(created_dt)
                ack_ist = _to_ist(ack_dt)
                if not created_ist or not ack_ist or ack_ist <= created_ist:
                    if not (created_ist and ack_ist):
                        continue
                created_minute = created_ist.replace(second=0, microsecond=0)
                ack_minute = ack_ist.replace(second=0, microsecond=0)
                resolved_ist = _to_ist(resolved_dt) if resolved_dt else None
                resolved_minute = resolved_ist.replace(second=0, microsecond=0) if resolved_ist else None
                is_blue_row = _row_has_blue_fill(row_idx)
                task_same_all = bool(resolved_minute and created_minute == ack_minute == resolved_minute)
                if not is_blue_row and not task_same_all:
                    continue

                created_src_now = debug_rows[list_index].get("CreatedSource", "") if list_index < len(debug_rows) else ""
                subject_norm = (state.get("subject_norm") or "").lower()
                parsed_created = isinstance(created_src_now, str) and created_src_now.startswith("PARSED_FROM_")
                if not parsed_created and not task_same_all:
                    continue
                if (ack_minute - created_minute) <= timedelta(minutes=16) and not task_same_all:
                    continue

                subject_core = _task_subject_core(subject_norm)
                if not subject_core:
                    continue
                row_id_tokens = _id_like_tokens(subject_norm)
                if not row_id_tokens:
                    row_id_tokens = _id_like_tokens(row_vals.get("Description") or state.get("description") or "")
                if not row_id_tokens:
                    continue
                same_all_lower_bound = ack_ist - timedelta(hours=24) if task_same_all else None

                def _task_pick_in_window(sent_ist, sent_minute):
                    if not sent_ist or not sent_minute:
                        return False
                    if sent_minute > ack_minute:
                        return False
                    if task_same_all:
                        return sent_ist >= same_all_lower_bound
                    return sent_ist.date() == ack_ist.date()

                candidate_picks = []

                def _push_task_pick(sent_ist, src, note, priority: int):
                    if not sent_ist:
                        return
                    sent_minute = sent_ist.replace(second=0, microsecond=0)
                    if not _task_pick_in_window(sent_ist, sent_minute):
                        return
                    if not task_same_all and sent_minute <= created_minute:
                        return
                    candidate_picks.append((sent_ist, priority, src, note))

                task_request_pool = (
                    task_msgs_by_core.get(subject_core, [])
                    if task_same_all
                    else task_msgs_by_core_day.get((subject_core, ack_ist.date()), [])
                )
                if task_request_pool:
                    anchor_cache_key = (
                        subject_core,
                        ack_ist.date().isoformat(),
                        bool(task_same_all),
                        tuple(sorted(row_id_tokens)),
                        ack_minute.isoformat(),
                    )
                    request_anchor = task_request_anchor_cache.get(anchor_cache_key)
                    if request_anchor is None:
                        request_anchor = _best_request_anchor_from_sources(
                            task_request_pool,
                            subject_norm,
                            _match_tokens(subject_norm),
                            row_id_tokens,
                            ack_ist,
                            timedelta(hours=48) if task_same_all else timedelta(hours=24),
                        )
                        task_request_anchor_cache[anchor_cache_key] = request_anchor
                    if request_anchor:
                        _push_task_pick(
                            request_anchor.get("when"),
                            request_anchor.get("src") or "TASK_REQUEST_ANCHOR",
                            "TaskRequestAnchor",
                            0,
                        )

                live_pick_pool = (
                    task_non_ess_msgs_by_core.get(subject_core, [])
                    if task_same_all
                    else task_non_ess_msgs_by_core_day.get((subject_core, ack_ist.date()), [])
                )
                for e in live_pick_pool:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    _push_task_pick(
                        e_ist,
                        e.sender_email or e.sender_name or "",
                        "TaskNearestPreAckRequest",
                        0,
                    )

                if row_id_tokens:
                    for subject_raw, sent_dt, sender_email, sender_name in _raw_task_candidates_for_ids(row_id_tokens):
                        sent_ist = _to_ist(sent_dt)
                        if sender_email and (
                            sender_email in ess_email_set
                            or ("@" in sender_email and sender_email.split("@", 1)[-1] in ess_domain_set)
                        ):
                            continue
                        s_core = _task_subject_core(subject_raw)
                        if not s_core or s_core != subject_core:
                            continue
                        _push_task_pick(
                            sent_ist,
                            sender_email or sender_name or "RAW_EML_SUBJECT_MATCH",
                            "TaskNearestPreAckRequestRaw",
                            2,
                        )

                def _quoted_from_line_is_ess(from_line: str):
                    return _quoted_from_line_is_ess_shared(from_line)

                quoted_live_pool = (
                    task_msgs_by_core.get(subject_core, [])
                    if task_same_all
                    else task_msgs_by_core_day.get((subject_core, ack_ist.date()), [])
                )
                for e in quoted_live_pool:
                    e_ist = _email_ist(e)
                    e_minute = e_ist.replace(second=0, microsecond=0) if e_ist else None
                    if not _task_pick_in_window(e_ist, e_minute):
                        continue
                    for from_line, sent_ist, q_subj in _get_quoted_blocks_with_subject_cached(e):
                        q_core = _task_subject_core(q_subj or "")
                        if not q_core or q_core != subject_core:
                            continue
                        is_ess = _quoted_from_line_is_ess(from_line)
                        if is_ess is not False:
                            continue
                        _push_task_pick(
                            sent_ist,
                            "PARSED_FROM_QUOTED_REQUEST",
                            "TaskQuotedPreAckRequest",
                            1,
                        )

                if subject_core and row_id_tokens:
                    raw_paths = _find_eml_paths_by_id(row_id_tokens)
                    for path in raw_paths:
                        header_summary = _get_eml_header_summary(path)
                        if not header_summary:
                            continue
                        s_core = _task_subject_core(header_summary.get("subject_raw") or "")
                        if not s_core or s_core != subject_core:
                            continue
                        for from_line, sent_ist, q_subj in _get_quoted_blocks_from_eml_path(path):
                            q_core = _task_subject_core(q_subj or "")
                            if not q_core or q_core != subject_core:
                                continue
                            is_ess = _quoted_from_line_is_ess(from_line)
                            if is_ess is not False:
                                continue
                            _push_task_pick(
                                sent_ist,
                                "PARSED_FROM_RAW_QUOTED_TASK",
                                "TaskQuotedPreAckRequestRaw",
                                3,
                            )

                best_pick_ist = None
                best_pick_src = ""
                best_pick_note = ""
                best_pick_minute = None
                if candidate_picks:
                    candidate_picks.sort(
                        key=lambda item: (
                            item[0].replace(second=0, microsecond=0),
                            -item[1],
                            item[0],
                        ),
                        reverse=True,
                    )
                    best_pick_ist, _best_priority, best_pick_src, best_pick_note = candidate_picks[0]
                    best_pick_minute = best_pick_ist.replace(second=0, microsecond=0)

                if not best_pick_ist:
                    continue
                if task_same_all:
                    if best_pick_minute >= ack_minute:
                        continue
                else:
                    if best_pick_minute <= created_minute or best_pick_minute > ack_minute:
                        continue

                t_c = _format_time(best_pick_ist)
                if not t_c:
                    continue
                row_vals["Created Date & Time"] = t_c
                ws.cell(row_idx, created_col).value = t_c
                if list_index < len(debug_rows):
                    debug_rows[list_index]["CreatedSource"] = best_pick_src
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; {best_pick_note}"
                if (ack_minute - best_pick_minute) <= timedelta(minutes=16):
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

            # Task local ack refinement:
            # For suspicious task rows, prefer a real in-between ESS ack from the
            # same local episode instead of leaving response collapsed to resolved
            # or to a later non-ack mail. Keep this cache-backed and day-local.
            for state in row_states:
                list_index = state.get("list_index")
                row_idx = state.get("row_index")
                if list_index is None or list_index >= len(automation_rows) or not row_idx:
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                requester = state.get("requester") or ""
                subject_norm = (state.get("subject_norm") or "").lower()
                created_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                ack_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                resolved_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                if not (created_dt and ack_dt and resolved_dt):
                    continue
                created_ist = _to_ist(created_dt)
                ack_ist = _to_ist(ack_dt)
                resolved_ist = _to_ist(resolved_dt)
                if not (created_ist and ack_ist and resolved_ist):
                    continue
                created_minute = created_ist.replace(second=0, microsecond=0)
                ack_minute = ack_ist.replace(second=0, microsecond=0)
                resolved_minute = resolved_ist.replace(second=0, microsecond=0)
                is_blue_row = _row_has_blue_fill(row_idx)
                if not (is_blue_row or ack_minute == resolved_minute):
                    continue
                if resolved_minute <= created_minute:
                    continue

                subject_core = _task_subject_core(subject_norm)
                if not subject_core:
                    continue

                episode_pool = task_msgs_by_core_day.get((subject_core, resolved_ist.date()), [])
                if not episode_pool:
                    episode_pool = task_msgs_by_core.get(subject_core, [])
                if not episode_pool:
                    continue

                ack_candidates = []
                for e in episode_pool:
                    e_ist = _email_ist(e)
                    if not e_ist:
                        continue
                    e_minute = e_ist.replace(second=0, microsecond=0)
                    if e_minute <= created_minute or e_minute > resolved_minute:
                        continue
                    if requester and _req_match(e, requester):
                        continue
                    if not _ess_sender(e):
                        continue
                    explicit_ack = _ack_like(e) or _ack_like_text_fallback(e)
                    short_ack = _ess_only_short_ack(e)
                    if not explicit_ack and not short_ack:
                        continue
                    ack_candidates.append((0 if explicit_ack else 1, e_ist, e))

                if not ack_candidates:
                    continue

                ack_candidates.sort(key=lambda item: (item[0], item[1]))
                _ack_rank, ack_pick_ist, ack_pick = ack_candidates[0]
                ack_pick_minute = ack_pick_ist.replace(second=0, microsecond=0)

                if ack_pick_minute >= resolved_minute:
                    continue
                if ack_pick_minute == ack_minute:
                    continue

                t_a = _format_time(ack_pick_ist)
                if not t_a:
                    continue
                if not _allow_guard_rewrite(
                    row_vals,
                    list_index,
                    created_ist,
                    ack_pick_ist,
                    resolved_ist or ack_pick_ist,
                    "TaskEpisodeAckRefine",
                    "task",
                ):
                    continue

                row_vals["Actual Response Date & Time"] = t_a
                ws.cell(row_idx, response_col).value = t_a
                if list_index < len(debug_rows):
                    debug_rows[list_index]["AckSource"] = ack_pick.sender_email or ack_pick.sender_name or "TASK_EPISODE_ACK"
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; TaskEpisodeAckRefine"
                if (ack_pick_minute - created_minute) <= timedelta(minutes=16):
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

            # Task rows should not finish with all three timestamps equal when we
            # can see a later same-episode reply. Recover response/resolved from
            # the first later non-requester message and the latest later thread
            # activity within a narrow window.
            for state in row_states:
                list_index = state.get("list_index")
                row_idx = state.get("row_index")
                if list_index is None or list_index >= len(automation_rows) or not row_idx:
                    continue
                if state.get("is_dep_req") or state.get("is_dep_succ"):
                    continue

                row_vals = automation_rows[list_index]
                requester = state.get("requester") or ""
                subject_norm = (state.get("subject_norm") or "").lower()
                created_dt = _parse_time_str(row_vals.get("Created Date & Time"))
                ack_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
                resolved_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
                created_ist = _to_ist(created_dt) if created_dt else None
                ack_ist = _to_ist(ack_dt) if ack_dt else None
                resolved_ist = _to_ist(resolved_dt) if resolved_dt else None
                if not (created_ist and ack_ist and resolved_ist):
                    continue
                created_minute = created_ist.replace(second=0, microsecond=0)
                ack_minute = ack_ist.replace(second=0, microsecond=0)
                resolved_minute = resolved_ist.replace(second=0, microsecond=0)
                if not (created_minute == ack_minute == resolved_minute):
                    continue

                subject_core = _task_subject_core(subject_norm)
                if not subject_core:
                    continue
                row_id_tokens = _id_like_tokens(subject_norm)
                if not row_id_tokens:
                    row_id_tokens = _id_like_tokens(row_vals.get("Description") or state.get("description") or "")
                if not row_id_tokens:
                    continue

                upper_bound = created_ist + timedelta(hours=48)
                later_msgs = []
                for e in task_msgs_by_core.get(subject_core, []):
                    e_ist = _email_ist(e)
                    if not e_ist or e_ist <= created_ist or e_ist > upper_bound:
                        continue
                    if _system_like_sender(e):
                        continue
                    later_msgs.append(e)

                if not later_msgs and row_id_tokens:
                    for subject_raw, sent_dt, sender_email, sender_name in _raw_task_candidates_for_ids(row_id_tokens):
                        sent_ist = _to_ist(sent_dt)
                        if not sent_ist or sent_ist <= created_ist or sent_ist > upper_bound:
                            continue
                        s_core = _task_subject_core(subject_raw)
                        if not s_core or s_core != subject_core:
                            continue
                        later_msgs.append(SimpleNamespace(
                            sent_time=sent_dt,
                            subject=subject_raw,
                            sender_email=sender_email,
                            sender_name=sender_name,
                            body="",
                            body_html="",
                        ))

                if not later_msgs:
                    continue

                later_msgs.sort(key=lambda e: _email_ist(e))
                response_pick = None
                for e in later_msgs:
                    if requester and _req_match(e, requester):
                        continue
                    response_pick = e
                    break
                if not response_pick:
                    continue

                response_pick_ist = _email_ist(response_pick)
                if not response_pick_ist:
                    continue

                resolved_pool = [
                    e for e in later_msgs
                    if _email_ist(e) and _email_ist(e) >= response_pick_ist
                ]
                resolved_pick = resolved_pool[-1] if resolved_pool else response_pick
                resolved_pick_ist = _email_ist(resolved_pick)
                if not resolved_pick_ist:
                    resolved_pick = response_pick
                    resolved_pick_ist = response_pick_ist

                t_a = _format_time(response_pick_ist)
                t_r = _format_time(resolved_pick_ist)
                if not t_a or not t_r:
                    continue
                if not _allow_guard_rewrite(
                    row_vals,
                    list_index,
                    created_ist,
                    response_pick_ist,
                    resolved_pick_ist,
                    "TaskSameAllRecovered",
                    "task",
                ):
                    continue

                row_vals["Actual Response Date & Time"] = t_a
                row_vals["Actual Resolved Date & Time"] = t_r
                ws.cell(row_idx, response_col).value = t_a
                ws.cell(row_idx, resolved_col).value = t_r
                if list_index < len(debug_rows):
                    debug_rows[list_index]["AckSource"] = response_pick.sender_email or response_pick.sender_name or "TASK_SAME_ALL_RECOVERY"
                    debug_rows[list_index]["ResolvedSource"] = resolved_pick.sender_email or resolved_pick.sender_name or "TASK_SAME_ALL_RECOVERY"
                    debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; TaskSameAllRecovered"
                if (response_pick_ist.replace(second=0, microsecond=0) - created_minute) <= timedelta(minutes=16):
                    _set_row_fill(row_idx, clear_fill)
                    if list_index < len(debug_rows):
                        debug_rows[list_index]["Notes"] = f"{debug_rows[list_index].get('Notes','')}; BlueClearedStrict"

            _stage_timer_stop("task_postprocess_passes", task_post_started_at, items=len(row_states) * 3)

        # Final blue cleanup validator:
        # If a row still has blue fill but its current Created->Response gap is no
        # longer blue-worthy, clear blue at the end so earlier stale markers do not
        # linger after later repairs.
        for state in row_states:
            list_index = state.get("list_index")
            row_idx = state.get("row_index")
            if list_index is None or list_index >= len(automation_rows) or not row_idx:
                continue
            if not _row_has_blue_fill(row_idx):
                continue
            row_vals = automation_rows[list_index]
            c_dt = _parse_time_str(row_vals.get("Created Date & Time"))
            a_dt = _parse_time_str(row_vals.get("Actual Response Date & Time"))
            r_dt = _parse_time_str(row_vals.get("Actual Resolved Date & Time"))
            if not (c_dt and a_dt):
                continue
            c_ist = _to_ist(c_dt)
            a_ist = _to_ist(a_dt)
            r_ist = _to_ist(r_dt) if r_dt else a_ist
            if not (c_ist and a_ist and r_ist):
                continue
            c_min = c_ist.replace(second=0, microsecond=0)
            a_min = a_ist.replace(second=0, microsecond=0)
            r_min = r_ist.replace(second=0, microsecond=0)
            if workbook_kind == "task_business" and c_min == a_min == r_min:
                continue
            if a_min < c_min or r_min < a_min:
                continue
            if (a_min - c_min) > timedelta(minutes=16):
                continue
            _set_row_fill(row_idx, clear_fill)
            if list_index < len(debug_rows):
                notes_now = debug_rows[list_index].get("Notes", "")
                if "BlueClearedFinalValidator" not in notes_now:
                    debug_rows[list_index]["Notes"] = f"{notes_now}; BlueClearedFinalValidator"

        for state in row_states:
            list_index = state.get("list_index")
            if list_index is None or list_index >= len(automation_rows):
                continue
            row_vals = automation_rows[list_index]
            _trace_focus_row(
                "final_row_state",
                state=state,
                row_vals=row_vals,
                list_index=list_index,
            )

    workbook_timer = time.perf_counter()
    logger.log(f"[INFO] Starting worksheet fill: {_workbook_label(workbook_kind)} ({template_path.name})")
    fill_started_at = time.perf_counter()
    fill_result = fill_template(
        template_path=template_path,
        output_path=template_path,
        row_resolver=resolve_row,
        logger=logger,
        post_process=_sequence_ordering_pass,
        sheet_name="LOG",
    )
    fill_elapsed = max(0.0, time.perf_counter() - fill_started_at)

    stem = _template_output_stem(template_path.name)
    write_csv(
        output_dir / f"automation_output_{stem}.csv",
        automation_rows,
        [
            "Description",
            "Requester",
            "Environment",
            "Category Type",
            "Created Date & Time",
            "Actual Response Date & Time",
            "Actual Resolved Date & Time",
            "ServiceRequest/Incident?",
            "ServiceRequest/Incident type?",
            "Interface Code",
        ],
    )
    write_csv(
        output_dir / f"debug_subjects_{stem}.csv",
        debug_rows,
        [
            "Description",
            "Requester",
            "SubjectKey",
            "MatchFound",
            "ThreadSize",
            "CreatedSource",
            "AckSource",
            "ResolvedSource",
            "Notes",
        ],
    )
    write_csv(
        output_dir / "debug_same_times.csv",
        same_time_rows,
        [
            "Description",
            "Requester",
            "SubjectKey",
            "MatchFound",
            "ThreadSize",
            "Created",
            "Response",
            "Resolved",
            "CreatedSource",
            "AckSource",
            "ResolvedSource",
            "Notes",
        ],
    )

    if stage_times_enabled and stage_time_stats:
        logger.log("[INFO] Stage timing summary:")
        for name, stat in sorted(stage_time_stats.items(), key=lambda kv: kv[1]["seconds"], reverse=True):
            avg_ms = (stat["seconds"] * 1000.0 / stat["items"]) if stat["items"] else 0.0
            logger.log(
                f"[INFO]   {name}: {stat['seconds']:.2f}s "
                f"(calls={stat['calls']}, items={stat['items']}, avg={avg_ms:.2f}ms/item)"
            )

    unique_subjects = {
        (row.get("SubjectKey") or "").strip().lower()
        for row in debug_rows
        if (row.get("SubjectKey") or "").strip()
    }
    matched_unique_subjects = {
        (row.get("SubjectKey") or "").strip().lower()
        for row in debug_rows
        if (row.get("SubjectKey") or "").strip() and (row.get("MatchFound") == "Y")
    }
    total_subjects = len(unique_subjects)
    matched_subjects = len(matched_unique_subjects)

    logger.log(f"[INFO] Finished worksheet fill: {_workbook_label(workbook_kind)} in {fill_elapsed:.2f}s")
    logger.log(f"[INFO] Subject match summary: {matched_subjects}/{total_subjects} unique subjects matched.")
    logger.log(
        f"[INFO] Completed. Filled rows: {fill_result.filled_count}, "
        f"Skipped (maintenance): {fill_result.maintenance_count}, "
        f"Marked unknown: {fill_result.unknown_count}"
    )

    return 0


if __name__ == "__main__":
    sys.exit(main())
