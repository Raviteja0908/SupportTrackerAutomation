import os
import re
import shutil
import time
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..output.run_logger import MarkingReason


EXPECTED_HEADERS = [
    "Service No",
    "Environment",
    "Module",
    "Issue Type",
    "Category",
    "Priority",
    "Created Date & Time",
    "Description",
    "Status",
    "Requester",
    "Consultant",
    "Location/ Branch",
    "Actual Response Date & Time",
    "Actual Resolved Date & Time",
    "SLA Start Date & Time",
    "SLA Target End Date & Time (Response Time)",
    "SLA Target End Date & Time (Resolution Time)",
    "SLA Met (Response)? (Auto)",
    "SLA Met (Resolution)? (Auto)",
    "Time Delay (days) (Response)",
    "Time Delay (days) (Resolution)",
    "Created Date (YYMM DD) (Auto)",
    "Target Response Date (YYMM DD) (Auto)",
    "Target Resolved Date (YYMM DD) (Auto)",
    "Actual Resolved Date (YYMM DD) (Auto)",
    "Comments",
    "Category Type",
    "ServiceRequest/Incident?",
    "ServiceRequest/Incident type?",
    "Issue occurred in",
    "Interface Code",
    "Time spent in minutes",
    "Time spent in hours",
    "Time spent in seconds",
]


def _normalize_header(text: str) -> str:
    if text is None:
        return ""
    norm = " ".join(str(text).replace("\n", " ").split()).strip().lower()
    norm = norm.replace("?", "")
    norm = norm.replace("/", " ")
    norm = norm.replace("(", " ").replace(")", " ")
    norm = " ".join(norm.split()).strip()
    aliases = {
        "service request incident": "servicerequest incident",
        "service request incident type": "servicerequest incident type",
    }
    return aliases.get(norm, norm)


@dataclass
class FillResult:
    filled_count: int
    maintenance_count: int
    unknown_count: int


def select_target_sheet(wb, logger=None, preferred_sheet_name="LOG"):
    if preferred_sheet_name and preferred_sheet_name in wb.sheetnames:
        if logger:
            logger.log(f"[INFO] Using worksheet: {preferred_sheet_name}")
        return wb[preferred_sheet_name]

    expected = {_normalize_header(h) for h in EXPECTED_HEADERS}
    best_ws = wb.active
    best_hits = -1
    for ws in wb.worksheets:
        hits = 0
        # Search up to 50 rows instead of just 15
        search_limit = min(50, ws.max_row)
        for row in range(1, search_limit + 1):
            row_values = [
                _normalize_header(ws.cell(row, c).value)
                for c in range(1, ws.max_column + 1)
            ]
            row_hits = sum(1 for v in row_values if v in expected and v)
            if row_hits > hits:
                hits = row_hits
        if hits > best_hits:
            best_hits = hits
            best_ws = ws

    if logger:
        logger.log(f"[INFO] Using worksheet: {best_ws.title}")
    return best_ws


def fill_template(template_path, output_path, row_resolver, logger, post_process=None, sheet_name="LOG"):
    template_path = Path(template_path)
    output_path = Path(output_path)

    wb = load_workbook(template_path)
    try:
        calc = getattr(wb, "calculation", None)
        if calc is not None:
            # openpyxl preserves the formulas themselves, but this template is in
            # manual calc mode. Force Excel to recalculate when the user opens it.
            calc.calcMode = "auto"
            calc.fullCalcOnLoad = True
            if hasattr(calc, "forceFullCalc"):
                calc.forceFullCalc = True
        ws = select_target_sheet(wb, logger, preferred_sheet_name=sheet_name)
        filter_subject = (os.environ.get("FILTER_SUBJECT") or "").strip().lower()
        filter_no_save = os.environ.get("FILTER_NO_SAVE") == "1"
        debug_stage_times = os.environ.get("DEBUG_STAGE_TIMES") == "1"
        resolver_total_seconds = 0.0
        resolver_calls = 0
        slowest_rows = []

        header_row = _find_header_row(ws, logger)
        col_map = _build_col_map(ws, header_row)

        if "description" not in col_map:
            raise RuntimeError("Description column not found in template.")

        comments_col = col_map.get("comments")

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        clear_fill = PatternFill(fill_type=None)

        filled = 0
        maintenance = 0
        unknown = 0

        for row in range(header_row + 1, ws.max_row + 1):
            description = ws.cell(row, col_map["description"]).value
            service_no = ws.cell(row, col_map.get("service no", 1)).value

            if _is_row_empty(description, service_no):
                continue

            row_context = _build_row_context(ws, row, col_map)

            desc_text = str(row_context.get("Description", "") or "")
            if filter_subject and filter_subject not in desc_text.lower():
                continue

            started_at = time.perf_counter() if debug_stage_times else None
            resolved = row_resolver(row_context)
            if debug_stage_times and started_at is not None:
                elapsed = max(0.0, time.perf_counter() - started_at)
                resolver_total_seconds += elapsed
                resolver_calls += 1
                slowest_rows.append(
                    (
                        elapsed,
                        row,
                        str(row_context.get("Service No", "") or ""),
                        desc_text,
                    )
                )
            mark_blue = resolved.pop("_MarkBlue", False)
            required_missing, reason = _is_unknown(resolved)
            if required_missing:
                _mark_row(ws, row, yellow_fill)
                _write_comment(ws, row, comments_col, reason or MarkingReason.unknown)
                unknown += 1
                continue

            if mark_blue:
                _mark_row(ws, row, blue_fill)
                _write_comment(ws, row, comments_col, MarkingReason.blue)
            else:
                _clear_row_if_fill_matches(ws, row, yellow_fill, clear_fill)

            _write_values(ws, row, col_map, resolved)
            filled += 1

        if post_process:
            post_process(ws, col_map, header_row)
            resolved_unknowns = _clear_filled_yellow_rows(ws, col_map, header_row, yellow_fill)
            if resolved_unknowns:
                unknown = max(0, unknown - resolved_unknowns)

        _normalize_datetime_cells(ws, col_map, header_row)
        resolved_unknowns, _blue_applied = _enforce_final_row_fills(
            ws,
            col_map,
            header_row,
            yellow_fill,
            blue_fill,
        )
        if resolved_unknowns:
            unknown = max(0, unknown - resolved_unknowns)

        if filter_no_save:
            logger.log("[INFO] FILTER_NO_SAVE=1; skipping Excel save.")
        else:
            safe_path = _resolve_output_path(output_path)
            saved_path = safe_path
            try:
                wb.save(safe_path)
                logger.log(f"[INFO] Excel saved: {safe_path}")
            except PermissionError:
                alt_path = _resolve_output_path(output_path, force_suffix=True)
                wb.save(alt_path)
                saved_path = alt_path
                logger.log(f"[WARNING] Output locked, saved to: {alt_path}")
            _restore_template_extensions(template_path, saved_path, ws.title, logger)

        if debug_stage_times and resolver_calls:
            slowest_rows.sort(key=lambda item: item[0], reverse=True)
            logger.log(
                f"[INFO] Row resolver timing: {resolver_total_seconds:.2f}s total across {resolver_calls} row(s)"
            )
            for elapsed, row_idx, service_no, description in slowest_rows[:10]:
                logger.log(
                    f"[INFO]   slow-row {elapsed:.2f}s | row={row_idx} | service={service_no or '-'} | desc={description[:120]}"
                )

        return FillResult(filled, maintenance, unknown)
    finally:
        # Ensure workbook is closed to release file handles
        try:
            wb.close()
        except Exception:
            pass


def _resolve_output_path(path: Path, force_suffix: bool = False) -> Path:
    path = Path(path)
    if force_suffix:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return path.with_name(path.stem + "_" + stamp + path.suffix)
    return path


# Namespaces used by worksheet extensions. They are declared directly on the
# copied <ext> element so the restored block is self-contained regardless of
# what the destination worksheet root happens to declare.
_EXT_KNOWN_NS = {
    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "xm": "http://schemas.microsoft.com/office/excel/2006/main",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
}


def _worksheet_path_for_title(zf, title):
    """Map a sheet display name to its xl/worksheets/*.xml path inside the zip."""
    try:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
    except KeyError:
        return None
    m = re.search(
        r'<sheet[^>]*\bname="' + re.escape(title) + r'"[^>]*\br:id="([^"]+)"', wb_xml
    )
    if not m:
        m = re.search(
            r'<sheet[^>]*\br:id="([^"]+)"[^>]*\bname="' + re.escape(title) + r'"', wb_xml
        )
    if not m:
        return None
    rid = m.group(1)
    try:
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    except KeyError:
        return None
    target = None
    for rel in re.findall(r"<Relationship\b[^>]*/?>", rels):
        idm = re.search(r'\bId="([^"]+)"', rel)
        if idm and idm.group(1) == rid:
            tm = re.search(r'\bTarget="([^"]+)"', rel)
            if tm:
                target = tm.group(1)
            break
    if target is None:
        return None
    if target.startswith("/"):
        return target.lstrip("/")
    return "xl/" + target.replace("../", "")


def _extract_worksheet_extlst(sheet_xml):
    """Return the worksheet-level <extLst> (the one right before </worksheet>)."""
    m = re.search(
        r"<extLst>(?:(?!<extLst>).)*?</extLst>\s*</worksheet>",
        sheet_xml,
        flags=re.DOTALL,
    )
    if not m:
        return None
    block = m.group(0)
    return block[: block.rindex("</extLst>") + len("</extLst>")]


def _ext_uris(extlst_xml):
    return set(re.findall(r'<ext\b[^>]*\buri="([^"]+)"', extlst_xml))


def _split_exts(extlst_xml):
    return re.findall(r"<ext\b.*?</ext>", extlst_xml, flags=re.DOTALL)


def _ext_uri(ext_xml):
    m = re.search(r'\buri="([^"]+)"', ext_xml)
    return m.group(1) if m else None


def _ensure_ext_namespaces(ext_xml):
    open_m = re.match(r"<ext\b[^>]*?>", ext_xml, flags=re.DOTALL)
    if not open_m:
        return ext_xml
    open_tag = open_m.group(0)
    used = set(re.findall(r"</?(\w+):", ext_xml))
    additions = "".join(
        f' xmlns:{p}="{_EXT_KNOWN_NS[p]}"'
        for p in used
        if p in _EXT_KNOWN_NS and f"xmlns:{p}=" not in open_tag
    )
    if not additions:
        return ext_xml
    new_open = open_tag[:-1] + additions + ">"
    return ext_xml.replace(open_tag, new_open, 1)


def _preserve_template_extensions(template_path, output_path, sheet_title):
    """Copy worksheet extensions (x14 data validations etc.) from the template
    into the saved output, merging by `ext` uri. Returns True if it changed the
    output file. Raises on unexpected I/O/XML errors (caller handles)."""
    template_path = Path(template_path)
    output_path = Path(output_path)

    with zipfile.ZipFile(template_path) as ztpl:
        tpl_ws = _worksheet_path_for_title(ztpl, sheet_title)
        if not tpl_ws or tpl_ws not in ztpl.namelist():
            return False
        tpl_sheet = ztpl.read(tpl_ws).decode("utf-8")
    tpl_extlst = _extract_worksheet_extlst(tpl_sheet)
    if not tpl_extlst:
        return False  # template has no extensions; nothing to restore

    with zipfile.ZipFile(output_path) as zout:
        out_ws = _worksheet_path_for_title(zout, sheet_title)
        if not out_ws or out_ws not in zout.namelist():
            return False
        names = zout.namelist()
        data = {n: zout.read(n) for n in names}

    out_sheet = data[out_ws].decode("utf-8")
    out_extlst = _extract_worksheet_extlst(out_sheet)
    existing = _ext_uris(out_extlst) if out_extlst else set()

    add = [
        _ensure_ext_namespaces(e)
        for e in _split_exts(tpl_extlst)
        if _ext_uri(e) and _ext_uri(e) not in existing
    ]
    if not add:
        return False

    if out_extlst is None:
        new_extlst = "<extLst>" + "".join(add) + "</extLst>"
        new_sheet = out_sheet.replace("</worksheet>", new_extlst + "</worksheet>", 1)
    else:
        merged = out_extlst[: -len("</extLst>")] + "".join(add) + "</extLst>"
        new_sheet = out_sheet.replace(out_extlst, merged, 1)

    if new_sheet == out_sheet:
        return False

    # Safety: never write a malformed sheet. If parsing fails, leave output as-is.
    try:
        ET.fromstring(new_sheet)
    except ET.ParseError:
        return False

    data[out_ws] = new_sheet.encode("utf-8")
    tmp = output_path.with_name(output_path.name + ".extfix.tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zw:
        for n in names:
            zw.writestr(n, data[n])
    shutil.move(str(tmp), str(output_path))
    return True


def _restore_template_extensions(template_path, output_path, sheet_title, logger):
    """Best-effort re-injection of worksheet extensions that openpyxl drops when
    it re-saves the template (e.g. the x14 data-validation dropdowns that trigger
    the "Data Validation extension is not supported and will be removed" warning).
    Any failure is swallowed so the already-saved output is never disturbed."""
    try:
        if _preserve_template_extensions(template_path, output_path, sheet_title):
            if logger:
                logger.log("[INFO] Restored template data-validation extensions in output.")
    except Exception as exc:  # noqa: BLE001 - never let this break the save
        if logger:
            logger.log(f"[WARNING] Could not restore template extensions: {exc}")


def _find_header_row(ws, logger):
    expected = {_normalize_header(h) for h in EXPECTED_HEADERS}
    best_row = 1
    best_hits = 0

    # Search up to 50 rows instead of just 15
    search_limit = min(50, ws.max_row)
    for row in range(1, search_limit + 1):
        row_values = [
            _normalize_header(ws.cell(row, c).value)
            for c in range(1, ws.max_column + 1)
        ]
        hits = sum(1 for v in row_values if v in expected and v)
        if hits > best_hits:
            best_hits = hits
            best_row = row

    if best_hits < 5:
        logger.log("[WARNING] Header row detection weak; using row 1.")
        return 1

    logger.log(f"[INFO] Header row detected at {best_row} with {best_hits} matches.")
    return best_row


def _build_col_map(ws, header_row):
    col_map = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        if raw is None:
            continue
        key = _normalize_header(raw)
        if key and key not in col_map:
            col_map[key] = col
    return col_map


def _build_row_context(ws, row, col_map):
    context = {}
    for key, col in col_map.items():
        value = ws.cell(row, col).value
        context[_title_case(key)] = value if value is not None else ""
    context["RowIndex"] = row
    return context


def _title_case(key: str) -> str:
    return " ".join(p.capitalize() for p in key.split())


def _is_row_empty(description, service_no):
    if description is None and service_no is None:
        return True
    if str(description).strip() == "" and str(service_no).strip() == "":
        return True
    return False


def _write_values(ws, row, col_map, values):
    for key, value in values.items():
        col = col_map.get(_normalize_header(key))
        if not col:
            continue
        ws.cell(row, col).value = value


def _normalize_datetime_cells(ws, col_map, header_row):
    datetime_keys = [
        "created date & time",
        "actual response date & time",
        "actual resolved date & time",
    ]
    target_cols = [col_map.get(key) for key in datetime_keys if col_map.get(key)]
    if not target_cols:
        return

    for row in range(header_row + 1, ws.max_row + 1):
        for col in target_cols:
            cell = ws.cell(row, col)
            value = cell.value
            parsed = _parse_datetime_cell(value)
            if not parsed:
                continue
            current_format = str(cell.number_format or "").strip()
            cell.value = parsed
            if not current_format or current_format.lower() == "general":
                cell.number_format = "DD-MM-YYYY HH:mm"


def _parse_datetime_cell(value):
    if isinstance(value, datetime):
        return value
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _write_comment(ws, row, comments_col, text):
    if not comments_col or text is None:
        return
    ws.cell(row, comments_col).value = str(text)


def _mark_row(ws, row, fill):
    for col in range(1, ws.max_column + 1):
        ws.cell(row, col).fill = fill


def _row_has_fill(ws, row, target_fill) -> bool:
    for col in range(1, ws.max_column + 1):
        if _fill_matches(ws.cell(row, col).fill, target_fill):
            return True
    return False


def _clear_row_if_fill_matches(ws, row, target_fill, clear_fill):
    if _row_has_fill(ws, row, target_fill):
        _mark_row(ws, row, clear_fill)


def _fill_matches(cell_fill, target_fill) -> bool:
    """Compare two cell fills more robustly, handling None and various format differences."""
    if not cell_fill or not target_fill:
        return False
    try:
        cell_type = str(getattr(cell_fill, "fill_type", "") or "")
        target_type = str(getattr(target_fill, "fill_type", "") or "")
        if cell_type != target_type:
            return False
        cell_color = getattr(cell_fill, "start_color", None)
        target_color = getattr(target_fill, "start_color", None)
        if not cell_color or not target_color:
            return False
        cell_rgb = str(getattr(cell_color, "rgb", "") or "").upper()
        target_rgb = str(getattr(target_color, "rgb", "") or "").upper()
        return bool(cell_rgb and target_rgb and cell_rgb == target_rgb)
    except Exception:
        return False


def _clear_filled_yellow_rows(ws, col_map, header_row, yellow_fill):
    cleared = 0
    clear_fill = PatternFill(fill_type=None)
    for row in range(header_row + 1, ws.max_row + 1):
        if not _row_has_fill(ws, row, yellow_fill):
            continue
        current_values = _build_row_context(ws, row, col_map)
        required_missing, _reason = _is_unknown(current_values)
        if required_missing:
            continue
        _clear_row_if_fill_matches(ws, row, yellow_fill, clear_fill)
        cleared += 1
    return cleared


def _context_value(resolved_values, *keys):
    direct_keys = list(keys)
    normalized_keys = {_normalize_header(key) for key in keys if key}
    for key in direct_keys:
        value = resolved_values.get(key, "")
        if value not in (None, ""):
            return value
    for key, value in resolved_values.items():
        if value in (None, ""):
            continue
        if _normalize_header(key) in normalized_keys:
            return value
    return ""


def _enforce_final_row_fills(ws, col_map, header_row, yellow_fill, blue_fill):
    cleared_yellow = 0
    clear_fill = PatternFill(fill_type=None)
    for row in range(header_row + 1, ws.max_row + 1):
        current_values = _build_row_context(ws, row, col_map)
        required_missing, _reason = _is_unknown(current_values)
        has_yellow = _row_has_fill(ws, row, yellow_fill)
        has_blue = _row_has_fill(ws, row, blue_fill)
        if required_missing:
            if has_blue:
                _clear_row_if_fill_matches(ws, row, blue_fill, clear_fill)
            continue

        if has_yellow:
            _clear_row_if_fill_matches(ws, row, yellow_fill, clear_fill)
            cleared_yellow += 1

        created = _parse_datetime_cell(_context_value(current_values, "Created Date & Time"))
        response = _parse_datetime_cell(_context_value(current_values, "Actual Response Date & Time"))
        if created and response and response >= created and (response - created) > timedelta(minutes=16):
            if not has_blue:
                _mark_row(ws, row, blue_fill)
        elif has_blue:
            _clear_row_if_fill_matches(ws, row, blue_fill, clear_fill)

    return cleared_yellow, True


def _is_unknown(resolved_values):
    created = _context_value(resolved_values, "Created Date & Time")
    resolved = _context_value(resolved_values, "Actual Resolved Date & Time")
    sr = _context_value(
        resolved_values,
        "ServiceRequest/Incident?",
        "ServiceRequest Incident",
    )
    sr_type = _context_value(
        resolved_values,
        "ServiceRequest/Incident type?",
        "ServiceRequest Incident type",
    )

    missing = []
    if not created:
        missing.append("Created Date & Time")
    if not resolved:
        missing.append("Actual Resolved Date & Time")
    if not sr:
        missing.append("ServiceRequest/Incident?")
    if not sr_type:
        missing.append("ServiceRequest/Incident type?")

    if missing:
        return True, "Missing: " + ", ".join(missing)

    return False, ""
