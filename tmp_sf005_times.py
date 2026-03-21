from pathlib import Path
from openpyxl import load_workbook

files = sorted(Path('/app/output').glob('*_filled*.xlsx'), key=lambda p: p.stat().st_mtime)
if not files:
    print('NO_FILLED_XLSX')
    raise SystemExit(1)

path = files[-1]
print('FILE:', path.name)
wb = load_workbook(path, data_only=False)
ws = wb.active
headers = [c.value for c in ws[1]]
for row in ws.iter_rows(min_row=2):
    vals = ['' if c.value is None else str(c.value) for c in row]
    if 'SF005-->RE: SF005 IDoc Failed at ES PROD --> 05-02-2026' not in ' | '.join(vals):
        continue
    print('ROW:', row[0].row)
    for idx, cell in enumerate(row, start=1):
        if isinstance(headers[idx-1], str) and 'date' in headers[idx-1].lower():
            print(f"  {headers[idx-1]} = {cell.value}")
    print('---')
