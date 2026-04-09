from pathlib import Path
from openpyxl import load_workbook
import csv
from datetime import datetime

out = Path('/app/output')
auto = sorted(out.glob('automation_output_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
debug = sorted(out.glob('debug_subjects_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
xlsx = sorted(out.glob('*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
if not auto or not debug or not xlsx:
    raise SystemExit('Missing latest output artifacts')
auto = auto[0]
debug = debug[0]
xlsx = xlsx[0]

def parse_dt(v):
    if not v:
        return None
    for fmt in ('%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M'):
        try:
            return datetime.strptime(str(v).strip(), fmt)
        except Exception:
            pass
    return None

auto_rows = list(csv.DictReader(auto.open('r', encoding='utf-8-sig', newline='')))
debug_rows = list(csv.DictReader(debug.open('r', encoding='utf-8-sig', newline='')))

all_same = 0
bad_order = 0
blue_note = 0
created_retained = 0
seed_primary = 0
for a, d in zip(auto_rows, debug_rows):
    c = parse_dt(a.get('Created Date & Time',''))
    r = parse_dt(a.get('Actual Response Date & Time',''))
    z = parse_dt(a.get('Actual Resolved Date & Time',''))
    notes = (d.get('Notes','') or '').lower()
    if c and r and z:
        if c == r == z:
            all_same += 1
        if not (c <= r <= z):
            bad_order += 1
    if 'blue' in notes:
        blue_note += 1
    if 'created retained (response anchor unreliable)' in notes:
        created_retained += 1
    if 'seedprimaryresolver' in notes or 'lanelocalinitialepisode' in notes:
        seed_primary += 1

wb = load_workbook(xlsx)
ws = wb['LOG'] if 'LOG' in wb.sheetnames else wb.active

def rgb(cell):
    return str(getattr(getattr(cell.fill, 'start_color', None), 'rgb', '') or '').upper()

fill_counts = {'blue': 0, 'yellow': 0, 'red': 0, 'clear': 0, 'other': 0}
for row in range(2, ws.max_row + 1):
    fill = rgb(ws.cell(row, 1))
    if fill == '00BDD7EE':
        fill_counts['blue'] += 1
    elif fill == '00FFF2CC':
        fill_counts['yellow'] += 1
    elif fill == '00FFC7CE':
        fill_counts['red'] += 1
    elif not fill or fill in {'00000000', '000000', 'NONE'}:
        fill_counts['clear'] += 1
    else:
        fill_counts['other'] += 1

print('=== WHOLE RUN SUMMARY ===')
print('auto_csv=', auto.name)
print('debug_csv=', debug.name)
print('workbook=', xlsx.name)
print('rows=', len(auto_rows))
print('all_same_rows=', all_same)
print('bad_order_rows=', bad_order)
print('blue_note_rows=', blue_note)
print('created_retained_rows=', created_retained)
print('seed_primary_tag_rows=', seed_primary)
for k, v in fill_counts.items():
    print(f'fill_{k}={v}')

sus = []
for i, (a, d) in enumerate(zip(auto_rows, debug_rows), start=2):
    c = parse_dt(a.get('Created Date & Time',''))
    r = parse_dt(a.get('Actual Response Date & Time',''))
    z = parse_dt(a.get('Actual Resolved Date & Time',''))
    notes = d.get('Notes','') or ''
    score = 0
    if c and r and z and c == r == z:
        score += 3
    if c and r and z and not (c <= r <= z):
        score += 5
    if 'Created retained (response anchor unreliable)' in notes:
        score += 4
    if 'Blue' in notes:
        score += 2
    if score:
        sus.append((score, i, a.get('Description',''), a.get('Created Date & Time',''), a.get('Actual Response Date & Time',''), a.get('Actual Resolved Date & Time',''), notes[:220]))

sus.sort(reverse=True)
print('=== TOP SUSPICIOUS ROWS ===')
for item in sus[:25]:
    score, row, desc, c, r, z, notes = item
    print('=' * 120)
    print('score=', score, 'row=', row)
    print('desc=', desc)
    print('triplet=', c, '/', r, '/', z)
    print('notes=', notes)
